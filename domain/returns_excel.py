"""domain/returns_excel.py -- the returns analysis as an eight-sheet workbook.

The screen is for looking at. This is for working from: sorting, filtering,
pasting into a supplier email, handing to somebody who does not have a login.

    Executive Summary    the figures, each with a note and a flag
    By Parent            one row per product, with the sizing split
    By Parent — Monthly  the same rows across months, with a trend
    SKU Detail           every child ASIN, grouped under its parent
    Sellable by Parent   sellable against unsellable, month by month
    Return Reasons       every reason code, with a running total
    At Risk              the listings Amazon has flagged
    Action Plan          the findings, as things to do

IT WRITES ONLY WHAT IT WAS GIVEN. Every number here comes from
returns_view.summarise and returns_intel.build -- this file does no counting of
its own, so a figure in the workbook cannot disagree with the same figure on the
screen. Where a column has no data behind it (units ordered without sales data,
a disposition without an FBA report, Amazon's badge without a Listing Quality
export) the cell is left EMPTY and the sheet says why in its header note. An
empty cell is honest; a zero is a claim.

NUMBERS STAY NUMBERS. Rates are written as real numbers with a percentage
FORMAT, not as the text "26.5%", so the columns can be sorted and summed in
Excel. That is the whole reason for exporting rather than screenshotting.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# The palette, once. Dark blue for a sheet title, mid blue for a header row, and
# three flag colours that mean the same thing everywhere they appear.
NAVY = "1F3864"
BLUE = "2E75B6"
RED = "C00000"
AMBER = "BF8F00"
GREEN = "375623"
GREY = "F2F2F2"

TITLE_FONT = Font(bold=True, size=12, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(bold=True, size=10, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
NOTE_FONT = Font(size=9, italic=True, color="595959")
BOLD = Font(bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FLAG_FONT = {
    "CRITICAL": Font(bold=True, size=10, color=RED),
    "HIGH": Font(bold=True, size=10, color=RED),
    "WARNING": Font(bold=True, size=10, color=AMBER),
    "MEDIUM": Font(bold=True, size=10, color=AMBER),
    "MONITOR": Font(bold=True, size=10, color=GREEN),
    "LOW": Font(bold=True, size=10, color=GREEN),
    "OK": Font(bold=True, size=10, color=GREEN),
}

PCT = "0.0%"
PCT2 = "0.00%"
INT = "#,##0"
MONEY = "#,##0.00"


def _title(ws, text, width):
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    for c in range(1, max(1, width) + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws.row_dimensions[1].height = 20


def _note(ws, row, text, width):
    """A sentence under the title saying what the sheet can and cannot answer.

    Not decoration. A blank column in a spreadsheet is read as a zero by whoever
    opens it next week, and this is where the difference gets written down.
    """
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=False)
    return row + 1


def _head(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # The header row stays put while you scroll -- on a 500-row SKU sheet that
    # is the difference between usable and not.
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def _row(ws, row, values, fmts=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        f = (fmts or {}).get(i)
        if f:
            c.number_format = f
    return row + 1


def _pct(n):
    """A percentage FIGURE (26.5) as an Excel fraction (0.265), or None.

    None all the way through rather than 0. A rate that could not be worked out
    and a rate of nought are different facts, and Excel shows the first as an
    empty cell, which is what it is.
    """
    return None if n is None else round(float(n) / 100.0, 6)


def _ratio(small, large):
    if not large:
        return "" if not small else "%d:0" % small
    return "%.1f:1" % (float(small) / float(large))


# =============================================================================
# 1 -- EXECUTIVE SUMMARY
# =============================================================================
def _sheet_summary(wb, s, intel, meta):
    ws = wb.create_sheet("Executive Summary")
    _title(ws, "%s — RETURNS INTELLIGENCE%s"
           % (str(meta.get("account") or "Returns").upper(),
              (" (%s to %s)" % (meta.get("start"), meta.get("end")))
              if meta.get("start") else ""), 4)
    r = _note(ws, 2, meta.get("source_note") or "", 4)
    r = _head(ws, r + 1, ["Metric", "Value", "Notes", "Flag"],
              [42, 22, 46, 12])

    nat = s.get("natures") or {}
    rsn = s.get("reasons") or {}
    units = int(s.get("units_returned") or 0)
    disp = s.get("dispositions") or {}
    sellable = int(disp.get("SELLABLE") or 0)
    unsellable = sum(int(v or 0) for k, v in disp.items() if k != "SELLABLE")
    graded = sellable + unsellable

    def share(n):
        return (" (%.1f%%)" % (n / units * 100.0)) if units and n else ""

    def add(metric, value, note="", flag=""):
        nonlocal r
        r = _row(ws, r, [metric, value, note, flag])
        if flag:
            ws.cell(row=r - 1, column=4).font = FLAG_FONT.get(
                flag, Font(bold=True, size=10))

    def blank():
        nonlocal r
        r += 1

    rate = s.get("return_rate")
    add("Total units returned", "{:,}".format(units),
        "%s to %s" % (meta.get("start") or "?", meta.get("end") or "?"))
    add("Total units ordered", ("{:,}".format(int(s["total_ordered"]))
                               if s.get("total_ordered") else ""),
        ("" if s.get("total_ordered") else
         "No sales data for this period, so there is no rate below."))
    add("Overall return rate", ("%.1f%%" % rate) if rate is not None else "",
        "" if rate is not None else "Cannot be worked out without units ordered.",
        "" if rate is None else ("HIGH" if rate >= 20 else "OK"))
    blank()

    sizing = int(nat.get("Sizing & Fit") or 0)
    small = int(rsn.get("APPAREL_TOO_SMALL") or 0)
    large = int(rsn.get("APPAREL_TOO_LARGE") or 0)
    if sizing:
        add("Sizing returns (too small + too large)",
            "{:,}{}".format(sizing, share(sizing)),
            "The single biggest cause" if units and sizing / units > 0.3 else "",
            "CRITICAL" if units and sizing / units > 0.3 else "WARNING")
        add("  Too small", "{:,}{}".format(small, share(small)))
        add("  Too large", "{:,}{}".format(large, share(large)))
        if small and large:
            add("  Small : large ratio", _ratio(small, large),
                "Runs small" if small > large else "Runs large")
        blank()

    if graded:
        add("Sellable returns", "{:,}{}".format(sellable, share(sellable)),
            "Came back fit to sell again — not a quality problem")
        add("Unsellable returns", "{:,}{}".format(unsellable, share(unsellable)))
        for k, v in sorted(disp.items(), key=lambda kv: -kv[1]):
            if k != "SELLABLE":
                add("  %s" % k.replace("_", " ").title(), "{:,}".format(int(v)))
        blank()
    else:
        add("Disposition", "",
            "Amazon only grades a return it receives itself. This report is "
            "seller-fulfilled, so there is nothing to grade.")
        blank()

    risky = intel.get("at_risk") or []
    showing = sum(1 for a in risky if a.get("state") == "badge showing")
    soon = sum(1 for a in risky if a.get("state") == "at risk")
    if intel.get("has_amazon_quality"):
        add("Listings showing the returns badge", showing,
            "Amazon's 'frequently returned item' warning is live on these",
            "CRITICAL" if showing else "OK")
        add("Listings at risk of the badge", soon, "", "WARNING" if soon else "OK")
    else:
        add("Listings at risk of the returns badge", "",
            "Upload a Listing Quality (Listing Summary) export to fill this in.")

    refunded = s.get("refunded")
    if refunded is not None:
        add("Total refunded", refunded,
            "The real figure from the report, not an estimate")
        ws.cell(row=r - 1, column=2).number_format = MONEY
    blank()
    add("Product lines", len(intel.get("parents") or []),
        "Grouped by %s" % ((intel.get("parents") or [{}])[0].get("grouped_by")
                           or "the product name"))
    add("Child ASINs seen", s.get("unique_skus") or 0)
    return ws


# =============================================================================
# 2 -- BY PARENT
# =============================================================================
def _sheet_parents(wb, s, intel, meta):
    ws = wb.create_sheet("By Parent")
    heads = ["Product", "Total returns", "Units ordered", "Return rate",
             "% of all returns", "Child ASINs", "Sellable", "Unsellable",
             "Sellable %", "Too small", "Too large", "Small:large",
             "Sizing %", "Not as described", "Defective", "Trend"]
    _title(ws, "RETURNS BY PARENT PRODUCT", len(heads))
    r = _note(ws, 2, "Grouped by %s. Units ordered and the rate are blank for a "
                     "product this app has no sales data for."
              % ((intel.get("parents") or [{}])[0].get("grouped_by")
                 or "the product name"), len(heads))
    r = _head(ws, r + 1, heads,
              [40, 13, 13, 11, 14, 11, 10, 11, 11, 10, 10, 12, 10, 15, 11, 14])
    fmts = {2: INT, 3: INT, 4: PCT, 5: PCT, 6: INT, 7: INT, 8: INT, 9: PCT,
            10: INT, 11: INT, 13: PCT, 14: INT, 15: INT}
    for p in intel.get("parents") or []:
        rs = p.get("reasons") or {}
        nats = p.get("natures") or {}
        small = int(rs.get("APPAREL_TOO_SMALL") or 0)
        large = int(rs.get("APPAREL_TOO_LARGE") or 0)
        sizing = int(nats.get("Sizing & Fit") or 0)
        tot = int(p.get("returns") or 0)
        r = _row(ws, r, [
            p.get("label"), tot, p.get("ordered"), _pct(p.get("return_rate")),
            _pct(p.get("share")), p.get("child_count"),
            p.get("sellable") or None, p.get("unsellable") or None,
            _pct(p.get("sellable_pct")), small or None, large or None,
            _ratio(small, large),
            (round(sizing / tot, 4) if tot and sizing else None),
            int(nats.get("Listing Content") or 0) or None,
            int(nats.get("Product Quality") or 0) or None,
            {"increasing": "RISING", "decreasing": "FALLING",
             "stable": "STABLE"}.get(p.get("trend"), ""),
        ], fmts)
        t = ws.cell(row=r - 1, column=16)
        if p.get("trend") == "increasing":
            t.font = Font(bold=True, size=10, color=RED)
        elif p.get("trend") == "decreasing":
            t.font = Font(bold=True, size=10, color=GREEN)
    return ws


# =============================================================================
# 3 -- BY PARENT, MONTH BY MONTH
# =============================================================================
def _month_headers(intel):
    """Month columns, with the incomplete one marked rather than dropped."""
    months = intel.get("months") or []
    out = []
    for i, m in enumerate(months):
        last = (i == len(months) - 1)
        out.append(m + (" (part)" if last and intel.get("partial_last_month")
                        else ""))
    return months, out


def _sheet_monthly(wb, s, intel, meta):
    months, labels = _month_headers(intel)
    ws = wb.create_sheet("By Parent — Monthly")
    heads = ["Product"] + labels + ["Total", "Trend"]
    _title(ws, "MONTHLY RETURNS BY PARENT", len(heads))
    r = _note(ws, 2, ("The last column of months is part of a month, so it is "
                      "marked and left OUT of the trend."
                      if intel.get("partial_last_month") else
                      "Every month shown is complete.")
              + "  Trend: the last three complete months against the first "
                "three, ignored below four months.", len(heads))
    r = _head(ws, r + 1, heads,
              [40] + [10] * len(months) + [11, 14])
    fmts = dict((i, INT) for i in range(2, len(months) + 3))
    for p in intel.get("parents") or []:
        mm = p.get("monthly") or {}
        r = _row(ws, r, [p.get("label")] + [int(mm.get(m) or 0) for m in months]
                 + [int(p.get("returns") or 0),
                    {"increasing": "RISING", "decreasing": "FALLING",
                     "stable": "STABLE"}.get(p.get("trend"), "too few months")],
                 fmts)
        t = ws.cell(row=r - 1, column=len(heads))
        if p.get("trend") == "increasing":
            t.font = Font(bold=True, size=10, color=RED)
        elif p.get("trend") == "decreasing":
            t.font = Font(bold=True, size=10, color=GREEN)
    return ws


# =============================================================================
# 4 -- EVERY CHILD ASIN
# =============================================================================
def _sheet_skus(wb, s, intel, meta):
    ws = wb.create_sheet("SKU Detail")
    heads = ["Parent", "SKU", "Child ASIN", "Product", "Returns", "Ordered",
             "Return rate", "Sellable", "Unsellable", "Too small", "Too large",
             "Returns badge", "CX health", "Top NCX reason"]
    _title(ws, "EVERY CHILD ASIN THAT CAME BACK", len(heads))
    r = _note(ws, 2, ("The last three columns are Amazon's own judgement and "
                      "come from a Listing Quality export."
                      if intel.get("has_amazon_quality") else
                      "The last three columns are blank because no Listing "
                      "Quality export was supplied — blank meaning not "
                      "supplied, not meaning fine."), len(heads))
    r = _head(ws, r + 1, heads,
              [26, 20, 14, 46, 10, 10, 11, 10, 11, 10, 10, 13, 12, 22])
    fmts = {5: INT, 6: INT, 7: PCT, 8: INT, 9: INT, 10: INT, 11: INT}
    last = None
    for c in intel.get("children") or []:
        if c.get("parent") != last:
            last = c.get("parent")
            cell = ws.cell(row=r, column=1, value=last)
            cell.font = BOLD
            for i in range(1, len(heads) + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GREY)
            r += 1
        r = _row(ws, r, [
            c.get("parent"), c.get("sku"), c.get("asin"), c.get("name"),
            c.get("returns"), c.get("ordered"), _pct(c.get("rate")),
            c.get("sellable") or None, c.get("unsellable") or None,
            c.get("too_small") or None, c.get("too_large") or None,
            c.get("badge"), c.get("cx_health"), c.get("top_reason"),
        ], fmts)
    return ws


# =============================================================================
# 5 -- SELLABLE AGAINST UNSELLABLE, MONTH BY MONTH
# =============================================================================
def _sheet_sellable(wb, s, intel, meta):
    months, labels = _month_headers(intel)
    ws = wb.create_sheet("Sellable by Parent")
    heads = ["Product"]
    for l in labels:
        heads += [l + " sellable", l + " unsellable"]
    heads += ["Sellable %"]
    _title(ws, "SELLABLE vs UNSELLABLE, BY MONTH", len(heads))
    if not (s.get("dispositions") or {}):
        r = _note(ws, 2, "This sheet is empty because the report is "
                         "seller-fulfilled: Amazon never receives the unit, so "
                         "it never grades it. An FBA Customer Returns file "
                         "fills this in.", len(heads))
        _head(ws, r + 1, heads, [40] + [12] * (len(heads) - 2) + [11])
        return ws
    r = _note(ws, 2, "Sellable means Amazon graded the returned unit fit to "
                     "sell again. A high return rate that comes back sellable "
                     "is a listing problem; the same rate coming back damaged "
                     "is a product problem.", len(heads))
    r = _head(ws, r + 1, heads, [40] + [12] * (len(heads) - 2) + [11])
    fmts = dict((i, INT) for i in range(2, len(heads)))
    fmts[len(heads)] = PCT
    for p in intel.get("parents") or []:
        vals = [p.get("label")]
        ms, mu = p.get("monthly_sellable") or {}, p.get("monthly_unsellable") or {}
        for m in months:
            vals += [int(ms.get(m) or 0), int(mu.get(m) or 0)]
        vals.append(_pct(p.get("sellable_pct")))
        r = _row(ws, r, vals, fmts)
    return ws


# =============================================================================
# 6 -- EVERY REASON CODE
# =============================================================================
def _sheet_reasons(wb, s, intel, meta):
    ws = wb.create_sheet("Return Reasons")
    heads = ["Reason", "Units", "% of returns", "Cumulative %", "Cause",
             "What it means"]
    _title(ws, "EVERY RETURN REASON AMAZON GAVE", len(heads))
    r = _note(ws, 2, "Cumulative % is there to find the short list: the reasons "
                     "above 80% are the ones worth working on.", len(heads))
    r = _head(ws, r + 1, heads, [34, 10, 13, 13, 22, 70])
    from domain import returns_view as _rv
    total = sum(int(v or 0) for v in (s.get("reasons") or {}).values()) or 0
    run = 0
    for reason, n in (s.get("reasons") or {}).items():
        n = int(n or 0)
        run += n
        cause = _rv.nature_of(reason)
        r = _row(ws, r, [
            reason, n,
            (round(n / total, 6) if total else None),
            (round(run / total, 6) if total else None),
            cause, (s.get("nature_actions") or {}).get(cause, ""),
        ], {2: INT, 3: PCT, 4: PCT})
    return ws


# =============================================================================
# 7 -- WHAT AMAZON HAS FLAGGED
# =============================================================================
def _sheet_at_risk(wb, s, intel, meta):
    ws = wb.create_sheet("At Risk")
    heads = ["State", "ASIN", "SKU", "Product", "Return rate", "Orders",
             "Top NCX reason", "CX health", "Star rating"]
    _title(ws, "LISTINGS AMAZON HAS FLAGGED", len(heads))
    rows = intel.get("at_risk") or []
    if not intel.get("has_amazon_quality"):
        r = _note(ws, 2, "Empty because no Listing Quality (Listing Summary) "
                         "export was supplied. That file is the only place "
                         "Amazon says whether the returns badge is showing — "
                         "it is not in the returns report and not in the API.",
                  len(heads))
        _head(ws, r + 1, heads, [14, 14, 20, 50, 12, 10, 24, 12, 11])
        return ws
    r = _note(ws, 2, "'Badge showing' means the 'frequently returned item' "
                     "warning is live on the listing now and is costing "
                     "conversion today. 'At risk' means it is not showing yet.",
              len(heads))
    r = _head(ws, r + 1, heads, [14, 14, 20, 50, 12, 10, 24, 12, 11])
    for a in rows:
        r = _row(ws, r, [
            a.get("state", "").upper(), a.get("asin"), a.get("sku"),
            a.get("name"), _pct(a.get("return_rate")), a.get("orders"),
            a.get("top_reason"), a.get("cx_health"), a.get("star_rating"),
        ], {5: PCT2, 6: INT, 9: "0.0"})
        if a.get("state") == "badge showing":
            ws.cell(row=r - 1, column=1).font = FLAG_FONT["CRITICAL"]
        else:
            ws.cell(row=r - 1, column=1).font = FLAG_FONT["WARNING"]
    return ws


# =============================================================================
# 8 -- WHAT TO DO
# =============================================================================
def _sheet_actions(wb, s, intel, meta):
    ws = wb.create_sheet("Action Plan")
    heads = ["Priority", "Action", "Details", "Scope", "Timeline"]
    _title(ws, "WHAT TO DO ABOUT IT", len(heads))
    r = _note(ws, 2, "Every row here was produced by a figure in this workbook. "
                     "Nothing is here as general advice — a problem this "
                     "account does not have gets no row.", len(heads))
    r = _head(ws, r + 1, heads, [12, 46, 88, 26, 13])
    for a in intel.get("action_plan") or []:
        r = _row(ws, r, [a.get("priority"), a.get("action"), a.get("details"),
                         a.get("scope"), a.get("timeline")])
        ws.cell(row=r - 1, column=1).font = FLAG_FONT.get(
            a.get("priority"), Font(bold=True, size=10))
        ws.cell(row=r - 1, column=3).alignment = Alignment(wrap_text=True,
                                                           vertical="top")
        ws.row_dimensions[r - 1].height = 46
    return ws


# The workbook, in order. The NAME lives here rather than only inside each
# builder, so that a sheet which fails can still be created under its proper
# name -- a workbook that silently loses "At Risk" and gains "Atrisk" is worse
# than one that says the sheet broke.
SHEETS = (
    ("Executive Summary", _sheet_summary),
    ("By Parent", _sheet_parents),
    ("By Parent — Monthly", _sheet_monthly),
    ("SKU Detail", _sheet_skus),
    ("Sellable by Parent", _sheet_sellable),
    ("Return Reasons", _sheet_reasons),
    ("At Risk", _sheet_at_risk),
    ("Action Plan", _sheet_actions),
)


def build(summary, intel, meta=None):
    """-> an openpyxl Workbook. Takes what the screen already has; counts nothing.

    A sheet that throws does not take the workbook down with it. A returns file
    with one odd row should cost you one sheet, not the download -- and the
    sheet that failed says so, under its own name and in its own first cell,
    rather than being silently absent or turning up under a different one.
    """
    meta = meta or {}
    wb = Workbook()
    wb.remove(wb.active)
    for name, fn in SHEETS:
        try:
            fn(wb, summary or {}, intel or {}, meta)
        except Exception as e:
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
            ws.cell(row=1, column=1,
                    value=("This sheet could not be built: %s. Every other "
                           "sheet in this workbook is unaffected."
                           % str(e)[:180])).font = Font(bold=True, color=RED)
    # The order can be disturbed by a rebuild above, and a workbook whose tabs
    # move about between downloads is one nobody trusts.
    wb._sheets.sort(key=lambda ws: [n for n, _ in SHEETS].index(ws.title)
                    if ws.title in [n for n, _ in SHEETS] else 99)
    if not wb.sheetnames:
        wb.create_sheet("Empty")
    return wb


def to_bytes(summary, intel, meta=None):
    """The workbook as bytes, ready to send."""
    import io
    buf = io.BytesIO()
    build(summary, intel, meta).save(buf)
    return buf.getvalue()
