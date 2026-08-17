"""domain/source_bulk.py -- attach suppliers to many SKUs from one uploaded sheet.

"the repricer tool give me an option to upload a sheet containing the sku's or
 original asins of the item, to add their suppliers through a sheet upload, i
 will write, i can use whether the asin whether the sku to tell the app that
 this supplier or source belongs to this asin or sku"

So a row identifies its listing by EITHER a SKU or an ASIN, and carries the
supplier link. Everything else about the row is optional.

WHY EITHER KEY, AND WHY THAT IS NOT AMBIGUOUS
A SKU names exactly one listing. An ASIN can name several -- the same product
relisted, or one ASIN carrying two of your SKUs -- so an ASIN row attaches the
supplier to every SKU that sits on it, and says how many that was. Silently
picking one would attach a real supplier to a listing nobody meant, and the
repricer would then price a listing from a cost that is not its own.

CLAUDE.md Rule 1 applies to the column: the ASIN in a generated SKU is the
COMPETITOR's, used to identify which of OUR listings was built from it. It is
never treated as our own ASIN, and nothing here writes it to Amazon.

WHAT IT REFUSES, AND WHY IT SAYS SO PER ROW
A bulk import that reports only a total is how twelve silently-skipped rows
become "the repricer is not working". Every row comes back with what happened
to it: attached, already had one, no such listing, or a link nothing can read.
The link itself is judged by domain/source_link.classify -- the same rule that
refuses an Amazon page as a supplier, so a sheet cannot introduce something the
automatic path would reject.

Nothing here enrols anything into LIVE pricing. Tracking is not pricing (see
the Repricer screen): a SKU added by sheet is in dry run like any other.
"""
import csv
import io
import re

from domain import source_link as _link
from domain import source_repo as _repo

# What a column might be called. Matched loosely -- a person's sheet says
# "Seller SKU", "sku", "SKU " or "Item SKU", and refusing all but one spelling
# is a way of making a working file look broken.
_SKU_COLS = ("sku", "seller sku", "sellersku", "seller-sku", "item sku",
             "merchant sku", "our sku")
_ASIN_COLS = ("asin", "competitor asin", "original asin", "amazon asin",
              "parent asin", "competitor_asin")
_URL_COLS = ("url", "link", "source", "source url", "source link", "supplier",
             "supplier url", "supplier link", "ebay", "ebay url", "ebay link")

_ASIN_RE = re.compile(r"\b(B0[A-Z0-9]{8})\b", re.I)


def _norm(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").strip().lower()).strip()


def _pick(headers, wanted):
    """The index of the first column whose name looks like one of `wanted`."""
    norm = [_norm(h) for h in headers]
    for w in wanted:
        if w in norm:
            return norm.index(w)
    # A looser pass, so "eBay Source Link" still finds the url column.
    for i, h in enumerate(norm):
        for w in wanted:
            if w and w in h:
                return i
    return -1


def read_table(data, filename=""):
    """(headers, rows, error) from an uploaded CSV or spreadsheet. Never raises.

    `error` is "" when the file was read. It is the third value because a file
    that cannot be read and a file with nothing in it are different answers, and
    returning ([], []) for both leaves the caller with nothing to tell anyone.
    """
    name = str(filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[("" if c is None else str(c)) for c in r]
                    for r in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            return [], [], ("that spreadsheet could not be read (%s). Save it as "
                            "CSV and try again." % str(e)[:80])
    else:
        try:
            txt = data.decode("utf-8-sig", "replace")
        except Exception:
            txt = str(data)
        # PLAIN COMMA FIRST, AND ONLY SNIFF IF THAT CLEARLY FAILS.
        #
        # This sniffed first, and the sniffer gets a real file wrong. MEASURED
        # on the cost sheet this app generates: one product is called
        #
        #     46-Piece 1/4" Drive Socket Ratchet Wrench Set with Storage Case,
        #     Metric Socket & Bit Set, ...
        #
        # -- a quoted field containing both a comma AND a double quote, which
        # csv.writer correctly escapes by doubling. The sniffer sees the doubled
        # quotes, picks a different quoting rule, and the field splits on its
        # internal commas. Every column after it shifts, and "Metric Socket &
        # Bit Set" arrives where the cost should be. It was caught only because
        # a cost that is not a number is refused; had the fragment been numeric
        # it would have been stored as somebody's cost price.
        #
        # A comma-separated file always yields more than one column under the
        # plain excel dialect. Sniffing is for the genuinely foreign file --
        # semicolons from a European Excel, or tabs -- so it is now the FALLBACK
        # rather than the first guess.
        rows = [r for r in csv.reader(io.StringIO(txt), csv.excel)]
        if not rows or max((len(r) for r in rows[:5]), default=0) < 2:
            try:
                dialect = csv.Sniffer().sniff(txt[:4000], delimiters=",;\t")
                rows = [r for r in csv.reader(io.StringIO(txt), dialect)]
            except Exception:
                pass
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return [], [], "that file has no rows in it"
    return rows[0], rows[1:], ""


def _sku_exists(config_path, workspace_id, marketplace, sku):
    """Does this account actually have that SKU -- as a listing, on Amazon, or
    already being tracked?

    Three places because a SKU can legitimately be in any of them: a draft this
    app made, something Amazon returned that has no draft here, or one already
    enrolled in the repricer (re-uploading a sheet must not orphan those).
    """
    from data import db as _db
    s = str(sku or "").strip()
    if not s:
        return False
    try:
        conn = _db.get_db(config_path)
        r = conn.execute("SELECT 1 FROM listings WHERE workspace_id=? AND "
                         "UPPER(sku)=UPPER(?) LIMIT 1", (workspace_id, s)).fetchone()
        if r:
            return True
    except Exception:
        pass
    try:
        from domain import live_snapshots as _ls
        rec = _ls.get(config_path, workspace_id, marketplace) or {}
        want = s.upper()
        for it in (rec.get("items") or []):
            if str(it.get("sku") or "").strip().upper() == want:
                return True
    except Exception:
        pass
    try:
        for row in _repo.enrolled(config_path, workspace_id, marketplace):
            if str(row.get("sku") or "").strip().upper() == s.upper():
                return True
    except Exception:
        pass
    return False


def skus_for_key(config_path, workspace_id, marketplace, sku="", asin=""):
    """Which of OUR SKUs a row is talking about.

    A SKU is taken as given. An ASIN is resolved to every SKU that sits on it --
    from the live catalogue Amazon returned, and from the competitor ASIN the
    generator wrote into the SKU itself.
    """
    from data import db as _db
    sku = str(sku or "").strip()
    if sku:
        # A TYPED SKU IS STILL CHECKED. Taking it on trust meant a typo -- or a
        # SKU from another account's sheet -- created an enrolment for a listing
        # that does not exist, which then sat in the repricer for ever reporting
        # that it could not be read. Caught by the upload test: a row keyed
        # "NO-SUCH-SKU-123" was attached without complaint.
        return [sku] if _sku_exists(config_path, workspace_id, marketplace, sku) else []
    a = str(asin or "").strip().upper()
    if not a:
        return []
    found, seen = [], set()

    # 1. Our own listings, by the competitor ASIN they were built from.
    try:
        conn = _db.get_db(config_path)
        for r in conn.execute(
                "SELECT sku FROM listings WHERE workspace_id=? AND "
                "UPPER(IFNULL(competitor_asin,''))=? AND IFNULL(sku,'')<>''",
                (workspace_id, a)):
            s = str(r["sku"]).strip()
            if s and s.upper() not in seen:
                seen.add(s.upper())
                found.append(s)
    except Exception:
        pass

    # 2. The live catalogue, where the ASIN is OURS on Amazon.
    try:
        from domain import live_snapshots as _ls
        rec = _ls.get(config_path, workspace_id, marketplace) or {}
        for it in (rec.get("items") or []):
            if str(it.get("asin") or "").strip().upper() != a:
                continue
            s = str(it.get("sku") or "").strip()
            if s and s.upper() not in seen:
                seen.add(s.upper())
                found.append(s)
    except Exception:
        pass
    return found


# The template's own column names, each an EXACT match for one of the lists
# above so the file that comes back is read without guesswork.
#
# ONE link column, not a "current" and a "new" pair. That pair was the obvious
# design and it is a trap: _pick falls back to a substring match, "current
# supplier link" contains "supplier link", and it appears first -- so the
# upload would have re-applied the old links and silently ignored every new one
# the person had just typed. One column, arriving pre-filled with whatever is
# attached now, is both unambiguous to the parser and the thing a person
# actually wants: you can see what is there and change it.
TEMPLATE_HEADERS = ["sku", "asin", "product", "supplier link"]


def template_rows(config_path, workspace_id, marketplace, enrolled,
                  catalogue=None, current=None):
    """The sheet to hand someone, already filled in with what we know.

    "when a user wants to update the supplier links through the sheet, give the
     user the template first filled by the asins enrolled for tracking in the
     repricer, the user will fill that template and upload it back"

    So the only empty column is the one they are there to fill. Handing over a
    blank sheet means typing forty SKUs by hand, and a SKU typed by hand is the
    NO-SUCH-SKU-123 that this module already has a check for -- the fix for
    which is not to make people type them at all.

    `enrolled` is [sku, ...]. `catalogue` is the shared lookup from
    domain/catalogue.py, so the product name shown here is the same one the rest
    of the app shows. `current` is {sku: url} for links already attached; those
    arrive filled in, so someone changing one supplier can see the other rows
    are done and leave them alone.

    A row whose link comes back unchanged is recognised as already attached and
    nothing happens to it, so uploading an untouched template is a no-op.
    """
    from domain import catalogue as _cat
    idx = catalogue if catalogue is not None else {}
    cur = current or {}
    out = []
    for sku in enrolled:
        s = str(sku or "").strip()
        if not s:
            continue
        got = _cat.look(idx, s)
        out.append([s,
                    # OUR ASIN if the catalogue knows it. Falling back to the
                    # one inside the SKU is falling back to the COMPETITOR's
                    # (Rule 1) -- which is still the right thing in a column
                    # headed "asin" here, because that is what the upload
                    # matches on and how the app already finds a listing from a
                    # pasted ASIN. Same reader source_link uses, not a second
                    # copy of the pattern (Rule 12).
                    got.get("asin") or _link._asin_from_sku(s),
                    got.get("title") or "",
                    str(cur.get(s) or "")])
    # ROWS THAT STILL NEED A LINK FIRST. The sheet is opened to do a job, and
    # the blank ones are that job -- putting them at the bottom of forty
    # already-filled rows is how they get missed.
    out.sort(key=lambda r: (bool(r[3]), (r[2] or "").lower(), r[0]))
    return out


def to_csv(headers, rows):
    """A CSV string ready for Excel. See domain/sheets.to_csv for the rules.

    Kept as a name here because this module's callers already use it, but the
    writing itself moved out the moment a SECOND template needed it -- costs, as
    well as supplier links. Two writers would have been two opinions about
    quoting and the byte-order mark (Rule 12).
    """
    from domain import sheets as _sheets
    return _sheets.to_csv(headers, rows)


def apply_rows(config_path, workspace_id, marketplace, headers, rows):
    """Attach every row's supplier. Returns a report, per row.

    Never raises: one malformed line must not lose the rest of the file.
    """
    i_sku = _pick(headers, _SKU_COLS)
    i_asin = _pick(headers, _ASIN_COLS)
    i_url = _pick(headers, _URL_COLS)

    out = {"ok": True, "attached": 0, "already": 0, "skipped": 0,
           "tracked": 0, "rows": [],
           "columns": {"sku": (headers[i_sku] if i_sku >= 0 else ""),
                       "asin": (headers[i_asin] if i_asin >= 0 else ""),
                       "url": (headers[i_url] if i_url >= 0 else "")}}
    if i_url < 0:
        out["ok"] = False
        out["error"] = ("no supplier link column found. Name one of the columns "
                        "'url', 'link', 'source' or 'supplier'.")
        return out
    if i_sku < 0 and i_asin < 0:
        out["ok"] = False
        out["error"] = ("no SKU or ASIN column found. Name a column 'sku' or "
                        "'asin' so each link can be matched to a listing.")
        return out

    def cell(r, i):
        return str(r[i]).strip() if (0 <= i < len(r)) else ""

    for n, r in enumerate(rows, start=2):        # row 1 is the header
        sku = cell(r, i_sku)
        asin = cell(r, i_asin)
        url = cell(r, i_url)
        rep = {"line": n, "sku": sku, "asin": asin, "url": url,
               "status": "", "note": ""}
        if not url:
            # An ASIN typed into the url column is a common slip; say so rather
            # than "no link".
            rep["status"] = "skipped"
            rep["note"] = "no supplier link on this row"
            out["skipped"] += 1
            out["rows"].append(rep)
            continue
        # An ASIN can be pasted anywhere in an Amazon URL; pull it out.
        if not asin and not sku:
            m = _ASIN_RE.search(url)
            if m:
                asin = m.group(1)
        kind, why = _link.classify(url)
        if not kind:
            rep["status"] = "skipped"
            rep["note"] = why
            out["skipped"] += 1
            out["rows"].append(rep)
            continue

        targets = skus_for_key(config_path, workspace_id, marketplace,
                               sku=sku, asin=asin)
        if not targets:
            rep["status"] = "skipped"
            rep["note"] = ("no listing in this account matches %s"
                           % (("SKU " + sku) if sku else ("ASIN " + asin)))
            out["skipped"] += 1
            out["rows"].append(rep)
            continue

        done = []
        for s in targets:
            try:
                _repo.enrol(config_path, workspace_id, marketplace, s, mode="dry_run")
                out["tracked"] += 1
                _sid, created = _repo.ensure_source(
                    config_path, workspace_id, marketplace, s, url,
                    kind=kind, label=url)
                if created:
                    out["attached"] += 1
                    done.append(s + " ✓")
                else:
                    out["already"] += 1
                    done.append(s + " (already had it)")
            except Exception as e:
                out["skipped"] += 1
                done.append("%s failed: %s" % (s, str(e)[:60]))
        rep["status"] = "attached"
        rep["matched"] = targets
        # An ASIN carrying more than one SKU is worth saying out loud: the same
        # supplier now prices both, and that is only right if it really is the
        # same product.
        rep["note"] = ("; ".join(done)
                       + ("  — this ASIN carries %d of your SKUs" % len(targets)
                          if len(targets) > 1 else ""))
        out["rows"].append(rep)
    return out
