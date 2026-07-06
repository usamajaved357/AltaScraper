"""
template_schema.py  --  Per-product-type attribute schema, built from LOCAL sources.

Replaces the blocked SP-API getDefinitionsProductType. Three working sources are
combined (AttributePTDMAP is corrupted in the unified template and is NOT used):

  1. Data Definitions sheet   -> requirement level + label + guidance per field
  2. valid_values.json[PT]    -> per-product-type ENUMERATED attributes + exact
                                 accepted strings (case-sensitive)
  3. UNIVERSAL_PHYSICAL set    -> free-text attributes virtually every physical
                                 product needs (count, components, dimensions, ...)

Per-product relevance is decided downstream by Claude (with the product images).
This module hands Claude a focused, per-type candidate list.
"""

import json
import openpyxl


def _field_key(field_id) -> str:
    if field_id is None:
        return ""
    return str(field_id).split("[")[0].split("#")[0].replace("::", "").strip().lower()


UNIVERSAL_PHYSICAL = {
    "number_of_items":      "Total identical items sold as one unit (e.g. 15 for a 15-piece set)",
    "included_components":  "What is in the box (read the product image)",
    "unit_count":           "Numeric count for products measured by count/volume/weight",
    "unit_count_type":      "The unit the count is expressed in (Count, Ounce, ...)",
    "number_of_boxes":      "How many physical boxes the item ships in",
    "country_of_origin":    "Country the product is made in",
    "special_features":     "Distinguishing features of the product",
    "item_shape":           "Physical shape of the item, if applicable",
    "is_fragile":           "Whether the item is fragile (Yes/No)",
}

_SKIP = (
    "contribution_sku", "product_type", "record_action", "item_name", "brand",
    "product_id", "recommended_browse_nodes", "parentage_level",
    "child_parent_sku_relationship", "variation_theme", "fulfillment_availability",
    "merchant_shipping_group", "purchasable_offer", "list_price", "product_tax_code",
    "bullet_point", "product_description", "generic_keyword", "condition_type",
    "package_contains_sku", "supplier_declared", "externally_assigned",
    "main_image", "other_image", "swatch_image", "map_policy", "offering_can_be",
    "merchant_suggested_asin", "item_type_keyword", "model_name", "item_sku",
    "parent_sku", "manufacturer", "model_number",
)

_NA_TOKENS = {"does not apply", "not applicable", "n/a", "na", "none"}


class TemplateSchema:
    def __init__(self, template_path: str, valid_values_path: str):
        self._defs = self._load_data_definitions(template_path)
        self._vv = self._load_valid_values(valid_values_path)

    @staticmethod
    def _load_data_definitions(template_path: str) -> dict:
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb["Data Definitions"]
        cols, hdr_row = {}, None
        for r in range(1, 6):
            rowvals = [str(ws.cell(row=r, column=c).value or "").strip().lower()
                       for c in range(1, ws.max_column + 1)]
            if "required?" in rowvals:
                hdr_row = r
                cols = {rowvals[c - 1]: c for c in range(1, ws.max_column + 1)}
                break
        out = {}
        if hdr_row:
            cf, cl = cols.get("field name"), cols.get("local label name")
            ca, cr = cols.get("accepted values"), cols.get("required?")
            for r in range(hdr_row + 1, ws.max_row + 1):
                k = _field_key(ws.cell(row=r, column=cf).value) if cf else ""
                if not k or k in out:
                    continue
                out[k] = {
                    "label":        str(ws.cell(row=r, column=cl).value or k).strip() if cl else k,
                    "requirement":  str(ws.cell(row=r, column=cr).value or "Optional").strip() if cr else "Optional",
                    "accepted_desc": str(ws.cell(row=r, column=ca).value or "").strip()[:160] if ca else "",
                }
        wb.close()
        return out

    @staticmethod
    def _load_valid_values(path: str) -> dict:
        try:
            raw = json.load(open(path, encoding="utf-8"))
        except Exception:
            return {}
        out = {}
        for pt, attrs in raw.items():
            if pt.startswith("_") or not isinstance(attrs, dict):
                continue
            out[pt] = {_field_key(a): v for a, v in attrs.items() if isinstance(v, list)}
        return out

    def _spec(self, key, pt_enums, fallback_desc=""):
        meta = self._defs.get(key, {})
        vals = pt_enums.get(key, [])
        is_enum = bool(vals)
        has_na = any(any(t in str(v).lower() for t in _NA_TOKENS) for v in vals)
        return {
            "key": key,
            "label": meta.get("label", key.replace("_", " ").title()),
            "requirement": meta.get("requirement", "Optional"),
            "is_enum": is_enum,
            "valid_values": vals,
            "has_not_applicable": has_na,
            "accepted_desc": meta.get("accepted_desc") or fallback_desc,
        }

    def attributes_for(self, product_type: str) -> list:
        pt_enums = self._vv.get(product_type, {})
        keys, specs = set(), []
        for k in pt_enums:
            if k in _SKIP or k.startswith(("battery_", "lithium_", "epr_",
                                           "item_package_weight", "item_volume_unit",
                                           "item_weight_unit", "package_")):
                continue
            if k in keys:
                continue
            keys.add(k); specs.append(self._spec(k, pt_enums))
        for k, desc in UNIVERSAL_PHYSICAL.items():
            if k in keys or k in _SKIP:
                continue
            keys.add(k); specs.append(self._spec(k, pt_enums, fallback_desc=desc))
        rank = {"required": 0, "conditionally required": 1, "recommended": 2, "optional": 3}
        specs.sort(key=lambda s: rank.get(s["requirement"].lower(), 3))
        return specs


if __name__ == "__main__":
    import sys
    tpl = sys.argv[1] if len(sys.argv) > 1 else "KITCHEN_COOKWARE_SET_FOOD_SPATULA_KITCHEN_KNIFE_HARDWARE__2_.xlsm"
    vvp = sys.argv[2] if len(sys.argv) > 2 else "valid_values.json"
    sc = TemplateSchema(tpl, vvp)
    for pt in ["KITCHEN", "COOKWARE_SET"]:
        attrs = sc.attributes_for(pt)
        print(f"\n{'='*62}\n{pt}: {len(attrs)} candidate attributes for Claude to fill")
        for s in attrs:
            kind = ("ENUM(" + str(len(s["valid_values"])) + ")") if s["is_enum"] else "free-text"
            na = " +N/A" if s["has_not_applicable"] else ""
            ex = ("  e.g. " + " | ".join(map(str, s["valid_values"][:3]))) if s["is_enum"] else ""
            print(f"   [{s['requirement'][:11]:11}] {s['label']:30} {kind:9}{na}{ex}")
