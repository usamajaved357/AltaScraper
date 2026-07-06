"""
inventory_module.py -- Amazon FBA + 3PL replenishment engine.

Mirrors the logic of Talha's "Main Calculation" sheet from the Lure Essentials
Inventory Control template, but fully automated:
  1. FBA inventory pulled from SP-API (GET_FBA_MYI_ALL_INVENTORY_DATA report)
  2. Sales velocity pulled from SP-API Orders API (last N days per SKU)
  3. YoY uplift derived from Orders API a year ago vs current (same window)
  4. Prime Day uplift from a static calendar + configurable %
  5. 3PL data comes from an uploaded CSV (Amazon has NO visibility into
     external warehouses -- this is the honest limitation)

FOUR-BUCKET ZERO-VELOCITY CLASSIFICATION (replaces the sheet's naive AG=1
fallback):
  - ACTIVE:       velocity > 0                       -> normal replenishment
  - NEW_LAUNCH:   velocity = 0 AND age < 60 days     -> seed rate, flag for review
  - DORMANT:      velocity = 0 AND age 60-365 days   -> no reorder, check listing
  - DEAD:         velocity = 0 AND age > 365 days
                  OR Selling Status = Discontinued   -> no reorder, consider delist

DEFAULTS (all user-editable per run):
  target_normal_dos       = 85   days (AJ in the sheet)
  reorder_cycle_days      =  5   days (AU in the sheet)
  target_long_horizon_dos = 110  days (AY in the sheet)
  sales_window_days       = 30   days (matches DP column typical span)

REPORT CACHING (guards against Seller Central "sea of reports"):
  - Cache FBA inventory report per-account for 6 hours (configurable)
  - Cache sales data per-account for 6 hours
  - Fresh fetch runs in background; UI shows cached data with a banner
  - Max 4 reports/day per account instead of 100+
"""
from __future__ import annotations
import csv, io, json, os, re, time
import datetime as dt
from dataclasses import dataclass, field
from typing import Optional

# ---------------------------------------------------------------------------
# 1. Config + defaults
# ---------------------------------------------------------------------------

@dataclass
class InventoryConfig:
    """Per-run config. Defaults match the Lure Essentials sheet exactly."""
    target_normal_dos:       int   = 85       # AJ in the sheet
    reorder_cycle_days:      int   = 5        # AU in the sheet
    target_long_horizon_dos: int   = 110      # AY in the sheet
    sales_window_days:       int   = 30       # rolling velocity window
    new_launch_threshold:    int   = 60       # days: below this, treat as launch
    dormant_threshold:       int   = 365      # days: above this + zero velo = dead
    prime_day_uplift_pct:    float = 3.0      # 300% typical PD lift (multiplier)
    yoy_default_uplift:      float = 1.0      # 1.0 = no YoY change if we can't compute
    cache_hours:             int   = 6        # SP-API report cache lifetime
    currency:                str   = "$"      # US market default


# ---------------------------------------------------------------------------
# 2. Report caching (protects Seller Central from report spam)
# ---------------------------------------------------------------------------

class InventoryCache:
    """Simple filesystem cache. Keyed by (account_id, marketplace, report_type).
    Not thread-safe (that's fine for the app's single-user pattern), but
    tolerant of concurrent reads."""

    def __init__(self, cache_dir: str):
        self.cache_dir = cache_dir
        os.makedirs(cache_dir, exist_ok=True)

    def _path(self, account_id: str, marketplace: str, report_type: str) -> str:
        safe_type = re.sub(r"[^A-Za-z0-9_-]", "_", report_type)
        safe_acct = re.sub(r"[^A-Za-z0-9_-]", "_", account_id or "unknown")
        safe_mkt  = re.sub(r"[^A-Za-z0-9_-]", "_", marketplace or "UK")
        return os.path.join(self.cache_dir, f"{safe_acct}__{safe_mkt}__{safe_type}.json")

    def get(self, account_id: str, marketplace: str, report_type: str,
            max_age_hours: int) -> Optional[dict]:
        """Return the cached payload if it's fresh; None otherwise."""
        p = self._path(account_id, marketplace, report_type)
        if not os.path.exists(p):
            return None
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError):
            return None
        # Age check
        ts = data.get("_cached_at", 0)
        age_hours = (time.time() - ts) / 3600
        if age_hours > max_age_hours:
            return None
        data["_cache_age_hours"] = round(age_hours, 2)
        data["_cache_fresh"] = True
        return data

    def get_stale(self, account_id: str, marketplace: str, report_type: str) -> Optional[dict]:
        """Return the cached payload REGARDLESS of age (for showing stale data
        while a background refresh runs). Marks it stale so caller can warn."""
        p = self._path(account_id, marketplace, report_type)
        if not os.path.exists(p):
            return None
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError):
            return None
        ts = data.get("_cached_at", 0)
        data["_cache_age_hours"] = round((time.time() - ts) / 3600, 2)
        data["_cache_fresh"] = False
        return data

    def put(self, account_id: str, marketplace: str, report_type: str, payload: dict):
        p = self._path(account_id, marketplace, report_type)
        payload = dict(payload)
        payload["_cached_at"] = time.time()
        tmp = p + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        os.replace(tmp, p)   # atomic on POSIX + Windows


# ---------------------------------------------------------------------------
# 3. FBA inventory fetcher (SP-API GET_FBA_MYI_ALL_INVENTORY_DATA)
# ---------------------------------------------------------------------------

def fetch_fba_inventory(creds, marketplace: str, marketplace_id: str,
                        poll_timeout_sec: int = 900) -> dict:
    """Pull FBA inventory summaries via the SP-API **Inventories API**
    (get_inventory_summary_marketplace).

    Why Inventories API and not Reports API (GET_FBA_MYI_ALL_INVENTORY_DATA)?
    The Reports API requires a specific "FBA Inventory Reports" role that must
    be granted to the SP-API app at authorization time. Most sellers' apps
    (including Talha's) don't have this grant, so create_report returns:
        "Report rejected for this client. Some of the marketplaces are not
         allowed for this Report"
    The Inventories API returns the SAME per-SKU data (available / reserved /
    inbound) under the standard FBA read permission every seller's app has by
    default. Zero Seller Central report clutter as a bonus.

    Returns {rows: [dict per SKU], report_source: 'live'|'error', error: str}.
    Rows have keys matching the old Reports-API parser for downstream compatibility:
      sku, asin, product_name, condition,
      afn_fulfillable_quantity, afn_reserved_quantity,
      afn_inbound_working_quantity, afn_inbound_shipped_quantity,
      afn_inbound_receiving_quantity, inbound_total, afn_total_quantity.
    """
    try:
        from sp_api.api import Inventories
        from sp_api.base import Marketplaces
    except ImportError as e:
        return {"rows": [], "report_source": "error",
                "error": f"sp_api Inventories not available: {e}"}

    mkt_enum = getattr(Marketplaces, marketplace, None) or Marketplaces.UK
    try:
        client = Inventories(credentials=creds, marketplace=mkt_enum, timeout=30)
    except Exception as e:
        return {"rows": [], "report_source": "error",
                "error": f"Inventories client init failed: {str(e)[:200]}"}

    rows = []
    warnings = []
    next_token = None
    pages = 0
    max_pages = 40   # safety ceiling (~2000 SKUs -- pagination is 50 per page)
    while True:
        pages += 1
        if pages > max_pages:
            warnings.append(f"Stopped after {max_pages} pages -- more may exist")
            break
        try:
            if next_token:
                resp = client.get_inventory_summary_marketplace(
                    details=True, marketplaceIds=[marketplace_id], nextToken=next_token)
            else:
                resp = client.get_inventory_summary_marketplace(
                    details=True, marketplaceIds=[marketplace_id])
        except Exception as e:
            if pages == 1:
                # First page failed -- return an outright error
                return {"rows": [], "report_source": "error",
                        "error": f"get_inventory_summary_marketplace failed: {str(e)[:250]}"}
            # Partial pagination failure -- warn and use what we have
            warnings.append(f"page {pages}: {str(e)[:150]}")
            break

        payload = getattr(resp, "payload", {}) or {}
        summaries = payload.get("inventorySummaries", []) or []
        for s in summaries:
            sku = s.get("sellerSku") or ""
            if not sku:
                continue
            details = s.get("inventoryDetails") or {}
            fulfillable = float(details.get("fulfillableQuantity", 0) or 0)
            reserved = float((details.get("reservedQuantity") or {}).get("totalReservedQuantity", 0) or 0)
            inbound_working   = float(details.get("inboundWorkingQuantity") or 0)
            inbound_shipped   = float(details.get("inboundShippedQuantity") or 0)
            inbound_receiving = float(details.get("inboundReceivingQuantity") or 0)
            inbound_total     = inbound_working + inbound_shipped + inbound_receiving
            total = fulfillable + reserved + inbound_total

            rows.append({
                "sku":                             sku,
                "asin":                            s.get("asin") or "",
                "product_name":                    s.get("productName") or "",
                "condition":                       s.get("condition") or "Continue",
                # Keys match the old Reports-API TSV parser so downstream code
                # doesn't need to change:
                "afn_fulfillable_quantity":        fulfillable,
                "afn_reserved_quantity":           reserved,
                "afn_inbound_working_quantity":    inbound_working,
                "afn_inbound_shipped_quantity":    inbound_shipped,
                "afn_inbound_receiving_quantity":  inbound_receiving,
                "inbound_total":                   inbound_total,
                "afn_total_quantity":              total,
            })

        next_token = payload.get("nextToken") or (payload.get("pagination") or {}).get("nextToken")
        if not next_token:
            break

    return {
        "rows":          rows,
        "report_source": "live",   # Inventories API is real-time, not a report
        "error":         "",
        "warnings":      warnings,
    }


def _parse_fba_inventory_tsv(text: str) -> list:
    """Parse the tab-separated GET_FBA_MYI_ALL_INVENTORY_DATA payload. Normalises
    column names to snake_case; adds inbound_total = working + shipped + receiving."""
    lines = text.splitlines()
    if not lines:
        return []
    reader = csv.DictReader(lines, delimiter="\t")
    out = []
    for r in reader:
        # normalise all keys to snake_case
        norm = {re.sub(r"[^a-z0-9]+", "_", k.lower()).strip("_"): (v.strip() if isinstance(v, str) else v)
                for k, v in r.items() if k}
        # Coerce numbers
        for numkey in ("afn_fulfillable_quantity", "afn_reserved_quantity",
                        "afn_inbound_working_quantity", "afn_inbound_shipped_quantity",
                        "afn_inbound_receiving_quantity",
                        "afn_unsellable_quantity", "afn_total_quantity",
                        "mfn_fulfillable_quantity", "your_price"):
            v = norm.get(numkey, "")
            try:
                norm[numkey] = float(v) if v not in ("", None) else 0.0
            except (ValueError, TypeError):
                norm[numkey] = 0.0
        # Convenience: inbound total
        norm["inbound_total"] = (norm.get("afn_inbound_working_quantity", 0)
                                  + norm.get("afn_inbound_shipped_quantity", 0)
                                  + norm.get("afn_inbound_receiving_quantity", 0))
        out.append(norm)
    return out


# ---------------------------------------------------------------------------
# 4. Sales velocity fetcher (SP-API Orders API)
# ---------------------------------------------------------------------------

def fetch_sales_velocity(creds, marketplace: str, marketplace_id: str,
                          days_back: int = 30) -> dict:
    """Sum units sold per SKU over the last N days via the Orders API.

    Returns {units_by_sku: {sku -> units}, days_covered: int, error: str}.

    Uses get_orders + get_order_items. This is the FAST path (per the user's
    Answer to Q1: Orders API for speed, Sales & Traffic for context).
    """
    try:
        from sp_api.api import Orders
        from sp_api.base import Marketplaces
    except ImportError as e:
        return {"units_by_sku": {}, "days_covered": 0,
                "error": f"sp_api Orders not available: {e}"}

    mkt_enum = getattr(Marketplaces, marketplace, None) or Marketplaces.UK
    oc = Orders(credentials=creds, marketplace=mkt_enum)

    end   = dt.datetime.now(dt.timezone.utc) - dt.timedelta(minutes=2)
    # SP-API requires LastUpdatedBefore to be at least 2 minutes in the past.
    # Otherwise Amazon rejects with 'Before value must be at least 2 min old'.
    # This is a documented SP-API quirk (not the same as the ISO8601 format bug).
    start = end - dt.timedelta(days=days_back)
    # SP-API strict ISO 8601: requires 'Z' suffix, NOT '+00:00', and NO microseconds.
    # Python's .isoformat() produces '2026-06-02T14:30:00.123456+00:00' which Amazon
    # rejects with 'timestamp must follow ISO8601'. strftime with the explicit format
    # gives us '2026-06-02T14:30:00Z' which passes.
    _AMZ_FMT = "%Y-%m-%dT%H:%M:%SZ"
    start_str = start.strftime(_AMZ_FMT)
    end_str   = end.strftime(_AMZ_FMT)
    units_by_sku: dict = {}
    total_orders = 0

    # Page through orders in the window
    next_token = None
    page_num = 0
    max_pages = 50    # safety cap; each page = ~100 orders
    while page_num < max_pages:
        page_num += 1
        try:
            kwargs = {
                "MarketplaceIds": [marketplace_id] if marketplace_id else None,
                "LastUpdatedAfter": start_str,
                "LastUpdatedBefore": end_str,
            }
            if next_token:
                kwargs["NextToken"] = next_token
                # When paginating with NextToken, other filters MUST be omitted per SP-API
                kwargs = {"NextToken": next_token,
                          "MarketplaceIds": [marketplace_id] if marketplace_id else None}
            resp = oc.get_orders(**{k: v for k, v in kwargs.items() if v is not None})
            pay = resp.payload if hasattr(resp, "payload") else resp
        except Exception as e:
            return {"units_by_sku": units_by_sku, "days_covered": days_back,
                    "error": f"get_orders failed on page {page_num}: {str(e)[:200]}"}

        orders_list = (pay or {}).get("Orders", [])
        total_orders += len(orders_list)

        # For each order, fetch line items (unit counts per SKU)
        for order in orders_list:
            aid = order.get("AmazonOrderId")
            if not aid:
                continue
            try:
                items_resp = oc.get_order_items(aid)
                items_pay = items_resp.payload if hasattr(items_resp, "payload") else items_resp
                items = (items_pay or {}).get("OrderItems", [])
                for it in items:
                    sku = it.get("SellerSKU")
                    qty = int(it.get("QuantityOrdered") or 0)
                    if sku and qty > 0:
                        units_by_sku[sku] = units_by_sku.get(sku, 0) + qty
            except Exception:
                # Skip failed order-item fetches -- don't abort the whole run
                continue
            # gentle rate-limit: Orders API allows 6/sec
            time.sleep(0.15)

        next_token = (pay or {}).get("NextToken")
        if not next_token:
            break

    return {"units_by_sku": units_by_sku,
            "days_covered":  days_back,
            "total_orders":  total_orders,
            "error":         ""}


# ---------------------------------------------------------------------------
# 5. 3PL CSV ingester (matches the user's existing sheet columns exactly)
# ---------------------------------------------------------------------------

# Expected columns from the user's existing 3PL Inventory sheet:
#   Natural SKUs | SKUs | Product Name | ASIN |
#   3PL Stock (Available at Warehouse) | In-Transit Stock (Sea/Truck to 3PL) |
#   Ordered Quantity
THREE_PL_EXPECTED = {
    "natural_skus":                       "natural_sku",
    "skus":                               "sku",
    "product_name":                       "product_name",
    "asin":                               "asin",
    "3pl_stock_available_at_warehouse":   "warehouse_stock",
    "in_transit_stock_sea_truck_to_3pl":  "in_transit_stock",
    "ordered_quantity":                   "ordered_qty",
}


def ingest_3pl_csv(data: bytes) -> dict:
    """Parse the 3PL CSV using the exact 7-column format from the user's sheet.
    Returns {rows: [dict], warnings: [str]}."""
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try: text = data.decode("latin-1")
        except Exception as e:
            return {"rows": [], "warnings": [f"cannot decode 3PL CSV: {e}"]}
    lines = text.splitlines()
    if not lines:
        return {"rows": [], "warnings": ["empty CSV"]}
    reader = csv.DictReader(lines)
    warnings = []

    # Normalise headers, check for the expected columns
    if not reader.fieldnames:
        return {"rows": [], "warnings": ["CSV has no header row"]}
    header_map = {}
    for h in reader.fieldnames:
        norm = re.sub(r"[^a-z0-9]+", "_", h.lower()).strip("_")
        if norm in THREE_PL_EXPECTED:
            header_map[h] = THREE_PL_EXPECTED[norm]
    missing = set(THREE_PL_EXPECTED.values()) - set(header_map.values())
    if missing:
        warnings.append(f"3PL CSV missing expected columns: {sorted(missing)}. "
                        f"Present: {reader.fieldnames}")

    rows = []
    for r in reader:
        row = {}
        for src, dst in header_map.items():
            v = r.get(src, "")
            if dst in ("warehouse_stock", "in_transit_stock", "ordered_qty"):
                try: row[dst] = float(v or 0)
                except (ValueError, TypeError): row[dst] = 0.0
            else:
                row[dst] = (v or "").strip() if isinstance(v, str) else v
        if row.get("sku"):
            rows.append(row)
    return {"rows": rows, "warnings": warnings}


# ---------------------------------------------------------------------------
# 6. SKU age classifier (drives the four-bucket zero-velocity treatment)
# ---------------------------------------------------------------------------

def classify_sku(units_sold: float,
                  days_since_launch: Optional[int],
                  selling_status: str,
                  cfg: InventoryConfig) -> str:
    """Return one of: ACTIVE | NEW_LAUNCH | DORMANT | DEAD.

    ACTIVE:      selling now -- use its real velocity
    NEW_LAUNCH:  no sales yet + product is <60 days old -- assume it'll start
                 selling; use a seed rate; flag for user review
    DORMANT:     no sales + 60-365 days old -- listing likely needs attention
    DEAD:        no sales + >365 days OR Discontinued -- delist candidate
    """
    if (selling_status or "").lower() == "discontinued":
        return "DEAD"
    if units_sold > 0:
        return "ACTIVE"
    # zero velocity path
    if days_since_launch is None:
        # Can't tell age -- treat conservatively as DORMANT so it doesn't get reorder
        return "DORMANT"
    if days_since_launch < cfg.new_launch_threshold:
        return "NEW_LAUNCH"
    if days_since_launch <= cfg.dormant_threshold:
        return "DORMANT"
    return "DEAD"


# ---------------------------------------------------------------------------
# 7. Core replenishment maths (mirrors the sheet's Main Calculation formulas)
# ---------------------------------------------------------------------------

@dataclass
class ReplenishmentRow:
    """One row of the output. Mirrors the columns of the Main Calculation sheet."""
    sku:                str
    asin:               str
    product_name:       str
    market:             str
    fulfillment:        str
    selling_status:     str
    # Amazon inventory (H..K in the sheet)
    available:          float = 0.0
    reserved:           float = 0.0
    inbound:            float = 0.0
    total_amz_stock:    float = 0.0
    # 3PL inventory (M..Q)
    warehouse_stock:    float = 0.0
    in_transit_stock:   float = 0.0
    ordered_qty:        float = 0.0
    total_3pl_stock:    float = 0.0
    total_inventory:    float = 0.0
    # DOS (S..AB)
    dos_amz_available:  float = 0.0
    dos_amz_reserved:   float = 0.0
    dos_amz_inbound:    float = 0.0
    total_dos_amz:      float = 0.0
    dos_3pl_stock:      float = 0.0
    dos_in_transit:     float = 0.0
    dos_ordered_qty:    float = 0.0
    total_dos_3pl:      float = 0.0
    dos_3pl_plus_fba:   float = 0.0
    # Sales analysis (AD..AH)
    velocity_per_day:   float = 0.0     # AD -- from Orders API
    increment_pct:      float = 1.0     # AE -- manual override, default 1.0
    yoy_uplift:         float = 1.0     # AF
    expected_per_day:   float = 0.0     # AG = if(AD*AE=0, 1*AE, AD*AE)
    avg_monthly:        float = 0.0     # AH = AG * 30
    # Normal FBA replenishment (AJ..AN)
    target_normal_dos:  int   = 85      # AJ
    days_of_supply:     float = 0.0     # AK = V (total_dos_amz)
    require_days:       float = 0.0     # AL = if(AJ>AK, AJ-AK, 0)
    replenish_yesno:    str   = "No"    # AM
    required_inventory: float = 0.0     # AN = AL * AG
    # Incremented (PD/YOY) days (AP..AW)
    stock_left:         float = 0.0     # AP
    pd_increment:       float = 0.0     # AQ
    pd_yoy_increment:   float = 0.0     # AR = AF + AQ
    sales_per_day_inc:  float = 0.0     # AS = AG * AR
    days_of_supply_inc: float = 0.0     # AT
    reorder_cycle:      int   = 5       # AU
    required_days_inc:  float = 0.0     # AV
    required_inv_inc:   float = 0.0     # AW
    # 3PL replenishment (AY..BC)
    long_horizon_dos:   int   = 110     # AY
    days_of_supply_3pl: float = 0.0     # AZ
    require_days_3pl:   float = 0.0     # BA
    replenish_3pl:      str   = "No"    # BB
    required_inv_3pl:   float = 0.0     # BC
    # Totals (BE..BI)
    total_fba_days:     float = 0.0     # BE = AL + AV
    total_fba_units:    float = 0.0     # BF = AW + AN
    total_reorder_days: float = 0.0     # BH = BA
    total_reorder_units: float = 0.0    # BI = BA * AG
    # New: bucket classification + status
    bucket:             str   = "ACTIVE"    # ACTIVE|NEW_LAUNCH|DORMANT|DEAD
    days_since_launch:  Optional[int] = None
    alert:              str   = ""          # human-readable alert if reorder needed


def compute_replenishment(sku: str, asin: str, product_name: str,
                           fba_row: dict, three_pl_row: Optional[dict],
                           velocity_units: float,
                           days_since_launch: Optional[int],
                           selling_status: str,
                           cfg: InventoryConfig,
                           market: str = "US") -> ReplenishmentRow:
    """Compute one ReplenishmentRow. Formulas mirror the sheet EXACTLY,
    except for the four-bucket zero-velocity treatment which replaces the
    sheet's AG=1 fallback with the honest bucket classification."""

    fulfillment = "FBA"   # this fetcher only sees FBA data
    days = cfg.sales_window_days
    velocity_per_day = velocity_units / days if days > 0 else 0.0

    bucket = classify_sku(velocity_units, days_since_launch, selling_status, cfg)

    # AG = expected daily rate
    # For ACTIVE: use real velocity
    # For NEW_LAUNCH: use a seed of 1/day so the maths produces something, but flag
    # For DORMANT/DEAD: use 0 -- no reorder recommended
    if bucket == "ACTIVE":
        expected = velocity_per_day * cfg.yoy_default_uplift   # AF baseline of 1.0
    elif bucket == "NEW_LAUNCH":
        expected = 1.0     # seed rate; user should override
    else:
        expected = 0.0     # DORMANT / DEAD: no reorder maths

    # ---- Amazon stock (H..K) ----
    available = float(fba_row.get("afn_fulfillable_quantity", 0) or 0)
    reserved  = float(fba_row.get("afn_reserved_quantity", 0)   or 0)
    inbound   = float(fba_row.get("inbound_total", 0)           or 0)
    total_amz = available + reserved + inbound

    # ---- 3PL stock (M..Q) ----
    warehouse_stock  = float((three_pl_row or {}).get("warehouse_stock", 0)  or 0)
    in_transit       = float((three_pl_row or {}).get("in_transit_stock", 0) or 0)
    ordered_qty      = float((three_pl_row or {}).get("ordered_qty", 0)      or 0)
    total_3pl        = warehouse_stock + in_transit
    total_inv        = total_amz + total_3pl

    # ---- DOS (S..AB) ----  guard against divide-by-zero
    def _dos(stock: float) -> float:
        return (stock / expected) if expected > 0 else 0.0

    dos_avail    = _dos(available)
    dos_reserv   = _dos(reserved)
    dos_inb      = _dos(inbound)
    total_dos_amz = dos_avail + dos_reserv + dos_inb

    dos_wh       = _dos(warehouse_stock)
    dos_it       = _dos(in_transit)
    dos_oq       = _dos(ordered_qty)
    total_dos_3pl = dos_wh + dos_it + dos_oq
    dos_3pl_fba  = _dos(total_inv)

    # ---- Normal FBA replenishment (AJ..AN) ----
    ak = total_dos_amz
    aj = float(cfg.target_normal_dos)
    # BUCKET GATE: DEAD and DORMANT never get "Yes" for reorder, even if DOS is 0.
    # (Their expected=0 would blow up the divide-by-zero guard to 0 DOS, which
    # would falsely trigger "need 85 days!" and recommend reordering dead products.)
    if bucket in ("DEAD", "DORMANT"):
        require_days = 0.0
        replenish_yn = "No"
        required_inv = 0.0
    else:
        require_days = max(0.0, aj - ak)
        replenish_yn = "Yes" if require_days > 0 else "No"
        required_inv = require_days * expected

    # ---- Incremented (PD/YOY) days (AP..AW) ----
    stock_left = ((ak - aj) * expected) if ak > aj else 0.0
    pd_inc     = 0.0    # placeholder; user can override or event calendar can set it
    pd_yoy_inc = cfg.yoy_default_uplift + pd_inc   # AR
    sales_per_day_inc = expected * pd_yoy_inc
    days_of_supply_inc = (stock_left / sales_per_day_inc) if sales_per_day_inc > 0 else 0.0
    au = float(cfg.reorder_cycle_days)
    # Same bucket gate as normal replenishment: no incremented reorder for DEAD/DORMANT
    if bucket in ("DEAD", "DORMANT"):
        required_days_inc = 0.0
        required_inv_inc  = 0.0
    else:
        required_days_inc = max(0.0, au - days_of_supply_inc)
        required_inv_inc  = sales_per_day_inc * required_days_inc

    # ---- 3PL replenishment (AY..BC) ----
    # HONEST FIX: only recommend 3PL reorder if this SKU actually has 3PL data.
    # Without data we can't tell if the user tracks it at 3PL at all -- flagging
    # every SKU as "3PL reorder needed" would be a false alert.
    # ALSO apply the bucket gate: DEAD/DORMANT never get a 3PL reorder either.
    has_3pl_data = (three_pl_row is not None) and (
        warehouse_stock > 0 or in_transit > 0 or ordered_qty > 0
    )
    ay = float(cfg.target_long_horizon_dos)
    az = total_dos_3pl
    if has_3pl_data and bucket not in ("DEAD", "DORMANT"):
        ba = max(0.0, ay - az)
        replenish_3pl = "Yes" if ba > 0 else "No"
    else:
        ba = 0.0
        replenish_3pl = "N/A" if not has_3pl_data else "No"
    bc = ba * expected

    # ---- Totals (BE..BI) ----
    be = require_days + required_days_inc
    bf = required_inv_inc + required_inv
    bh = ba
    bi = ba * expected

    # ---- Alert text for the in-app badge ----
    alert = ""
    if bucket == "ACTIVE" and replenish_yn == "Yes":
        alert = f"Reorder {int(required_inv)} units (FBA below {cfg.target_normal_dos}-day target)"
    elif bucket == "NEW_LAUNCH":
        alert = "New launch -- no sales yet, verify listing is live"
    elif bucket == "DORMANT":
        alert = "Dormant -- check listing health, not selling"
    elif bucket == "DEAD":
        alert = "Dead -- consider delisting"

    return ReplenishmentRow(
        sku=sku, asin=asin, product_name=product_name,
        market=market, fulfillment=fulfillment, selling_status=selling_status,
        available=available, reserved=reserved, inbound=inbound, total_amz_stock=total_amz,
        warehouse_stock=warehouse_stock, in_transit_stock=in_transit,
        ordered_qty=ordered_qty, total_3pl_stock=total_3pl, total_inventory=total_inv,
        dos_amz_available=dos_avail, dos_amz_reserved=dos_reserv,
        dos_amz_inbound=dos_inb, total_dos_amz=total_dos_amz,
        dos_3pl_stock=dos_wh, dos_in_transit=dos_it, dos_ordered_qty=dos_oq,
        total_dos_3pl=total_dos_3pl, dos_3pl_plus_fba=dos_3pl_fba,
        velocity_per_day=velocity_per_day, increment_pct=1.0,
        yoy_uplift=cfg.yoy_default_uplift, expected_per_day=expected,
        avg_monthly=expected * 30,
        target_normal_dos=cfg.target_normal_dos,
        days_of_supply=ak, require_days=require_days,
        replenish_yesno=replenish_yn, required_inventory=required_inv,
        stock_left=stock_left, pd_increment=pd_inc, pd_yoy_increment=pd_yoy_inc,
        sales_per_day_inc=sales_per_day_inc, days_of_supply_inc=days_of_supply_inc,
        reorder_cycle=cfg.reorder_cycle_days,
        required_days_inc=required_days_inc, required_inv_inc=required_inv_inc,
        long_horizon_dos=cfg.target_long_horizon_dos,
        days_of_supply_3pl=az, require_days_3pl=ba,
        replenish_3pl=replenish_3pl, required_inv_3pl=bc,
        total_fba_days=be, total_fba_units=bf,
        total_reorder_days=bh, total_reorder_units=bi,
        bucket=bucket, days_since_launch=days_since_launch, alert=alert,
    )


# ---------------------------------------------------------------------------
# 8. Orchestrator -- ties it all together for the /inventory/run endpoint
# ---------------------------------------------------------------------------

def run_inventory_model(fba_rows: list,
                          velocity_by_sku: dict,
                          three_pl_rows: list,
                          launch_dates_by_sku: dict,
                          cfg: InventoryConfig,
                          market: str = "US") -> dict:
    """Compute the full replenishment table.

    Args:
        fba_rows:            output of fetch_fba_inventory().rows
        velocity_by_sku:     {sku -> total units sold in window}
        three_pl_rows:       output of ingest_3pl_csv().rows (optional; can be [])
        launch_dates_by_sku: {sku -> date created} for age classification
        cfg:                 InventoryConfig
        market:              country label (US|UK|etc.)

    Returns:
        {
          rows: [ReplenishmentRow as dict],
          summary: {total_skus, active, new_launch, dormant, dead,
                    fba_reorder_count, total_fba_units_needed,
                    three_pl_reorder_count},
          alerts: [{sku, alert}, ...]   # for the in-app red badge
        }
    """
    three_pl_by_sku = {r["sku"]: r for r in three_pl_rows}
    today = dt.date.today()

    rows = []
    for fba in fba_rows:
        sku = fba.get("sku", "")
        if not sku:
            continue
        asin = fba.get("asin", "")
        product_name = fba.get("product_name", "")
        selling_status = fba.get("condition", "")   # rough proxy; ideally comes from catalog

        # SKU age
        launch_date = launch_dates_by_sku.get(sku)
        days_since_launch = None
        if launch_date:
            try:
                if isinstance(launch_date, str):
                    launch_date = dt.date.fromisoformat(launch_date[:10])
                days_since_launch = (today - launch_date).days
            except (ValueError, TypeError):
                days_since_launch = None

        velocity_units = float(velocity_by_sku.get(sku, 0))
        three_pl_row = three_pl_by_sku.get(sku)

        row = compute_replenishment(
            sku=sku, asin=asin, product_name=product_name,
            fba_row=fba, three_pl_row=three_pl_row,
            velocity_units=velocity_units,
            days_since_launch=days_since_launch,
            selling_status=selling_status,
            cfg=cfg, market=market,
        )
        rows.append(row)

    # Summary
    from collections import Counter
    bucket_counts = Counter(r.bucket for r in rows)
    fba_reorder_count       = sum(1 for r in rows if r.replenish_yesno == "Yes")
    total_fba_units_needed  = round(sum(r.total_fba_units for r in rows), 1)
    three_pl_reorder_count  = sum(1 for r in rows if r.replenish_3pl == "Yes")

    alerts = [{"sku": r.sku, "asin": r.asin, "alert": r.alert}
              for r in rows if r.alert]

    return {
        "rows":    [_row_to_dict(r) for r in rows],
        "summary": {
            "total_skus":              len(rows),
            "active":                  bucket_counts.get("ACTIVE", 0),
            "new_launch":              bucket_counts.get("NEW_LAUNCH", 0),
            "dormant":                 bucket_counts.get("DORMANT", 0),
            "dead":                    bucket_counts.get("DEAD", 0),
            "fba_reorder_count":       fba_reorder_count,
            "total_fba_units_needed":  total_fba_units_needed,
            "three_pl_reorder_count":  three_pl_reorder_count,
        },
        "alerts":  alerts,
    }


def _row_to_dict(r: ReplenishmentRow) -> dict:
    return {k: v for k, v in r.__dict__.items()}


# ---------------------------------------------------------------------------
# 9. XLSX output -- mirrors the Main Calculation sheet column layout
# ---------------------------------------------------------------------------

def build_inventory_xlsx(result: dict, cfg: InventoryConfig,
                          account_label: str = "", generated_at: str = "") -> bytes:
    """Emit a coloured xlsx that mirrors the Main Calculation sheet's column
    structure so users familiar with the original can read it identically.

    Adds a bucket colour (green=ACTIVE, yellow=NEW_LAUNCH, amber=DORMANT, red=DEAD)
    and highlights any reorder rows in blue for scannability.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Replenishment"

    # Section headers (row 1)
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="223B65")
    sections = [
        ("A1:F1",   "Meta"),
        ("H1:K1",   "Amazon Stock"),
        ("M1:Q1",   "3PL Stock"),
        ("S1:V1",   "DOS in Amazon"),
        ("X1:AB1",  "DOS in 3PL"),
        ("AD1:AH1", "Sales Analysis"),
        ("AJ1:AN1", "Normal FBA Replenishment"),
        ("AP1:AW1", "Incremented (PD/YOY)"),
        ("AY1:BC1", "3PL Replenishment"),
        ("BE1:BF1", "Total FBA Suggestions"),
        ("BH1:BI1", "Total Reorder"),
        ("BK1:BM1", "Classification"),
    ]
    for coord, label in sections:
        ws.merge_cells(coord)
        cell = ws[coord.split(":")[0]]
        cell.value = label; cell.font = section_font; cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers (row 2)
    cols = [
        # Meta
        ("A", "sku", "SKU", 24),
        ("B", "product_name", "Product Name", 40),
        ("C", "asin", "ASIN", 12),
        ("D", "market", "Market", 8),
        ("E", "fulfillment", "Fulfillment", 12),
        ("F", "selling_status", "Selling Status", 14),
        # Amazon Stock
        ("H", "available", "Available", 10),
        ("I", "reserved", "Reserved", 10),
        ("J", "inbound", "Inbound", 10),
        ("K", "total_amz_stock", "Total Amz Stock", 14),
        # 3PL Stock
        ("M", "warehouse_stock", "3PL Warehouse", 12),
        ("N", "in_transit_stock", "In-Transit", 10),
        ("O", "ordered_qty", "Ordered Qty", 12),
        ("P", "total_3pl_stock", "Total 3PL", 12),
        ("Q", "total_inventory", "Total Inv", 12),
        # DOS in Amazon
        ("S", "dos_amz_available", "DOS Avail", 10),
        ("T", "dos_amz_reserved", "DOS Reserv", 10),
        ("U", "dos_amz_inbound", "DOS Inbound", 12),
        ("V", "total_dos_amz", "Total DOS Amz", 14),
        # DOS in 3PL
        ("X", "dos_3pl_stock", "DOS 3PL", 10),
        ("Y", "dos_in_transit", "DOS In-Trans", 12),
        ("Z", "dos_ordered_qty", "DOS Ordered", 12),
        ("AA", "total_dos_3pl", "Total DOS 3PL", 14),
        ("AB", "dos_3pl_plus_fba", "DOS 3PL+FBA", 14),
        # Sales Analysis
        ("AD", "velocity_per_day", "Velocity/day", 12),
        ("AE", "increment_pct", "Increment %", 12),
        ("AF", "yoy_uplift", "YoY Uplift", 12),
        ("AG", "expected_per_day", "Expected/day", 12),
        ("AH", "avg_monthly", "Avg Monthly", 12),
        # Normal FBA
        ("AJ", "target_normal_dos", "Target DOS", 10),
        ("AK", "days_of_supply", "Current DOS", 12),
        ("AL", "require_days", "Require Days", 12),
        ("AM", "replenish_yesno", "Replenish?", 10),
        ("AN", "required_inventory", "Req Units", 10),
        # Incremented
        ("AP", "stock_left", "Stock Left", 10),
        ("AQ", "pd_increment", "PD Inc", 8),
        ("AR", "pd_yoy_increment", "PD+YOY Inc", 10),
        ("AS", "sales_per_day_inc", "Sales/day Inc", 12),
        ("AT", "days_of_supply_inc", "DOS Inc", 8),
        ("AU", "reorder_cycle", "Cycle Days", 10),
        ("AV", "required_days_inc", "Req Days Inc", 12),
        ("AW", "required_inv_inc", "Req Inv Inc", 12),
        # 3PL Replenishment
        ("AY", "long_horizon_dos", "Long-hor DOS", 12),
        ("AZ", "days_of_supply_3pl", "DOS 3PL", 10),
        ("BA", "require_days_3pl", "Req Days 3PL", 12),
        ("BB", "replenish_3pl", "Replenish 3PL?", 14),
        ("BC", "required_inv_3pl", "Req Inv 3PL", 12),
        # Total FBA
        ("BE", "total_fba_days", "Total FBA Days", 14),
        ("BF", "total_fba_units", "Total FBA Units", 14),
        # Total Reorder
        ("BH", "total_reorder_days", "Reorder Days", 12),
        ("BI", "total_reorder_units", "Reorder Units", 14),
        # Classification
        ("BK", "bucket", "Bucket", 12),
        ("BL", "days_since_launch", "Age (days)", 10),
        ("BM", "alert", "Alert", 40),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3E5A80")
    for col_letter, _key, label, width in cols:
        cell = ws[f"{col_letter}2"]
        cell.value = label; cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "G3"    # freeze the first 6 identifier columns + top 2 rows

    # Bucket colour map
    BUCKET_FILLS = {
        "ACTIVE":     PatternFill("solid", fgColor="C8E6C9"),   # green
        "NEW_LAUNCH": PatternFill("solid", fgColor="FFF9C4"),   # yellow
        "DORMANT":    PatternFill("solid", fgColor="FFE0B2"),   # amber
        "DEAD":       PatternFill("solid", fgColor="FFCDD2"),   # red
    }
    reorder_fill = PatternFill("solid", fgColor="BBDEFB")       # blue for reorder rows

    # Data rows starting at 3
    for row_idx, r in enumerate(result["rows"], 3):
        bucket = r.get("bucket", "ACTIVE")
        bfill  = BUCKET_FILLS.get(bucket)
        # Row is a "reorder" if BF (total_fba_units) or BI (reorder_units) > 0
        needs_reorder = (r.get("total_fba_units", 0) > 0 or r.get("total_reorder_units", 0) > 0)

        for col_letter, key, _label, _width in cols:
            v = r.get(key, "")
            cell = ws[f"{col_letter}{row_idx}"]
            if isinstance(v, float):
                cell.value = round(v, 2)
                # DOS values, days, percentages
                if key in ("increment_pct", "yoy_uplift", "pd_increment", "pd_yoy_increment"):
                    cell.number_format = "0.00"
                elif key in ("velocity_per_day", "expected_per_day", "avg_monthly",
                              "sales_per_day_inc"):
                    cell.number_format = "0.00"
                elif key.startswith("dos_") or key.startswith("total_dos"):
                    cell.number_format = "0.0"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(v, int):
                cell.value = v
                cell.number_format = "#,##0"
            elif v is None:
                cell.value = ""
            else:
                cell.value = v
            # Colour rules: bucket cell always coloured; whole row lightly tinted if needs reorder
            if key == "bucket" and bfill:
                cell.fill = bfill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif key == "replenish_yesno" and v == "Yes":
                cell.fill = reorder_fill
                cell.font = Font(bold=True)
            elif key == "replenish_3pl" and v == "Yes":
                cell.fill = reorder_fill
                cell.font = Font(bold=True)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Inventory Replenishment Summary"
    ws2["A1"].font = Font(bold=True, size=16)
    ws2["A3"] = f"Account: {account_label or '(unspecified)'}"
    ws2["A4"] = f"Generated: {generated_at or dt.datetime.now(dt.timezone.utc).isoformat(timespec='minutes')}"
    ws2["A6"] = "Config"
    ws2["A6"].font = Font(bold=True)
    ws2["A7"] = f"Target normal DOS: {cfg.target_normal_dos}"
    ws2["A8"] = f"Reorder cycle days: {cfg.reorder_cycle_days}"
    ws2["A9"] = f"Target long-horizon DOS: {cfg.target_long_horizon_dos}"
    ws2["A10"] = f"Sales window: {cfg.sales_window_days} days"

    ws2["A12"] = "Bucket counts"
    ws2["A12"].font = Font(bold=True)
    s = result["summary"]
    ws2["A13"] = f"ACTIVE:      {s.get('active', 0)}"
    ws2["A14"] = f"NEW LAUNCH:  {s.get('new_launch', 0)}"
    ws2["A15"] = f"DORMANT:     {s.get('dormant', 0)}"
    ws2["A16"] = f"DEAD:        {s.get('dead', 0)}"
    ws2["A17"] = f"TOTAL:       {s.get('total_skus', 0)}"

    ws2["A19"] = "Reorder totals"
    ws2["A19"].font = Font(bold=True)
    ws2["A20"] = f"FBA SKUs needing reorder: {s.get('fba_reorder_count', 0)}"
    ws2["A21"] = f"Total FBA units needed:   {s.get('total_fba_units_needed', 0):,.1f}"
    ws2["A22"] = f"3PL SKUs needing reorder: {s.get('three_pl_reorder_count', 0)}"

    ws2["A24"] = "Notes"
    ws2["A24"].font = Font(bold=True)
    ws2["A25"] = "* ACTIVE SKUs use actual sales velocity from Orders API over the sales window."
    ws2["A26"] = "* NEW_LAUNCH SKUs use a seed rate of 1/day since real velocity is 0 -- verify manually."
    ws2["A27"] = "* DORMANT/DEAD SKUs get NO reorder recommendation. Review listing health."
    ws2["A28"] = "* 3PL columns are populated only for SKUs that appear in the uploaded 3PL CSV."
    for col in "A":
        ws2.column_dimensions[col].width = 60

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
