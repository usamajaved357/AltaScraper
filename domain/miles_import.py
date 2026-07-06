"""
miles_import.py  --  Miles Lubricants supplier-site -> internal product-dict adapter.

Build #3 of the multi-source listing feature for amazon_listing_generator.py.

WHY THIS EXISTS
---------------
Some suppliers (here: mileslubricants.com) have no Shopify export and no usable
CSV. The product data lives only on the live website + the PDF spec/SDS sheets
attached to each product page. This module harvests that data so the SAME
downstream pipeline (compliance gate, IP scan, Claude copy, sheet write, export)
can produce an Amazon DRAFT listing from it.

THE LOOP (confirmed with the user, mirrors the site exactly)
------------------------------------------------------------
Input  : a CSV/XLSX with ONE column of item numbers (e.g. M001000603).
         Each item number == one product in one size == one Amazon listing.
Per item number:
  1. Search the site by item number  (/search/showResult, "by Item Number").
  2. In the results table, match the row whose Product Number == the item number
     EXACTLY (no spaces / formatting).
       - 0 matches    -> flag NOT_FOUND, skip, create nothing.
       - 2+ matches   -> flag NEEDS_REVIEW, skip, create nothing.  (user's rule)
       - 1 match      -> open its /product/show/... page.
  3. Scrape the product page: description text, the specification table, and the
     drum / pail pricing.
  4. Download EVERY PDF attached on the page (Spec/TDS, SDS, + any third).
  5. Upload those PDFs to Google Drive under  <MASTER_FOLDER>/<item_number>/  and
     also extract their text locally so Claude can use it for the copy.
  6. Emit an internal product-dict (same shape the generator consumes) carrying
     the page text + spec table + PDF text, so compliance/IP/copy run unchanged.

NOTHING here publishes to Amazon. It produces the harvest bundle; the generator
turns it into a DRAFT with all the existing checks applied.

HAZMAT NOTE
-----------
These are industrial lubricants / hydraulic fluids -> hazmat / dangerous goods on
Amazon. The SDS PDF carries the GHS classification the compliance fields need;
the bundle keeps SDS text separate (key: 'sds_text') so the prompt can lean on it.
"""

import io
import re
import csv
from pathlib import Path

BASE_URL      = "https://mileslubricants.com"
SEARCH_URL    = BASE_URL + "/search/showResult"
MASTER_FOLDER_ID = "1OgdTSsz1mutoQyGh4ZdNDYULO7xiNepj"   # Miles folder in Full Circle Shared Drive

# Status flags the harvester can return per item number.
OK           = "OK"
NOT_FOUND    = "NOT_FOUND"
NEEDS_REVIEW = "NEEDS_REVIEW"
ERROR        = "ERROR"


# =============================================================================
# INPUT: read the item-number column from CSV / XLSX
# =============================================================================

def read_item_numbers(path: str) -> list:
    """Return a clean list of item numbers from a CSV or XLSX.

    Accepts a single-column file or a multi-column one with a header that names
    the item-number column (item, item number, sku, product number, etc.).
    Whitespace is stripped; blanks dropped; order + de-dup preserved.
    """
    p = Path(path)
    if not p.exists():
        return []
    rows = []
    if p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(p), read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in r])
        except Exception:
            return []
    else:
        # CSV / TSV with encoding + delimiter tolerance
        raw = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                raw = p.read_text(encoding=enc)
                break
            except Exception:
                continue
        if raw is None:
            return []
        delim = ";" if raw.count(";") > raw.count(",") else ","
        rows = [row for row in csv.reader(io.StringIO(raw), delimiter=delim)]

    if not rows:
        return []

    # Decide which column holds the item numbers.
    header = [str(c).strip().lower() for c in rows[0]]
    name_keys = ("item number", "item_number", "item", "sku", "product number",
                 "product_number", "itemno", "item no", "number")
    col_idx = 0
    has_header = False
    for i, h in enumerate(header):
        if h in name_keys:
            col_idx = i
            has_header = True
            break
    # If the first cell looks like an actual item number (not a header word),
    # treat the file as headerless and read column 0.
    if not has_header and _looks_like_item(header[0] if header else ""):
        data_rows = rows
    else:
        data_rows = rows[1:] if has_header else rows

    out, seen = [], set()
    for row in data_rows:
        if col_idx >= len(row):
            continue
        val = str(row[col_idx]).strip()
        if not val or val.lower() in name_keys:
            continue
        if val not in seen:
            seen.add(val)
            out.append(val)
    return out


def _looks_like_item(s: str) -> bool:
    """Heuristic: Miles item numbers are alphanumeric codes like M001000603."""
    s = (s or "").strip()
    return bool(re.match(r"^[A-Za-z]{1,4}\d{4,}$", s)) or bool(re.match(r"^\d{6,}$", s))


# =============================================================================
# SEARCH: item number -> product-page URL (exact Product Number match)
# =============================================================================

def find_product_url(search_html: str, item_number: str):
    """Given the rendered search-results HTML, return (status, url_or_msg).

    Matches the results-table row whose Product Number cell equals item_number
    EXACTLY. Enforces the user's rule: 0 -> NOT_FOUND, 2+ -> NEEDS_REVIEW.
    """
    item = (item_number or "").strip()
    if not item:
        return ERROR, "empty item number"

    # The results table pairs a Product Number with a product link. Pull every
    # /product/show/... href and the text around it, then match by exact number.
    # We look for the item number appearing as its own token next to a product
    # link. Two passes: structured (anchor) then loose (number proximity).
    hrefs = re.findall(r'href=["\'](/product/show/[^"\']+)["\']', search_html, re.I)
    hrefs = [h for h in dict.fromkeys(hrefs)]   # de-dup, keep order

    # Find rows: split on table-row boundaries so each chunk is one result.
    rows = re.split(r"</tr>", search_html, flags=re.I)
    matches = []
    for chunk in rows:
        # exact item number as a standalone token in this row?
        if re.search(r"(?<![A-Za-z0-9])" + re.escape(item) + r"(?![A-Za-z0-9])", chunk):
            m = re.search(r'href=["\'](/product/show/[^"\']+)["\']', chunk, re.I)
            if m:
                matches.append(m.group(1))

    # de-dup row matches
    matches = list(dict.fromkeys(matches))

    if not matches and hrefs:
        # Fallback: the page returned product links but our row-parse missed the
        # exact-number cell (markup variance). If there is exactly ONE product
        # link total, and the exact number appears anywhere, accept it.
        if len(hrefs) == 1 and re.search(
                r"(?<![A-Za-z0-9])" + re.escape(item) + r"(?![A-Za-z0-9])", search_html):
            matches = hrefs

    if not matches:
        return NOT_FOUND, f"no exact match for item {item}"
    if len(matches) > 1:
        return NEEDS_REVIEW, (f"item {item} returned {len(matches)} product matches "
                              f"-- manual review required")
    url = matches[0]
    if url.startswith("/"):
        url = BASE_URL + url
    return OK, url


# =============================================================================
# PRODUCT PAGE: parse text, spec table, pricing, and PDF links
# =============================================================================

def parse_product_page(page_html: str, page_markdown: str = "") -> dict:
    """Extract the structured product data from a /product/show/ page.

    Returns:
      title, description (text), spec_table (list[dict] rows), pricing (dict of
      pack->price), pdf_links (list of absolute URLs), volume (e.g. '5 Gal. / PAIL').
    """
    out = {
        "title": "", "description": "", "spec_table": [], "pricing": {},
        "pdf_links": [], "volume": "", "images": [],
    }

    # --- og:image / twitter:image: the canonical product photo (highest quality)
    for pat in (r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
                r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
                r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']'):
        mi = re.search(pat, page_html, re.I)
        if mi:
            u = mi.group(1).strip()
            absimg = u if u.startswith("http") else (
                BASE_URL + u if u.startswith("/") else BASE_URL + "/" + u)
            if absimg not in out["images"]:
                out["images"].insert(0, absimg)   # canonical -> first priority
            break

    # --- title: the big product heading -------------------------------------
    m = re.search(r"<h1[^>]*>(.*?)</h1>", page_html, re.I | re.S)
    if not m:
        m = re.search(r"<h2[^>]*>(.*?)</h2>", page_html, re.I | re.S)
    if m:
        out["title"] = _clean_text(m.group(1))

    # --- volume / pack label (e.g. 'vol: 5 Gal. / PAIL') --------------------
    mv = re.search(r"vol\s*:\s*([^<\n]+)", page_html, re.I)
    if mv:
        out["volume"] = _clean_text(mv.group(1))

    # --- pricing tiers: 'Single Pack $140', '10 Pack $1330' etc. ------------
    # We keep them all; downstream only the single/drum/pail unit is used.
    for label, price in re.findall(
            r"(Single Pack|\d+\s*Pack|Drum|Pail)\s*\$?\s*([\d,]+(?:\.\d+)?)",
            page_html, re.I):
        out["pricing"][_clean_text(label)] = float(price.replace(",", ""))

    # --- specification table -------------------------------------------------
    out["spec_table"] = _parse_first_table(page_html)

    # --- description text ----------------------------------------------------
    # Prefer the markdown (cleaner); fall back to stripped HTML body text.
    desc = page_markdown or ""
    if not desc:
        # take text after the tabs region
        body = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", page_html)
        desc = _clean_text(body)
    out["description"] = desc[:6000]

    # --- PDF / download links (Spec/TDS, SDS, + any third) ------------------
    # Miles' Download buttons are not always plain <a href=...pdf>. They can be
    # onclick handlers, data-* attributes, form actions, or links to a download
    # route. Cast a wide net across all of these.
    found = []
    # 1) direct .pdf anywhere in an attribute value
    found += re.findall(r'["\']([^"\']+\.pdf[^"\']*)["\']', page_html, re.I)
    # 2) href/data-*/action attrs pointing at download-ish endpoints
    found += re.findall(r'(?:href|data-url|data-file|data-href|data-download|action)\s*=\s*["\']([^"\']+)["\']',
                        page_html, re.I)
    # 3) onclick / window.open / location = '...'  (JS-triggered downloads)
    found += re.findall(r'(?:window\.open|location(?:\.href)?\s*=|downloadFile|download)\s*\(?\s*["\']([^"\']+)["\']',
                        page_html, re.I)
    # 4) any path that looks like a file route on this site
    found += re.findall(r'["\'](/[^"\']*(?:download|getfile|file|sds|tds|spec|document|upload)[^"\']*)["\']',
                        page_html, re.I)

    seen = set()
    _raw_links = []
    for u in found:
        u = u.strip()
        if not u or u.startswith("#") or u.startswith("javascript:"):
            continue
        if "<" in u or ">" in u:        # leaked table markup, not a URL
            continue
        low = u.lower()
        if low.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".css", ".js")):
            # Images: not documents, but capture likely PRODUCT photos so the
            # auto-main-image feature has a real source to restyle. Skip obvious
            # asset/icon/logo files.
            if (low.endswith((".png", ".jpg", ".jpeg", ".webp"))
                    and not any(skip in low for skip in
                                ("logo", "icon", "favicon", "sprite", "banner",
                                 "header", "footer", "bg-", "background", "btn",
                                 "thumb-", "/css/", "/js/", "loader", "spinner"))):
                absimg = u if u.startswith("http") else (
                    BASE_URL + u if u.startswith("/") else BASE_URL + "/" + u)
                if absimg not in out["images"]:
                    out["images"].append(absimg)
            continue                     # not a document either way
        # keep only things that plausibly resolve to a downloadable document
        if not (low.endswith(".pdf") or "msds" in low or "sds" in low or "tds" in low
                or "download" in low or "getfile" in low or "/document" in low
                or ("/uploads/" in low and "doc" in low)):
            continue
        absu = u if u.startswith("http") else (BASE_URL + u if u.startswith("/") else BASE_URL + "/" + u)
        if absu not in seen:
            seen.add(absu)
            _raw_links.append(absu)

    # Prefer real files over bare directory URLs. If we have any link that ends in
    # a real filename (has a '.' in its last path segment), drop the bare-folder
    # links (e.g. '.../msds_docs/') that would only 403. Keep bare folders ONLY if
    # they're all we found (so the retry/diagnostic still records them).
    def _has_filename(u):
        tail = u.split("?")[0].rstrip("/").split("/")[-1]
        return bool(tail) and "." in tail
    _real = [u for u in _raw_links if _has_filename(u)]
    out["pdf_links"] = _real if _real else _raw_links

    # Recovery + diagnostic. On some pages the real .pdf href isn't captured
    # (the filename may carry an apostrophe/odd char, or the anchor splits the
    # folder and name). If we only have bare-folder links, try to recover the
    # real filename two ways before giving up.
    only_bare = out["pdf_links"] and all(
        ("." not in u.split("?")[0].rstrip("/").split("/")[-1]) for u in out["pdf_links"])
    if (not out["pdf_links"]) or only_bare:
        recovered = []
        # (a) grab full anchor tags that point into /uploads/ and pull BOTH the
        #     href and the visible text (some pages put the filename in the text).
        for am in re.finditer(r'<a\b[^>]*href\s*=\s*["\']([^"\']*uploads[^"\']*)["\'][^>]*>(.*?)</a>',
                              page_html, re.I | re.S):
            href, text = am.group(1), _clean_text(am.group(2))
            cand = href
            # if href is a bare folder but the text looks like a filename, join them
            if "." not in href.split("/")[-1] and text:
                if re.search(r"\.pdf\b", text, re.I):
                    fname = re.search(r"([^\s/][^/]*\.pdf)\b", text, re.I)
                    if fname:
                        cand = href.rstrip("/") + "/" + fname.group(1).strip()
            recovered.append(cand)
        # (b) any /uploads/....pdf anywhere in the page, even unquoted (lenient)
        for fm in re.finditer(r'/uploads/\w+_docs/[^"\'<>]+?\.pdf', page_html, re.I):
            recovered.append(fm.group(0))
        # absolutise + keep only ones that now have a real filename
        fixed = []
        for u in recovered:
            absu = u if u.startswith("http") else (BASE_URL + u if u.startswith("/") else BASE_URL + "/" + u)
            tail = absu.split("?")[0].rstrip("/").split("/")[-1]
            if "." in tail and absu not in fixed:
                fixed.append(absu)
        if fixed:
            out["pdf_links"] = fixed

    # Diagnostic: if STILL no real file, capture the raw HTML around the uploads
    # links so the exact markup is visible in the log for a targeted fix.
    still_bare = (not out["pdf_links"]) or all(
        ("." not in u.split("?")[0].rstrip("/").split("/")[-1]) for u in out["pdf_links"])
    if still_bare:
        snips = re.findall(r'.{0,120}uploads/\w+_docs.{0,200}', page_html, re.I)
        out["_download_debug"] = " || ".join(_clean_text(s) for s in snips[:4])[:1200]

    return out


def _parse_first_table(html_str: str) -> list:
    """Parse the first real <table> into a list of row-dicts keyed by header.
    Used for the AW-32..AW-220 property grid. Returns [] if no usable table."""
    tm = re.search(r"<table[^>]*>(.*?)</table>", html_str, re.I | re.S)
    if not tm:
        return []
    table = tm.group(1)
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", table, re.I | re.S)
    if not rows:
        return []
    def cells(r):
        return [_clean_text(c) for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", r, re.I | re.S)]
    header = cells(rows[0])
    if not header:
        return []
    out = []
    for r in rows[1:]:
        cs = cells(r)
        if not cs or not any(cs):
            continue
        # pad/truncate to header length
        cs = (cs + [""] * len(header))[:len(header)]
        out.append(dict(zip(header, cs)))
    return out


def _clean_text(s: str) -> str:
    import html as _html
    s = re.sub(r"(?is)<[^>]+>", " ", s or "")
    s = _html.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# =============================================================================
# DRIVE: per-item folder + upload (reuses brand_profile.build_drive_service auth)
# =============================================================================

DRIVE_RW_SCOPES = ["https://www.googleapis.com/auth/drive"]


def build_drive_rw(config: dict, base_dir: Path):
    """Authenticated Drive v3 service with READ+WRITE scope (for folder create +
    file upload). Reuses the owner's service-account JSON, same as the rest of
    the app. Returns (service, None) or (None, reason)."""
    try:
        from google.oauth2 import service_account as _sa
        from googleapiclient.discovery import build as _build
    except Exception as e:
        return None, (f"Google API libraries missing ({e}). "
                      f"pip install google-api-python-client google-auth")
    sa_path = Path(config.get("google_service_account_json", "service_account.json"))
    if not sa_path.is_absolute():
        sa_path = Path(base_dir) / sa_path
    if not sa_path.exists():
        return None, f"Service account file not found: {sa_path}"
    try:
        creds = _sa.Credentials.from_service_account_file(str(sa_path), scopes=DRIVE_RW_SCOPES)
        service = _build("drive", "v3", credentials=creds, cache_discovery=False)
        return service, None
    except Exception as e:
        return None, f"Drive auth failed: {type(e).__name__}: {str(e)[:160]}"


def item_has_drive_files(drive_service, item_number: str, master_id: str = MASTER_FOLDER_ID,
                         log=None) -> bool:
    """True if <master>/<item_number> exists in Drive AND contains at least one
    file (not counting sub-folders). Used so that if the user DELETES an item's
    PDFs from Drive, the app re-harvests it instead of skipping it as 'already
    done' based only on the local history file.

    Logs WHAT it found so a wrong skip/keep can be diagnosed instead of guessed."""
    def _log(m):
        if log:
            try: log(m)
            except Exception: pass
    if not drive_service:
        _log(f"    [drive-check {item_number}] no drive service -> treat as MISSING")
        return False
    try:
        safe = item_number.replace("'", "")
        q = (f"name = '{safe}' and '{master_id}' in parents "
             f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false")
        res = drive_service.files().list(q=q, fields="files(id,name)",
                                         supportsAllDrives=True,
                                         includeItemsFromAllDrives=True).execute()
        folders = res.get("files", [])
        if not folders:
            _log(f"    [drive-check {item_number}] no folder under master -> MISSING (will re-harvest)")
            return False               # folder itself was deleted -> re-harvest
        fid = folders[0]["id"]
        q2 = (f"'{fid}' in parents and trashed = false "
              f"and mimeType != 'application/vnd.google-apps.folder'")
        res2 = drive_service.files().list(q=q2, fields="files(id,name)", pageSize=10,
                                          supportsAllDrives=True,
                                          includeItemsFromAllDrives=True).execute()
        files = res2.get("files", [])
        _log(f"    [drive-check {item_number}] folder exists, {len(files)} file(s) inside -> "
             f"{'PRESENT (skip)' if files else 'EMPTY (will re-harvest)'}")
        return len(files) > 0
    except Exception as e:
        # If the check itself errors, treat as MISSING so the user can always
        # force a re-harvest by deleting from Drive. (Previously this returned
        # True, which silently skipped items the user had deleted.)
        _log(f"    [drive-check {item_number}] check failed ({type(e).__name__}: "
             f"{str(e)[:80]}) -> treating as MISSING so it re-harvests")
        return False

def list_all_item_folders(drive_service, master_id: str = MASTER_FOLDER_ID,
                          with_files_only: bool = False, log=None) -> list:
    """Return the item-number NAMES of every folder directly under the master
    Drive folder -- i.e. every item that has EVER been harvested to Drive.

    This is the source of truth for "what should have a listing": generation runs
    over these, and run_miles skips any whose SKU is already in the output sheet.
    Paginated so large libraries (100s of items) are fully enumerated, not capped.

    with_files_only=True adds one extra list call per folder to drop empty folders
    (nothing to ground a listing on). Default False for speed -- run_miles already
    filters out items whose Drive folder yields no SDS/spec text."""
    def _log(m):
        if log:
            try: log(m)
            except Exception: pass
    if not drive_service:
        _log("    [drive-scan] no drive service -> 0 folders")
        return []
    names, page_token = [], None
    try:
        while True:
            q = (f"'{master_id}' in parents and trashed = false "
                 f"and mimeType = 'application/vnd.google-apps.folder'")
            res = drive_service.files().list(
                q=q, fields="nextPageToken, files(id,name)", pageSize=1000,
                supportsAllDrives=True, includeItemsFromAllDrives=True,
                pageToken=page_token).execute()
            for f in res.get("files", []):
                nm = (f.get("name") or "").strip()
                if not nm:
                    continue
                if with_files_only:
                    q2 = (f"'{f['id']}' in parents and trashed = false "
                          f"and mimeType != 'application/vnd.google-apps.folder'")
                    r2 = drive_service.files().list(
                        q=q2, fields="files(id)", pageSize=1,
                        supportsAllDrives=True,
                        includeItemsFromAllDrives=True).execute()
                    if not r2.get("files"):
                        _log(f"    [drive-scan] {nm}: empty folder -> skip")
                        continue
                names.append(nm)
            page_token = res.get("nextPageToken")
            if not page_token:
                break
    except Exception as e:
        _log(f"    [drive-scan] list failed: {type(e).__name__}: {str(e)[:120]}")
    # de-dup preserve order
    seen, out = set(), []
    for n in names:
        if n not in seen:
            seen.add(n); out.append(n)
    return out



def ensure_item_folder(drive_service, item_number: str, master_id: str = MASTER_FOLDER_ID):
    """Return the Drive folder id for <master>/<item_number>, creating it if
    needed. Returns (folder_id, None) or (None, reason)."""
    if not drive_service:
        return None, "no drive service"
    try:
        safe = item_number.replace("'", "")
        q = (f"name = '{safe}' and '{master_id}' in parents "
             f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false")
        res = drive_service.files().list(q=q, fields="files(id,name)",
                                         supportsAllDrives=True,
                                         includeItemsFromAllDrives=True).execute()
        files = res.get("files", [])
        if files:
            return files[0]["id"], None
        meta = {"name": item_number, "parents": [master_id],
                "mimeType": "application/vnd.google-apps.folder"}
        folder = drive_service.files().create(body=meta, fields="id",
                                              supportsAllDrives=True).execute()
        return folder["id"], None
    except Exception as e:
        return None, f"folder create failed: {type(e).__name__}: {str(e)[:160]}"


def upload_bytes_to_drive(drive_service, folder_id: str, filename: str,
                          data: bytes, mime: str = "application/pdf"):
    """Upload raw bytes as a file into a Drive folder. Returns (file_id, None)
    or (None, reason)."""
    if not drive_service:
        return None, "no drive service"
    try:
        from googleapiclient.http import MediaIoBaseUpload
        media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime, resumable=False)
        meta = {"name": filename, "parents": [folder_id]}
        f = drive_service.files().create(body=meta, media_body=media, fields="id",
                                         supportsAllDrives=True).execute()
        return f["id"], None
    except Exception as e:
        msg = str(e)
        low = msg.lower()
        if "storage quota" in low or ("service account" in low and "quota" in low) \
           or "storagequotaexceeded" in low:
            return None, ("Drive 403: service accounts have no personal storage. "
                          "The target folder must be inside a SHARED DRIVE (not My "
                          "Drive). Detail: " + msg[:120])
        if "403" in msg or "permission" in low or "insufficientpermissions" in low:
            return None, ("Drive 403: the service account is not allowed to write "
                          "here. Add ebay-to-amz-ds-link-automation@...gserviceaccount.com "
                          "as a MEMBER of the Shared Drive with role Content Manager "
                          "(Viewer/Commenter can't upload). Detail: " + msg[:120])
        return None, f"upload failed: {type(e).__name__}: {msg[:200]}"


# =============================================================================
# PDF: extract text for the AI (uses pdfplumber/PyPDF2 if present)
# =============================================================================

def extract_pdf_text(data: bytes) -> str:
    """Best-effort text extraction from PDF bytes. Tries pdfplumber, then
    pypdf/PyPDF2. Returns '' if none available or extraction fails."""
    # pdfplumber (best quality)
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            return "\n".join((pg.extract_text() or "") for pg in pdf.pages).strip()
    except Exception:
        pass
    for mod in ("pypdf", "PyPDF2"):
        try:
            _m = __import__(mod)
            reader = _m.PdfReader(io.BytesIO(data))
            return "\n".join((pg.extract_text() or "") for pg in reader.pages).strip()
        except Exception:
            continue
    return ""


def classify_pdf(filename: str, text: str = "") -> str:
    """Label a downloaded PDF: 'sds' | 'tds' | 'spec' | 'other' so the bundle can
    keep the SDS separate (it drives hazmat/compliance fields)."""
    n = (filename or "").lower()
    t = (text or "")[:2000].lower()
    if "sds" in n or "safety data" in n or "safety_data" in n or "safety data sheet" in t:
        return "sds"
    if "tds" in n or "technical data" in n:
        return "tds"
    if "spec" in n or "specification" in n:
        return "spec"
    return "other"


# =============================================================================
# BUNDLE -> internal product-dict (the shape the generator consumes)
# =============================================================================

def to_product_dict(item_number: str, product_url: str, parsed: dict,
                    pdf_bundle: list) -> dict:
    """Shape the harvested data into the internal product-dict.

    pdf_bundle: list of {filename, kind, text, drive_file_id}.
    The generator reads amazon_url/ebay_url today; for Miles we have neither, so
    we carry the source as 'source_url' and the rich data in dedicated keys. The
    caller (generator) routes these into the prompt + attributes.
    """
    sds_text  = "\n".join(b["text"] for b in pdf_bundle if b.get("kind") == "sds" and b.get("text"))
    spec_text = "\n".join(b["text"] for b in pdf_bundle
                          if b.get("kind") in ("tds", "spec") and b.get("text"))
    other_text = "\n".join(b["text"] for b in pdf_bundle
                           if b.get("kind") == "other" and b.get("text"))

    # Flatten the spec table into attribute-style key/values for the prompt.
    attrs = {}
    for row in (parsed.get("spec_table") or []):
        # row like {'PROPERTY': 'Viscosity, cSt/40C (D445)', 'AW-46': '45.3', ...}
        prop = ""
        for k, v in row.items():
            if k.lower() in ("property", "properties", ""):
                prop = v
                break
        if not prop:
            # first column as the property name
            vals = list(row.values())
            prop = vals[0] if vals else ""
        for k, v in row.items():
            if v and k and k.lower() not in ("property", "properties", "") and v != prop:
                attrs[f"{prop} ({k})"[:80]] = str(v)[:120]

    return {
        "source":        "miles",
        "item_number":   item_number,
        "source_url":    product_url,
        # mirror the comp_data shape where it helps downstream:
        "amazon_url":    "",                       # none -- not an arbitrage row
        "ebay_url":      "",
        "title":         parsed.get("title", "") or item_number,
        "description":   parsed.get("description", ""),
        "attributes":    attrs,
        "images":        parsed.get("images", [])[:5],   # product photos for auto main image
        "pricing_tiers": parsed.get("pricing", {}),
        "volume":        parsed.get("volume", ""),
        # PDF-derived text, kept separated so the prompt can weight them:
        "sds_text":      sds_text[:8000],          # hazmat/GHS source
        "spec_text":     spec_text[:8000],         # technical data sheet
        "other_pdf_text": other_text[:4000],
        "pdf_files":     [{"filename": b["filename"], "kind": b["kind"],
                           "drive_file_id": b.get("drive_file_id", "")}
                          for b in pdf_bundle],
        "product_type":  "",                       # let schema/preview resolve; lubricant
    }

def bundle_from_drive(drive_service, item_number: str,
                      master_id: str = MASTER_FOLDER_ID, log=print) -> dict:
    """Build a bundle dict for ONE already-harvested item by reading its PDFs from
    the item's Drive folder (Drive is the source of truth -- NO website scraping).
    Reuses extract_pdf_text + classify_pdf + to_product_dict, so the entry is the
    SAME shape harvest_item produces. Returns the bundle dict, or None if the
    folder is missing or holds no readable PDFs."""
    item_number = (item_number or "").strip()
    if not item_number or not drive_service:
        return None
    folder_id = find_item_folder_id(drive_service, item_number, master_id)
    if not folder_id:
        log(f"  [{item_number}] no Drive folder -- skip")
        return None
    q = (f"'{folder_id}' in parents and trashed = false "
         f"and mimeType != 'application/vnd.google-apps.folder'")
    try:
        res = drive_service.files().list(q=q, fields="files(id,name)", pageSize=100,
                                         supportsAllDrives=True,
                                         includeItemsFromAllDrives=True).execute()
        files = res.get("files", [])
    except Exception as e:
        log(f"  [{item_number}] Drive list failed: {type(e).__name__}: {str(e)[:80]}")
        return None
    pdf_bundle = []
    for f in files:
        fname = (f.get("name", "") or "")
        if not fname.lower().endswith(".pdf"):
            continue
        data, why = download_drive_file_bytes(drive_service, f["id"])
        if not data:
            log(f"  [{item_number}] could not download {fname}: {why}")
            continue
        text = extract_pdf_text(data)
        kind = classify_pdf(fname, text)
        pdf_bundle.append({"filename": fname, "kind": kind, "text": text,
                           "drive_file_id": f["id"]})
        log(f"  [{item_number}] {kind.upper()}: {fname} ({len(text)} chars from Drive)")
    if not pdf_bundle:
        log(f"  [{item_number}] Drive folder holds no readable PDFs -- skip")
        return None
    # No product page in this path: pass an empty 'parsed' so to_product_dict fills
    # title from the item number and leaves page-only fields blank. sds_text /
    # spec_text / other_pdf_text + pdf_files come from the Drive PDFs above.
    parsed = {"title": item_number, "description": "", "spec_table": [],
              "images": [], "pricing": {}, "volume": ""}
    return to_product_dict(item_number, "", parsed, pdf_bundle)


def download_drive_file_bytes(drive_service, file_id: str):
    """Download a Drive file's raw bytes (shared-drive safe). Returns (bytes, err)."""
    try:
        data = drive_service.files().get_media(
            fileId=file_id, supportsAllDrives=True).execute()
        return data, None
    except Exception as e:
        return None, f"{type(e).__name__}: {str(e)[:120]}"


def find_item_folder_id(drive_service, item_number: str,
                        master_id: str = MASTER_FOLDER_ID):
    """Return the Drive folder id for <master>/<item_number>, or None. Same lookup
    the harvest path uses (item_has_drive_files / ensure_item_folder)."""
    if not drive_service or not item_number:
        return None
    q = (f"'{master_id}' in parents and name = '{item_number}' "
         f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false")
    try:
        res = drive_service.files().list(q=q, fields="files(id,name)",
                                         supportsAllDrives=True,
                                         includeItemsFromAllDrives=True).execute()
        fs = res.get("files", [])
        return fs[0]["id"] if fs else None
    except Exception:
        return None



# =============================================================================
# SCRAPER: fetch a page's raw HTML + markdown (browser-rendered via crawl4ai)
# =============================================================================

async def _fetch_page(url: str, timeout: int = 30000, delay: float = 3.0,
                      click_tabs: bool = False):
    """Return (html, markdown) for a URL using crawl4ai's browser renderer.
    We need raw HTML (for PDF links + the spec table) AND markdown (clean text).
    When click_tabs=True, click each product tab (Specifications / SDS Sheet /
    Additional Info) before capturing, so files behind inactive tabs are present
    in the HTML. Returns ('', '') on failure."""
    try:
        from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
    except Exception:
        return "", ""
    cfg = BrowserConfig(headless=True, verbose=False)
    # JS to click every tab so all tab-panes render into the DOM.
    tab_js = """
    (async () => {
      const clickable = Array.from(document.querySelectorAll(
        'a, button, li, .nav-link, .tab, [role=tab], [data-toggle=tab]'));
      const wanted = ['specification','sds','additional','technical','data'];
      for (const el of clickable) {
        const t = (el.textContent||'').toLowerCase();
        if (wanted.some(w => t.includes(w))) {
          try { el.click(); } catch(e) {}
          await new Promise(r => setTimeout(r, 350));
        }
      }
      await new Promise(r => setTimeout(r, 600));
    })();
    """
    run_kwargs = dict(
        word_count_threshold=1, remove_overlay_elements=True,
        page_timeout=timeout, delay_before_return_html=delay,
    )
    if click_tabs:
        run_kwargs["js_code"] = tab_js
    run_cfg = CrawlerRunConfig(**run_kwargs)
    try:
        async with AsyncWebCrawler(config=cfg) as crawler:
            result = await crawler.arun(url=url, config=run_cfg)
            html = (getattr(result, "html", "") or
                    getattr(result, "cleaned_html", "") or
                    getattr(result, "fit_html", "") or "")
            md = ""
            _m = getattr(result, "markdown", "")
            if isinstance(_m, str):
                md = _m
            else:
                md = getattr(_m, "raw_markdown", "") or getattr(_m, "fit_markdown", "") or str(_m or "")
            return html, md
    except Exception:
        return "", ""


async def _search_item(item_number: str, timeout: int = 30000, diag: list = None):
    """Search the Miles site for an item number and return the results HTML.

    The site's search is a form POST to /search/showResult with a text field and
    a 'by Item Number' mode. We don't know the exact field names, so we try the
    most likely ones via POST first (urllib), then fall back to GET query params,
    then a rendered-browser GET. Every attempt's outcome is appended to `diag`
    (a list) so the caller can show the user exactly what was tried.
    """
    import urllib.parse as _up
    import urllib.request as _ur
    diag = diag if diag is not None else []

    def _try_post(fields: dict):
        try:
            data = _up.urlencode(fields).encode()
            req = _ur.Request(SEARCH_URL, data=data, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
                "Content-Type": "application/x-www-form-urlencoded",
                "Referer": BASE_URL + "/",
            })
            with _ur.urlopen(req, timeout=timeout // 1000 or 30) as r:
                body = r.read().decode("utf-8", "ignore")
            return body
        except Exception as e:
            diag.append(f"POST {list(fields)[:2]} failed: {type(e).__name__}")
            return ""

    # The Miles site's search now responds to GET query params (POST endpoints
    # were removed/changed and return URLError). Try GET FIRST -- it's what works
    # -- and only fall back to POST variants if GET finds nothing. This removes
    # the noisy 'POST ... failed: URLError' lines and speeds the harvest up.
    for param in ("keyword", "q", "search", "item", "itemNumber"):
        url = f"{SEARCH_URL}?{_up.urlencode({param: item_number})}"
        html, _ = await _fetch_page(url, timeout=timeout)
        has_prod = "/product/show/" in (html or "")
        diag.append(f"GET ?{param}= -> {len(html or '')} bytes, "
                    f"{'HAS' if has_prod else 'no'} product links")
        if has_prod:
            return html

    # Fallback: POST attempts (older form behavior). Failures here are expected if
    # the site is GET-only, so they're logged compactly, not as alarming errors.
    post_field_sets = [
        {"keyword": item_number, "searchType": "item"},
        {"keyword": item_number, "type": "item"},
        {"search": item_number, "searchType": "item"},
        {"q": item_number, "by": "item"},
        {"keyword": item_number, "searchBy": "itemNumber"},
        {"item": item_number},
        {"keyword": item_number},
    ]
    _post_fail = 0
    for fs in post_field_sets:
        body = _try_post(fs)
        if body:
            has_prod = "/product/show/" in body
            diag.append(f"POST {list(fs)[0]}='{item_number}' -> {len(body)} bytes, "
                        f"{'HAS' if has_prod else 'no'} product links")
            if has_prod:
                return body
        else:
            _post_fail += 1
    if _post_fail:
        diag.append(f"({_post_fail} POST search variant(s) not supported by the site -- normal)")

    # last resort: rendered browser GET of the search page
    html, _ = await _fetch_page(f"{SEARCH_URL}?keyword={_up.quote(item_number)}", timeout=timeout)
    diag.append(f"browser GET -> {len(html or '')} bytes")
    return html or ""


# =============================================================================
# FILE DOWNLOAD: fetch a PDF's bytes
# =============================================================================

def download_file_bytes(url: str, timeout: int = 30):
    """Download a file (PDF) and return (bytes, reason).
    On success: (data, ""). On failure: (b"", reason_string).
    Tries with full browser-like headers + referer. The Miles site rejects bare
    requests, so we mimic a real browser and set the product domain as referer.
    SSL verification is disabled because the supplier site's cert fails to verify
    on some Windows/Python setups (CERTIFICATE_VERIFY_FAILED) -- this is a known
    site we are deliberately fetching from, so it is safe here."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
        "Accept": "application/pdf,application/octet-stream,*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": BASE_URL + "/",
        "Connection": "keep-alive",
    }
    # 1) try requests if available (handles cookies/redirects better)
    try:
        import requests
        try:
            import urllib3
            urllib3.disable_warnings()
        except Exception:
            pass
        r = requests.get(url, headers=headers, timeout=timeout,
                         allow_redirects=True, verify=False)
        if r.status_code == 200 and r.content:
            return r.content, ""
        return b"", f"HTTP {r.status_code}"
    except ImportError:
        pass
    except Exception as e:
        _req_err = f"requests: {type(e).__name__}"
    # 2) urllib fallback with an unverified SSL context
    try:
        import urllib.request, ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
            data = resp.read()
            return (data, "") if data else (b"", "empty response")
    except Exception as e:
        return b"", f"{type(e).__name__}: {str(e)[:80]}"


# =============================================================================
# ORCHESTRATOR: full per-item harvest loop
# =============================================================================

async def harvest_item(item_number: str, drive_service, log=print,
                       master_id: str = MASTER_FOLDER_ID) -> dict:
    """Harvest ONE item number end to end. Returns a result dict:
       {status, item_number, product, message}
    where status is OK / NOT_FOUND / NEEDS_REVIEW / ERROR and `product` is the
    internal product-dict (only on OK)."""
    item_number = (item_number or "").strip()
    if not item_number:
        return {"status": ERROR, "item_number": item_number, "message": "empty item number"}

    log(f"  [{item_number}] searching...")
    _diag = []
    search_html = await _search_item(item_number, diag=_diag)
    for d in _diag:
        log(f"  [{item_number}]   search: {d}")
    if not search_html:
        return {"status": ERROR, "item_number": item_number,
                "message": "search returned nothing (site unreachable or bot-check). "
                           "Tried: " + "; ".join(_diag[-3:])}

    status, url_or_msg = find_product_url(search_html, item_number)
    if status != OK:
        log(f"  [{item_number}] {status}: {url_or_msg}")
        return {"status": status, "item_number": item_number,
                "message": url_or_msg + " | search diag: " + "; ".join(_diag[-3:])}

    product_url = url_or_msg
    log(f"  [{item_number}] product page: {product_url}")
    page_html, page_md = await _fetch_page(product_url, click_tabs=True)
    log(f"  [{item_number}]   page loaded: {len(page_html)} bytes html, {len(page_md)} bytes text")
    if not page_html:
        return {"status": ERROR, "item_number": item_number,
                "message": "could not load product page (empty HTML)"}

    parsed = parse_product_page(page_html, page_md)
    # If the page title came back weak (just the item number, a bare number/price,
    # or blank), derive a readable name. Try the description's "Series MILES X"
    # pattern first, then the URL slug.
    _title = (parsed.get("title", "") or "").strip()
    _weak = (not _title) or _title == item_number or re.fullmatch(r"[\d.,$]+", _title or "")
    if _weak:
        better = ""
        desc = parsed.get("description", "") or ""
        m = re.search(r"\b(MILES\s+[A-Z][A-Za-z]*\s*[-\s]?\s*\d+[A-Za-z]?)", desc)
        if not m:
            m = re.search(r"\bSeries\s+(MILES\s+[A-Z][\w\s-]{2,30})", desc)
        if m:
            better = re.sub(r"\s+", " ", m.group(1)).strip()
        if not better:
            slug = product_url.rstrip("/").split("/product/show/")[-1]
            if slug and not slug.isdigit():
                better = slug.replace("_", " ").strip()
        if better:
            parsed["title"] = better
    log(f"  [{item_number}] '{parsed.get('title','')[:40]}' | "
        f"{len(parsed.get('spec_table', []))} spec rows | "
        f"{len(parsed.get('pdf_links', []))} pdf(s)")
    if not parsed.get("pdf_links") and parsed.get("_download_debug"):
        log(f"  [{item_number}]   no PDFs found. Download region looked like: "
            f"{parsed['_download_debug'][:200]}")
    if not parsed.get("pdf_links") and parsed.get("_download_debug"):
        log(f"  [{item_number}]   no PDFs found. Download region looked like: "
            f"{parsed['_download_debug'][:200]}")

    # --- Drive folder for this item ------------------------------------------
    folder_id = None
    if drive_service:
        folder_id, ferr = ensure_item_folder(drive_service, item_number, master_id)
        if ferr:
            log(f"  [{item_number}] Drive folder warning: {ferr}")

    # --- download + upload + extract each PDF --------------------------------
    pdf_bundle = []
    import urllib.parse as _upq
    _links = parsed.get("pdf_links", [])
    log(f"  [{item_number}] {len(_links)} file link(s) detected:")
    for purl in _links:
        log(f"  [{item_number}]     - {purl[:90]}")
    _dbg = parsed.get("_download_debug", "")
    if _dbg:
        log(f"  [{item_number}]   PDF-link HTML (for diagnosis): {_dbg[:400]}")
    for purl in _links:
        # reject genuine HTML-fragment leakage, but NOT spaces (Miles filenames
        # like 'TDS - MILES STRATUS D.pdf' contain spaces -- encode them instead).
        if "<" in purl or ">" in purl:
            log(f"  [{item_number}] skipped (looks like markup): {purl[:60]}")
            continue
        # skip bare directory URLs with no filename (e.g. .../msds_docs/) -- these
        # 403 and carry no file; the page just didn't expose the real PDF name.
        _tail = purl.rstrip().split("?")[0].rstrip("/").split("/")[-1]
        if not _tail or "." not in _tail:
            log(f"  [{item_number}] skipped (no filename in link): {purl[:70]}")
            continue
        # percent-encode spaces and other unsafe chars in the path/filename
        safe_url = _upq.quote(purl, safe=":/?#[]@!$&'()*+,;=%~")
        data, why = download_file_bytes(safe_url)
        if not data and safe_url != purl:
            data, why = download_file_bytes(purl)   # retry raw in case encoding hurt
        if not data:
            log(f"  [{item_number}] could not download {purl[:70]} -- {why}")
            continue
        # verify it's really a PDF (some blocks return an HTML error page)
        if data[:4] != b"%PDF" and not purl.lower().endswith(".pdf"):
            log(f"  [{item_number}] skipped {purl[:60]} -- not a PDF (got {len(data)} bytes)")
            continue
        fname = purl.split("/")[-1].split("?")[0] or "file.pdf"
        fname = _upq.unquote(fname)   # restore readable filename
        if not fname.lower().endswith(".pdf"):
            fname += ".pdf"
        text = extract_pdf_text(data)
        kind = classify_pdf(fname, text)
        drive_id = ""
        if drive_service and folder_id:
            drive_id, uerr = upload_bytes_to_drive(drive_service, folder_id, fname, data)
            if uerr:
                log(f"  [{item_number}] Drive upload warning ({fname}): {uerr}")
        pdf_bundle.append({"filename": fname, "kind": kind, "text": text,
                           "drive_file_id": drive_id})
        log(f"  [{item_number}] {kind.upper()}: {fname} "
            f"({len(text)} chars text{', uploaded to Drive' if drive_id else ''})")

    product = to_product_dict(item_number, product_url, parsed, pdf_bundle)
    return {"status": OK, "item_number": item_number, "product": product,
            "message": f"{len(pdf_bundle)} file(s)"}


async def harvest_batch(item_numbers: list, config: dict, base_dir,
                        log=print, master_id: str = MASTER_FOLDER_ID) -> dict:
    """Harvest a whole batch. Returns:
       {products: [...], needs_review: [...], not_found: [...], errors: [...]}
    Builds the Drive service once and reuses it."""
    from pathlib import Path as _P
    base_dir = _P(base_dir) if not isinstance(base_dir, _P) else base_dir

    drive_service, derr = build_drive_rw(config, base_dir)
    if derr:
        log(f"  Drive unavailable -- files will NOT be saved to Drive: {derr}")
        drive_service = None

    out = {"products": [], "needs_review": [], "not_found": [], "errors": []}
    total = len(item_numbers)
    for i, item in enumerate(item_numbers, 1):
        log(f"[{i}/{total}] {item}")
        try:
            res = await harvest_item(item, drive_service, log=log, master_id=master_id)
        except Exception as e:
            out["errors"].append({"item": item, "message": f"{type(e).__name__}: {str(e)[:120]}"})
            log(f"  [{item}] ERROR: {type(e).__name__}: {str(e)[:120]}")
            continue
        st = res.get("status")
        if st == OK:
            out["products"].append(res["product"])
        elif st == NEEDS_REVIEW:
            out["needs_review"].append({"item": item, "message": res.get("message", "")})
        elif st == NOT_FOUND:
            out["not_found"].append({"item": item, "message": res.get("message", "")})
        else:
            out["errors"].append({"item": item, "message": res.get("message", "")})

    log("")
    log(f"Harvest summary: {len(out['products'])} ok | "
        f"{len(out['needs_review'])} need review | "
        f"{len(out['not_found'])} not found | {len(out['errors'])} errors")
    return out
