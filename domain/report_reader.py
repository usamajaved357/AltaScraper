"""domain/report_reader.py -- turn an uploaded report into rows, in ONE place.

WHY THIS EXISTS

Three features now take an Amazon report as a file upload -- Returns
Intelligence, the PPC search-term analytics, and the weekly KPI pack -- and each
had grown its own copy of the same two jobs:

    read the bytes            CSV, TSV, or XLSX, with whatever encoding
    find the columns          by NAME, because column ORDER changes

routes/returns_routes.py does it with csv.reader over decoded text and handles
no XLSX at all. domain/ppc_module.py does it again with its own decode and its
own fuzzy header match. A third copy was about to be written for the KPI pack,
which is exactly what CLAUDE.md Rule 12 forbids.

WHY BY NAME AND NEVER BY POSITION

The sheet this replaces sums COLUMN LETTERS -- `SUM('PPC US'!V:V)` for spend,
`SUM('Child Sales US'!AD:AD)` for new-to-brand orders. That is why it is wrong:
the Business Report tab has 16 columns, so column AD does not exist and the
new-to-brand figure has been a silent zero in every weekly pack ever sent.

A letter is a guess about a layout. A name is what the report says it is.

WHAT IT DOES NOT DO

It does not know what any report MEANS. It returns headers and rows; deciding
that "Ordered Product Sales" is revenue belongs to the feature that cares.
"""
import csv
import io
import re


def _norm(h):
    """A header, reduced to what identifies it.

    Amazon punctuates the same column differently between reports and between
    marketplaces -- "Sessions - Total", "Sessions – Total", "Sessions—Total".
    Everything that is not a letter or a digit goes.
    """
    return re.sub(r"[^a-z0-9]+", "", str(h or "").strip().lower())


def _decode(data):
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", "replace")


def _sniff(text):
    """Tab or comma? Decided on the header line, where it matters.

    Amazon ships the same report as .csv and as tab-separated .txt, and a
    tab-separated file read as CSV becomes one enormous column -- which then
    fails detection with "this is not the right report", sending someone back to
    Seller Central for a file that was already correct.
    """
    first = (text.splitlines() or [""])[0]
    return "\t" if first.count("\t") > first.count(",") else ","


def is_xlsx(data):
    # XLSX is a zip. Checking the bytes beats trusting a file extension, which
    # is whatever the browser guessed.
    return bool(data) and data[:2] == b"PK"


def read(data, filename=""):
    """bytes -> {"headers": [...], "rows": [[...]], "format": "csv|tsv|xlsx"}.

    Never raises: returns an `error` key instead, because every caller is a
    file-upload route and a traceback tells the person holding the file nothing.
    """
    if not data:
        return {"headers": [], "rows": [], "format": "", "error": "empty file"}

    if is_xlsx(data):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                        data_only=True)
            ws = wb[wb.sheetnames[0]]
            grid = [[("" if c is None else c) for c in row]
                    for row in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            return {"headers": [], "rows": [], "format": "xlsx",
                    "error": "could not read the spreadsheet: %s" % str(e)[:160]}
        grid = [r for r in grid if any(str(c).strip() for c in r)]
        if not grid:
            return {"headers": [], "rows": [], "format": "xlsx",
                    "error": "the spreadsheet has no rows in it"}
        return {"headers": [str(c) for c in grid[0]], "rows": grid[1:],
                "format": "xlsx", "error": ""}

    text = _decode(data)
    delim = _sniff(text)
    try:
        grid = list(csv.reader(io.StringIO(text), delimiter=delim))
    except Exception as e:
        return {"headers": [], "rows": [], "format": "csv",
                "error": "could not read the file: %s" % str(e)[:160]}
    grid = [r for r in grid if any(str(c).strip() for c in r)]
    if not grid:
        return {"headers": [], "rows": [], "format": "csv",
                "error": "the file has no rows in it"}
    return {"headers": [str(c) for c in grid[0]], "rows": grid[1:],
            "format": ("tsv" if delim == "\t" else "csv"), "error": ""}


def index(headers, wanted):
    """{field: column position} for the columns a caller needs.

    `wanted` is {field: (name, name, ...)} -- the alternatives Amazon has been
    seen to use for the same column. An exact normalised match wins; a
    containment match is the fallback, so "Sessions - Total" still answers to
    "sessions total" when Amazon adds a suffix.

    A field with no match is ABSENT from the result rather than mapped to 0.
    Zero is a real number and a missing column is not zero -- that confusion is
    exactly what made new-to-brand orders read as 0 for months.
    """
    norm = [_norm(h) for h in (headers or [])]
    out = {}
    for field, names in (wanted or {}).items():
        if isinstance(names, str):
            names = (names,)
        alts = [_norm(n) for n in names if str(n or "").strip()]
        hit = None
        for a in alts:                       # exact first
            if a in norm:
                hit = norm.index(a)
                break
        if hit is None:                      # then containment
            for a in alts:
                for i, h in enumerate(norm):
                    if a and (a in h or h in a):
                        hit = i
                        break
                if hit is not None:
                    break
        if hit is not None:
            out[field] = hit
    return out


def cell(row, idx, field, default=""):
    """One field out of a row, safely. `idx` is what index() returned."""
    i = idx.get(field)
    if i is None or i >= len(row):
        return default
    v = row[i]
    return default if v is None else v


_NUM = re.compile(r"-?\d+(?:\.\d+)?")

# The symbols Amazon writes in a money column, and the code each means.
_CCY = (("$", "USD"), ("£", "GBP"), ("€", "EUR"), ("¥", "JPY"),
        ("C$", "CAD"), ("A$", "AUD"), ("kr", "SEK"), ("zł", "PLN"))


def currency_of(values):
    """The currency the money in these cells is written in, or "".

    WHY THIS IS READ FROM THE DATA AND NOT FROM THE ACCOUNT

    An agency runs a client's US reports while a UK workspace happens to be
    open, and every dollar figure is then drawn with a pound sign. The number is
    right and the label is a lie, which is worse than either being wrong on its
    own -- MEASURED: uploading Naturealm's US pack into jack_uk showed
    "£61,843.59" for $61,843.59.

    Amazon writes the symbol into the cell -- "$36,195.02" -- so the report
    already knows. Returns "" when nothing says, and the caller falls back to
    the account, which is the right guess when there is nothing better.
    """
    for v in (values or []):
        s = str(v or "")
        if not s:
            continue
        for sym, code in _CCY:
            if sym in s:
                return code
    return ""


def num(v, default=None):
    """A number out of whatever Amazon wrote.

    Handles "$36,195.02", "35.86%", "1,791", "(108.57)" for negatives, and the
    empty string. Returns `default` -- None by design -- when there is no number
    in there at all, because unknown is not zero and every screen in this app
    depends on being able to tell them apart.
    """
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return default
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace(",", "").replace(" ", "")
    m = _NUM.search(s)
    if not m:
        return default
    n = float(m.group(0))
    if "%" in str(v):
        n = n / 100.0
    return -n if neg else n
