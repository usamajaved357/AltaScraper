"""The uploaded spreadsheet is not asked how big it is.

    "the campaign export parser reads only 1 row"

WHAT WAS HAPPENING

An .xlsx carries a <dimension> element declaring the rectangle it occupies --
"A1:L1" means one row. In read-only mode openpyxl BELIEVES that element: it
iterates the declared rectangle and stops, without looking at what is actually
in the file. Amazon's Campaign Manager export is built server-side and stamps a
dimension of a single row, so a 26-campaign export was read as the header plus
one campaign.

WHY IT WAS INVISIBLE

Nothing failed. The header row is real, so detect() still said "campaign
manager"; parse_campaigns() still returned rows; the pack still built and still
rendered. Spend, sales, RoAS and the branded split were simply computed over one
campaign out of twenty-six -- a wrong number that arrives quietly, which is the
same failure mode as the four spreadsheet defects this feature was built to
replace.

MEASURED, NOT ASSUMED. Every test below builds a real .xlsx, rewrites only the
dimension element inside the zip, and reads it back. The honest file must give
the same answer either way, or the fix would be trading one wrong count for
another.

This tests the SHARED reader, not the KPI pack -- Returns Intelligence and the
PPC search-term upload go through the same function and had the same hole.
"""
import io
import os
import re
import sys
import zipfile

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

FAILS = []


def check(label, got, want):
    ok = got == want
    if not ok:
        FAILS.append(label)
    print("  %-66s %s" % (label, "OK" if ok else
                          "FAIL got=%r want=%r" % (got, want)))


def truthy(label, got):
    check(label, bool(got), True)


import openpyxl

from domain import report_reader as _rr
from domain import weekly_kpi as _wk

CAMPAIGN_HDR = ["State", "Campaign name", "Status", "Type", "Targeting",
                "Impressions", "Clicks", "Total cost (converted)", "Purchases",
                "Sales (converted)", "Purchases (new to brand)",
                "Sales (new to brand) (converted)"]


def build(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def lie_about_size(data, ref="A1:L1"):
    """Same file, same rows, one honest element replaced with a lie."""
    zin = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
    rewritten = False
    for item in zin.infolist():
        body = zin.read(item.filename)
        if item.filename == "xl/worksheets/sheet1.xml":
            body, n = re.subn(rb'<dimension ref="[^"]+"/>',
                              ('<dimension ref="%s"/>' % ref).encode(), body)
            rewritten = bool(n)
        zout.writestr(item, body)
    zout.close()
    zin.close()
    assert rewritten, "the fixture did not actually rewrite the dimension"
    return out.getvalue()


def campaign_rows(n):
    return [["enabled", "[Laurence] B0DP2V7GR%d - Co" % i, "Delivering",
             "Sponsored Products", "Auto", 1000 + i, 40 + i, "$12.00", 3,
             "$88.00", 1, "$30.00"] for i in range(1, n + 1)]


HONEST = build(CAMPAIGN_HDR, campaign_rows(26))
LIAR = lie_about_size(HONEST)

print("=== the fixture reproduces the bug before anything is asserted ===")
# If this ever stops being true the test below proves nothing, so it is checked
# rather than believed.
_wb = openpyxl.load_workbook(io.BytesIO(LIAR), read_only=True, data_only=True)
_ws = _wb[_wb.sheetnames[0]]
_naive = sum(1 for _ in _ws.iter_rows(values_only=True))
_wb.close()
check("believing the file's own dimension gives one row", _naive, 1)
check("  though 27 rows are in there",
      len(openpyxl.load_workbook(io.BytesIO(LIAR))[  # not read-only
          openpyxl.load_workbook(io.BytesIO(LIAR)).sheetnames[0]].rows and
          list(openpyxl.load_workbook(io.BytesIO(LIAR)).active.rows)), 27)


print("\n=== the reader counts the rows itself ===")
got = _rr.read(LIAR, "campaigns.xlsx")
check("the header is still the header", got["headers"][:2], ["State", "Campaign name"])
check("every campaign under it is read", len(got["rows"]), 26)
check("  and it is read as a spreadsheet", got["format"], "xlsx")
check("  with no error", got["error"], "")
# THE HONEST FILE MUST NOT MOVE. A fix that changes a correct answer is not a
# fix.
honest = _rr.read(HONEST, "campaigns.xlsx")
check("an honest file reads the same as it always did", len(honest["rows"]), 26)
check("  same headers", honest["headers"], got["headers"])
check("  same first row", honest["rows"][0], got["rows"][0])
check("  same last row", honest["rows"][-1], got["rows"][-1])


print("\n=== a lie in the other direction is harmless too ===")
# A dimension claiming more rows than exist must not invent blank campaigns.
BIG = lie_about_size(HONEST, "A1:L500")
big = _rr.read(BIG, "campaigns.xlsx")
check("no empty rows are invented", len(big["rows"]), 26)
truthy("  and none of them are blank",
       all(any(str(c).strip() for c in r) for r in big["rows"]))


print("\n=== the pack is right, which is the point ===")
table = _rr.read(LIAR, "campaigns.xlsx")
check("the report is still recognised", _wk.detect(table["headers"]), _wk.CAMPAIGN)
parsed = _wk.parse_campaigns(table)
check("every campaign is parsed", len(parsed["rows"]), 26)
k = _wk.kpis({"rows": []}, parsed, ["laurence"])
# 26 x $12.00, not 1 x $12.00. The old reader made this 12.00 and nothing said so.
check("spend is the whole export", round(k["ad_spend"], 2), 312.00)
check("  sales likewise", round(k["ad_sales"], 2), 2288.00)
check("  and clicks", k["ad_clicks"], sum(40 + i for i in range(1, 27)))
check("  and the pack says what it counted", k["campaigns_counted"], 26)
check("the branded split sees them all", k["br_campaigns"], 26)


print("\n=== the other two uploads went through the same door ===")
# Rule 12: Returns Intelligence and the PPC search-term upload call this same
# function, so they were truncated too and are fixed by the same line.
RET = build(["Order ID", "ASIN", "Return reason", "Quantity"],
            [["203-%d" % i, "B0DP2V7GR2", "Too small", 1] for i in range(1, 41)])
check("a returns export is not truncated either",
      len(_rr.read(lie_about_size(RET, "A1:D1"), "returns.xlsx")["rows"]), 40)
PPC = build(["Customer Search Term", "Impressions", "Clicks", "Spend"],
            [["boot laces %d" % i, 10, 2, "1.00"] for i in range(1, 16)])
check("nor a search-term export",
      len(_rr.read(lie_about_size(PPC, "A1:D1"), "terms.xlsx")["rows"]), 15)


print("\n=== ragged rows still read safely ===")
# reset_dimensions lets rows come back at their true widths rather than padded
# to a declared one. cell() already guards a short row; this proves it.
wb = openpyxl.Workbook()
ws = wb.active
ws.append(CAMPAIGN_HDR)
ws.append(["enabled", "Short row campaign"])          # 2 of 12 columns
ws.append(["enabled", "Full row campaign", "Delivering", "Sponsored Products",
           "Auto", 500, 20, "$5.00", 1, "$40.00", 0, "$0.00"])
buf = io.BytesIO()
wb.save(buf)
ragged = _rr.read(lie_about_size(buf.getvalue(), "A1:L1"), "x.xlsx")
check("both rows arrive", len(ragged["rows"]), 2)
p = _wk.parse_campaigns(ragged)
check("  both parse", len(p["rows"]), 2)
check("  the short one reads zero, not a crash", p["rows"][0]["clicks"], 0)
check("  and the full one reads its real figure", p["rows"][1]["clicks"], 20)


print("\n=== the fix is in the shared reader, not copied (Rule 12) ===")
SRC = open("domain/report_reader.py", encoding="utf-8").read()
# The CALL, not the name -- the docstring above it explains the fix using the
# same word, and counting the string counted the explanation as a second call.
check("reset_dimensions is called once, in the shared reader",
      len([l for l in SRC.split("\n") if l.strip() == "ws.reset_dimensions()"]), 1)
truthy("  and it says why", "declares its own extent" in SRC)


def live_py():
    """Every .py the app actually runs.

    Baselines and the _merge_ archive are copies of old code kept on purpose;
    counting them would make this assertion permanently red for a reason that is
    not a defect.
    """
    skip = (".git", "__pycache__", "_backups", "_merge_", ".claude", "venv")
    for dp, dn, fn in os.walk("."):
        if any(s in dp for s in skip):
            continue
        for f in fn:
            if (f.endswith(".py") and not f.startswith("test_")
                    and not f.endswith(".baseline.py")):
                p = os.path.join(dp, f).replace("\\", "/")
                try:
                    yield p, open(p, encoding="utf-8").read()
                except Exception:
                    continue


_ro = sorted(p for p, t in live_py() if "read_only=True" in t)
check("only the shared reader opens a workbook read-only", _ro,
      ["./domain/report_reader.py"])
# A read-only sheet is the only one that lies about its size, so anything that
# walks rows must come through the shared reader rather than open its own.
_walk = sorted(p for p, t in live_py()
               if "iter_rows(" in t and "report_reader" not in p)
check("nothing else walks a workbook's rows on its own", _walk, [])
# The two template readers were fixed a DIFFERENT way and the difference
# matters: reset_dimensions() sets max_column to None, so using it there would
# have mapped no columns at all. MEASURED before the edit, not assumed.
for p in ("./domain/unified_export.py", "./amazon_listing_generator.py"):
    t = open(p, encoding="utf-8").read()
    truthy("%s reads its template in full" % os.path.basename(p),
           "read_only=True" not in t)
    truthy("  and records why reset_dimensions was wrong there",
           "reset_dimensions" in t or "max_column" in t)

print("\nFAILURES: %d" % len(FAILS))
for f in FAILS:
    print("  - " + f)
raise SystemExit(1 if FAILS else 0)
