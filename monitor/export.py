"""monitor/export.py — build an Excel (.xlsx) of the current ASIN Monitor results.

Two sheets:
  Summary  — the headline counts + one row per ASIN x marketplace (sellers, buy-box, unknowns).
  Sellers  — one row per seller offer (ASIN, marketplace, seller, type, buy-box, price, feedback,
             storefront) with UNKNOWN third-party sellers highlighted amber -- the hijacker view.

Read-only: uses the stored snapshots via checker.overview(); no live Amazon calls.
"""
import io
import datetime

from monitor import checker as _chk
from monitor import asin_monitor as _store


def build_xlsx(config_path, cfg):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ov = _chk.overview(config_path, cfg)
    summary = ov.get("summary", {})
    sku_by_asin = {r.get("asin"): r.get("sku", "") for r in _store.list_asins(config_path)}

    HEAD = Font(bold=True, color="FFFFFF")
    HEAD_FILL = PatternFill("solid", fgColor="2F3A4D")
    AMBER = PatternFill("solid", fgColor="FDE7C6")     # unknown-seller rows
    TEAL = PatternFill("solid", fgColor="D9F2EA")      # your own offers
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.font = HEAD; cell.fill = HEAD_FILL; cell.alignment = Alignment(vertical="center")

    wb = openpyxl.Workbook()

    # ---- Summary sheet ----
    s = wb.active; s.title = "Summary"
    s.append(["ASIN Monitor export", now])
    s.cell(row=1, column=1).font = Font(bold=True, size=13)
    s.append([])
    for label, key in (("ASINs tracked", "tracked"), ("Clean (only known sellers)", "clean"),
                       ("With unknown sellers", "with_unknown"), ("Unknown sellers total", "total_unknown")):
        s.append([label, summary.get(key, 0)])
        s.cell(row=s.max_row, column=1).font = Font(bold=True)
    s.append([])
    # a seller's display NAME, always falling back to the seller ID so a cell is never blank
    def _nm(x):
        return x.get("label") or x.get("id") or ""
    s.append(["ASIN", "Label", "Marketplace", "# Sellers", "Buy Box",
              "Sellers (★ = Buy Box)", "Unknown sellers", "Unknown #", "State", "Last checked"])
    _style_header(s, 10)
    for row in ov.get("rows", []):
        for m in row.get("per_marketplace", []):
            if m.get("checked"):
                sellers = m.get("sellers", [])           # already Buy-Box-first, unknown-last
                bb = next((x for x in sellers if x.get("buybox")), None)
                unk = sum(1 for x in sellers if x.get("kind") == "unknown")
                # EVERY seller by name (not just the Buy-Box holder), ★ marks who holds the Buy Box
                names_all = ", ".join(_nm(x) + (" ★" if x.get("buybox") else "") for x in sellers)
                # just the unrecognised ones, so threats can be scanned without filtering me/Amazon
                names_unknown = ", ".join(_nm(x) for x in sellers if x.get("kind") == "unknown")
                state = "skipped" if m.get("skipped") else "live"
                s.append([row["asin"], row.get("label", ""), m["marketplace"], m.get("seller_count", 0),
                          (_nm(bb) if bb else ""), names_all, names_unknown, unk, state, m.get("ts", "")])
            else:
                state = "skipped — not listed here" if m.get("skipped") else "not checked yet"
                s.append([row["asin"], row.get("label", ""), m["marketplace"], "", "", "", "", "", state, m.get("last_checked", "")])
            if m.get("checked") and sum(1 for x in m.get("sellers", []) if x.get("kind") == "unknown"):
                for c in range(1, 11):                    # any unknown seller -> highlight the row amber
                    s.cell(row=s.max_row, column=c).fill = AMBER
    s.freeze_panes = "A9"

    # ---- Sellers sheet (detailed) ----
    d = wb.create_sheet("Sellers")
    d.append(["ASIN", "Label", "SKU", "Marketplace", "Last checked", "Seller ID", "Seller name",
              "Type", "Buy Box", "Fulfilment", "Price", "Currency", "Feedback %", "Feedback count", "Storefront"])
    _style_header(d, 15)
    for row in ov.get("rows", []):
        asin = row["asin"]; label = row.get("label", ""); sku = sku_by_asin.get(asin, "")
        for m in row.get("per_marketplace", []):
            if not m.get("checked"):
                continue
            for sel in m.get("sellers", []):
                # "Seller name" uses the resolved label (manual name for known unknowns, e.g.
                # La Casa Siesta / StarPlanet GmbH / Woux LLC), falling back to the ID if unresolved
                d.append([asin, label, sku, m["marketplace"], m.get("ts", ""),
                          sel.get("id", ""), sel.get("label") or sel.get("id", ""), sel.get("kind", ""),
                          "Yes" if sel.get("buybox") else "", "FBA" if sel.get("fba") else "FBM",
                          sel.get("price"), sel.get("currency", ""),
                          sel.get("feedback_pct"), sel.get("feedback_count"), sel.get("storefront", "")])
                kind = sel.get("kind")
                if kind == "unknown":
                    for c in range(1, 16):
                        d.cell(row=d.max_row, column=c).fill = AMBER
                elif kind == "me":
                    for c in range(1, 16):
                        d.cell(row=d.max_row, column=c).fill = TEAL
    d.freeze_panes = "A2"

    # column widths
    for ws, widths in ((s, [16, 24, 12, 9, 22, 42, 30, 9, 22, 18]),
                       (d, [14, 24, 14, 12, 18, 16, 24, 12, 9, 11, 10, 9, 11, 13, 40])):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()
