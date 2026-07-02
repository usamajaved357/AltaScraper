"""
unified_export.py  --  Drop-in UNIFIED-template export for amazon_listing_generator.py

WHY THIS EXISTS
---------------
The original script exported to TWO old-style templates (FILE1 / FILE2) using
HARDCODED column positions (price at col 302 / 396, etc.). Amazon moved you to
ONE new unified template (761 columns, feedType=256) with a completely different
column order. Hardcoded positions land every value in the wrong column.

This module changes ONLY two things and reuses everything else from your script:
  1. Columns are located DYNAMICALLY by reading the template's field-ID row (row 5).
  2. Output is written to a LOCAL .xlsm copy of your template -- row 1 (account
     signature) and the VBA macros are preserved, so Seller Central accepts the
     upload (no CONTRIBUTOR_FROM_OTHER_ACCOUNT).

It does NOT touch generation. Generate/retry paths are unchanged.

CONFIG KEYS USED (config.json)
------------------------------
  unified_template_path : blank template downloaded from the CORRECT Seller Central
                          account (row 1 signature must be valid).
  unified_output_path   : where to write the filled file
                          (default: filled_unified_template.xlsm)
"""

import re
import shutil
from pathlib import Path

import openpyxl


# Marketplace tag present in every localized UK field ID:
UK_MKT = "A1F83G8C2ARO7P"

# Row indices in the template (1-based, as openpyxl uses):
ROW_FIELD_NAMES = 4
ROW_FIELD_IDS   = 5   # <-- we match logical fields against this row
ROW_EXAMPLE     = 6
ROW_DATA_START  = 7

# Regex patterns to locate each logical field by its field ID in row 5.
# Verified against the real unified template column dump.
FIELD_ID_PATTERNS = {
    "SKU":                  [r"^contribution_sku#1\.value"],
    "Product Type":         [r"^product_type#1\.value"],
    "Listing Action":       [r"record_action"],
    "Item Name":            [r"^item_name\[marketplace_id=" + UK_MKT],
    "Brand Name":           [r"^brand\[marketplace_id=" + UK_MKT],
    "Product Id Type":      [r"product_id_type"],
    "Product Id":           [r"product_id_value"],
    "Browse Node 1":        [r"^recommended_browse_nodes\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "Manufacturer":         [r"^manufacturer\[marketplace_id=" + UK_MKT],
    "Model Number":         [r"^model_number\[marketplace_id=" + UK_MKT],
    "Product Description":  [r"^product_description\[marketplace_id=" + UK_MKT],
    "Bullet Point 1":       [r"^bullet_point\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "Bullet Point 2":       [r"^bullet_point\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#2\.value"],
    "Bullet Point 3":       [r"^bullet_point\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#3\.value"],
    "Bullet Point 4":       [r"^bullet_point\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#4\.value"],
    "Bullet Point 5":       [r"^bullet_point\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#5\.value"],
    "Generic Keyword":      [r"^generic_keyword\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "Material":             [r"^material\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "Colour":               [r"^color\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "Size":                 [r"^size\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "Number of Items":      [r"^number_of_items\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "Target Gender":        [r"^target_gender\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "Age Range Description":[r"^age_range_description\[marketplace_id=" + UK_MKT],
    "Item Condition":       [r"^condition_type\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "List Price with Tax":  [r"^list_price\[marketplace_id=" + UK_MKT + r"\]#1\.value_with_tax"],
    "Product Tax Code":     [r"^product_tax_code#1\.value"],
    "Fulfillment Channel Code (UK)": [r"^fulfillment_availability#1\.fulfillment_channel_code"],
    "Quantity (UK)":        [r"^fulfillment_availability#1\.quantity"],
    "Handling Time (UK)":   [r"^fulfillment_availability#1\.lead_time_to_ship_max_days"],
    # NEW unified pricing: actual selling price = purchasable_offer.our_price
    "Your Price GBP":       [r"^purchasable_offer\[marketplace_id=" + UK_MKT + r"\]\[audience=ALL\]#1\.our_price"],
    "Country of Origin":    [r"^country_of_origin\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "Are batteries required?": [r"^batteries_required\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "Are batteries included?": [r"^batteries_included\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    # --- Pillars 3-4: generatable attributes (mapped from Attributes JSON) -----
    "unit_count":           [r"^unit_count\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "unit_count_type":      [r"^unit_count\[marketplace_id=" + UK_MKT + r"\]#1\.type", r"^unit_count_type"],
    "included_components":  [r"^included_components\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "number_of_boxes":      [r"^number_of_boxes\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "is_fragile":           [r"^is_fragile\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "item_shape":           [r"^item_shape\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "special_features":     [r"^special_feature\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "is_oven_safe":         [r"^is_oven_safe\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "is_dishwasher_safe":   [r"^is_dishwasher_safe\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "closure_material":     [r"^closure_material\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "material_type_free":   [r"^material_type_free\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "handle_material":      [r"^handle_material\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "with_lid":             [r"^with_lid\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "finish_type":          [r"^finish_type\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "pattern":              [r"^pattern\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "number_of_pieces":     [r"^number_of_pieces\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "number_of_packs":      [r"^number_of_packs\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "model_name":           [r"^model_name\[marketplace_id=" + UK_MKT + r"\]\[language_tag=en_GB\]#1\.value"],
    "part_number":          [r"^part_number\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "item_length":          [r"^item_length_width\[marketplace_id=" + UK_MKT + r"\]#1\.length\.value", r"^item_length\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "item_weight":          [r"^item_weight\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
    "item_diameter":        [r"^item_diameter\[marketplace_id=" + UK_MKT + r"\]#1\.decimal_value"],
    "merchant_shipping_group": [r"^merchant_shipping_group\[marketplace_id=" + UK_MKT + r"\]#1\.value"],
}


def build_field_map(template_path: str) -> dict:
    """
    Read row 5 (field IDs) and resolve each logical field name to its 0-based
    column index. Returns {logical_name: col_index, ..., 'TOTAL_COLS': N}.
    """
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb[wb.sheetnames[0]]
    total_cols = ws.max_column
    field_ids = []
    for c in range(1, total_cols + 1):
        v = ws.cell(row=ROW_FIELD_IDS, column=c).value
        field_ids.append(str(v) if v is not None else "")
    wb.close()

    col_map = {}
    for logical, patterns in FIELD_ID_PATTERNS.items():
        found = None
        for pat in patterns:
            rx = re.compile(pat, re.I)
            for idx, fid in enumerate(field_ids):
                if fid and rx.search(fid):
                    found = idx
                    break
            if found is not None:
                break
        if found is not None:
            col_map[logical] = found
    col_map["TOTAL_COLS"] = total_cols

    # --- Generic pass: map EVERY column by its normalised field key ------------
    # So any attribute Claude generates (e.g. dangerous_goods_regulations,
    # contains_liquid_contents) can be located, not just the named patterns above.
    # Named entries already set above win; generic keys fill the rest.
    def _fkey(fid: str) -> str:
        return str(fid).split("[")[0].split("#")[0].replace("::", "").strip().lower()
    for idx, fid in enumerate(field_ids):
        if not fid:
            continue
        k = _fkey(fid)
        if k and k not in col_map:
            col_map[k] = idx

    # --- Dimension map: scope -> group_key -> axis -> {"value":[cols],"unit":[cols]}
    # Tracked PER FIELD-GROUP (not just per-axis) so the writer can fill a group
    # only when EVERY axis it requires has a value. Amazon rejects a partially
    # filled group (e.g. item_depth_width_height with height+width but no depth
    # -> error 99022), so an incomplete group must be left entirely blank.
    # ITEM and PACKAGE are separate scopes so values never cross over.
    _ITEM_GROUPS = {"item_depth_width_height", "item_length_width_height",
                    "item_length_width", "item_length", "item_weight",
                    "item_display_weight", "item_dimensions"}
    _PKG_GROUPS  = {"item_package_dimensions", "item_package_weight",
                    "package_dimensions", "package_weight"}
    dim_groups = {"item": {}, "package": {}}
    for idx, fid in enumerate(field_ids):
        if not fid:
            continue
        low = str(fid).lower()
        key = low.split("[")[0].split("#")[0].strip()
        scope = ("package" if key in _PKG_GROUPS else
                 "item" if key in _ITEM_GROUPS else None)
        if scope is None:
            continue
        tail = low.split("#", 1)[1] if "#" in low else ""
        slot = "unit" if tail.endswith("unit") else "value"
        axis = None
        for a in ("length", "width", "height", "depth"):
            if f".{a}." in tail or tail.startswith(a + "."):
                axis = a
                break
        if axis is None:                      # bare item_length / *_weight
            if "weight" in key:
                axis = "weight"
            elif key in ("item_length",):
                axis = "length"
        if axis:
            g = dim_groups[scope].setdefault(key, {})
            g.setdefault(axis, {"value": [], "unit": []})[slot].append(idx)
    col_map["_DIM_GROUPS"] = dim_groups

    required = ["SKU", "Product Type", "Listing Action", "Item Name", "Brand Name",
                "Product Id Type", "Product Id", "Browse Node 1",
                "Bullet Point 1", "Item Condition", "Your Price GBP",
                "Fulfillment Channel Code (UK)", "Quantity (UK)", "Handling Time (UK)"]
    missing = [f for f in required if f not in col_map]
    if missing:
        print(f"  [unified_export] WARNING: could not locate columns for: {missing}")
    return col_map


def write_local_template(template_path: str, output_path: str, data_rows: list) -> str:
    """
    Copy the template to output_path (preserving row 1 signature + VBA),
    clear the example/stale rows, and write data_rows starting at row 7.
    data_rows: list of lists, each laid out to TOTAL_COLS width.
    """
    src = Path(template_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(src, dst)

    wb = openpyxl.load_workbook(dst, keep_vba=True)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb[wb.sheetnames[0]]
    total_cols = ws.max_column

    max_clear = max(ws.max_row, ROW_DATA_START + len(data_rows) + 5)
    for r in range(ROW_EXAMPLE, max_clear + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).value = None

    for i, row_vals in enumerate(data_rows):
        r = ROW_DATA_START + i
        for c0, val in enumerate(row_vals):
            if val is None or val == "":
                continue
            ws.cell(row=r, column=c0 + 1).value = val

    wb.save(dst)
    wb.close()
    return str(dst)
