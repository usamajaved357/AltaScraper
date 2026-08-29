"""routes/input_upload_routes.py -- put a CSV or Excel file into the product queue.

    POST /input/upload            a file in, queued products out
    GET  /input/upload/template   a blank CSV with the right headers

ONE OF THE TWO WAYS IN, the other being the "Add a product" form. Both write
into the LISTINGS STORE with status=QUEUED, through data/queued_store.add_queued,
and the generator picks those rows up. There is no separate queue table any
more: one table, one source of truth, and a row that was uploaded is the same
kind of thing as a row that was typed or generated.

THE SKU IS REAL FROM THE MOMENT THE ROW EXISTS -- built by the generator's own
build_sku, in the generator's own format. See data/input_row.to_listing_row for
why a temporary id would have had to be renamed later, and why that rename is
not something the store can do.

WHAT IT WILL NOT DO
  * Replace the queue. Every upload ADDS. A second file adds to the first, and
    nothing here can empty the queue -- /input/clear is still the only way work
    leaves it. That is the same promise the sheet import makes, for the same
    reason: silently dropping work because a file changed is not a behaviour
    worth having.
  * Stop at the first bad row. A file is usually mostly fine, and one unparsable
    line is not a reason to refuse the other two hundred. Each row is caught on
    its own and reported.
  * Reach Amazon, eBay or Google. This writes rows to the local queue. Nothing
    is generated and nothing is sent.

WHICH HEADERS IT UNDERSTANDS is not decided here -- see data/input_row.py, which
holds the one alias table, shared with the reader the generator itself uses.
"""
import csv
import io

from flask import request, jsonify, Response


# A file bigger than this is not a product list, and reading it into memory to
# find that out is how a local app stops responding. 8 MB is roughly 40,000
# rows of the shape this queue takes.
MAX_BYTES = 8 * 1024 * 1024

# Stop describing failures after this many. A file whose columns are wrong
# produces one error per row, and ten identical complaints say everything two
# hundred would.
MAX_ERRORS = 10

PREVIEW_ROWS = 5


# ---- reading the file -------------------------------------------------------
#
# MODULE LEVEL, not closures inside register(). They need no app, no config and
# no request -- they turn bytes into rows -- and as closures the only way to
# exercise them was to stand up Flask and post a file, which is why parsing bugs
# in this shape of code are usually found by a user rather than a test.


def read_csv_rows(data, filename=""):
    """Rows out of a CSV or TSV file, as lists of strings.

    utf-8-sig, because a CSV exported from Excel begins with a byte-order mark
    and the first header would otherwise be "﻿sku", matching nothing. The
    delimiter is sniffed rather than assumed: a European Excel writes
    semicolons, and every row arriving as a single column is the usual sign that
    the separator was guessed wrong.
    """
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        # A spreadsheet saved as "CSV (Windows)" is cp1252. Falling back beats
        # refusing a file over one curly apostrophe.
        text = data.decode("cp1252", errors="replace")
    name = (filename or "").lower()
    delim = "\t"
    if not name.endswith(".tsv"):
        try:
            delim = csv.Sniffer().sniff(text[:8000], delimiters=",;\t|").delimiter
        except Exception:
            delim = ","
    return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]


def read_xlsx_rows(data):
    """Rows out of the first worksheet of an .xlsx.

    read_only and values_only: this is a product list, not a document, and
    loading formatting for thousands of rows is the difference between a moment
    and a minute. Dates and numbers arrive as their Python types and become text
    in row_to_product.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
    finally:
        try:
            wb.close()
        except Exception:
            pass


def rows_of(data, filename):
    """(rows, error). Never raises -- a broken file is an answer, not a 500.

    .xlsm IS read: it is a normal xlsx with macros attached, openpyxl opens it,
    and the macros are never executed because nothing here runs the workbook.

    .xls IS NOT, and cannot be. The brief asked for it alongside .xlsx and
    .xlsm "via openpyxl", but .xls is the pre-2007 BIFF format, a different
    file format entirely, and openpyxl states it does not support it -- passing
    one in raises rather than returning rows. Reading it would mean adding xlrd
    as a dependency for a format Excel has not written by default since 2007.
    So it gets the one thing better than a silent failure: a refusal that says
    exactly what to do about it.
    """
    name = (filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xlsm")):
            return read_xlsx_rows(data), ""
        if name.endswith((".csv", ".tsv", ".txt")):
            return read_csv_rows(data, name), ""
        if name.endswith(".xls"):
            return [], ("That is an old-format .xls file, which this cannot "
                        "read. Open it in Excel and use Save As to make a "
                        ".xlsx or .csv, then upload that.")
        return [], ("Give a .csv, .tsv, .xlsx or .xlsm file — that one is "
                    "%s." % (name.rsplit(".", 1)[-1] if "." in name
                             else "of no recognisable type"))
    except Exception as e:
        return [], "That file could not be read: %s" % str(e)[:200]


def register(app, *, CONFIG_PATH, _state):
    """Attach /input/upload* to the app."""

    def _wsid():
        return str(_state.get("active_account_id", "") or "") or "_no_account"

    # ---- the upload ---------------------------------------------------------

    @app.route("/input/upload", methods=["POST"])
    def input_upload():
        from data import input_row as _ir
        from data import queued_store as _qs

        f = request.files.get("file")
        if f is None or not (f.filename or "").strip():
            return jsonify({"ok": False, "error": "No file was sent."}), 400
        filename = str(f.filename)

        data = f.read(MAX_BYTES + 1)
        if len(data) > MAX_BYTES:
            return jsonify({"ok": False, "error": (
                "That file is over %d MB. A product list this large is usually "
                "an export of something else — split it, or clear the columns "
                "you do not need." % (MAX_BYTES // (1024 * 1024)))}), 400
        if not data:
            return jsonify({"ok": False, "error": "That file is empty."}), 400

        rows, err = rows_of(data, filename)
        if err:
            return jsonify({"ok": False, "error": err, "filename": filename}), 400
        # A header row on its own is a file with no products in it, which is a
        # different problem from a file that would not parse.
        if not rows:
            return jsonify({"ok": False, "filename": filename,
                            "error": "There are no rows in that file."}), 400

        headers = [("" if h is None else str(h)) for h in rows[0]]
        mapping, matched, ignored = _ir.map_headers(headers)
        # THEIR HEADER -> OUR COLUMN, in their spelling. A list of our column
        # names answers "what did it understand"; this answers "what did it
        # think MY column meant", which is the question someone has when the
        # cost landed somewhere unexpected -- see the note on a bare "price" in
        # data/input_row.py.
        mapped_columns = {headers[i]: col for i, col in mapping.items()
                          if i < len(headers)}
        if not mapping:
            # SAY WHAT WAS IN THE FILE. "No columns matched" with nothing else
            # leaves the reader guessing at spelling; the headers we DID find
            # are the whole diagnosis.
            return jsonify({"ok": False, "filename": filename,
                            "found_columns": [h for h in headers if h.strip()],
                            "error": (
                                "None of that file's columns were recognised. "
                                "It needs a header row naming at least one of: "
                                "a source link, an Amazon link or ASIN, or a "
                                "product name.")}), 400

        wsid = _wsid()
        added = skipped = 0
        errors, preview = [], []
        stopped = False

        # ROWS GO STRAIGHT INTO THE LISTINGS STORE AS status=QUEUED. There is no
        # separate queue table any more: one table, one source of truth, and the
        # generator picks QUEUED rows up from it.
        #
        # The SKU set is read ONCE and grows as rows are added, so two rows in
        # one file that would build the same SKU get _2 rather than the second
        # quietly overwriting the first -- (workspace, sku) is UNIQUE and an
        # upsert would treat the repeat as an update.
        taken = _qs.taken_skus(CONFIG_PATH, wsid)

        for n, row in enumerate(rows[1:], start=2):     # 2 = first row after headers
            if len(errors) >= MAX_ERRORS:
                stopped = True
                break
            try:
                if not any(str(c).strip() for c in row if c is not None):
                    continue                            # a blank line, not a failure
                product = _ir.row_to_product(row, mapping)
                # NOT AN ERROR. A list with a heading row, a totals row or a
                # note in it is normal, and calling those failures buries the
                # rows that really did break.
                if not _ir.is_generatable(product):
                    skipped += 1
                    continue
                extras = _qs.add_queued(CONFIG_PATH, wsid, product, taken=taken)
                added += 1
                if len(preview) < PREVIEW_ROWS:
                    preview.append(dict(product, sku=extras["sku"]))
            except Exception as e:
                errors.append("Row %d: %s" % (n, str(e)[:160]))

        if stopped:
            errors.append("(stopped after %d errors — the rest of the file was "
                          "not read)" % MAX_ERRORS)

        return jsonify({
            "ok": True,
            "filename": filename,
            "workspace": wsid,
            "added": added,
            "skipped": skipped,
            "errors": errors,
            # {their header: our column} -- see above.
            "mapped_columns": mapped_columns,
            # Our column names, in COLUMNS order, for anything that wants the
            # canonical list rather than the file's wording.
            "matched": matched,
            "unmatched_columns": ignored,
            "preview": preview,
            **_qs.queued_count(CONFIG_PATH, wsid),
        })

    # ---- the blank file to fill in ------------------------------------------

    @app.route("/input/upload/template")
    def input_upload_template():
        """A blank CSV with the headers this understands, and two rows to delete.

        Headers in the order a person fills them in -- where you buy it, what it
        is, what it costs -- rather than the storage order of COLUMNS.

        The instruction is a ROW rather than a leading "#" comment line, because
        a "#" line is just a row to every spreadsheet program and would become
        the header the moment the file was opened and saved. Both rows say to
        delete themselves, and neither is harmless if left: each has a link or a
        name, so each would queue.
        """
        headers = ["ebay_url", "amazon_url", "item_name", "source_cost",
                   "selling_price", "upc", "handling_time"]
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerow([
            "https://www.ebay.co.uk/itm/1234567890",
            "https://www.amazon.co.uk/dp/B0EXAMPLE1",
            "Stainless Steel Garlic Press",
            "4.20", "12.99", "5012345678900", "3",
        ])
        w.writerow([
            "", "",
            "DELETE BOTH EXAMPLE ROWS. Only ebay_url OR amazon_url is "
            "required, not both. Leave selling_price empty and the app prices "
            "it from the cost and Amazon's fees.",
            "", "", "", "",
        ])
        w.writerow([""] * len(headers))
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition":
                     'attachment; filename="product-queue-template.csv"'})
