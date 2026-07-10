"""
Amazon Listing Generator v7.0 UK
==================================
UK Marketplace | Brand-New Listings | Flat File Export

WORKFLOW:
  Step 1 -- SP-API: competitor specs, live price, exact fees, product type schema
  Step 2 -- crawl4ai: real browser review scraping for VOC (demographics, pain points)
  Step 3 -- Autocomplete: keyword research from amazon.co.uk (no Brand Registry needed)
  Step 4 -- Claude: generates title, bullets, description, attributes from all data above
  Step 5 -- Google Sheet: write for your review (Status = APPROVED / SKIP / NEEDS_REVIEW)
  Step 6 -- Export: one command fills Amazon UK flat file template in Google Sheets

COMMANDS:
  py -3.11 amazon_listing_generator.py          # generate listings
  py -3.11 amazon_listing_generator.py retry    # re-process NEEDS_REVIEW rows
  py -3.11 amazon_listing_generator.py export   # fill template from APPROVED rows
"""

import asyncio
import base64
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

# --- Windows console encoding fix -------------------------------------------
# Amazon PDP pages and reviews contain Unicode (arrows like U+2193, em-dashes,
# bullets). On Windows the default console is cp1252, so printing or processing
# these characters raises 'charmap codec can't encode'. Force UTF-8 on the
# standard streams so scraping/printing never crashes on a stray glyph.
for _stream in ("stdout", "stderr"):
    try:
        getattr(sys, _stream).reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import anthropic
import gspread
from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
from google.oauth2.service_account import Credentials
from rich.console import Console

# --- brand-listing feature ---------------------------------------------------
import shopify_import
import brand_profile
import brand_listing
from sp_api.api import (
    CatalogItemsV20220401 as CatalogItems,
    ProductFees,
    ProductTypeDefinitions,
)
from sp_api.api import ProductsV0 as ProductPricing
from sp_api.base import Marketplaces
# ListingsItemsV20210801 / Sellers are imported lazily inside the api-mode
# functions so an older sp_api install can't break generate/export/retry.

import unified_export

console     = Console()
CONFIG_PATH = Path(os.environ.get("CONFIG_PATH", str(Path(__file__).parent / "config.json")))

MARKETPLACE_ID = "A1F83G8C2ARO7P"          # UK (default); MUTABLE — reassigned on marketplace switch
# US_MARKETPLACE_ID moved to listing/constants.py in Phase 5 (immutable, shared). Imported
# here so amazon_listing_generator.US_MARKETPLACE_ID still resolves for every use below.
from listing.constants import US_MARKETPLACE_ID

# When True (set by --minimal), build_api_attributes keeps ONLY the fields
# Amazon strictly requires (its `required` list) plus the offer essentials
# (price/quantity/condition/images/identifiers). Everything optional is dropped
# so a listing can be created fast, then enriched later in Seller Central. NOTE:
# Amazon escalates some conditionally-required fields (e.g. the lithium-battery
# group when it detects a lithium cell) to hard errors -- those still apply.
MINIMAL_MODE = False


def _cur_code() -> str:
    """ISO currency code for the ACTIVE marketplace (reads the live global)."""
    return "USD" if MARKETPLACE_ID == US_MARKETPLACE_ID else "GBP"


def _cur_sym() -> str:
    """Currency symbol for the active marketplace."""
    return "$" if MARKETPLACE_ID == US_MARKETPLACE_ID else "£"


def _amazon_domain() -> str:
    """The amazon.* domain for the active marketplace, used by all scrapers so a
    US run never scrapes amazon.co.uk (which caused catalogue mismatches)."""
    return "amazon.com" if MARKETPLACE_ID == US_MARKETPLACE_ID else "amazon.co.uk"

def sp_creds(config: dict, marketplace: str = "UK") -> dict:
    """Return SP-API LWA creds for the requested marketplace.

    If the dashboard scoped this run to a specific ACCOUNT (--account-id), use
    THAT account's credentials so a submit always publishes to the correct seller
    account in the active workspace -- never a stale top-level default.

    UK (default) reads the top-level sp_api_* keys.
    US reads a `us_spapi` block if present:
        "us_spapi": {
            "lwa_client_id":     "...",
            "lwa_client_secret": "...",
            "refresh_token":     "...",
            "seller_id":         "...",
            "marketplace_id":    "ATVPDKIKX0DER"
        }
    Falls back to the UK keys if the US block is missing (so nothing breaks),
    and prints a clear warning so the user knows to add us_spapi.
    """
    # ACCOUNT-SCOPED creds take priority (set by --account-id at startup).
    _acc_creds = config.get("_account_creds")
    if _acc_creds and _acc_creds.get("lwa_client_secret") and _acc_creds.get("refresh_token"):
        return {
            "lwa_app_id":        _acc_creds.get("lwa_app_id") or _acc_creds.get("lwa_client_id", ""),
            "lwa_client_secret": _acc_creds["lwa_client_secret"],
            "refresh_token":     _acc_creds["refresh_token"],
        }
    if str(marketplace).upper() == "US":
        us = config.get("us_spapi") or {}
        if us.get("lwa_client_secret") and us.get("refresh_token"):
            return {
                "lwa_app_id":        us.get("lwa_client_id") or us.get("lwa_app_id", ""),
                "lwa_client_secret": us["lwa_client_secret"],
                "refresh_token":     us["refresh_token"],
            }
        console.print("[yellow]Marketplace is US but no complete 'us_spapi' block "
                      "in config.json -- falling back to UK credentials. Add a "
                      "us_spapi block to publish to the US marketplace.[/yellow]")
    return {
        "lwa_app_id":        config["sp_api_client_id"],
        "lwa_client_secret": config["sp_api_client_secret"],
        "refresh_token":     config["sp_api_refresh_token"],
    }


def marketplace_id_for(marketplace: str = "UK") -> str:
    """Resolve the Amazon marketplace ID from a 'UK'/'US' label."""
    return US_MARKETPLACE_ID if str(marketplace).upper() == "US" else MARKETPLACE_ID


def seller_id_for(config: dict, marketplace: str = "UK") -> str:
    """Seller/Merchant token for the marketplace (US block can carry its own)."""
    _acc_creds = config.get("_account_creds")
    if _acc_creds and _acc_creds.get("seller_id"):
        return str(_acc_creds["seller_id"]).strip()
    if str(marketplace).upper() == "US":
        us = config.get("us_spapi") or {}
        if us.get("seller_id"):
            return str(us["seller_id"]).strip()
    return str(config.get("seller_id") or "").strip()


def _safe_records(ws):
    """get_all_records() that tolerates blank / duplicate header cells.
    gspread's own get_all_records() raises when the header row repeats a value
    (including empty strings from trailing blank columns) -- which otherwise
    crashes generate, retry, export and the API preview. Reads raw values and
    builds the dicts directly, keeping the first occurrence of each named header."""
    vals = ws.get_all_values()
    if not vals:
        return []
    headers = vals[0]
    cols, seen = [], set()
    for i, h in enumerate(headers):
        name = (h or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        cols.append((i, name))
    out = []
    for row in vals[1:]:
        out.append({name: (row[i] if i < len(row) else "") for i, name in cols})
    return out
MARKETPLACE    = Marketplaces.UK
MIN_MARGIN     = 20.0
OUTPUT_TAB     = "Listings v7.0 UK"

# =============================================================================
# CONFIG + TIMER
# =============================================================================

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        console.print("[red]config.json not found.[/red]")
        sys.exit(1)
    with open(CONFIG_PATH) as f:
        return json.load(f)

class Timer:
    def __init__(self): self.t = time.time()
    def elapsed(self): return round(time.time() - self.t, 1)

def sp_creds_LEGACY_REMOVED(config: dict) -> dict:
    # superseded by the marketplace-aware sp_creds() defined earlier
    return sp_creds(config, "UK")

# =============================================================================
# SP-API -- COMPETITOR DATA
# =============================================================================

# =============================================================================
# EBAY BROWSE API -- OPTIONAL ENRICHMENT
# =============================================================================

_EBAY_TOKEN_CACHE = {"token": None, "expires_at": 0}


def _get_ebay_token(app_id: str, cert_id: str) -> str:
    """OAuth client_credentials token from eBay. Cached per script run."""
    if _EBAY_TOKEN_CACHE["token"] and time.time() < _EBAY_TOKEN_CACHE["expires_at"] - 60:
        return _EBAY_TOKEN_CACHE["token"]
    try:
        creds_b64 = base64.b64encode(f"{app_id}:{cert_id}".encode()).decode()
        req = urllib.request.Request(
            "https://api.ebay.com/identity/v1/oauth2/token",
            data=b"grant_type=client_credentials&scope=https%3A%2F%2Fapi.ebay.com%2Foauth%2Fapi_scope",
            headers={
                "Authorization": f"Basic {creds_b64}",
                "Content-Type":  "application/x-www-form-urlencoded",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            payload = json.loads(r.read().decode("utf-8"))
        _EBAY_TOKEN_CACHE["token"]      = payload.get("access_token", "")
        _EBAY_TOKEN_CACHE["expires_at"] = time.time() + payload.get("expires_in", 7200)
        return _EBAY_TOKEN_CACHE["token"]
    except Exception as e:
        console.print(f"  [yellow]eBay token fetch failed: {str(e)[:80]}[/yellow]")
        return ""


def _extract_ebay_item_id(url: str) -> str:
    m = re.search(r"/itm/(?:[^/?]+/)?(\d{9,15})", url)
    return m.group(1) if m else ""


def fetch_ebay_supplement(ebay_url: str, app_id: str, cert_id: str) -> dict:
    """
    Pull a product's data from eBay Browse API as additional context.
    Returns dict with: title, description, price, item_specifics (dict),
    image_count, condition, category_path. Empty if anything fails or
    credentials are missing.
    """
    empty = {"title": "", "description": "", "price": "", "item_specifics": {},
             "image_count": 0, "condition": "", "category_path": ""}
    if not ebay_url or not app_id or not cert_id:
        return empty
    item_id = _extract_ebay_item_id(ebay_url)
    if not item_id:
        return empty

    token = _get_ebay_token(app_id, cert_id)
    if not token:
        return empty

    try:
        url = f"https://api.ebay.com/buy/browse/v1/item/get_item_by_legacy_id?legacy_item_id={item_id}"
        req = urllib.request.Request(
            url,
            headers={
                "Authorization":             f"Bearer {token}",
                "X-EBAY-C-MARKETPLACE-ID":   "EBAY_GB",
                "Accept":                    "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        # 400/404 usually means the listing ENDED, was removed, or never existed.
        # eBay's Browse API doesn't serve ended listings, BUT the item page HTML
        # often stays viewable for ~90 days -- we don't scrape here (adds
        # complexity), but the honest log line makes clear WHY it failed so the
        # user knows the eBay side is unrecoverable (not a bug on our end).
        if e.code in (400, 404):
            console.print(f"  [yellow]eBay item {item_id}: ended or removed (HTTP {e.code}). "
                          f"Amazon competitor data is still authoritative -- continuing.[/yellow]")
        else:
            console.print(f"  [yellow]eBay fetch ({item_id}): HTTP {e.code} {e.reason}[/yellow]")
        return empty
    except Exception as e:
        console.print(f"  [yellow]eBay fetch ({item_id}): {str(e)[:80]}[/yellow]")
        return empty

    # Parse item specifics (list of {name, value} into a flat dict)
    specifics = {}
    for asp in data.get("localizedAspects", []) or []:
        name = asp.get("name", "")
        val  = asp.get("value", "")
        if name and val:
            specifics[name.strip()] = val.strip()

    # eBay catalogue product aspects are often richer than the seller's item
    # specifics -- merge them in (without overwriting seller-provided values).
    _prod = data.get("product") or {}
    for _an, _av in (_prod.get("aspects") or {}).items():
        if not _an or not _av:
            continue
        _joined = " | ".join(str(x) for x in _av) if isinstance(_av, list) else str(_av)
        if _joined.strip():
            specifics.setdefault(_an.strip(), _joined.strip()[:300])
    # Top-level eBay catalogue identifiers / attributes
    for _fld, _label in (("brand", "Brand"), ("mpn", "MPN"), ("gtin", "GTIN"),
                         ("color", "Colour"), ("material", "Material"), ("pattern", "Pattern"),
                         ("size", "Size"), ("sizeType", "Size Type")):
        _v = data.get(_fld) or _prod.get(_fld)
        if _v and str(_v).strip():
            specifics.setdefault(_label, str(_v).strip())

    # Description: capture BOTH short and full description for maximum input.
    desc = data.get("description", "") or data.get("shortDescription", "") or ""
    if desc:
        desc = re.sub(r"<[^>]+>", " ", desc)
        desc = re.sub(r"\s+", " ", desc).strip()[:3000]
    cond_desc = data.get("conditionDescription", "") or ""
    if cond_desc:
        cond_desc = re.sub(r"\s+", " ", cond_desc).strip()[:500]

    price_obj = data.get("price", {}) or {}
    price_str = ""
    if price_obj.get("value"):
        price_str = f"{price_obj.get('currency', '')} {price_obj.get('value', '')}".strip()

    images = []
    primary = data.get("image", {}).get("imageUrl") if isinstance(data.get("image"), dict) else None
    if primary:
        images.append(primary)
    for img in (data.get("additionalImages") or [])[:9]:
        if isinstance(img, dict) and img.get("imageUrl"):
            images.append(img["imageUrl"])

    return {
        "title":          data.get("title", ""),
        "description":    desc,
        "condition_description": cond_desc,
        "price":          price_str,
        "item_specifics": specifics,
        "image_count":    len(images),
        "images":         images,
        "condition":      data.get("condition", ""),
        "category_path":  " > ".join(data.get("categoryPath", "").split("|"))
                          if data.get("categoryPath") else "",
    }


# =============================================================================
# SP-API -- COMPETITOR DATA
# =============================================================================

def _flatten_attr_value(entry) -> str:
    """Turn one SP-API attribute entry into a clean human-readable string.
    Handles {value}, {value,unit}, {displayValue}, and nested dimension shapes
    like {length:{value,unit}, width:..., height:...} -- so we never dump a raw
    Python dict (e.g. \"{'length': {'value': 30...}}\") into the attribute set."""
    if not isinstance(entry, dict):
        return str(entry).strip()
    if entry.get("value") not in (None, ""):
        unit = str(entry.get("unit") or entry.get("unit_of_measure") or "").strip()
        return f"{entry['value']} {unit}".strip()
    for k in ("displayValue", "amount", "name"):
        if entry.get(k) not in (None, ""):
            return str(entry[k]).strip()
    parts = []
    for axis in ("length", "width", "height", "depth", "weight"):
        sub = entry.get(axis)
        if isinstance(sub, dict) and sub.get("value") not in (None, ""):
            unit = str(sub.get("unit", "")).strip()
            parts.append(f"{axis} {sub['value']} {unit}".strip())
    return ", ".join(parts)


def get_competitor_asin_data(asin: str, creds: dict) -> dict:
    empty = {"asin": asin, "title": "", "brand": "", "product_type": "",
             "item_type_keyword": "", "attributes": {}, "browse_nodes": [],
             "sales_rank": [], "images": []}
    if not asin:
        return empty
    # Longer timeout + retry: the default sp-api timeout is short (~5s), which
    # ConnectTimeout's a lot from slow/congested links (e.g. Pakistan -> EU SP-API
    # endpoint). A 30s timeout with 3 retries catches transient network hiccups
    # instead of collapsing to the scraper on the very first stall.
    import time as _t_retry
    _last_err = None
    for _attempt in range(3):
        try:
            cat = CatalogItems(credentials=creds, marketplace=MARKETPLACE, timeout=30)
            _full_included = ["attributes", "dimensions", "summaries", "productTypes",
                              "salesRanks", "identifiers", "images"]
            try:
                res = cat.get_catalog_item(asin=asin, includedData=_full_included,
                                           marketplaceIds=[MARKETPLACE_ID])
            except Exception as _full_err:
                # Some roles can read the catalogue but are denied the richer blocks
                # (attributes/dimensions/identifiers). Amazon then denies the WHOLE
                # call. Retry with the minimal, always-available set so we still get
                # the title + product type instead of falling through to scraping.
                _le = (type(_full_err).__name__ + " " + str(_full_err)).lower()
                if "forbidden" in _le or "unauthorized" in _le or "accessdenied" in _le or "access to requested" in _le:
                    console.print("  [yellow]Full catalogue blocked -- retrying with minimal "
                                  "fields (summaries, productTypes)...[/yellow]")
                    res = cat.get_catalog_item(asin=asin,
                                               includedData=["summaries", "productTypes"],
                                               marketplaceIds=[MARKETPLACE_ID])
                else:
                    raise
            _last_err = None
            break   # success
        except Exception as _e:
            _last_err = _e
            _en = (type(_e).__name__ + " " + str(_e)).lower()
            # only retry TRANSIENT network errors; auth/notfound stop immediately
            if any(k in _en for k in ("timeout", "timed out", "connection", "reset",
                                       "getaddrinfo", "temporarily")) and _attempt < 2:
                console.print(f"  [yellow]Catalogue timed out (attempt {_attempt+1}/3), retrying...[/yellow]")
                _t_retry.sleep(2 + 2 * _attempt)
                continue
            raise
    try:
        p            = res.payload
        summaries    = p.get("summaries", [{}])
        title        = summaries[0].get("itemName", "") if summaries else ""
        brand        = summaries[0].get("brand",    "") if summaries else ""
        pt_list      = p.get("productTypes", [{}])
        product_type = pt_list[0].get("productType", "") if pt_list else ""
        item_type_kw = product_type.replace("_", " ").title() if product_type else ""

        sales_ranks = []
        for sr in p.get("salesRanks", []):
            for rank in sr.get("classificationRanks", []) + sr.get("displayGroupRanks", []):
                rt = rank.get("title", "") or rank.get("classificationName", "")
                rv = rank.get("rank",  0)  or rank.get("value", 0)
                if rt:
                    sales_ranks.append({"category": rt, "rank": rv})

        attributes = {}
        for key, val_list in p.get("attributes", {}).items():
            if not isinstance(val_list, list) or not val_list:
                continue
            vals = []
            for entry in val_list:                 # keep EVERY value, not just [0]
                fv = _flatten_attr_value(entry)
                if fv and fv not in vals:
                    vals.append(fv)
            if vals:
                attributes[key] = " | ".join(vals)[:400]

        # --- Summaries: pull every useful scalar (model, part no, colour, size,
        #     style, manufacturer, package qty, classification) -- not just name.
        summ = summaries[0] if summaries else {}
        for _sk, _dest in (("manufacturer", "manufacturer"), ("modelNumber", "model_number"),
                           ("partNumber", "part_number"), ("colorName", "color"),
                           ("color", "color"), ("size", "size"), ("sizeName", "size"),
                           ("style", "style"), ("styleName", "style"),
                           ("packageQuantity", "package_quantity")):
            _sv = summ.get(_sk)
            if _sv and str(_sv).strip():
                attributes.setdefault(_dest, str(_sv).strip())
        _bc = summ.get("browseClassification") or {}
        if isinstance(_bc, dict) and _bc.get("displayName"):
            attributes.setdefault("amazon_browse_classification", str(_bc["displayName"]).strip())

        # --- Identifiers: UPC / EAN / GTIN etc. (requested but never extracted) -
        for _idblock in p.get("identifiers", []):
            if _idblock.get("marketplaceId") not in (MARKETPLACE_ID, None, ""):
                continue
            for _ident in _idblock.get("identifiers", []):
                _it = str(_ident.get("identifierType") or "").strip().lower()
                _iv = _ident.get("identifier")
                if _it and _iv:
                    attributes.setdefault(f"identifier_{_it}", str(_iv).strip())
            break

        # --- Structured dimensions block: exact item + package L/W/H/weight -----
        # This is the cleanest source for physical size; previously not requested.
        for dblock in p.get("dimensions", []):
            if dblock.get("marketplaceId") not in (MARKETPLACE_ID, None, ""):
                continue
            for scope in ("item", "package"):
                dims = dblock.get(scope, {})
                if not isinstance(dims, dict):
                    continue
                pre = "" if scope == "item" else "package_"
                for axis in ("length", "width", "height", "depth", "weight"):
                    sub = dims.get(axis)
                    if isinstance(sub, dict) and sub.get("value") not in (None, ""):
                        unit = str(sub.get("unit", "")).strip()
                        attributes.setdefault(f"item_{pre}{axis}",
                                              f"{sub['value']} {unit}".strip())
            break   # only the first matching marketplace

        images = []
        for img_set in p.get("images", []):
            if img_set.get("marketplaceId") == MARKETPLACE_ID:
                for img in img_set.get("images", [])[:4]:
                    images.append(img.get("link", ""))

        return {"asin": asin, "title": title, "brand": brand,
                "product_type": product_type, "item_type_keyword": item_type_kw,
                "attributes": attributes, "browse_nodes": [],
                "sales_rank": sales_ranks, "images": images}
    except Exception as e:
        _etype = type(e).__name__
        _emsg  = str(e)
        _low   = (_etype + " " + _emsg).lower()
        if "throttl" in _low or "quotaexceeded" in _low or "429" in _low:
            _why = ("THROTTLED by Amazon (SP-API rate limit). The ASIN is fine; "
                    "the call was rejected for firing too fast. Re-run this SKU on "
                    "its own, or space out generation.")
        elif "notfound" in _low or "not found" in _low or "404" in _low:
            _why = ("ASIN NOT FOUND in this marketplace's catalogue. It may have "
                    "been removed/suppressed since you saved it, or the ASIN is for "
                    "a different marketplace than this run (--marketplace).")
        elif "forbidden" in _low or "unauthorized" in _low or "accessdenied" in _low or "403" in _low:
            _why = ("FORBIDDEN. Auth works but this SP-API app's role isn't granted "
                    "Catalog Items (or the requested includedData). Check the app's "
                    "roles in Seller Central > Develop Apps.")
        else:
            _why = "unclassified -- see the type/message above."
        console.print(f"  [red]Catalogue fetch failed: {_etype}: {_emsg[:200]}[/red]")
        console.print(f"  [yellow]Reason: {_why}[/yellow]")
        return empty


def get_pricing_data(asin: str, creds: dict) -> dict:
    """Return the best available price for a competitor ASIN.

    Cascades through THREE SP-API sources so an out-of-stock or Buy-Box-less
    listing still yields a usable price:
      1. Buy Box (getCompetitivePricing) -- the ideal case.
      2. Any active offer (getItemOffers) -- covers the "offers exist but
         nobody won the Buy Box" case that used to short-circuit to scraping.
      3. Retail-price / list-price (getPricing) -- covers OOS listings whose
         list price is still recorded on the catalogue.
    result['price_source'] tells the caller WHICH tier actually produced the
    number so the honest log line can show it.
    """
    result = {"buy_box_price": 0.0, "offer_count": 0, "price_source": ""}
    if not asin:
        return result
    import time as _t_retry

    def _retryable(e):
        _en = (type(e).__name__ + " " + str(e)).lower()
        return any(k in _en for k in ("timeout", "timed out", "connection", "reset",
                                       "getaddrinfo", "temporarily"))

    # --- Tier 1: Buy Box via getCompetitivePricing ---
    for _attempt in range(3):
        try:
            pricing = ProductPricing(credentials=creds, marketplace=MARKETPLACE, timeout=30)
            comp    = pricing.get_competitive_pricing_for_asins(asin_list=[asin])
            items   = comp.payload if isinstance(comp.payload, list) else []
            for item in items:
                product = item.get("Product", {})
                for cp in product.get("CompetitivePricing", {}).get("CompetitivePrices", []):
                    if cp.get("CompetitivePriceId") == "1":
                        result["buy_box_price"] = float(
                            cp.get("Price", {}).get("LandedPrice", {}).get("Amount", 0))
                offers = product.get("CompetitivePricing", {}).get("NumberOfOfferListings", [])
                result["offer_count"] = sum(
                    o.get("Count", 0) for o in offers if o.get("condition", "").lower() == "new")
            if result["buy_box_price"] > 0:
                result["price_source"] = "Buy Box"
                return result
            break
        except Exception as e:
            if _retryable(e) and _attempt < 2:
                console.print(f"  [yellow]Pricing (Buy Box) timed out (attempt {_attempt+1}/3), retrying...[/yellow]")
                _t_retry.sleep(2 + 2 * _attempt)
                continue
            console.print(f"  [yellow]Pricing (Buy Box) failed: {str(e)[:80]}[/yellow]")
            break

    # --- Tier 2: any active offer via getItemOffers ---
    # Runs when Buy Box was 0 (no winner) or absent. Reads the lowest new-condition
    # offer so we still get a real price even without a Buy Box winner.
    try:
        pricing = ProductPricing(credentials=creds, marketplace=MARKETPLACE, timeout=30)
        offers_resp = pricing.get_item_offers(asin=asin, item_condition="New")
        offers_payload = offers_resp.payload if isinstance(offers_resp.payload, dict) else {}
        lowest = 0.0
        for offer in (offers_payload.get("Offers") or []):
            price = float(offer.get("ListingPrice", {}).get("Amount", 0) or 0)
            ship  = float(offer.get("Shipping", {}).get("Amount", 0) or 0)
            landed = price + ship
            if landed > 0 and (lowest == 0.0 or landed < lowest):
                lowest = landed
        summary_offers = (offers_payload.get("Summary", {}) or {}).get("NumberOfOffers", [])
        if not result["offer_count"]:
            result["offer_count"] = sum(
                o.get("OfferCount", 0) for o in summary_offers if o.get("condition","").lower() == "new")
        if lowest > 0:
            result["buy_box_price"] = lowest
            result["price_source"] = "lowest offer (no Buy Box winner)"
            return result
    except Exception as e:
        # Not fatal -- just try tier 3
        _msg = str(e)[:80]
        if "not found" not in _msg.lower() and "notfound" not in _msg.lower():
            console.print(f"  [yellow]Pricing (getItemOffers) failed: {_msg}[/yellow]")

    # --- Tier 3: retail/list price via getPricing ---
    # Some ASINs (permanently OOS, but still catalogued) still expose a list
    # price. This is a hint, not a real market price -- flag it so the caller
    # can decide whether to trust it.
    try:
        pricing = ProductPricing(credentials=creds, marketplace=MARKETPLACE, timeout=30)
        pp = pricing.get_product_pricing_for_asins(asin_list=[asin], item_condition="New")
        items = pp.payload if isinstance(pp.payload, list) else []
        for item in items:
            for offer in (item.get("Product", {}).get("Offers", []) or []):
                lp = float(offer.get("BuyingPrice", {}).get("LandedPrice", {}).get("Amount", 0) or 0)
                if lp > 0:
                    result["buy_box_price"] = lp
                    result["price_source"] = "list price (competitor may be OOS)"
                    return result
    except Exception:
        pass

    result["price_source"] = "no price found (no Buy Box, no offers, no list price)"
    return result


def get_fees(asin: str, price: float, creds: dict) -> dict:
    default = {"referral_fee": round(price * 0.15, 2),
               "variable_closing": 0.0,
               "total_amazon_fees": round(price * 0.15, 2),
               "fee_source": "estimated (15%)"}
    if not asin or price <= 0:
        return default
    try:
        fees_api = ProductFees(credentials=creds, marketplace=MARKETPLACE, timeout=30)
        _fee_cur = "USD" if MARKETPLACE_ID == US_MARKETPLACE_ID else "GBP"
        res = fees_api.get_product_fees_estimate_for_asin(
            asin=asin, price=price, currency=_fee_cur,
            is_fba=False, marketplace_id=MARKETPLACE_ID)
        fee_detail = (res.payload.get("FeesEstimateResult", {})
                      .get("FeesEstimate", {}).get("FeeDetailList", []))
        referral = var_closing = 0.0
        for fee in fee_detail:
            amt = float(fee.get("FinalFee", {}).get("Amount", 0))
            if fee.get("FeeType") == "ReferralFee":
                referral = amt
            elif fee.get("FeeType") == "VariableClosingFee":
                var_closing = amt
        return {"referral_fee": referral, "variable_closing": var_closing,
                "total_amazon_fees": round(referral + var_closing, 2),
                "fee_source": "SP-API (exact)"}
    except Exception as e:
        console.print(f"  [yellow]Fees API: {str(e)[:60]}[/yellow]")
        return default


# --- User's pricing rule -----------------------------------------------------
# Selling price = max(floor, competitor Buy Box)
# where floor = source_cost + Amazon fees + shipping_label(£3) + ads_margin(£2)
#              + min_profit(£1)
# If the competitor is HIGHER than the floor -> use competitor (more profit).
# If the competitor is LOWER or MISSING -> use the floor (never sell at a loss).
# Defaults come from constants below but can be overridden per-account/product.

PRICING_RULE_SHIPPING_LABEL = 3.00   # £ per unit -- Royal Mail Tracked 48 baseline
PRICING_RULE_ADS_MARGIN     = 2.00   # £ per unit -- estimated CPA / PPC budget
PRICING_RULE_MIN_PROFIT     = 1.00   # £ per unit -- absolute minimum to accept an order


def compute_selling_price(source_cost: float,
                          amazon_fees: float,
                          competitor_price: float,
                          shipping_label: float = PRICING_RULE_SHIPPING_LABEL,
                          ads_margin:     float = PRICING_RULE_ADS_MARGIN,
                          min_profit:     float = PRICING_RULE_MIN_PROFIT) -> dict:
    """Apply the user's pricing rule.

    Returns dict with:
      selling_price -- what to charge on the listing
      floor         -- the calculated cost-plus floor
      rule_source   -- 'competitor' (matched Buy Box) | 'floor' (used cost formula)
      breakdown     -- component list for the log line

    NOTE: Amazon fees are price-sensitive (referral fee is a % of selling price),
    so a naive floor with a fixed fee estimate under-prices. This is handled by
    the caller: it computes an initial fee at a reasonable seed price, calls this
    function, then re-fetches fees at the new price and calls this again once.
    Two passes is enough to converge for standard referral rates.
    """
    floor = round(source_cost + amazon_fees + shipping_label + ads_margin + min_profit, 2)
    competitor_price = float(competitor_price or 0)
    if competitor_price > floor:
        chosen = competitor_price
        source = "competitor (higher than floor)"
    else:
        chosen = floor
        source = "floor (competitor missing or below floor)"
    return {
        "selling_price": chosen,
        "floor":         floor,
        "rule_source":   source,
        "breakdown":     (f"cost {source_cost:.2f} + fees {amazon_fees:.2f} "
                          f"+ ship {shipping_label:.2f} + ads {ads_margin:.2f} "
                          f"+ profit {min_profit:.2f} = floor {floor:.2f}"),
    }


def calculate_financials(source_cost: float, selling_price: float,
                          shipping_cost: float, fees: dict) -> dict:
    total_costs = round(source_cost + shipping_cost + fees["total_amazon_fees"], 2)
    profit      = round(selling_price - total_costs, 2)
    margin      = round((profit / selling_price) * 100, 1) if selling_price > 0 else 0
    roi         = round((profit / source_cost)    * 100, 1) if source_cost  > 0 else 0
    return {
        "source_cost":       source_cost,
        "shipping_cost":     shipping_cost,
        "referral_fee":      fees["referral_fee"],
        "variable_closing":  fees["variable_closing"],
        "total_amazon_fees": fees["total_amazon_fees"],
        "total_costs":       total_costs,
        "selling_price":     selling_price,
        "profit":            profit,
        "margin_pct":        f"{margin}%",
        "roi_pct":           f"{roi}%",
        "viable":            "YES" if margin >= MIN_MARGIN else "LOW MARGIN",
        "fee_source":        fees["fee_source"],
    }


def get_product_type_schema(product_type: str, creds: dict, marketplace: str = None) -> dict:
    empty = {"required": {}, "optional": {}, "all": {}}
    if not product_type:
        return empty
    # Default to the ACTIVE marketplace (set by --marketplace) instead of a
    # hardcoded UK, so a US run fetches the US schema with the US locale.
    if not marketplace:
        marketplace = "US" if MARKETPLACE_ID == US_MARKETPLACE_ID else "UK"
    mkt = str(marketplace or "UK").upper()
    mkt_id = marketplace_id_for(mkt)
    locale = "en_US" if mkt == "US" else "en_GB"
    try:
        mkt_enum = Marketplaces.US if mkt == "US" else Marketplaces.UK
    except Exception:
        mkt_enum = MARKETPLACE

    def _parse(schema: dict) -> dict:
        props   = schema.get("properties", {})
        req_set = set(schema.get("required", []))
        skip    = {"purchasable_offer", "fulfillment_availability",
                   "main_offer_image_locator", "main_product_image_locator"}
        result  = {"required": {}, "optional": {}, "all": {}}
        for field, prop in props.items():
            if field in skip:
                continue
            items      = prop.get("items", {})
            item_props = items.get("properties", {}) if isinstance(items, dict) else {}
            val_prop   = item_props.get("value", {})
            allowed    = (val_prop.get("enum") or item_props.get("enum") or
                          items.get("enum") or prop.get("enum") or [])
            meta = {"allowed":     [str(a) for a in allowed[:20]],
                    "description": (prop.get("title") or field.replace("_", " ").title()),
                    "required":    field in req_set}
            result["all"][field] = meta
            if field in req_set:
                result["required"][field] = meta
            else:
                result["optional"][field] = meta
        return result

    for attempt in range(2):
        try:
            ptd  = ProductTypeDefinitions(credentials=creds, marketplace=mkt_enum)
            resp = ptd.get_definitions_product_type(
                productType=product_type, requirements="LISTING",
                requirementsEnforced="ENFORCED", locale=locale,
                marketplaceIds=[mkt_id])
            link = resp.payload.get("schema", {}).get("link", {}).get("resource", "")
            if not link:
                return empty
            # Download the schema JSON (large file, slow CDN) with a generous
            # timeout + retry so a transient stall doesn't fail the fetch.
            raw = None
            _derr = ""
            for _dl in range(3):
                try:
                    _sreq = urllib.request.Request(link, headers={"Accept": "application/json"})
                    with urllib.request.urlopen(_sreq, timeout=60) as r:
                        raw = json.loads(r.read().decode("utf-8"))
                    break
                except Exception as _de:
                    _derr = str(_de)
                    if _dl < 2:
                        time.sleep(2)
                        continue
            if raw is None:
                console.print(f"  [yellow]schema download slow -- using local attribute data ({_derr[:60]})[/yellow]")
                return empty
            result = _parse(raw)
            req = len(result["required"])
            opt = len(result["optional"])
            console.print(f"  Schema: {req} required + {opt} optional = {req+opt} attributes")
            return result
        except Exception as e:
            err = str(e)
            if attempt == 0 and any(x in err for x in ["11001", "getaddrinfo", "timeout"]):
                console.print("  [yellow]Schema DNS error -- retrying in 4s[/yellow]")
                time.sleep(4)
                continue
            _low_err = err.lower()
            if "unauthorized" in _low_err or "forbidden" in _low_err or "denied" in _low_err:
                console.print("  [dim]Product-type schema needs the Listings/Definitions "
                              "role (optional) -- continuing without it[/dim]")
            else:
                console.print(f"  [yellow]Schema unavailable: {err[:80]}[/yellow]")
            return empty
    return empty


# =============================================================================
# CRAWL4AI -- REVIEW SCRAPING
# =============================================================================

BROWSER_CFG = BrowserConfig(
    headless=True, verbose=False,
    headers={
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-GB,en;q=0.9",
    },
    # Pin the amazon.co.uk delivery location to a UK postcode BEFORE any page
    # loads. Without this, a non-UK visitor (e.g. from Pakistan) gets served a
    # location-fallback view: prices hidden, no Buy Box, "cannot ship to your
    # location" banners -- exactly why the scraper was returning thin data on
    # UK PDPs when SP-API fell back. SW1A 1AA = London postcode.
    cookies=[
        {"name": "lc-main",     "value": "en_GB", "domain": ".amazon.co.uk", "path": "/"},
        {"name": "i18n-prefs",  "value": "GBP",   "domain": ".amazon.co.uk", "path": "/"},
        {"name": "sp-cdn",      "value": "L5Z9:GB", "domain": ".amazon.co.uk", "path": "/"},
    ],
)

NOISE_RE = re.compile(
    r"(https?://|amazon|basket|account|sign.in|navigation|menu|header|footer|"
    r"breadcrumb|sponsored|advertisement|cookies|privacy|\.co\.uk|\.com|www\.|"
    r"customer.service|help|returns|orders|prime|javascript|void\(0\))",
    re.IGNORECASE
)

REVIEW_CSS = "[data-hook='review-body'], .review-text-content, [data-hook='review']"


async def _scrape(url: str, css: str = None, timeout: int = 25000,
                  delay: float = 2.0) -> str:
    run_cfg = CrawlerRunConfig(
        css_selector=css, word_count_threshold=15,
        remove_overlay_elements=True, exclude_external_links=True,
        page_timeout=timeout, delay_before_return_html=delay,
        excluded_tags=["nav", "header", "footer", "script", "style"] if not css else [],
    )
    async def _run():
        async with AsyncWebCrawler(config=BROWSER_CFG) as crawler:
            result = await crawler.arun(url=url, config=run_cfg)
            return (result.markdown or result.cleaned_html or "").strip()
    # Hard ceiling: the page_timeout above is crawl4ai-internal and can still
    # hang on browser launch/navigation. Kill the whole attempt a few seconds
    # past the page timeout so a stuck browser can never freeze the run.
    try:
        return await asyncio.wait_for(_run(), timeout=(timeout / 1000.0) + 8)
    except asyncio.TimeoutError:
        return ""


def _extract_reviews(content: str) -> list:
    reviews = []
    for line in content.split("\n"):
        line = line.strip()
        if len(line) < 30 or len(line) > 500:
            continue
        if NOISE_RE.search(line):
            continue
        if re.match(r"^[\d\*#]", line):
            continue
        if re.search(r"\d+\s*out\s*of\s*\d+\s*stars", line, re.IGNORECASE):
            continue
        if re.search(r"Reviewed in|Verified Purchase|people found", line, re.IGNORECASE):
            continue
        if sum(c.isalpha() or c == " " for c in line) / len(line) < 0.6:
            continue
        reviews.append(line)
    return reviews[:25]


async def scrape_pdp(asin: str) -> tuple:
    """
    Fallback when SP-API is unavailable. Scrapes Amazon UK PDP for the same
    fields SP-API would have returned. Returns (catalog_dict, pricing_dict).
    Both dicts have the same shape as get_competitor_asin_data and
    get_pricing_data return, so they're drop-in replacements.

    PROGRESS: each stage prints a labelled line so a stuck run tells you which
    stage is hung (browser launch / page fetch / parse) instead of silently
    sitting on '[FETCH]...' for minutes with no signal.
    """
    catalog_empty = {"asin": asin, "title": "", "brand": "", "product_type": "",
                     "item_type_keyword": "", "attributes": {}, "browse_nodes": [],
                     "sales_rank": [], "images": []}
    pricing_empty = {"buy_box_price": 0.0, "offer_count": 0}

    if not asin:
        return catalog_empty, pricing_empty

    t       = Timer()
    pdp_url = f"https://www.{_amazon_domain()}/dp/{asin}"
    console.print(f"  [cyan]  PDP scrape: launching browser + fetching {pdp_url}[/cyan]")

    # Background heartbeat: prints every 10s so a slow scrape doesn't look
    # frozen. Cancelled the instant _scrape returns.
    async def _heartbeat():
        elapsed = 0
        while True:
            try:
                await asyncio.sleep(10)
            except asyncio.CancelledError:
                return
            elapsed += 10
            console.print(f"  [dim]  PDP scrape still running ({elapsed}s elapsed, max ~40s)...[/dim]")

    _hb_task = asyncio.create_task(_heartbeat())
    try:
        content = await _scrape(pdp_url, css=None, timeout=30000, delay=3.0)
    except Exception as e:
        console.print(f"  [red]PDP scrape failed: {str(e)[:80]}[/red]")
        return catalog_empty, pricing_empty
    finally:
        _hb_task.cancel()
        try:
            await _hb_task
        except Exception:
            pass

    if not content:
        console.print(f"  [red]PDP scrape returned nothing ({t.elapsed()}s) -- browser timed out or was blocked[/red]")
        return catalog_empty, pricing_empty
    if len(content) < 500:
        console.print(f"  [yellow]PDP scrape returned too little content ({len(content)} chars, {t.elapsed()}s)[/yellow]")
        return catalog_empty, pricing_empty
    console.print(f"  [cyan]  PDP fetched {len(content)} chars in {t.elapsed()}s -- parsing...[/cyan]")

    # --- Title: first substantial non-link, non-UI line ------------------
    # Filters: skip Amazon's screen-reader hints and UI elements that
    # frequently appear at the top of PDP markdown.
    UI_NOISE_TITLE = re.compile(
        r"(keyboard shortcut|shift\s*\+|alt\s*\+|opt\s*\+|press\s+enter|"
        r"visit the|see more|see all|skip to|deliver to|back to|"
        r"close menu|department|all departments|hello sign|cart\b|"
        r"customer reviews|out of \d+ stars|prime\s*$|see (?:less|more) options)",
        re.IGNORECASE)
    title = ""
    for line in content.split("\n")[:120]:
        line = line.strip()
        if not (30 < len(line) < 220):
            continue
        if line.startswith(("#", "*", "[", "!", "-", ">", "|")):
            continue
        if UI_NOISE_TITLE.search(line):
            continue
        # Real product titles contain letters and at least one space; skip
        # lines that are mostly symbols or short kbd-shortcut-shaped tokens.
        letters = sum(c.isalpha() for c in line)
        if letters < 15:
            continue
        title = line
        break

    # --- Brand: "Visit the X Store" or "Brand: X" ------------------------
    brand = ""
    bm = re.search(r"Visit the ([A-Za-z0-9 &'\.-]+?) Store", content)
    if bm:
        brand = bm.group(1).strip()
    else:
        bm = re.search(r"\bBrand[:\s]+([A-Za-z0-9 &'\.-]{2,40})", content)
        if bm:
            brand = bm.group(1).strip()

    # --- Attributes: key:value lines from the spec tables ----------------
    attributes = {}
    for line in content.split("\n"):
        line = line.strip()
        kv = re.match(r"^\|?\s*([A-Z][A-Za-z &/\(\)-]{2,40})\s*[:\|]\s*(.{2,200})\s*\|?$", line)
        if kv:
            key = kv.group(1).strip().lower().replace(" ", "_").replace("/", "_")
            val = kv.group(2).strip().strip("|").strip()
            if val and val.lower() not in ("n/a", "none", "-", "see below") and not val.startswith("http"):
                attributes[key] = val[:200]

    # --- Bullets: lines starting with * or • (feature bullets) -----------
    bullets = []
    for line in content.split("\n"):
        line = line.strip()
        if line.startswith(("*", "•")) and 20 < len(line) < 500:
            clean = line.lstrip("*•- ").strip()
            if clean and not NOISE_RE.search(clean) and len(clean) > 25:
                bullets.append(clean)
    if bullets:
        attributes["bullet_points"] = " | ".join(bullets[:5])

    # --- Browse nodes: breadcrumb-shaped patterns ------------------------
    browse_nodes = []
    bc = re.search(r"([A-Z][A-Za-z &'-]{2,30}(?:\s*[›>]\s*[A-Z][A-Za-z &'-]{2,30}){1,5})", content)
    if bc:
        browse_nodes = [s.strip() for s in re.split(r"[›>]", bc.group(0)) if s.strip()]
    item_type_kw = browse_nodes[-1] if browse_nodes else ""

    # --- Sales rank ------------------------------------------------------
    sales_ranks = []
    for line in content.split("\n"):
        sr = re.search(r"#([\d,]+)\s*in\s*([A-Z][A-Za-z &'-]{3,60})", line)
        if sr:
            try:
                sales_ranks.append({"category": sr.group(2).strip(),
                                    "rank":     int(sr.group(1).replace(",", ""))})
            except ValueError:
                pass

    # --- Images: amazon media URLs ---------------------------------------
    images = []
    for m in re.finditer(r"https://[^\s\)\"\']+?\.(?:jpg|jpeg|png|webp)", content):
        url = m.group(0)
        if ("m.media-amazon" in url or "images-amazon" in url) and url not in images:
            images.append(url)
            if len(images) >= 5:
                break

    # --- Price: £ pattern ------------------------------------------------
    buy_box = 0.0
    _cur_sym = "$" if MARKETPLACE_ID == US_MARKETPLACE_ID else "£"
    pm = re.search(re.escape(_cur_sym) + r"\s*(\d{1,5}(?:[\.,]\d{2})?)", content)
    if pm:
        try:
            buy_box = float(pm.group(1).replace(",", "."))
        except ValueError:
            pass

    catalog = {"asin": asin, "title": title, "brand": brand, "product_type": "",
               "item_type_keyword": item_type_kw, "attributes": attributes,
               "browse_nodes": browse_nodes, "sales_rank": sales_ranks,
               "images": images}
    pricing = {"buy_box_price": buy_box, "offer_count": 0}

    console.print(f"  [yellow]PDP scrape ({t.elapsed()}s): "
                  f"title={'Y' if title else 'N'} | brand={'Y' if brand else 'N'} | "
                  f"{len(attributes)} attrs | {len(images)} imgs | price={_cur_code()}{buy_box:.2f}[/yellow]")

    return catalog, pricing


async def get_voc_data(asin: str, product_name: str, core_term: str) -> dict:
    result = {"reviews": [], "source": "none", "review_count": 0}
    t      = Timer()

    try:
        review_url = f"https://www.{_amazon_domain()}/product-reviews/{asin}?sortBy=recent&pageNumber=1"
        content    = await _scrape(review_url, css=REVIEW_CSS, delay=2.0)
        reviews    = _extract_reviews(content)
        blocks     = len(re.findall(r"Verified Purchase|Reviewed in", content, re.IGNORECASE))
        if reviews and blocks >= 3:
            console.print(f"  ({t.elapsed()}s) {len(reviews)} reviews from primary ASIN")
            return {"reviews": reviews[:20], "source": f"primary ({asin})",
                    "review_count": len(reviews)}
        else:
            console.print(f"  ({t.elapsed()}s) primary blocked or sparse -- trying search")
    except Exception as e:
        console.print(f"  ({t.elapsed()}s) primary scrape: {str(e)[:50]}")

    if t.elapsed() < 25:
        try:
            search_url  = f"https://www.{_amazon_domain()}/s?k={urllib.parse.quote(core_term)}&s=review-rank"
            s_content   = await _scrape(search_url,
                                         css="h2.a-size-mini span.a-text-normal",
                                         delay=1.5, timeout=20000)
            found_asins = list(dict.fromkeys(re.findall(r"/dp/([A-Z0-9]{10})", s_content)))
            found_asins = [a for a in found_asins if a != asin][:3]
            all_reviews = []
            for alt in found_asins:
                if t.elapsed() > 25:
                    break
                try:
                    alt_url     = f"https://www.{_amazon_domain()}/product-reviews/{alt}?sortBy=recent"
                    alt_content = await _scrape(alt_url, css=REVIEW_CSS, delay=1.5)
                    alt_reviews = _extract_reviews(alt_content)
                    if alt_reviews:
                        all_reviews.extend(alt_reviews[:8])
                except Exception:
                    continue
            if all_reviews:
                console.print(f"  ({t.elapsed()}s) {len(all_reviews)} reviews from category search")
                return {"reviews": all_reviews[:20], "source": f"category ({core_term})",
                        "review_count": len(all_reviews)}
        except Exception as e:
            console.print(f"  ({t.elapsed()}s) category search: {str(e)[:50]}")

    console.print(f"  ({t.elapsed()}s) No reviews scraped -- Claude uses category knowledge")
    return result


# =============================================================================
# KEYWORDS -- amazon.co.uk AUTOCOMPLETE
# =============================================================================

INFORMATIONAL = ["what is", "how does", "why is", "history of",
                  "difference between", "meaning of"]


def get_autocomplete_keywords(core_term: str) -> list:
    variations = [core_term, f"best {core_term}", f"{core_term} set",
                  f"{core_term} for", f"buy {core_term}"]
    headers    = {
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "en-GB,en;q=0.9",
        "Accept":          "application/json, text/javascript, */*",
    }
    seen, all_kws = set(), []
    for v in variations:
        enc = urllib.parse.quote(v)
        url = (f"https://completion.amazon.co.uk/search/complete"
               f"?method=completion&q={enc}&search-alias=aps&mkt=3&x=String")
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=5) as r:
                data = json.loads(r.read().decode("utf-8"))
            if isinstance(data, list) and len(data) > 1:
                for pos, s in enumerate(data[1]):
                    if isinstance(s, str) and 3 < len(s) < 100:
                        kl = s.lower().strip()
                        if kl not in seen and not any(inf in kl for inf in INFORMATIONAL):
                            seen.add(kl)
                            all_kws.append({"keyword":   kl,
                                            "vol_score": round(max(0, 1 - pos / 15), 2)})
        except Exception:
            continue
    all_kws.sort(key=lambda x: x["vol_score"], reverse=True)
    return all_kws[:30]


def extract_core_search_term(item_name: str) -> str:
    noise   = r"\b(\d+|pcs|pc|pack|piece|inch|lbs|lot|uk|usa|new|best|buy|get|the|and|with|for)\b"
    cleaned = re.sub(noise, "", item_name.lower(), flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    words   = [w for w in cleaned.split() if len(w) > 2][:5]
    return " ".join(words) if words else item_name[:30]


# =============================================================================
# CLAUDE -- LISTING GENERATION
# =============================================================================

SYSTEM_PROMPT = (
    "You are a senior Amazon UK marketplace specialist expert in:\n"
    "- Voice of Customer (VOC) copywriting from real review data\n"
    "- Amazon listing policy compliance (UK marketplace)\n"
    "- Consumer psychology and buyer demographics\n"
    "- Conversion-optimised product copy that targets pain points and purchase triggers\n\n"
    "You generate accurate, policy-compliant, conversion-focused Amazon UK listings.\n"
    "Use British English spelling throughout (colour, maximise, aluminium, organise, etc.).\n"
    "You always respond with valid JSON only -- no preamble, no markdown fences."
)


def _build_message_content(prompt: str, images: list) -> list:
    """Build Claude message content. Images are validated by magic bytes and
    size before being attached; bad images are skipped silently."""
    # Anthropic-supported formats
    MAGIC = {
        b"\xff\xd8\xff":           "image/jpeg",
        b"\x89PNG\r\n\x1a\n":      "image/png",
        b"GIF87a":                 "image/gif",
        b"GIF89a":                 "image/gif",
        b"RIFF":                   "image/webp",   # checked further below
    }
    MAX_IMG_BYTES = 4_500_000   # ~4.5 MB (Anthropic limit is 5 MB)
    MIN_IMG_BYTES = 1_000       # reject tiny 1x1 trackers / 0-byte responses

    content = []
    for img_url in images[:2]:
        if not img_url or not img_url.startswith("http"):
            continue
        try:
            req = urllib.request.Request(img_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=8) as r:
                img_bytes = r.read()
        except Exception:
            continue
        if not img_bytes or len(img_bytes) < MIN_IMG_BYTES:
            continue
        if len(img_bytes) > MAX_IMG_BYTES:
            continue
        # Detect actual media type by magic bytes (ignore URL extension - it lies)
        media_type = None
        for sig, mt in MAGIC.items():
            if img_bytes.startswith(sig):
                if sig == b"RIFF":
                    # WebP files: 'RIFF' + 4-byte size + 'WEBP'
                    if len(img_bytes) >= 12 and img_bytes[8:12] == b"WEBP":
                        media_type = "image/webp"
                else:
                    media_type = mt
                break
        if not media_type:
            continue
        img_b64 = base64.b64encode(img_bytes).decode("utf-8")
        content.append({"type": "image",
                         "source": {"type": "base64",
                                    "media_type": media_type,
                                    "data": img_b64}})
    if content:
        content.append({"type": "text",
                         "text": ("Above: product images of the competitor item. "
                                  "Use them to visually confirm: material, colour, "
                                  "handle material, finish type.\n\n" + prompt)})
    else:
        content.append({"type": "text", "text": prompt})
    return content


def build_prompt(comp_data: dict, pricing: dict, financials: dict,
                 keywords: list, voc_data: dict, schema: dict,
                 brand_name: str, manufacturer: str,
                 selling_price: float, handling_time: str,
                 item_name: str,
                 static_vv: dict = None) -> str:

    comp_brand   = comp_data.get("brand",        "")
    product_type = comp_data.get("product_type", "")
    cat_attrs    = comp_data.get("attributes",   {})
    sales_rank   = comp_data.get("sales_rank",   [])

    spec_lines    = [f"  {k}: {v}" for k, v in list(cat_attrs.items())[:200]
                     if v and str(v).strip().upper() not in ("N/A", "NONE", "")]
    specs_section = "\n".join(spec_lines) if spec_lines else "  (no specs available)"

    # Surface physical dimensions explicitly -- they may fall past the 30-spec cap
    # above, and Amazon flags dimension columns as required. Pull from the Amazon
    # catalogue first (now includes the structured dimensions block), then fall
    # back to eBay item specifics.
    _ebay_sp = {str(k).strip().lower(): v
                for k, v in ((comp_data.get("_ebay_supplement") or {}).get("item_specifics") or {}).items()}
    _cat_low = {str(k).strip().lower(): v for k, v in cat_attrs.items()}
    dim_lines = []
    for _label, _keys in (
        ("Item Length", ["item_length", "length", "length (longer edge)", "item length"]),
        ("Item Width",  ["item_width", "width", "item width"]),
        ("Item Height", ["item_height", "height", "depth", "item height"]),
        ("Item Weight", ["item_weight", "weight", "net weight", "item weight"]),
        ("Item Package Length", ["item_package_length", "package length"]),
        ("Item Package Width",  ["item_package_width", "package width"]),
        ("Item Package Height", ["item_package_height", "package height"]),
        ("Item Package Weight", ["item_package_weight", "package weight", "shipping weight"]),
    ):
        for _src in (_cat_low, _ebay_sp):
            _hit = next((str(_src[k]).strip() for k in _keys
                         if k in _src and str(_src[k]).strip()), None)
            if _hit:
                dim_lines.append(f"  {_label}: {_hit}")
                break
    dims_section = ("\nKNOWN PHYSICAL DIMENSIONS (from Amazon catalogue / eBay -- copy the "
                    "value AND unit VERBATIM into product_attributes; do NOT convert units. "
                    "Item = the product itself; Item Package = the shipping box):\n"
                    + "\n".join(dim_lines)) if dim_lines else ""

    # --- Optional eBay cross-reference data -----------------------------------
    ebay_supp = comp_data.get("_ebay_supplement") or {}
    if ebay_supp.get("title") or ebay_supp.get("item_specifics"):
        ebay_specs = ebay_supp.get("item_specifics") or {}
        ebay_spec_lines = [f"  {k}: {v}" for k, v in list(ebay_specs.items())[:200]]
        ebay_section = (
            "\nEBAY CROSS-REFERENCE (same product listed on eBay UK -- use to validate "
            "and fill attribute gaps from Amazon competitor data, but Amazon data takes "
            "precedence where they conflict):"
            f"\nEBAY TITLE: {ebay_supp.get('title','')[:200]}"
            f"\nEBAY PRICE: {ebay_supp.get('price','')}"
            f"\nEBAY CONDITION: {ebay_supp.get('condition','')}"
            f"\nEBAY CATEGORY: {ebay_supp.get('category_path','')}"
            f"\nEBAY ITEM SPECIFICS:\n" + ("\n".join(ebay_spec_lines) if ebay_spec_lines else "  (none)")
        )
        if ebay_supp.get("condition_description"):
            ebay_section += f"\nEBAY CONDITION NOTES: {ebay_supp['condition_description']}"
        if ebay_supp.get("description"):
            ebay_section += f"\nEBAY DESCRIPTION: {ebay_supp['description']}"
    else:
        ebay_section = ""

    bsr_line = ""
    if sales_rank:
        top      = sales_rank[0]
        bsr_line = f"\nBSR: #{top['rank']:,} in {top['category']}"

    fin_line = (
        f"Buy Box: {_cur_code()}{pricing['buy_box_price']:.2f} | "
        f"Our Price: {_cur_code()}{financials['selling_price']:.2f} | "
        f"Fees: {_cur_code()}{financials['total_amazon_fees']:.2f} ({financials['fee_source']}) | "
        f"Profit: {_cur_code()}{financials['profit']:.2f} | "
        f"Margin: {financials['margin_pct']} | ROI: {financials['roi_pct']}"
    )

    kw_lines   = [f"  {kw['keyword']} (vol_score: {kw['vol_score']})" for kw in keywords[:15]]
    kw_section = (
        "\nAUTOCOMPLETE KEYWORDS (real Amazon UK searches, ranked by frequency):\n"
        + "\n".join(kw_lines)
    ) if kw_lines else ""

    if voc_data["reviews"]:
        reviews_text = "\n".join(f'- "{r}"' for r in voc_data["reviews"][:20])
        voc_section  = (
            f"\nREAL CUSTOMER REVIEWS (source: {voc_data['source']} | "
            f"{voc_data['review_count']} reviews):\n---\n"
            f"{reviews_text}\n---\n\n"
            "From these reviews identify and USE:\n"
            "1. TOP COMPLAINTS -- address them directly in bullets\n"
            "2. TOP PRAISE -- amplify in bullets\n"
            "3. EXACT CUSTOMER LANGUAGE -- mirror their words\n"
            "4. PURCHASE CONTEXT -- why did they buy?\n"
            "5. DEMOGRAPHIC SIGNALS -- lifestyle, age, situation clues"
        )
    else:
        voc_section = "\nREVIEWS: None scraped. Use category knowledge for VOC."

    req_fields  = list(schema.get("required", {}).keys())
    req_section = (
        f"\nREQUIRED ATTRIBUTES for {product_type}: {', '.join(req_fields)}"
        if req_fields else ""
    )
    opt_fields  = list(schema.get("optional", {}).keys())[:35]
    opt_section = (
        "\nOPTIONAL ATTRIBUTES (fill from confirmed specs):\n" + ", ".join(opt_fields)
    ) if opt_fields else ""

    # ENFORCED VALUES: prefer Amazon's LIVE schema enums (authoritative +
    # complete) for each attribute Claude writes; fall back to the static XLSM
    # list only when the live schema doesn't ship an enum for that field. This
    # is what stops Claude from inventing a value Amazon will reject.
    enforced_section = ""
    relevant = [
        "material", "material_type_free", "blade_material", "handle_material",
        "closure_material", "base_material", "shade_material",
        "colour", "color", "colour_map", "color_map", "shade_color",
        "size", "size_name", "size_map",
        "pattern", "pattern_name", "style", "style_name",
        "shape", "shape_name", "finish", "finish_type",
        "fabric_type", "outer_material", "inner_material",
        "target_gender", "department_name", "age_range_description",
        "scent_name", "flavour", "flavor",
        "included_components", "compatible_devices",
        "is_oven_safe", "is_dishwasher_safe", "is_microwave_safe",
        "with_lid", "special_features", "special_feature",
        "item_condition", "condition_type", "country_of_origin",
        "skin_type", "hair_type", "skin_tone",
        "form_factor", "connectivity_technology",
        "power_source_type", "lighting_method",
        "occasion_type",
    ]
    _live_all = schema.get("all", {}) if isinstance(schema, dict) else {}
    _static_pt = (static_vv or {}).get(product_type, {}) if static_vv else {}
    lines = []
    for k in relevant:
        vals = []
        # 1) live schema enum (ground truth)
        meta = _live_all.get(k)
        if meta and meta.get("allowed"):
            vals = list(meta["allowed"])[:20]
        # 2) fallback to static XLSM list
        elif k in _static_pt and _static_pt[k]:
            vals = list(_static_pt[k])[:20]
        if vals:
            lines.append(f"  {k}: " + " | ".join(str(v) for v in vals))
    if lines:
        enforced_section = (
            f"\nAMAZON-ENFORCED ATTRIBUTE VALUES for {product_type} "
            "(these are Amazon's OWN allowed values from the live schema -- use ONLY "
            "these EXACT strings, case-sensitive; Amazon rejects anything else. If none "
            "fits the product, omit that attribute rather than inventing a value):\n"
            + "\n".join(lines)
        )

    # SAFETY & COMPLIANCE -- pull the EXACT accepted strings straight from the
    # live SP-API schema (authoritative + complete, unlike the cached XLSM
    # lists) for the columns Amazon marks required under Safety & Compliance.
    # Claude must answer each; for an ordinary non-hazardous retail product the
    # answer is the "Not Applicable" / "No" option.
    schema_all   = schema.get("all", {})
    safety_keys  = [
        "supplier_declared_dg_hz_regulation", "contains_liquid_contents",
        "ghs", "ghs_classification_class", "hazmat",
        "batteries_required", "batteries_included",
        "supplier_declared_material_regulation",
    ]
    safety_lines = []
    for k in safety_keys:
        meta = schema_all.get(k)
        if meta and meta.get("allowed"):
            safety_lines.append(f"  {k}: " + " | ".join(meta["allowed"][:20]))
    safety_section = ""
    if safety_lines:
        safety_section = (
            f"\nSAFETY & COMPLIANCE FIELDS for {product_type} (Amazon REQUIRES a value in each "
            "of these columns -- output a key for EVERY one below, choosing ONLY an exact string "
            "shown; for an ordinary non-hazardous, non-battery, non-liquid retail item pick the "
            "option meaning Not Applicable / No):\n"
            + "\n".join(safety_lines)
            + "\nGUIDANCE: supplier_declared_dg_hz_regulation -> the value meaning 'Not Applicable' "
            "unless the product is genuinely a regulated dangerous good (aerosol, flammable liquid, "
            "lithium battery, pressurised, corrosive). contains_liquid_contents -> 'No' unless the "
            "product itself holds/ships a liquid. batteries_required / batteries_included -> "
            "'No' unless it runs on batteries. ghs / hazmat -> the Not-Applicable / none option."
        )

    return (
        f"MARKETPLACE: Amazon UK\n"
        f"BRAND: {brand_name}\n"
        f"MANUFACTURER: {manufacturer}\n"
        f"PRODUCT TYPE: {product_type}\n"
        f"COMPETITOR TITLE: {comp_data.get('title', '')}\n"
        f"LABEL FROM INPUT: {item_name}\n"
        f"HANDLING TIME: {handling_time}\n"
        f"FINANCIALS: {fin_line}{bsr_line}\n"
        f"\nCOMPETITOR SPECS:\n{specs_section}"
        f"{dims_section}"
        f"{ebay_section}"
        f"{kw_section}"
        f"{voc_section}"
        f"{req_section}"
        f"{opt_section}"
        f"{enforced_section}\n"
        f"{safety_section}\n"
        "\n===================================\n"
        "ACCURACY RULES -- MANDATORY\n"
        "===================================\n"
        f"- Only state specs confirmed in competitor data -- never invent\n"
        f"- If a spec is uncertain -- write N/A\n"
        f"- Brand is always \"{brand_name}\" -- ignore competitor brand \"{comp_brand}\"\n"
        f"- No health/medical claims without certification\n"
        f"- Country of origin: CN (China)\n"
        f"- Use British English spelling throughout\n"
        "\n===================================\n"
        "INTELLECTUAL PROPERTY RULES -- ABSOLUTE\n"
        "===================================\n"
        f"- DO NOT mention ANY brand name except \"{brand_name}\". Not in title, not in bullets, not in description, not in search terms.\n"
        "- DO NOT use comparative phrases referencing other brands or products:\n"
        "  forbidden: \"compatible with X\", \"replacement for X\", \"alternative to X\",\n"
        "  \"equivalent to X\", \"works with X\", \"better than X\", \"same as X\",\n"
        "  \"OEM approved\", \"factory approved\", \"meets X specifications\".\n"
        "- DO NOT mention trademarked specification codes (e.g. ANY \"J-20C\", \"M2C134\", \"MAT 3505\", \"TES-295\", \"229.51\").\n"
        "- Refer to compatibility generically: \"for kitchen use\", \"for outdoor applications\", \"for industrial settings\".\n"
        "- Use generic descriptive nouns -- write \"cookware set\" not \"replacement cookware set for Tefal\".\n"
        "- Use industry standard codes ONLY (ISO, SAE, AGMA, NLGI, CE, UKCA, USB, LED) -- these are generic.\n"
        "- Materials, colours, and dimensions are FINE -- they are generic, not brand claims.\n"
        "\n===================================\n"
        "AMAZON UK COMPLIANCE\n"
        "===================================\n"
        "CONTENT POLICY -- CLAIMS (Amazon rejects listings that break these):\n"
        "- REALISTIC SPECS ONLY. Even if the source/eBay/competitor data states a\n"
        "  number, DO NOT repeat physically implausible or exaggerated specs. Sanity-\n"
        "  check every number against reality for the product category. If a spec is\n"
        "  implausible, OMIT the number and use a qualitative phrase instead\n"
        "  (e.g. 'bright, high-output beam' rather than a specific lumen figure).\n"
        "  Category sanity caps (typical real ranges -- never exceed without proof):\n"
        "    * flashlight/torch lumens: realistically 100-5000; NEVER write 5-6 digit\n"
        "      lumens (e.g. 100,000 lumens is impossible for a handheld torch -- omit).\n"
        "    * beam distance: a few hundred metres at most for handheld; do not claim km.\n"
        "    * battery life / runtime: state only if plausible and source-supported.\n"
        "- NO CALLS TO ACTION. Forbidden anywhere in title/bullets/description:\n"
        "  'add to basket', 'add to cart', 'buy now', 'order today', 'shop now',\n"
        "  'don't miss out', 'get yours', 'limited time', 'sale', 'hurry'.\n"
        "- NO EXTERNAL LINKS or site references (no URLs, no 'visit our store',\n"
        "  no 'see our website', no other-platform mentions).\n"
        "- NO SELLER SELF-PROMOTION / TRUST CLAIMS in bullets or description: no\n"
        "  'sold by', 'sold exclusively by', 'quality you can trust', 'trusted brand',\n"
        "  'we/our', or any company plug. Describe the PRODUCT, not the seller/brand.\n"
        "- NO SHIPPING / FULFILMENT / AVAILABILITY claims anywhere in title/bullets/\n"
        "  description: no 'fast/free/next-day delivery', 'handling', 'dispatch',\n"
        "  'despatch', 'ships from', 'processed promptly', 'in stock', 'arrives in ...\n"
        "  condition', 'unboxing experience', 'without a long wait'. Amazon rejects these.\n"
        "- NO UNVERIFIABLE SUPERLATIVES / QUALITY CLAIMS / GUARANTEES: 'best', '#1',\n"
        "  'number one', 'world's most powerful', 'guaranteed', 'lifetime guarantee',\n"
        "  '100% satisfaction', 'perfect', 'ultimate', 'premium', 'high-quality',\n"
        "  'top-quality', 'quality you can trust', 'impressive', 'pristine', 'exclusive',\n"
        "  'exclusively' -- unless objectively certifiable.\n"
        "- NO false/medical/efficacy claims; describe features and honest benefits.\n"
        "- It is BETTER to omit a doubtful detail than to risk a policy rejection.\n"
        "\n"
        f"TITLE (target 70-75 chars, fill close to the 75 cap): Do NOT start with the brand name \"{brand_name}\" or any brand; lead with the product's key descriptive terms (type, key feature, material/size). Front-load the most important keywords in the first ~70 chars (mobile truncates there). No ALL CAPS, no !, no &, no ~\n"
        f"ITEM HIGHLIGHTS (target 110-125 chars, fill close to the 125 cap): one short standout-feature line summarising the product's key selling point.\n"
        f"BULLETS (5, target 400-500 chars EACH, fill them fully): Sentence case (capitalise only the FIRST word and proper feature names). NEVER use Title Case mid-sentence. Benefit + verified spec, no health claims. Amazon indexes only the first ~1000 bytes across ALL 5 bullets combined, so put the most important keywords in the first 1-2 bullets, then keep adding genuine detail to fill each bullet.\n"
        f"DESCRIPTION (HTML, target 1700-2000 chars incl tags, fill it richly): <p> <ul> <li> <b> <br> only, max 2000 characters including the HTML tags\n"
        f"SEARCH TERMS (fill close to 249 BYTES without exceeding): single spaces between words, NO punctuation (no commas/semicolons), lowercase, no title repeats, no brand names. Exceeding 249 bytes voids the whole field.\n"
        f"UNIT COUNT: plain integer only (e.g. 7 for a 7-piece set)\n"
        "\n===================================\n"
        "PRODUCT ATTRIBUTES -- FILL COMPREHENSIVELY (read the attached images)\n"
        "===================================\n"
        "Product images ARE attached above. Examine them together with the title and specs,\n"
        "then populate \"product_attributes\" with EVERY attribute that applies to this product:\n"
        "- Include a key for each attribute named in AMAZON-ENFORCED ATTRIBUTE VALUES above,\n"
        "  using ONLY the exact accepted string shown (case-sensitive).\n"
        "- Include a key for EVERY attribute named in SAFETY & COMPLIANCE FIELDS above, picking the\n"
        "  exact string that means Not Applicable / No for an ordinary non-hazardous product.\n"
        "- ALSO include when applicable: number_of_items, included_components, unit_count,\n"
        "  unit_count_type, number_of_boxes, country_of_origin, is_fragile, item_shape,\n"
        "  special_features, item_condition, colour, material, size.\n"
        "HOW TO DETERMINE VALUES:\n"
        "- number_of_items: COUNT from the title/image (e.g. \"15-in-1\" -> 15; a visible set of 4 -> 4).\n"
        "- included_components: list what is actually shown in the box in the images.\n"
        "- colour / material / item_shape: read from the images when not stated in text.\n"
        "IRRELEVANCE (must stay upload-safe):\n"
        "- A free-text attribute that does not apply -> \"N/A\".\n"
        "- An enumerated attribute that does not apply -> use a Not-Applicable accepted value if one is\n"
        "  listed above; otherwise OMIT the key. NEVER write \"N/A\" into an enumerated attribute.\n"
        "DEFAULTS when genuinely undeterminable: is_fragile=\"No\", item_condition=\"New\", country_of_origin=\"CN\".\n"
        "Inference of observable facts and standard category benefits is encouraged; NEVER invent\n"
        "specific performance numbers, certifications, or efficacy claims.\n"
        "\n===================================\n"
        "OUTPUT -- valid JSON only, no markdown\n"
        "===================================\n"
        "{{\n"
        '  "amazon_category":      "Top-level Amazon UK category",\n'
        '  "amazon_subcategory":   "Specific subcategory",\n'
        '  "target_demographic":   "1-2 sentences: who buys this and why",\n'
        '  "pain_points":          "Top 2-3 pain points from reviews or category knowledge",\n'
        '  "purchase_trigger":     "Primary reason customer buys this product",\n'
        '  "title":                "[product] [key feature] [material or size] -- max 75 chars incl spaces, do NOT include any brand name",\n'
        '  "item_highlights":      "single standout-feature line, max 125 chars",\n'
        '  "bullet_1":             "BENEFIT: verified spec detail",\n'
        '  "bullet_2":             "BENEFIT: verified spec detail",\n'
        '  "bullet_3":             "BENEFIT: verified spec detail",\n'
        '  "bullet_4":             "BENEFIT: verified spec detail",\n'
        '  "bullet_5":             "BENEFIT: verified spec detail",\n'
        '  "description":          "<p>Hook.</p>\\n<ul>\\n<li><b>Feature:</b> Detail.</li>\\n</ul>\\n<p>Close.</p>",\n'
        '  "search_terms":         "keyword1 keyword2 keyword3 -- single spaces no punctuation max 249 bytes",\n'
        '  "material":             "Exact material from specs or N/A",\n'
        '  "colour":               "Exact colour from specs or N/A",\n'
        '  "size":                 "Exact dimensions/count from specs or N/A",\n'
        '  "number_of_items":      "Integer count (e.g. 7) or N/A",\n'
        '  "target_gender":        "Male / Female / Unisex",\n'
        '  "age_range":            "All Ages / Adult / Child",\n'
        '  "compliance_notes":     "Battery / electrical / chemical flags or None",\n'
        '  "product_attributes": {{\n'
        '    "material":            "exact accepted string or N/A",\n'
        '    "color":               "exact accepted string or N/A",\n'
        '    "number_of_items":     7,\n'
        '    "included_components": "what is in the box, read from the images",\n'
        '    "unit_count":          "numeric value or N/A",\n'
        '    "unit_count_type":     "Count or other accepted unit, or N/A",\n'
        '    "number_of_boxes":     1,\n'
        '    "is_fragile":          "No",\n'
        '    "item_shape":          "shape from image or N/A",\n'
        '    "special_features":    "from specs/images or N/A",\n'
        '    "item_condition":      "New",\n'
        '    "item_type_keyword":   "exact lowercase-hyphenated string e.g. cookware-sets",\n'
        '    "country_of_origin":   "CN",\n'
        '    "supplier_declared_dg_hz_regulation": "Not Applicable",\n'
        '    "contains_liquid_contents":           "No",\n'
        '    "batteries_required":                 "No",\n'
        '    "batteries_included":             "No",\n'
        '    "item_length":           "value + unit VERBATIM from KNOWN PHYSICAL DIMENSIONS, else omit",\n'
        '    "item_width":            "value + unit VERBATIM, else omit",\n'
        '    "item_height":           "value + unit VERBATIM, else omit",\n'
        '    "item_weight":           "value + unit VERBATIM, else omit",\n'
        '    "item_package_length":   "package value + unit VERBATIM, else omit",\n'
        '    "item_package_width":    "package value + unit VERBATIM, else omit",\n'
        '    "item_package_height":   "package value + unit VERBATIM, else omit",\n'
        '    "item_package_weight":   "package value + unit VERBATIM, else omit"\n'
        "  }}\n"
        "}}"
    )


def _scrub_listing_claims(listing: dict) -> dict:
    """SAFETY NET (runs even if the AI ignores the prompt rules): strip the
    Amazon-policy-violating patterns from the generated copy BEFORE it is saved.
    Catches: implausible lumen specs, calls-to-action, external links, and
    unverifiable superlatives. Conservative -- edits wording, never deletes a
    whole field, so the listing stays usable. Logs what it changed."""
    if not isinstance(listing, dict):
        return listing
    import re as _re
    text_keys = ["title", "item_name", "item_highlights",
                 "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5",
                 "description", "product_description"]
    changed = []

    def _fix(s):
        if not isinstance(s, str) or not s.strip():
            return s
        orig = s
        s = _re.sub(r"\b\d{1,3}(?:,\d{3})+\s*(?:-)?\s*lumens?\b", "high-output", s, flags=_re.I)
        s = _re.sub(r"\b\d{5,}\s*(?:-)?\s*lumens?\b", "high-output", s, flags=_re.I)
        s = _re.sub(r"\bup to high-output\b", "high-output", s, flags=_re.I)
        s = _re.sub(r"\b(?:up to\s*)?\d{1,3}(?:,\d{3})+\s*(?:ft|feet|m|metres|meters)\b",
                    "a long distance", s, flags=_re.I)
        s = _re.sub(r"\b(?:up to\s*)?\d{3,}\s*(?:km|kilometres|kilometers)\b",
                    "a long distance", s, flags=_re.I)
        cta = (r"\b(add to (?:basket|cart)|buy now|order (?:now|today)|shop now|"
               r"don'?t miss out|get yours|limited time|hurry|while stocks last|"
               r"add to your basket today)\b[^.!]*[.!]?")
        s = _re.sub(cta, "", s, flags=_re.I)
        s = _re.sub(r"https?://\S+", "", s)
        s = _re.sub(r"\b(visit our (?:store|website|shop)|see our website)\b[^.!]*[.!]?",
                    "", s, flags=_re.I)
        s = _re.sub(r"\b(world'?s most powerful|#\s*1|number one|guaranteed|"
                    r"lifetime guarantee|100%\s*satisfaction)\b", "reliable", s, flags=_re.I)
        s = _re.sub(r"\s{2,}", " ", s).strip()
        s = _re.sub(r"\s+--\s*$", "", s).strip()
        s = _re.sub(r"--\s*--", "--", s)
        return s if s != orig else orig

    for k in text_keys:
        if k in listing and isinstance(listing[k], str):
            new = _fix(listing[k])
            if new != listing[k]:
                listing[k] = new
                changed.append(k)
    if changed:
        try:
            console.print(f"  [yellow]content scrubber adjusted: {', '.join(changed)} "
                          f"(removed fake specs / CTAs / links / superlatives)[/yellow]")
        except Exception:
            pass
    return listing


def generate_listing(client, comp_data: dict, pricing: dict, financials: dict,
                     keywords: list, voc_data: dict, schema: dict,
                     brand_name: str, manufacturer: str,
                     selling_price: float, handling_time: str,
                     item_name: str, static_vv: dict = None) -> dict:
    t       = Timer()
    prompt  = build_prompt(comp_data, pricing, financials, keywords, voc_data,
                           schema, brand_name, manufacturer, selling_price,
                           handling_time, item_name, static_vv)
    images  = comp_data.get("images", [])
    content = _build_message_content(prompt, images)

    console.print("    Sending to Claude...", end=" ")
    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=5000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content}],
        )
    except Exception as e:
        # Print the full error so we know what Anthropic rejected
        err_str = str(e)
        console.print(f"[red]Failed[/red] ({t.elapsed()}s)")
        console.print(f"      Full error: {err_str[:500]}")
        # If images were attached and the error mentions image/messages/source,
        # retry text-only -- this rescues most "bad image" 400s.
        has_images = any(c.get("type") == "image" for c in content)
        if has_images and ("image" in err_str.lower() or "messages" in err_str.lower()
                            or "source" in err_str.lower() or "400" in err_str):
            console.print(f"      Retrying text-only (without images)...", end=" ")
            t2 = Timer()
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=5000,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": [{"type": "text", "text": prompt}]}],
            )
            console.print(f"Done ({t2.elapsed()}s)")
        else:
            raise
    else:
        console.print(f"Done ({t.elapsed()}s)")

    raw = response.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.MULTILINE).strip()
    try:
        return _scrub_listing_claims(json.loads(raw))
    except json.JSONDecodeError:
        for end in range(len(raw), max(len(raw) - 300, 0), -1):
            try:
                candidate = raw[:end]
                if not candidate.endswith("}"):
                    candidate += "}"
                return _scrub_listing_claims(json.loads(candidate))
            except Exception:
                continue
        raise


# =============================================================================
# GOOGLE SHEETS
# =============================================================================

FIXED_HEADERS = [
    "Competitor ASIN",       # 0
    "Source URL",            # 1
    "UPC",                   # 2  (column stores EAN or UPC; either is accepted)
    "SKU",                   # 3
    "Platform",              # 4
    "Buy Box Price (GBP)",   # 5
    "Our Price (GBP)",       # 6
    "Amazon Fees (GBP)",     # 7
    "Fee Source",            # 8
    "Profit (GBP)",          # 9
    "Margin %",              # 10
    "ROI %",                 # 11
    "Viable?",               # 12
    "Product Type",          # 13
    "Amazon Category",       # 14
    "Subcategory",           # 15
    "VOC Source",            # 16
    "VOC Review Count",      # 17
    "Target Demographic",    # 18
    "Pain Points",           # 19
    "Purchase Trigger",      # 20
    "Title",                 # 21
    "Bullet 1",              # 22
    "Bullet 2",              # 23
    "Bullet 3",              # 24
    "Bullet 4",              # 25
    "Bullet 5",              # 26
    "Description (HTML)",    # 27
    "Search Terms / KW",     # 28
    "Autocomplete Keywords", # 29
    "Material",              # 30
    "Colour",                # 31
    "Size",                  # 32
    "Number of Items",       # 33
    "Target Gender",         # 34
    "Age Range",             # 35
    "Compliance Notes",      # 36
    "Handling Time",         # 37
    "Handling Days",         # 38
    "Status",                # 39
    "Date Processed",        # 40
    "Brand",                 # 41  - chosen per-product or globally
    "Model Number",          # 42  - blank if category does not require it
    "Notes",                 # 43  - duplicate ASIN warnings, compliance flags, IP findings
    "Compliance Risk",       # 44  - HIGH / MEDIUM / BASELINE / "" - filter to spot regulated items
    "IP Risk",               # 45  - HIGH / "" - filter to spot brand-name / comparative-claim violations
    "Attributes JSON",       # 46  - full per-type attribute object (JSON) for unified export; ignored by legacy export
    "Item Highlights",       # 47  - short standout-feature line (max 125 chars), app-wide
    "API Payload JSON",      # 48  - EXACT body sent to Amazon on the last Preview/Submit (read-only debug view)
]
FIXED_COUNT = len(FIXED_HEADERS)

# --- shared word-count / format helpers (app-wide listing limits) -----------
TITLE_MAX_CHARS   = 75     # incl spaces
HIGHLIGHTS_MAX    = 125
BULLET_MAX_CHARS  = 500
DESC_MAX_CHARS    = 2000   # incl HTML tags
SEARCH_TERMS_MAX_BYTES = 249

def clean_search_terms(st: str) -> str:
    """Backend search terms: strip ALL punctuation, collapse to single spaces,
    lowercase, then byte-cap at 249 (Amazon ignores the whole field if over).
    Spaces are kept (Amazon tokenises on them); only punctuation is removed."""
    if not st:
        return ""
    import re as _re
    # replace any punctuation/separators with a space, then collapse spaces
    st = _re.sub(r"[^\w\s]", " ", st, flags=_re.UNICODE)
    st = _re.sub(r"\s+", " ", st).strip().lower()
    b = st.encode("utf-8")
    if len(b) > SEARCH_TERMS_MAX_BYTES:
        st = b[:SEARCH_TERMS_MAX_BYTES].decode("utf-8", "ignore")
        # don't end mid-word
        if " " in st:
            st = st[:st.rfind(" ")].strip()
    return st

def cap_chars(s: str, n: int) -> str:
    """Trim a string to n characters without an ellipsis hack when possible."""
    s = s or ""
    if len(s) <= n:
        return s
    cut = s[:n]
    if " " in cut[: n]:
        cut = cut[: cut.rfind(" ")].rstrip()
    return cut

COUNTER_PATH      = CONFIG_PATH.parent / "model_number_counter.json"
COMPLIANCE_PATH   = Path(__file__).parent / "compliance_rules.json"
IP_RULES_PATH     = Path(__file__).parent / "ip_rules.json"
VALID_VALUES_PATH = Path(__file__).parent / "valid_values.json"


# =============================================================================
# SKU / BRAND / MODEL NUMBER HELPERS
# =============================================================================

def build_sku(source_cost: float, handling_days: str, comp_asin: str,
              taken_skus: set) -> tuple:
    """
    SKU format: {source_price}_{N}Days_{COMP_ASIN}
    e.g. 7.99_3Days_B0XYZ12345
    If the resulting SKU is already taken in this run or previous runs,
    append _2, _3, etc. Returns (sku, was_duplicate).
    """
    price_part = f"{source_cost:.2f}" if source_cost > 0 else "0.00"
    days_part  = f"{handling_days}Days" if handling_days else "3Days"
    base       = f"{price_part}_{days_part}_{comp_asin}"
    if base not in taken_skus:
        return base, False
    n = 2
    while f"{base}_{n}" in taken_skus:
        n += 1
    return f"{base}_{n}", True


def prompt_for_brand() -> str:
    """Ask user once for the brand to use across this run.
    Empty string = auto-pick per category from schema."""
    try:
        entered = input("Enter brand name (or press Enter to auto-pick per category): ").strip()
    except EOFError:
        entered = ""
    return entered


# pick_brand_for_product moved to listing/brand_validator.py in Phase 5 (self-contained;
# behaviour unchanged). Imported so amazon_listing_generator.pick_brand_for_product resolves.
from listing.brand_validator import pick_brand_for_product


def derive_product_type_code(product_type: str) -> str:
    """First 3 letters of product_type, uppercased. e.g. COOKWARE_SET -> COO."""
    if not product_type:
        return "GEN"
    cleaned = re.sub(r"[^A-Za-z]", "", product_type).upper()
    return (cleaned[:3] or "GEN")


# Keyword -> Amazon product_type inference. Used ONLY when SP-API and the PDP
# scrape both fail to provide a product type (e.g. the SP-API app lacks the
# Catalog Items role). Ordered: the FIRST matching rule wins, so put more
# specific patterns before generic ones.
_PT_INFER_RULES = [
    (r"\bflash\s?light|\btorch\b|\bhead\s?lamp|\blantern\b|\bwork\s?light", "FLASHLIGHT"),
    (r"\bstring\s?light|\bfairy\s?light|\bfestoon", "STRING_LIGHT"),
    (r"\bdesk\s?lamp|\btable\s?lamp|\bfloor\s?lamp|\bbedside\s?lamp", "LAMP"),
    (r"\bceiling\s?light|\bwall\s?light|\bpendant|\bchandelier|\bsconce", "LIGHT_FIXTURE"),
    (r"\bbulb|\bled\s?light|\blighting|\blamp\b", "HOME_LIGHTING_AND_LAMPS"),
    (r"\bsecurity\s?camera|\bcctv|\bsurveillance|\bdoorbell\s?cam", "SECURITY_CAMERA"),
    (r"\bknife|\bknives|\bcleaver|\bchef'?s?\s?knife", "KITCHEN_KNIFE"),
    (r"\bcookware|\bpan\s?set|\bpot\s?set|\bsaucepan", "COOKWARE_SET"),
    (r"\bspatula|\bturner\b", "FOOD_SPATULA"),
    (r"\bglobe\b", "GLOBE"),
    (r"\bserum|\bcleanser|\bmoisturi|\bsunscreen|\bskincare|\bskin\s?care|\bcosmetic|\bface\s?cream", "BEAUTY"),
    (r"\bsupplement|\bvitamin|\bnebuli|\binhaler|\bthermometer|\bblood\s?pressure", "HEALTH_PERSONAL_CARE"),
    (r"\bhelmet|\bknee\s?pad|\belbow\s?pad|\bprotective\s?gear|\bguard\b", "POWERSPORTS_PROTECTIVE_GEAR"),
    (r"\bnet\b|\bgoal\s?net|\bsports\s?net", "SPORT_NET"),
    (r"\btarget\b|\bdart\s?board|\barchery", "SPORT_TARGET"),
    (r"\bdrill|\bwrench|\bscrewdriver|\bplier|\bhand\s?tool|\bpower\s?tool", "TOOLS"),
    (r"\bscrew|\bbolt|\bnut\b|\bbracket|\bhinge|\bfastener|\bhardware", "HARDWARE"),
    (r"\bart\s?(kit|set)|\bcraft\s?(kit|set)|\bpainting\s?set", "ART_CRAFT_KIT"),
    (r"\bfigure\b|\baction\s?figure|\bcollectible|\bfigurine", "TOY_FIGURE"),
]


def infer_product_type(comp_data: dict, item_name: str = "",
                       valid_types: dict = None) -> str:
    """Best-effort product type when none came from SP-API or the scrape.
    Matches keywords from the title + item_type_keyword + breadcrumbs against
    known Amazon types. Returns a valid product type, or 'HOME' as a safe
    generic that exists in the schema (never the invalid literal 'PRODUCT')."""
    haystack = " ".join(str(x) for x in [
        item_name,
        comp_data.get("title", ""),
        comp_data.get("item_type_keyword", ""),
        " ".join(comp_data.get("browse_nodes", []) or []),
        " ".join(f"{k} {v}" for k, v in (comp_data.get("attributes") or {}).items()),
    ]).lower()

    for pat, ptype in _PT_INFER_RULES:
        if re.search(pat, haystack):
            # only return it if the schema actually knows this type (when we have
            # the valid_values map); otherwise still return it -- SP-API will
            # validate at export and Claude uses it as a strong hint.
            if not valid_types or ptype in valid_types or ptype == "HOME":
                return ptype
            return ptype
    return "HOME"


def load_model_counter() -> dict:
    if not COUNTER_PATH.exists():
        return {}
    try:
        with open(COUNTER_PATH) as f:
            return json.load(f) or {}
    except Exception:
        return {}


def save_model_counter(data: dict):
    try:
        with open(COUNTER_PATH, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        console.print(f"  [yellow]Could not persist model counter: {e}[/yellow]")


def next_model_number(brand: str, product_type: str, counter: dict) -> str:
    """Generate model number: {first 4 of brand}-{3-letter category code}-{seq:03d}.
    Mutates the counter dict in place. Caller is responsible for saving."""
    prefix = re.sub(r"[^A-Za-z0-9]", "", brand or "Unb")[:4].title() or "Unbr"
    code   = derive_product_type_code(product_type)
    key    = f"{prefix}-{code}"
    counter[key] = counter.get(key, 0) + 1
    return f"{prefix}-{code}-{counter[key]:03d}"


def is_model_number_required(schema: dict) -> bool:
    """True iff schema marks any of model_number / model / part_number as required."""
    required = schema.get("required", {}) or {}
    for field in ("model_number", "model", "part_number"):
        if field in required:
            return True
    return False


# =============================================================================
# COMPLIANCE RULES
# =============================================================================

def load_compliance_rules() -> dict:
    """Load compliance_rules.json. Returns empty dict if file missing or invalid."""
    if not COMPLIANCE_PATH.exists():
        console.print(f"  [yellow]compliance_rules.json not found -- compliance check disabled[/yellow]")
        return {}
    try:
        with open(COMPLIANCE_PATH, encoding="utf-8") as f:
            data = json.load(f)
        # Drop meta entry from the working dict
        return {k: v for k, v in data.items() if not k.startswith("_")}
    except Exception as e:
        console.print(f"  [red]compliance_rules.json invalid: {e}[/red]")
        return {}


_RISK_PRIORITY = {"HIGH": 3, "MEDIUM": 2, "BASELINE": 1, "": 0}


def check_compliance(item_name: str, listing: dict, rules: dict) -> dict:
    """
    Match the product's title + bullets + description against compliance keywords.
    Returns {
      "matched_categories": [list of category keys],
      "highest_risk":       "HIGH" | "MEDIUM" | "BASELINE" | "",
      "summary":            short string for the Notes column,
      "requirements":       full requirements list (for richer reporting)
    }
    """
    if not rules:
        return {"matched_categories": [], "highest_risk": "",
                "summary": "", "requirements": []}

    haystack = " ".join([
        item_name or "",
        listing.get("title", ""),
        listing.get("bullet_1", ""), listing.get("bullet_2", ""),
        listing.get("bullet_3", ""), listing.get("bullet_4", ""),
        listing.get("bullet_5", ""),
        listing.get("description", ""),
        listing.get("search_terms", ""),
    ]).lower()

    matched     = []
    all_reqs    = []
    highest     = ""
    # Some keywords are too generic and cause false positives: a hand tool that
    # "weighs 540g" is not sports/fitness; copy that says "safe around children"
    # is not a children's TOY; "powerful" is not "mains-powered electrical". For
    # these high-false-positive categories, require a STRONGER signal: the match
    # must be a clear category keyword, AND for electrical we additionally require
    # a genuine electrical token (not just "power"/"light" used as adjectives).
    _weak_kw = {
        "sports_fitness": {"weight", "weights", "net", "swing", "training", "resistance"},
        "toys_children":  {"play", "game", "child", "children", "kids", "educational"},
        "electrical":     {"power", "light", "lighting"},
    }
    _electrical_strong = ("plug", "mains", "240v", "230v", "voltage", "volt", "watt",
                          "wattage", "rechargeable", "battery", "batteries", "usb",
                          "charger", "charging", "corded", "cordless", "led", "bulb",
                          "lamp", "socket", "adapter", "adaptor", "power supply",
                          "power cable", "power cord", "electric motor")
    for cat_key, rule in rules.items():
        if cat_key == "general":
            continue
        kws = rule.get("keywords", [])
        if not kws:
            continue
        _matched_kw = ""
        for kw in kws:
            if not kw:
                continue
            _k = kw.lower()
            try:
                _hit = re.search(rf"(?<![\w-]){re.escape(_k)}(?![\w-])", haystack) is not None
            except re.error:
                _hit = _k in haystack          # fallback if a keyword isn't a valid regex
            if _hit:
                _matched_kw = _k
                break
        if _matched_kw:
            # context guard: drop the match if the ONLY thing that matched is a
            # weak/generic keyword for a false-positive-prone category.
            _weak = _weak_kw.get(cat_key, set())
            if _matched_kw in _weak:
                if cat_key == "electrical":
                    # only keep electrical if a genuine electrical token is present
                    if not any(re.search(rf"(?<![\w-]){re.escape(t)}(?![\w-])", haystack)
                               for t in _electrical_strong):
                        continue
                else:
                    # toys/sports: a weak keyword alone isn't enough; need a real
                    # category keyword too (a non-weak keyword also present)
                    _strong_hit = any(
                        (kw.lower() not in _weak) and kw and
                        re.search(rf"(?<![\w-]){re.escape(kw.lower())}(?![\w-])", haystack)
                        for kw in kws)
                    if not _strong_hit:
                        continue
            matched.append(cat_key)
            all_reqs.extend(rule.get("requirements", []))
            risk = rule.get("risk_level", "")
            if _RISK_PRIORITY.get(risk, 0) > _RISK_PRIORITY.get(highest, 0):
                highest = risk

    # General requirements always apply
    if "general" in rules:
        all_reqs.extend(rules["general"].get("requirements", []))
        if not highest:
            highest = rules["general"].get("risk_level", "BASELINE")

    # De-duplicate requirements while preserving order
    seen = set()
    deduped_reqs = []
    for r in all_reqs:
        if r not in seen:
            seen.add(r)
            deduped_reqs.append(r)

    # MARKETPLACE-AWARE: compliance_rules.json is written for the UK (UKCA, BS 1363,
    # WEEE, UK Batteries Regs). Showing those on a US listing is misleading. For a
    # non-UK marketplace, drop the UK-law-specific lines and keep only the universal
    # ones (lithium handling, documentation), plus a clear marketplace note. The
    # RISK category is still flagged (electrical/lithium are sensitive everywhere).
    _is_uk = (MARKETPLACE_ID == "A1F83G8C2ARO7P")
    if not _is_uk and deduped_reqs:
        _uk_markers = ("ukca", "bs 1363", "weee", "uk batteries", "uk market",
                       "great britain", "gb after", "ce mark", "ce marking",
                       "ukni", "emc testing certificate", "producer registration",
                       "uk electromagnetic", "uk ", "(uk")
        _universal = []
        for _r in deduped_reqs:
            _low = _r.lower()
            if any(m in _low for m in _uk_markers):
                continue  # UK-specific legal line -> not relevant to this marketplace
            _universal.append(_r)
        _mkt_label = "US" if MARKETPLACE_ID == US_MARKETPLACE_ID else "this marketplace"
        if MARKETPLACE_ID == US_MARKETPLACE_ID:
            _universal.insert(0, "🇺🇸 US market: ensure FCC compliance for electronics, "
                                 "UL/ETL safety where applicable, and Prop 65 warnings if sold in California.")
        deduped_reqs = _universal

    if matched:
        summary = f"COMPLIANCE [{highest}]: {', '.join(matched)}"
    else:
        summary = ""

    return {"matched_categories": matched,
            "highest_risk":       highest,
            "summary":            summary,
            "requirements":       deduped_reqs}


# =============================================================================
# IP / TRADEMARK PROTECTION
# =============================================================================

def load_ip_rules() -> dict:
    """Load ip_rules.json. Returns empty dict if missing or invalid."""
    if not IP_RULES_PATH.exists():
        console.print(f"  [yellow]ip_rules.json not found -- IP check disabled[/yellow]")
        return {}
    try:
        with open(IP_RULES_PATH, encoding="utf-8") as f:
            data = json.load(f)
        # Build a fast lookup set from the safe word list (case-preserved)
        safe_set = set(data.get("safe_capitalised_words", []))
        return {
            "forbidden_phrases":   data.get("forbidden_phrases", []),
            "safe_capitalised":    safe_set,
            "safe_capitalised_lc": {w.lower() for w in safe_set},
            "max_unrecognised":    int(data.get("max_allowed_caps_words_unrecognised", 2)),
        }
    except Exception as e:
        console.print(f"  [red]ip_rules.json invalid: {e}[/red]")
        return {}


def _strip_html(text: str) -> str:
    return re.sub(r"<[^>]+>", " ", text or "")


# _split_brand_words moved to listing/compliance.py in Phase 5 (used only by
# check_ip_violations). Imported back so it still resolves.
from listing.compliance import _split_brand_words


# check_ip_violations moved to listing/compliance.py in Phase 5 (behaviour unchanged).
from listing.compliance import check_ip_violations


# =============================================================================
# EXISTING SKU / DUP HELPERS
# =============================================================================

def load_existing_skus_and_asins(ws_out) -> tuple:
    """Read the output sheet once at start, return (set of SKUs, set of ASINs)
    already present so we can detect duplicates from previous runs."""
    try:
        all_rows = _safe_records(ws_out)
    except Exception as e:
        console.print(f"  [yellow]Could not read existing rows: {str(e)[:80]}[/yellow]")
        return set(), set()
    skus  = {str(r.get("SKU", "")).strip()             for r in all_rows if str(r.get("SKU", "")).strip()}
    # An ASIN only counts as "already done" if its row also has a SKU (a complete
    # listing). A row you cleared (ASIN left behind but SKU/content blank) must NOT
    # block regeneration -- it gets refilled in place instead.
    asins = {str(r.get("Competitor ASIN", "")).strip() for r in all_rows
             if str(r.get("Competitor ASIN", "")).strip() and str(r.get("SKU", "")).strip()}
    return skus, asins



def _open_sheet_retry(gc, key: str, what: str = "sheet", tries: int = 5):
    """Open a Google spreadsheet, retrying on transient server errors (HTTP
    500/502/503/429). Google occasionally returns 503 'service unavailable' for
    a few seconds -- that should NOT crash the whole run."""
    import time as _t
    last = None
    for i in range(tries):
        try:
            return gc.open_by_key(key)
        except Exception as e:
            last = e
            _code = ""
            try:
                _code = str(getattr(e, "response", None).status_code)
            except Exception:
                _code = ""
            _msg = str(e)
            _transient = (_code in ("500", "502", "503", "429")
                          or "503" in _msg or "500" in _msg or "502" in _msg
                          or "unavailable" in _msg.lower() or "rate" in _msg.lower())
            if _transient and i < tries - 1:
                _wait = 2 * (i + 1)
                console.print(f"  [yellow]Google {what} temporarily unavailable "
                              f"(attempt {i+1}/{tries}) -- retrying in {_wait}s...[/yellow]")
                _t.sleep(_wait)
                continue
            raise
    if last:
        raise last


def init_sheets(config: dict):
    scopes = ["https://spreadsheets.google.com/feeds",
              "https://www.googleapis.com/auth/drive"]
    creds  = Credentials.from_service_account_file(
        config["google_service_account_json"], scopes=scopes)
    gc     = gspread.authorize(creds)

    sh_out = _open_sheet_retry(gc, config["google_spreadsheet_id"], "output sheet")
    # Prefer resolving the exact tab by gid (from the account's sheet URL); fall
    # back to the tab title / OUTPUT_TAB name.
    out_gid = str(config.get("_output_tab_gid") or "").strip()
    ws_out = None
    if out_gid.isdigit():
        try:
            ws_out = sh_out.get_worksheet_by_id(int(out_gid))
            console.print(f"  Output tab (by gid {out_gid}): '[bold]{ws_out.title}[/bold]'")
        except Exception:
            ws_out = None
    if ws_out is None:
        try:
            ws_out = sh_out.worksheet(OUTPUT_TAB)
            # Don't force the 48-col FIXED_HEADERS when Miles mode owns this tab
            # (it writes its own column layout).
            if not config.get("_miles_mode") and len(ws_out.row_values(1)) < FIXED_COUNT:
                ws_out.delete_rows(1)
                ws_out.insert_row(FIXED_HEADERS, 1)
            console.print(f"  Output tab: '[bold]{OUTPUT_TAB}[/bold]'")
        except gspread.WorksheetNotFound:
            ws_out = sh_out.add_worksheet(title=OUTPUT_TAB, rows=2000, cols=100)
            if not config.get("_miles_mode"):
                ws_out.append_row(FIXED_HEADERS, value_input_option="RAW")
                ws_out.format("1:1", {"textFormat": {"bold": True},
                                       "backgroundColor": {"red": 0.27, "green": 0.51, "blue": 0.71}})
                ws_out.freeze(rows=1)
            console.print(f"  Created new tab: '[bold]{OUTPUT_TAB}[/bold]'")

    sh_in = _open_sheet_retry(gc, config["input_spreadsheet_id"], "input sheet")
    in_gid = str(config.get("_input_tab_gid") or "").strip()
    ws_in = None
    if in_gid.isdigit():
        try:
            ws_in = sh_in.get_worksheet_by_id(int(in_gid))
            console.print(f"  Input tab (by gid {in_gid}): '[bold]{ws_in.title}[/bold]'")
        except Exception:
            ws_in = None
    if ws_in is None:
        ws_in = sh_in.get_worksheet(0)
    return gc, ws_in, ws_out


def read_input_sheet(ws_in) -> list:
    rows = ws_in.get_all_values()
    if not rows:
        return []
    headers  = [h.strip().lower().replace(" ", "_") for h in rows[0]]
    products = []
    for row in rows[1:]:
        if not any(row):
            continue
        row  = row + [""] * max(0, len(headers) - len(row))
        item = dict(zip(headers, row))
        norm = {
            "ebay_url":      item.get("ebay_link",     item.get("ebay_url",      "")),
            "source_cost":   item.get("ebay_price",    item.get("ebay_cost",     "")),
            "amazon_url":    item.get("amazon_link",   item.get("amazon_url",    "")),
            "selling_price": item.get("amazon_price",  item.get("selling_price", "")),
            "item_name":     item.get("item_name",     ""),
            "handling_time": item.get("delivery_time", item.get("handling_time", "")),
            "upc":           item.get("ean",           item.get("upc",            "")),
        }
        if norm["amazon_url"].strip():
            products.append(norm)
    return products


def _extract_asin(url: str) -> str:
    m = re.search(r"/(?:dp|gp/product|gp/aw/d)/([A-Z0-9]{10})", url)
    return m.group(1) if m else ""


def _extract_ebay_item(url: str) -> str:
    """eBay item number = the digits after /itm/ in an eBay URL."""
    m = re.search(r"/itm/(?:[^/]*?/)?(\d{6,})", str(url))
    if m:
        return m.group(1)
    # some eBay URLs carry it as ?item=12345 or /itm/12345?...
    m = re.search(r"[?&]item=(\d{6,})", str(url))
    return m.group(1) if m else ""


def select_rows(products: list, raw: str, sel_type: str = "auto"):
    """Filter input-sheet products down to the user's selection.

    Returns (filtered_list, error_message). On success error_message is "".
    On a problem (duplicate / no match / bad input) returns ([], message) so the
    caller can print it and stop -- never silently generate the wrong rows.

    sel_type: 'row' | 'asin' | 'ebay_item' | 'auto'
      - A pasted URL always auto-detects (ignores sel_type): amazon.* -> ASIN,
        ebay.* -> item number.
      - 'row'       -> comma-separated 1-based row numbers (input-sheet data rows).
      - 'asin'      -> match ASIN parsed from each row's amazon_url.
      - 'ebay_item' -> match item number parsed from each row's ebay_url.
    """
    raw = (raw or "").strip()
    if not raw:
        return products, ""   # empty -> generate all (unchanged)

    # --- URL pasted: auto-detect platform regardless of sel_type --------------
    low = raw.lower()
    if "http://" in low or "https://" in low or "amazon." in low or "ebay." in low:
        if "amazon." in low:
            asin = _extract_asin(raw)
            if not asin:
                return [], f"Couldn't read an ASIN from that Amazon URL: {raw[:60]}"
            hits = [(i, p) for i, p in enumerate(products, 1)
                    if _extract_asin(p.get("amazon_url", "")) == asin]
            return _finish_match(hits, f"ASIN {asin}")
        if "ebay." in low:
            item = _extract_ebay_item(raw)
            if not item:
                return [], f"Couldn't read an item number from that eBay URL: {raw[:60]}"
            hits = [(i, p) for i, p in enumerate(products, 1)
                    if _extract_ebay_item(p.get("ebay_url", "")) == item]
            return _finish_match(hits, f"eBay item {item}")
        if "docs.google." in low or "/spreadsheets/" in low or "drive.google." in low:
            return [], ("That's your Google Sheet link, not a product to select. "
                        "Leave the Generate box EMPTY to make every input-sheet row, "
                        "or type a row number (e.g. 1), or paste a single Amazon/eBay "
                        "product URL.")
        return [], f"Couldn't tell if that URL is Amazon or eBay: {raw[:60]}"

    # --- Row numbers ----------------------------------------------------------
    if sel_type == "row":
        nums = []
        for tok in re.split(r"[,\s]+", raw):
            tok = tok.strip()
            if not tok:
                continue
            if not tok.isdigit():
                return [], (f"'{tok}' is not a row number. For rows, enter digits "
                            f"like 2, 5, 7.")
            nums.append(int(tok))
        picked, bad = [], []
        for n in nums:
            if 1 <= n <= len(products):
                picked.append(products[n - 1])
            else:
                bad.append(n)
        if bad:
            return [], (f"Row(s) {', '.join(map(str, bad))} are out of range "
                        f"(sheet has {len(products)} data rows).")
        if not picked:
            return [], "No valid rows in that selection."
        return picked, ""

    # --- Bare ASIN ------------------------------------------------------------
    if sel_type == "asin":
        asin = raw.upper()
        hits = [(i, p) for i, p in enumerate(products, 1)
                if _extract_asin(p.get("amazon_url", "")) == asin]
        return _finish_match(hits, f"ASIN {asin}")

    # --- Bare eBay item number ------------------------------------------------
    if sel_type == "ebay_item":
        item = re.sub(r"\D", "", raw)
        hits = [(i, p) for i, p in enumerate(products, 1)
                if _extract_ebay_item(p.get("ebay_url", "")) == item]
        return _finish_match(hits, f"eBay item {item}")

    return [], f"Unknown selection type '{sel_type}'."


def _finish_match(hits: list, label: str):
    """hits = list of (row_number, product). Enforce the duplicate rule."""
    if not hits:
        return [], (f"No row found matching {label}. Check the value or the input "
                    f"sheet.")
    if len(hits) > 1:
        rows = ", ".join(str(i) for i, _ in hits)
        return [], (f"{label} appears in rows {rows} of the input sheet. Switch to "
                    f"Row number and enter the exact row you want.")
    return [hits[0][1]], ""


def _attrs_with_images(pa: dict, comp_data: dict) -> dict:
    """Stash the competitor's primary (+ additional) image URLs into the attribute
    dict so the dashboard can preview them and the API submit can use them as the
    product images. eBay images already take priority inside comp_data['images'].

    Also writes a `_provenance` map {attr_key: 'ebay'|'amazon'|'ai'} so the
    dashboard can tag each field with where its value came from. Source-supplied
    keys keep their eBay/Amazon tag; any attribute the AI produced (present in
    `pa` but not in the source map) is tagged 'ai'.
    """
    out = dict(pa or {})
    imgs = [u for u in (comp_data.get("images") or []) if u][:5]
    if imgs:
        out.setdefault("main_product_image_locator", imgs[0])
        for i, u in enumerate(imgs[1:5], start=1):
            out.setdefault(f"other_product_image_locator_{i}", u)
    # provenance: start from the eBay/Amazon source map, tag the rest as AI
    _src = dict((comp_data.get("_provenance") or {}))
    _prov = {}
    for _k in out.keys():
        if _k.startswith("main_product_image_locator") or _k.startswith("other_product_image_locator_"):
            continue  # images aren't attribute facts
        if _k in _src:
            _prov[_k] = _src[_k]
        else:
            _prov[_k] = "ai"   # the AI produced this value
    if _prov:
        out["_provenance"] = _prov
    return out


def build_sheet_row(comp_asin: str, row: dict, listing: dict,
                    comp_data: dict, financials: dict, pricing: dict,
                    voc_data: dict, keywords: list,
                    handling_time: str, handling_days: str,
                    sku: str, status: str,
                    brand: str = "", model_number: str = "", notes: str = "",
                    compliance_risk: str = "", ip_risk: str = "",
                    marketplace: str = "UK") -> list:
    # Belt-and-suspenders: strip any policy-risky promo copy (seller self-promotion,
    # shipping claims, links, unverifiable superlatives) the model may have slipped past
    # the generation prompt, BEFORE it's written to the sheet. Transparent -- logs removals.
    try:
        from listing.compliance import scrub_listing_copy
        _scrub_notes = scrub_listing_copy(listing)
        if _scrub_notes:
            console.print(f"  [yellow]Copy scrubber cleaned {len(_scrub_notes)} field(s) of policy-risky text:[/yellow]")
            for _sn in _scrub_notes:
                console.print(f"    [dim]• {_sn[:160]}[/dim]")
    except Exception:
        pass
    kw_str = ", ".join(k["keyword"] for k in keywords[:8])
    _mkt = str(marketplace or "UK").upper()
    _cur = "USD" if _mkt == "US" else "GBP"
    _plat = "AMAZON US" if _mkt == "US" else "AMAZON UK"
    out = [
        comp_asin,
        row.get("ebay_url",       ""),
        row.get("upc",            ""),
        sku,
        _plat,
        f"{_cur}{pricing['buy_box_price']:.2f}",
        f"{_cur}{financials['selling_price']:.2f}",
        f"{_cur}{financials['total_amazon_fees']:.2f}",
        financials["fee_source"],
        f"{_cur}{financials['profit']:.2f}",
        financials["margin_pct"],
        financials["roi_pct"],
        financials["viable"],
        comp_data.get("product_type",     ""),
        listing.get("amazon_category",    ""),
        listing.get("amazon_subcategory", ""),
        voc_data.get("source",            "none"),
        str(voc_data.get("review_count",  0)),
        listing.get("target_demographic", ""),
        listing.get("pain_points",        ""),
        listing.get("purchase_trigger",   ""),
        listing.get("title",              ""),
        listing.get("bullet_1",           ""),
        listing.get("bullet_2",           ""),
        listing.get("bullet_3",           ""),
        listing.get("bullet_4",           ""),
        listing.get("bullet_5",           ""),
        listing.get("description",        ""),
        listing.get("search_terms",       ""),
        kw_str,
        listing.get("material",           ""),
        listing.get("colour",             ""),
        listing.get("size",               ""),
        str(listing.get("number_of_items", "")),
        listing.get("target_gender",      ""),
        listing.get("age_range",          ""),
        listing.get("compliance_notes",   "None"),
        handling_time,
        handling_days,
        status,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        brand,                          # NEW
        model_number,                   # NEW
        notes,                          # NEW
        compliance_risk,                # NEW
        ip_risk,                        # NEW
        json.dumps(_attrs_with_images(listing.get("product_attributes", {}), comp_data), ensure_ascii=False),  # full attribute object (+competitor images) for unified export
        cap_chars(listing.get("item_highlights", ""), HIGHLIGHTS_MAX),  # 47 Item Highlights
        "",                             # 48 API Payload JSON - filled later on Preview/Submit
    ]
    assert len(out) == FIXED_COUNT, f"Row length {len(out)} != {FIXED_COUNT}"
    return out


def _find_target_row(ws, comp_asin: str):
    """Decide where a generated row should go so listings refill the row you
    cleared (or the first blank gap) instead of always appending at the bottom.
    Priority:
      1) a row with this exact Competitor ASIN but no SKU  (the row you cleared);
      2) the first fully-blank data row (SKU, Title, Competitor ASIN, Product Type all empty);
      3) None  -> caller appends.
    Returns a 1-based sheet row number, or None.
    """
    try:
        vals = ws.get_all_values()
    except Exception:
        return None
    if not vals:
        return None
    headers = vals[0]
    cidx = lambda name: headers.index(name) if name in headers else -1
    a_i, s_i = cidx("Competitor ASIN"), cidx("SKU")
    t_i, p_i = cidx("Title"),           cidx("Product Type")
    cell = lambda rv, i: (str(rv[i]).strip() if (0 <= i < len(rv)) else "")
    # 1) the row you cleared for this ASIN (ASIN still there, SKU gone)
    if comp_asin and a_i >= 0:
        for r in range(1, len(vals)):
            if cell(vals[r], a_i) == comp_asin and not cell(vals[r], s_i):
                return r + 1
    # 2) first fully-blank data row
    keyi = [i for i in (s_i, t_i, a_i, p_i) if i >= 0]
    for r in range(1, len(vals)):
        if all(not cell(vals[r], i) for i in keyi):
            return r + 1
    return None


def sheet_write_row(ws, row_data: list, comp_asin: str = ""):
    target = _find_target_row(ws, comp_asin)
    for attempt in range(1, 4):
        try:
            if target:                                    # refill in place (keeps position)
                rng = f"A{target}:{_col_letter(len(row_data) - 1)}{target}"
                ws.update([row_data], rng, value_input_option="USER_ENTERED")
            else:                                         # no gap -> append at bottom
                ws.append_row(row_data, value_input_option="USER_ENTERED")
            return True
        except Exception as e:
            if attempt == 3:
                console.print(f"  [red]Sheet write failed: {str(e)[:60]}[/red]")
                return False
            time.sleep(attempt * 5)
    return False


def get_retry_rows(ws_out, products: list) -> list:
    try:
        all_rows   = _safe_records(ws_out)
        retry_urls = {str(r.get("Source URL", "")).strip()
                      for r in all_rows
                      if str(r.get("Status", "")).strip().upper() in ("NEEDS_REVIEW", "ERROR")}
        if not retry_urls:
            console.print("[yellow]No NEEDS_REVIEW or ERROR rows found.[/yellow]")
            return []
        retry = [p for p in products if p.get("ebay_url", "").strip() in retry_urls]
        console.print(f"[yellow]Retry: {len(retry)} row(s) (NEEDS_REVIEW + ERROR)[/yellow]")
        return retry
    except Exception as e:
        console.print(f"[red]Retry read failed: {e}[/red]")
        return products


# =============================================================================
# FLAT FILE EXPORT -- DROPDOWN VALIDATION
# =============================================================================

FILE1_COLS = {
    "SKU":                              0,
    "Product Type":                     1,
    "Listing Action":                   2,
    "Item Name":                        6,
    "Brand Name":                       7,
    "Product Id Type":                  8,
    "Product Id":                       9,
    "Browse Node 1":                   10,
    "Model Number":                    18,
    "Manufacturer":                    19,
    "Product Description":             33,
    "Bullet Point 1":                  34,
    "Bullet Point 2":                  35,
    "Bullet Point 3":                  36,
    "Bullet Point 4":                  37,
    "Bullet Point 5":                  38,
    "Generic Keyword":                 39,
    "Material":                        45,
    "Number of Items":                 50,
    "Colour":                          52,
    "Size":                            53,
    "Target Gender":                  129,
    "Age Range Description":          130,
    "Item Condition":                 300,
    "List Price with Tax":            302,
    "Product Tax Code":               303,
    "Fulfillment Channel Code (UK)":  325,
    "Quantity (UK)":                  326,
    "Handling Time (UK)":             327,
    "Your Price GBP":                 330,
    "Country of Origin":              393,
    "Are batteries required?":        395,
    "Are batteries included?":        396,
    "TOTAL_COLS":                     489,
}

FILE2_COLS = {
    "SKU":                              0,
    "Product Type":                     1,
    "Listing Action":                   2,
    "Item Name":                        6,
    "Brand Name":                       7,
    "Product Id Type":                  8,
    "Product Id":                       9,
    "Browse Node 1":                   10,
    "Model Number":                    18,
    "Manufacturer":                    19,
    "Product Description":             38,
    "Bullet Point 1":                  39,
    "Bullet Point 2":                  40,
    "Bullet Point 3":                  41,
    "Bullet Point 4":                  42,
    "Bullet Point 5":                  43,
    "Generic Keyword":                 44,
    "Material":                        45,
    "Colour":                          51,
    "Size":                            52,
    "Number of Items":                 85,
    "Target Gender":                  161,
    "Age Range Description":          284,
    "Item Condition":                 394,
    "List Price with Tax":            396,
    "Product Tax Code":               397,
    "Fulfillment Channel Code (UK)":  419,
    "Quantity (UK)":                  420,
    "Handling Time (UK)":             421,
    "Your Price GBP":                 424,
    "Country of Origin":              485,
    "Are batteries required?":        486,
    "Are batteries included?":        487,
    "TOTAL_COLS":                     612,
}

PRODUCT_ROUTES = [
    (["cookware", "saucepan", "pots and pans", "frying pan", "casserole", "pan set"],
     "FILE1", "COOKWARE_SET", "11715891"),
    (["floor lamp", "standing lamp", "corner lamp", "rgb led lamp", "mood lamp"],
     "FILE1", "LAMP", "10709381"),
    (["light bar", "rgb light", "led bar", "tv backlight", "gaming light", "backlights"],
     "FILE1", "LAMP", "3764800031"),
    (["solar light", "security light", "outdoor light", "motion sensor"],
     "FILE1", "LAMP", "13679891"),
    (["shelf bracket", "floating shelf", "wall bracket", "mount bracket"],
     "FILE1", "HARDWARE", "1938668031"),
    (["changeover switch", "rotary cam", "cam switch", "electrical switch",
      "bearing puller", "gear puller", "extractor"],
     "FILE1", "HARDWARE", "1938353031"),
    (["golf", "chipping net", "practice net", "swing trainer"],
     "FILE1", "SPORT_TARGET", "26971320031"),
    (["teeth whitening", "whitening powder", "whitening strips"],
     "FILE2", "HEALTH_PERSONAL_CARE", "74136031"),
    (["night cream", "day cream", "face cream", "skin care", "moisturi", "collagen",
      "sleeping mask", "serum"],
     "FILE2", "BEAUTY", "18918424031"),
    (["body lotion", "body cream", "glutathione", "whitening lotion"],
     "FILE2", "BEAUTY", "344269031"),
    (["hair fibre", "hair fiber", "hair loss", "hair growth", "elixir"],
     "FILE2", "HEALTH_PERSONAL_CARE", "2867979031"),
    (["shampoo", "conditioner", "curl cream", "hair spray", "scalp scrub"],
     "FILE2", "HEALTH_PERSONAL_CARE", "18918425031"),
    (["hair dryer", "blow dryer"],
     "FILE2", "HEALTH_PERSONAL_CARE", "2868092031"),
    (["straightener", "hair straighten", "heated brush", "curling iron", "curler"],
     "FILE2", "HEALTH_PERSONAL_CARE", "74099031"),
    (["body spray", "perfume", "fragrance", "body mist"],
     "FILE2", "BEAUTY", "2790134031"),
    (["garlic press", "mandoline", "slicer", "chopper", "kitchen tool", "kitchen gadget"],
     "FILE2", "KITCHEN", "3187111031"),
    (["blender", "juicer", "food processor", "deep fryer", "air fryer"],
     "FILE2", "KITCHEN", "3538310031"),
    (["mop", "bucket set", "shelving unit", "shelf unit", "storage rack", "clothes rail"],
     "FILE2", "HOME", "3579745031"),
    (["extension lead", "power strip", "plug socket"],
     "FILE2", "HOME", "3538310031"),
    (["security camera", "cctv", "indoor camera", "surveillance"],
     "FILE2", "HOME", "3538310031"),
    (["massager", "shiatsu", "back massager"],
     "FILE2", "HEALTH_PERSONAL_CARE", "3360475031"),
]


# Product types THIS unified template accepts (from its Valid Values tab).
TEMPLATE_PRODUCT_TYPES = {
    "KITCHEN", "CORRECTIVE_EYEGLASSES", "GLOBE", "COOKWARE_SET", "AUTO_BATTERY",
    "CAR_ELECTRONICS", "FOOD_SPATULA", "HEALTH_PERSONAL_CARE", "KITCHEN_KNIFE",
    "HANDBAG", "AUTO_ACCESSORY", "HARDWARE", "SPORT_TARGET", "BEAUTY",
    "SUNGLASSES", "SECURITY_CAMERA", "LAMP", "SNOW_GLOBE", "HOME",
}

# Best-effort browse node when the sheet's type is trusted (blank is acceptable;
# recommended_browse_nodes is not a required field).
PT_DEFAULT_NODE = {
    "COOKWARE_SET": "11715891", "LAMP": "10709381", "HARDWARE": "1938668031",
    "SPORT_TARGET": "26971320031", "HEALTH_PERSONAL_CARE": "66280031",
    "BEAUTY": "18918424031", "KITCHEN": "3187111031", "HOME": "3579745031",
}


def _norm_pt(s: str) -> str:
    return re.sub(r"[^A-Z0-9_]", "", str(s).strip().upper().replace(" ", "_"))


# When a product's exact type isn't in this template, map it to the NEAREST
# available type. Order matters (first match wins); HOME is the final catch-all.
# Matching is whole-word on alphanumeric-tokenised text, so 'chair' never hits
# 'hair' and 'lightweight' never hits 'light'.
_PT_FALLBACK_RULES = [
    (["snow globe"], "SNOW_GLOBE"),
    (["globe", "atlas"], "GLOBE"),
    (["sunglasses", "sunglass"], "SUNGLASSES"),
    (["eyeglasses", "spectacles", "reading glasses", "prescription glasses", "optical frame"], "CORRECTIVE_EYEGLASSES"),
    (["cctv", "security camera", "surveillance camera", "ip camera", "webcam", "doorbell camera", "dash cam", "dashcam"], "SECURITY_CAMERA"),
    (["lamp", "lamps", "bulb", "bulbs", "chandelier", "sconce", "lantern", "lighting",
      "downlight", "spotlight", "floodlight", "light fixture", "ceiling light", "wall light",
      "pendant light", "led light", "string light", "night light", "desk light",
      "wall lamp", "desk lamp", "floor lamp", "table lamp"], "LAMP"),
    (["knife", "cleaver", "kitchen knife", "chef knife", "paring knife"], "KITCHEN_KNIFE"),
    (["spatula", "turner", "ladle"], "FOOD_SPATULA"),
    (["cookware", "saucepan", "frying pan", "casserole", "wok", "stockpot", "pots and pans"], "COOKWARE_SET"),
    (["blender", "juicer", "mixer", "peeler", "grater", "slicer", "chopper", "food processor", "air fryer", "kettle", "toaster", "whisk", "utensil", "kitchen gadget", "kitchen tool"], "KITCHEN"),
    (["handbag", "purse", "tote", "backpack", "satchel", "clutch", "shoulder bag", "crossbody"], "HANDBAG"),
    (["car battery", "vehicle battery", "leisure battery"], "AUTO_BATTERY"),
    (["car stereo", "head unit", "car audio", "car speaker"], "CAR_ELECTRONICS"),
    (["car", "automotive", "vehicle", "number plate", "seat cover", "floor mat", "wing mirror", "wiper"], "AUTO_ACCESSORY"),
    (["serum", "moisturiser", "moisturizer", "face cream", "body cream", "night cream", "lotion", "cosmetic", "skincare", "makeup", "fragrance", "perfume", "mascara", "lipstick", "face mask"], "BEAUTY"),
    (["massager", "supplement", "trimmer", "shaver", "toothbrush", "grooming", "scalp", "manicure"], "HEALTH_PERSONAL_CARE"),
    (["dartboard", "archery", "practice net", "chipping net", "golf net", "shooting target", "target board"], "SPORT_TARGET"),
    (["tool", "tools", "bracket", "fixing", "screw", "drill", "wrench", "hardware", "mount", "hinge", "hook", "fastener", "clamp"], "HARDWARE"),
]


def _fallback_pt(text: str) -> str:
    """Map an unsupported product to the NEAREST available template type.
    Returns 'HOME' (the generic catch-all) when nothing more specific fits."""
    t = " " + re.sub(r"[^a-z0-9]+", " ", text.lower()).strip() + " "
    for keywords, pt in _PT_FALLBACK_RULES:
        if any(f" {kw} " in t for kw in keywords):
            return pt
    return "HOME"


def detect_route(title: str, category: str, product_type: str) -> tuple:
    """
    Return (file_id, product_type, browse_node).
      1) Trust the sheet's Product Type when this template accepts it.
      2) Else keyword-route with LEFT word-boundary matching, so 'chair' can no
         longer match 'hair'.
      3) Else return ('', '', '') -- a SKIP signal; the caller skips and flags the
         row instead of forcing it into HOME.
    """
    text     = f"{title} {category} {product_type}".lower()
    pt_sheet = _norm_pt(product_type)

    # 1) Trust the explicit sheet value if the template supports it.
    if pt_sheet in TEMPLATE_PRODUCT_TYPES:
        node = PT_DEFAULT_NODE.get(pt_sheet, "")
        for keywords, file_id, pt, browse_node in PRODUCT_ROUTES:
            if pt == pt_sheet and any(re.search(r"\b" + re.escape(kw), text) for kw in keywords):
                node = browse_node
                break
        fid = "FILE1" if pt_sheet in {"COOKWARE_SET", "LAMP", "HARDWARE", "SPORT_TARGET"} else "FILE2"
        return fid, pt_sheet, node

    # 2) Keyword routing with left word-boundary matching.
    for keywords, file_id, pt, browse_node in PRODUCT_ROUTES:
        if any(re.search(r"\b" + re.escape(kw), text) for kw in keywords):
            if pt in TEMPLATE_PRODUCT_TYPES:
                return file_id, pt, browse_node

    # 3) Unsupported type -> map to the NEAREST available type (never skip).
    fb  = _fallback_pt(text)
    fid = "FILE1" if fb in {"COOKWARE_SET", "LAMP", "HARDWARE", "SPORT_TARGET"} else "FILE2"
    return fid, fb, PT_DEFAULT_NODE.get(fb, "")


def _parse_field_key(field_id: str) -> str:
    return field_id.split("[")[0].split("#")[0].strip().lower()


# Field key alias map:
# Amazon uses different field ID names per product type in the Dropdown Lists tab.
# Each entry lists aliases to try in order so we never miss a valid dropdown list.
FIELD_KEY_ALIASES = {
    "size":                ["size", "item_size", "item_package_quantity",
                            "item_display_dimensions", "volume_capacity_name"],
    "color":               ["color", "color_name", "colour", "item_color_name",
                            "exterior_color_name", "color_map"],
    "material":            ["material", "material_type", "item_material_type",
                            "outer_material_type"],
    "target_gender":       ["target_gender", "department", "department_name"],
    "age_range":           ["age_range_description", "age_range", "age_range_name"],
    "condition_type":      ["condition_type", "condition"],
    "country_of_origin":   ["country_of_origin", "country_of_manufacture"],
    "product_tax_code":    ["product_tax_code"],
    "batteries_required":  ["batteries_required", "are_batteries_required"],
    "batteries_included":  ["batteries_included", "are_batteries_included"],
    "fulfillment_channel": ["fulfillment_availability#1.fulfillment_channel_code",
                            "fulfillment_channel_code"],
}


def _smart_vlist(field_name: str, valid_values: dict) -> list:
    """Try all aliases for a field name, return first non-empty list found."""
    for alias in FIELD_KEY_ALIASES.get(field_name, [field_name]):
        result = valid_values.get(alias, [])
        if result:
            return result
    return []


def load_static_valid_values() -> dict:
    """
    Load valid_values.json - the authoritative Amazon UK enumerated valid-values
    dictionary extracted from the public XLSM flat-file templates.

    Structure: {product_type: {attribute_snake_case: [sorted list of valid strings]}}

    These take precedence over runtime-loaded Google Sheet template values
    because they are extracted directly from Amazon's published flat-file
    template binaries.
    """
    if not VALID_VALUES_PATH.exists():
        console.print("  [yellow]valid_values.json not found -- "
                      "using runtime Google Sheet template only[/yellow]")
        return {}
    try:
        with open(VALID_VALUES_PATH, encoding="utf-8") as f:
            data = json.load(f)
        # Strip _meta entries; we only want product-type keys
        return {k: v for k, v in data.items() if not k.startswith("_")}
    except Exception as e:
        console.print(f"  [red]valid_values.json invalid: {e}[/red]")
        return {}


def merge_static_into_runtime(runtime_vv: dict, static_vv: dict) -> dict:
    """
    Overlay static valid-values on top of runtime-loaded ones.
    For each product type covered by static_vv, replace the runtime values
    so the script uses Amazon-published enumerations as the source of truth.
    Product types only present in runtime_vv are preserved (they may exist
    in the Google Sheet template but not in our two XLSM files).
    """
    if not static_vv:
        return runtime_vv
    merged = dict(runtime_vv) if runtime_vv else {}
    for pt, attrs in static_vv.items():
        if pt not in merged:
            merged[pt] = {}
        for attr, values in attrs.items():
            merged[pt][attr] = list(values)  # static wins
    return merged


def load_dropdown_values(gc, sheet_id: str, label: str) -> dict:
    """Reads valid dropdown values from template Google Sheet at runtime."""
    console.print(f"  Reading dropdowns from {label}...", end=" ")
    try:
        sh       = gc.open_by_key(sheet_id)
        ws       = sh.worksheet("Dropdown Lists")
        all_rows = ws.get_all_values()
    except Exception as e:
        console.print(f"[red]FAIL {str(e)[:60]}[/red]")
        return {}

    if len(all_rows) < 4:
        console.print("[yellow]Too few rows[/yellow]")
        return {}

    pt_row     = all_rows[1]
    field_row  = all_rows[2]
    value_rows = all_rows[3:]
    result     = {}

    for col in range(2, len(field_row)):
        pt  = pt_row[col].strip()    if col < len(pt_row)    else ""
        fid = field_row[col].strip() if col < len(field_row) else ""
        if not pt or not fid:
            continue
        fk = _parse_field_key(fid)
        if not fk:
            continue
        if pt not in result:
            result[pt] = {}
        if fk not in result[pt]:
            result[pt][fk] = []
        for row in value_rows:
            val = row[col].strip() if col < len(row) else ""
            if val and val not in result[pt][fk]:
                result[pt][fk].append(val)

    total = sum(len(v) for v in result.values())
    console.print(f"[green]OK[/green] {len(result)} product types, {total} field lists")
    return result


def snap_to_valid(value: str, valid_list: list) -> str:
    """5-strategy fuzzy match to Amazon's exact valid dropdown value."""
    if not value or not valid_list:
        return ""
    v = value.strip()
    if v in valid_list:
        return v
    v_lower = v.lower()
    for item in valid_list:
        if item.lower() == v_lower:
            return item
    for item in valid_list:
        if item.lower() in v_lower:
            return item
    for item in valid_list:
        if v_lower in item.lower():
            return item
    v_words   = set(v_lower.split())
    best, best_score = "", 0
    for item in valid_list:
        score = len(v_words & set(item.lower().split()))
        if score > best_score:
            best_score, best = score, item
    return best if best_score >= 1 else ""


def _col_letter(col_0: int) -> str:
    result = ""
    col    = col_0 + 1
    while col > 0:
        col, r = divmod(col - 1, 26)
        result  = chr(65 + r) + result
    return result


# _clean_price moved to listing/builder.py in Phase 5 (self-contained; behaviour unchanged).
from listing.builder import _clean_price


def _strip_html(html: str) -> str:
    text = re.sub(r"<br\s*/?>", " ",   html, flags=re.IGNORECASE)
    text = re.sub(r"<li>",      " - ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>",  "",    text)
    return re.sub(r"\s+", " ", text).strip()


def _clean_days(row: dict) -> str:
    days = str(row.get("Handling Days", "")).strip()
    if days.isdigit():
        return days
    nums = re.findall(r"\d+", str(row.get("Handling Time", "")))
    return nums[0] if nums else "3"


# _has_battery moved to listing/hazmat.py in Phase 5 (self-contained; behaviour unchanged).
from listing.hazmat import _has_battery


# Normalises a measurement unit to the EXACT string Amazon accepts (its unit
# dropdowns are case-sensitive: "kilograms" is REJECTED, "Kilograms" is accepted).
# Amazon's catalogue dimensions block returns lowercase forms, so we must map up.
_DIM_UNIT_NORM = {
    "cm": "Centimeters", "cms": "Centimeters", "centimeter": "Centimeters",
    "centimetre": "Centimeters", "centimetres": "Centimeters", "centimeters": "Centimeters",
    "mm": "Millimeters", "millimeter": "Millimeters", "millimetre": "Millimeters",
    "millimetres": "Millimeters", "millimeters": "Millimeters",
    "m": "Meters", "meter": "Meters", "metre": "Meters", "metres": "Meters", "meters": "Meters",
    "in": "Inches", "ins": "Inches", "inch": "Inches", "inches": "Inches", '"': "Inches",
    "ft": "Feet", "foot": "Feet", "feet": "Feet",
    "g": "Grams", "gm": "Grams", "gms": "Grams", "gram": "Grams", "grams": "Grams",
    "kg": "Kilograms", "kgs": "Kilograms", "kilogram": "Kilograms", "kilograms": "Kilograms",
    "lb": "Pounds", "lbs": "Pounds", "pound": "Pounds", "pounds": "Pounds",
    "oz": "Ounces", "ounce": "Ounces", "ounces": "Ounces",
    "mg": "Milligrams", "milligram": "Milligrams", "milligrams": "Milligrams",
}


def _norm_dim_unit(raw: str) -> str:
    u = str(raw or "").strip().lower().rstrip(".")
    if not u:
        return ""
    u = u.split()[0]                          # "centimeters (cm)" -> "centimeters"
    return _DIM_UNIT_NORM.get(u, u)


# Safety & compliance attribute keys whose values are taken verbatim from the
# live SP-API schema enum (injected into the generation prompt). These are
# written to the flat file WITHOUT fuzzy snapping, because the static
# valid-values lists for these columns are frequently incomplete and snapping
# would blank a correct "No"/"Not Applicable" answer or match the wrong option.
_COMPLIANCE_PASSTHROUGH = {
    "supplier_declared_dg_hz_regulation",
    "contains_liquid_contents",
    "ghs",
    "ghs_classification_class",
    "hazmat",
    "batteries_required",
    "batteries_included",
    "supplier_declared_material_regulation",
    "pesticide_marking",
    "california_proposition_65_compliance_type",
}

# ---------------------------------------------------------------------------
# GLOBAL SAFE-DEFAULTS for safety/compliance fields.
# Strategy: for any compliance field, pick the option Amazon allows that asks the
# FEWEST follow-up questions -- i.e. the "not applicable / none / not regulated"
# branch -- and NEVER an option that triggers additional required sub-fields
# (e.g. dg regulation "ghs", which forces a GHS hazard class). Driven by the live
# schema each time, so it works on product types we've never seen.
#
# Two field families need OPPOSITE treatment:
#   1. Hazard/regulatory questions (ghs, hazmat, pesticide, prop65, material reg,
#      dg regulation): a non-chemical gadget genuinely has none -> not_applicable.
#   2. Battery facts: a product physically either has a battery or not. We can't
#      "not applicable" a real battery (that would be a false declaration). So we
#      fill the STANDARD BATTERY PROFILE below when the item is battery-powered.

# The single standard battery profile (choice 1a). Typical small rechargeable
# consumer gadget. Used for every battery product unless its own data overrides.
STANDARD_BATTERY_PROFILE = {
    "are_batteries_included": True,           # ships with the battery installed
    "battery_cell_composition": "lithium_ion",
    "battery_type": "lithium_ion",
    "number_of_cells": 1,
    "number_of_lithium_ion_cells": 1,
    "number_of_lithium_metal_cells": 0,
    "lithium_battery_packaging": "batteries_contained_in_equipment",
    "battery_weight_grams": 30,               # nominal; small device cell
    "battery_watt_hours": 5,                  # nominal small cell (<100Wh, ships fine)
}

# Compliance safe-defaults moved to listing/compliance.py in Phase 5 (behaviour
# unchanged): _enum_for, _pick_not_applicable, apply_compliance_safe_defaults + the two
# private constants they use. Functions imported back so they still resolve.
from listing.compliance import _enum_for, _pick_not_applicable, apply_compliance_safe_defaults











def build_flat_row(sheet_row: dict, brand: str, manufacturer: str,
                   cols_map: dict, valid_values: dict,
                   product_type: str, browse_node: str,
                   shipping_group: str = "") -> list:
    total    = cols_map["TOTAL_COLS"]
    out      = [""] * total
    title    = str(sheet_row.get("Title",                ""))[:200]
    upc      = str(sheet_row.get("UPC",                  "")).strip()
    asin     = str(sheet_row.get("Competitor ASIN",      "")).strip()
    sku      = str(sheet_row.get("SKU",                  "")).strip()
    price    = _clean_price(sheet_row.get("Our Price (GBP)", ""))
    desc     = _strip_html(str(sheet_row.get("Description (HTML)", "")))[:2000]
    keywords = str(sheet_row.get("Search Terms / KW",    ""))[:249]
    handling = _clean_days(sheet_row)
    battery  = _has_battery(sheet_row)

    def vv(field_name: str) -> list:
        return _smart_vlist(field_name, valid_values)

    # Constrained fields -- all snapped from live dropdown lists
    material = snap_to_valid(
        str(sheet_row.get("Material", "")).split(",")[0].strip().replace("N/A", ""),
        vv("material"))

    colour_raw = str(sheet_row.get("Colour", "")).replace("N/A", "")
    colour     = snap_to_valid(colour_raw, vv("color")) if vv("color") else colour_raw

    size_raw = str(sheet_row.get("Size", "")).replace("N/A", "")
    # Leave blank when no valid list -- raw values fail dropdown validation
    size     = snap_to_valid(size_raw, vv("size")) if vv("size") else ""

    gender   = snap_to_valid(
        str(sheet_row.get("Target Gender", "Unisex")).replace("N/A", "Unisex"),
        vv("target_gender")) or "Unisex"

    age      = snap_to_valid(
        str(sheet_row.get("Age Range", "Adult")).replace("N/A", "Adult"),
        vv("age_range")) or "Adult"

    condition = snap_to_valid("New",         vv("condition_type"))     or "New"
    fulfill   = snap_to_valid("DEFAULT",     vv("fulfillment_channel")) or "DEFAULT"
    country   = snap_to_valid("China",       vv("country_of_origin"))  or "China"
    batt_yes  = snap_to_valid("Yes",         vv("batteries_required")) or "Yes"
    batt_no   = snap_to_valid("No",          vv("batteries_required")) or "No"
    tax_code  = snap_to_valid("A_GEN_NOTAX", vv("product_tax_code"))   or "A_GEN_NOTAX"

    # Product Id: real barcode if the sheet provides one, else BLANK (GTIN-exempt).
    # Never write the competitor's ASIN -- you cannot list a new product under it.
    digits_only = re.sub(r"[^\d]", "", upc)
    if len(digits_only) == 13:
        prod_id_type = "EAN"
        prod_id      = digits_only
    elif len(digits_only) == 12:
        prod_id_type = "UPC"
        prod_id      = digits_only
    else:
        prod_id_type = ""        # no barcode -> leave blank; requires GTIN exemption on the account
        prod_id      = ""

    # Per-row brand from sheet wins; fall back to the export-level default.
    row_brand        = str(sheet_row.get("Brand", "")).strip()
    effective_brand  = row_brand or brand
    # Per-row model number: blank means category does not require it.
    row_model_number = str(sheet_row.get("Model Number", "")).strip()

    # Title must NOT lead with the brand -- strip it from the start if present
    # (covers rows already generated under the old brand-first prompt).
    title_clean = title
    _bn = effective_brand.strip()
    if _bn and title_clean.lower().startswith(_bn.lower()):
        title_clean = title_clean[len(_bn):].lstrip(" -\u2013\u2014:|,").strip()

    def s(key: str, val):
        idx = cols_map.get(key)
        if idx is not None and idx < total:
            out[idx] = str(val) if val is not None else ""

    s("SKU",                           sku)
    s("Product Type",                  product_type)
    s("Listing Action",                "Create or Replace (Full Update)")
    s("Item Name",                     title_clean)
    s("Brand Name",                    effective_brand)
    s("Product Id Type",               prod_id_type)
    s("Product Id",                    prod_id)
    s("Browse Node 1",                 browse_node)
    if row_model_number:
        s("Model Number",              row_model_number)
        s("model_name",                row_model_number)   # own-brand: mirror model number
        s("part_number",               row_model_number)
    s("Manufacturer",                  manufacturer)
    s("Product Description",           desc)
    s("Bullet Point 1",                str(sheet_row.get("Bullet 1", ""))[:500])
    s("Bullet Point 2",                str(sheet_row.get("Bullet 2", ""))[:500])
    s("Bullet Point 3",                str(sheet_row.get("Bullet 3", ""))[:500])
    s("Bullet Point 4",                str(sheet_row.get("Bullet 4", ""))[:500])
    s("Bullet Point 5",                str(sheet_row.get("Bullet 5", ""))[:500])
    s("Generic Keyword",               keywords)
    s("Material",                      material)
    s("Colour",                        colour)
    s("Size",                          size)
    s("Number of Items",               str(sheet_row.get("Number of Items", "1")) or "1")
    s("Target Gender",                 gender)
    s("Age Range Description",         age)
    s("Item Condition",                condition)
    s("List Price with Tax",           price)
    s("Product Tax Code",              tax_code)
    s("Fulfillment Channel Code (UK)", fulfill)
    s("Quantity (UK)",                 "99")
    s("Handling Time (UK)",            handling)
    s("Your Price GBP",                price)
    s("Country of Origin",             country)
    s("Are batteries required?",       batt_yes if battery else batt_no)
    s("Are batteries included?",       batt_yes if battery else batt_no)
    if shipping_group:
        s("merchant_shipping_group",   shipping_group)

    # --- Pillars 3-4: map the full attribute object generated for this product --
    # Reads the "Attributes JSON" column. Enumerated values are snapped to Amazon's
    # accepted strings (left BLANK if no clean match -- never writes an invalid enum
    # or "N/A" into a dropdown). Free-text values are written as-is. Skips fields
    # already written above so we never double-write.
    _already = {"material", "color", "colour", "size", "number_of_items",
                "country_of_origin", "item_condition", "item_type_keyword",
                "item_length", "item_width", "item_height", "item_depth",
                "item_weight", "length", "width", "height", "depth", "weight",
                "item_package_length", "item_package_width", "item_package_height",
                "item_package_weight", "package_length", "package_width",
                "package_height", "package_weight"}
    try:
        _gen = json.loads(str(sheet_row.get("Attributes JSON", "") or "{}"))
    except Exception:
        _gen = {}
    if isinstance(_gen, dict):
        for _ak, _av in _gen.items():
            _akl = str(_ak).strip().lower()
            if _akl in _already or _av is None or str(_av).strip() == "":
                continue
            if _akl not in cols_map:
                continue                      # template has no column for this attribute
            # Safety & compliance fields: their value comes from the live SP-API
            # schema enum injected into the generation prompt, so it is already a
            # valid Amazon string. Write it directly -- snapping it against the
            # (sometimes incomplete) static valid-values list would wrongly blank
            # a correct answer like "No" or "Not Applicable", or fuzzy-match it to
            # the wrong option (e.g. "Not Applicable" -> "GHS").
            if _akl in _COMPLIANCE_PASSTHROUGH:
                s(_akl, str(_av).strip()[:120])
                continue
            _vlist = vv(_akl)
            if _vlist:                        # enumerated: snap; blank if no clean match
                _snapped = snap_to_valid(str(_av).replace("N/A", "").strip(), _vlist)
                if _snapped:
                    s(_akl, _snapped)
            else:                             # free-text: write value, or N/A (accepted as text)
                _clean = str(_av).strip()
                if _clean:
                    s(_akl, _clean[:500])

    # --- Dimensions: fill a field-GROUP only when every axis it needs has a value.
    # Amazon errors on a partially filled group (e.g. depth/width/height with depth
    # missing), so we gather the measurements we actually have, then for each
    # template group write it ONLY if all its required axes are covered. Item and
    # package scopes are independent.
    _dim_groups = cols_map.get("_DIM_GROUPS") or {}

    def _split_dim(_raw):
        m = re.match(r"\s*(-?[\d.]+)\s*(.*)$", str(_raw).strip())
        if not m:
            return None, None
        return m.group(1).rstrip("."), _norm_dim_unit(m.group(2))

    _have = {"item": {}, "package": {}}       # scope -> axis -> (value, unit)
    _DIM_SOURCES = (
        ("item",    "height", ("item_height", "height")),
        ("item",    "length", ("item_length", "length")),
        ("item",    "width",  ("item_width", "width")),
        ("item",    "depth",  ("item_depth", "depth")),
        ("item",    "weight", ("item_weight", "weight")),
        ("package", "height", ("item_package_height", "package_height")),
        ("package", "length", ("item_package_length", "package_length")),
        ("package", "width",  ("item_package_width", "package_width")),
        ("package", "weight", ("item_package_weight", "package_weight")),
    )
    for _scope, _axis, _src_keys in _DIM_SOURCES:
        for _sk in _src_keys:
            _raw = _gen.get(_sk)
            if _raw and str(_raw).strip():
                _v, _u = _split_dim(_raw)
                if _v is not None:
                    _have[_scope][_axis] = (_v, _u)
                break

    for _scope, _groups in _dim_groups.items():
        for _gkey, _axes in _groups.items():
            _need = [a for a, slots in _axes.items() if slots.get("value")]
            if not _need or not all(a in _have[_scope] for a in _need):
                continue                      # incomplete group -> leave blank
            for _a in _need:
                _v, _u = _have[_scope][_a]
                for _ci in _axes[_a].get("value", []):
                    if not out[_ci]:
                        out[_ci] = _v
                if _u:
                    for _ci in _axes[_a].get("unit", []):
                        if not out[_ci]:
                            out[_ci] = _u

    # --- Compliance safety net: these are near-universal for ordinary retail
    # goods and Amazon BLOCKS the listing when a required one is missing. If the
    # generation step didn't emit them, write the safe default so we never ship a
    # row that fails on an empty compliance dropdown.
    for _ck, _default in (("supplier_declared_dg_hz_regulation", "Not Applicable"),
                          ("contains_liquid_contents", "No")):
        _ci = cols_map.get(_ck)
        if _ci is not None and not out[_ci]:
            out[_ci] = _default

    return out


def write_to_template_sheet(gc, sheet_id: str, data_rows: list,
                              label: str, total_cols: int):
    sh       = gc.open_by_key(sheet_id)
    ws       = sh.worksheet("Template")
    last_col = _col_letter(total_cols - 1)

    console.print(f"  Clearing {label} row 7+...", end=" ")
    try:
        ws.batch_clear([f"A7:{last_col}2000"])
        console.print("[green]OK[/green]")
    except Exception as e:
        console.print(f"[yellow]WARN {str(e)[:50]}[/yellow]")

    if not data_rows:
        console.print(f"  [yellow]No rows for {label}[/yellow]")
        return

    console.print(f"  Writing {len(data_rows)} row(s) to {label}...", end=" ")
    end_row    = 6 + len(data_rows)
    range_name = f"A7:{last_col}{end_row}"
    for attempt in range(1, 4):
        try:
            ws.update(data_rows, range_name, value_input_option="USER_ENTERED")
            console.print("[green]OK[/green]")
            return
        except gspread.exceptions.APIError as e:
            if attempt == 3:
                console.print(f"[red]FAIL {str(e)[:60]}[/red]")
                raise
            time.sleep(attempt * 5)


def run_export(config: dict, gc, status_filter: str = "APPROVED"):
    console.print(f"\n[bold cyan]{'='*55}[/bold cyan]")
    console.print(f"[bold cyan]  FLAT FILE EXPORT -- Status: {status_filter}[/bold cyan]")
    console.print(f"[bold cyan]{'='*55}[/bold cyan]\n")

    brand        = config.get("brand_name",   "Pollinecfecto")
    manufacturer = config.get("manufacturer",  brand)
    t1_id        = config.get("template1_spreadsheet_id")
    t2_id        = config.get("template2_spreadsheet_id")

    if not t1_id or not t2_id:
        console.print("[red]template1_spreadsheet_id or template2_spreadsheet_id missing[/red]")
        return

    console.print("[bold]Step 1:[/bold] Loading dropdown values from templates")
    valid1 = load_dropdown_values(gc, t1_id, "Template 1 (Cookware/Lamp/Hardware)")
    valid2 = load_dropdown_values(gc, t2_id, "Template 2 (Health/Beauty/Kitchen)")

    # Overlay Amazon-published XLSM template values (authoritative)
    static_vv = load_static_valid_values()
    if static_vv:
        before1 = sum(len(a) for a in valid1.values())
        before2 = sum(len(a) for a in valid2.values())
        valid1 = merge_static_into_runtime(valid1, static_vv)
        valid2 = merge_static_into_runtime(valid2, static_vv)
        after1 = sum(len(a) for a in valid1.values())
        after2 = sum(len(a) for a in valid2.values())
        console.print(f"  Static XLSM values merged: "
                      f"{len(static_vv)} product types, "
                      f"T1 attrs {before1} -> {after1}, T2 attrs {before2} -> {after2}")

    console.print("\n[bold]Step 2:[/bold] Reading approved listings")
    sh       = gc.open_by_key(config["google_spreadsheet_id"])
    ws       = sh.worksheet(OUTPUT_TAB)
    all_rows = _safe_records(ws)
    rows     = [r for r in all_rows
                if str(r.get("Status", "")).upper().startswith(status_filter.upper())]
    console.print(f"  {len(all_rows)} total -> [bold]{len(rows)}[/bold] '{status_filter}'")

    if not rows:
        console.print(f"[yellow]No '{status_filter}' rows. Set Status=APPROVED in sheet.[/yellow]")
        return

    console.print(f"\n[bold]Step 3:[/bold] Routing {len(rows)} product(s)")
    file1_rows, file2_rows = [], []
    seen_types = set()
    for row in rows:
        title  = str(row.get("Title",          ""))
        cat    = str(row.get("Amazon Category", ""))
        pt_raw = str(row.get("Product Type",   ""))
        file_id, prod_type, node = detect_route(title, cat, pt_raw)
        if file_id == "FILE1":
            file1_rows.append((row, prod_type, node))
        else:
            file2_rows.append((row, prod_type, node))
        vv = valid1.get(prod_type, {}) if file_id == "FILE1" else valid2.get(prod_type, {})
        if prod_type not in seen_types:
            console.print(f"  [dim]Keys for [{prod_type}]: {sorted(vv.keys())}[/dim]")
            seen_types.add(prod_type)

    console.print(f"  Template 1: [bold]{len(file1_rows)}[/bold] | Template 2: [bold]{len(file2_rows)}[/bold]")

    console.print(f"\n[bold]Step 4:[/bold] Building flat file rows")
    built1 = [build_flat_row(row, brand, manufacturer, FILE1_COLS, valid1.get(pt, {}), pt, node)
              for row, pt, node in file1_rows]
    built2 = [build_flat_row(row, brand, manufacturer, FILE2_COLS, valid2.get(pt, {}), pt, node)
              for row, pt, node in file2_rows]

    console.print(f"\n[bold]Step 5:[/bold] Writing to template sheets")
    if t1_id and built1:
        write_to_template_sheet(gc, t1_id, built1, "Template 1", FILE1_COLS["TOTAL_COLS"])
    if t2_id and built2:
        write_to_template_sheet(gc, t2_id, built2, "Template 2", FILE2_COLS["TOTAL_COLS"])

    console.print(f"\n[bold green]Export complete![/bold green]")
    if t1_id:
        console.print(f"  Template 1: https://docs.google.com/spreadsheets/d/{t1_id}")
    if t2_id:
        console.print(f"  Template 2: https://docs.google.com/spreadsheets/d/{t2_id}")
    console.print("  Open Template tab -> File -> Download -> Tab-separated values (.tsv)")
    console.print("  Seller Central UK -> Inventory -> Add Products via Upload\n")


# =============================================================================
# PROCESS ONE PRODUCT
# =============================================================================

async def process_row(row: dict, client, ws_out,
                       creds: dict, config: dict,
                       idx: int, total: int,
                       user_brand: str,
                       taken_skus: set, seen_asins: set,
                       model_counter: dict,
                       compliance_rules: dict,
                       ip_rules: dict,
                       static_vv: dict,
                       skip_existing: bool = True) -> bool:
    t_total      = Timer()
    # Prefer the brand resolved for THIS run (CLI --brand / the account's own
    # brand). Only fall back to global config when nothing was passed.
    fallback_brand = (user_brand or "").strip() or config.get("brand_name", "Pollinecfecto")
    manufacturer   = (user_brand or "").strip() or config.get("manufacturer", fallback_brand)
    shipping       = float(config.get("shipping_cost", "3.99"))

    amazon_url    = str(row.get("amazon_url",    "")).strip()
    comp_asin     = _extract_asin(amazon_url)
    upc           = str(row.get("upc",           "")).strip()
    item_name     = str(row.get("item_name",     "")).strip()
    source_cost   = float(re.sub(r"[^\d.]", "", str(row.get("source_cost",  "0"))) or 0)
    selling_str   = re.sub(r"[^\d.]", "", str(row.get("selling_price", "0")).split("-")[0])
    selling_price = float(selling_str or 0)
    handling_str  = str(row.get("handling_time", "3")).strip()
    # Range "3-5 Days" -> pick middle (round up), "3 Days" -> 3, blank -> 3
    _nums = [int(n) for n in re.findall(r"\d+", handling_str)]
    if len(_nums) >= 2:
        handling_days = str((_nums[0] + _nums[1] + 1) // 2)   # ceil of average
    elif len(_nums) == 1:
        handling_days = str(_nums[0])
    else:
        handling_days = "3"

    # --- Build SKU in the new format: {price}_{N}Days_{ASIN} -----------------
    # Already processed this competitor ASIN (previous run, or a duplicate input
    # row)? In generate mode, skip it entirely -- don't regenerate or create a _2
    # row. ('retry' passes skip_existing=False so held/errored rows can rebuild.)
    if skip_existing and comp_asin and comp_asin in seen_asins:
        console.print(f"\n[{idx}/{total}] {item_name or comp_asin}")
        console.print(f"  [yellow]SKIP -- ASIN {comp_asin} is already in the sheet "
                      f"(already processed). Delete its row or use 'retry' to rebuild.[/yellow]")
        return None

    notes_parts = []
    if source_cost <= 0:
        notes_parts.append("Missing source price -- SKU price part defaulted to 0.00.")

    sku, was_dup_sku = build_sku(source_cost, handling_days, comp_asin, taken_skus)
    if was_dup_sku and "Duplicate competitor ASIN" not in " ".join(notes_parts):
        notes_parts.append("SKU collision with existing/queued row -- uniqueness suffix appended.")

    if comp_asin:
        seen_asins.add(comp_asin)
    taken_skus.add(sku)

    console.print(f"\n{'='*60}")
    console.print(f"[bold cyan][{idx}/{total}][/bold cyan] {item_name or comp_asin}")
    console.print(f"  ASIN: {comp_asin} | SKU: {sku} | UPC: {upc or 'N/A'}")
    console.print(f"{'='*60}")

    if not comp_asin:
        console.print("[yellow]No ASIN in Amazon URL -- will rely on eBay as the "
                      "source (Amazon used only for price/fees/schema where available).[/yellow]")

    # =====================================================================
    # DATA SOURCING PRIORITY
    #   eBay link present  -> eBay is the SOURCE OF TRUTH for content (title,
    #                         specs, images), source price, and shipping time.
    #                         Amazon (the competitor ASIN) supplies ONLY selling
    #                         price, fees, product type + schema, and gap-fills
    #                         fields eBay didn't provide.
    #   no eBay link       -> Amazon-primary (original behaviour).
    # =====================================================================
    ebay_app_id  = config.get("ebay_app_id",  "")
    ebay_cert_id = config.get("ebay_cert_id", "")
    ebay_url     = str(row.get("ebay_url", "")).strip()
    ebay_first   = bool(ebay_url)   # eBay drives content when a link exists

    ebay_supp = {}
    if ebay_url:
        console.print("[bold]PRE-FLIGHT 1/3:[/bold] Fetching eBay source data (primary)")
        t = Timer()
        ebay_supp = fetch_ebay_supplement(ebay_url, ebay_app_id, ebay_cert_id)
        if ebay_supp.get("title"):
            console.print(f"  [cyan]eBay source: '{ebay_supp['title'][:50]}' | "
                           f"{len(ebay_supp.get('item_specifics', {}))} specs | "
                           f"{ebay_supp.get('image_count', 0)} imgs | {ebay_supp.get('price','')}[/cyan]"
                           f" ({t.elapsed()}s)")
        else:
            console.print(f"  [yellow]eBay returned no usable data for this URL -- "
                           f"falling back to Amazon as source.[/yellow]")
            ebay_first = False

    console.print("[bold]PRE-FLIGHT 1/3:[/bold] Fetching competitor ASIN data"
                  if not ebay_first else
                  "  Fetching Amazon data (price / fees / schema / gap-fill)")
    t            = Timer()
    comp_data    = get_competitor_asin_data(comp_asin, creds)
    pricing      = get_pricing_data(comp_asin, creds)

    # --- Scraper fallback ONLY if Catalog itself returned nothing ------------
    # Old logic fell back on missing Buy Box, but an OOS listing still has a
    # perfectly good Catalog record (title, specs, images) -- pricing is a
    # separate concern. New rule: if Catalog gave us a title, we KEEP its data
    # and only try the scraper to fill in a missing price. If Catalog itself
    # was empty (endpoint stall / denied), scrape the whole page as before.
    _title_ok = bool(comp_data.get("title"))
    _price_ok = pricing.get("buy_box_price", 0) > 0
    _price_src = pricing.get("price_source", "") or ""

    if not _title_ok:
        # Catalog itself failed -- full scrape needed
        console.print(f"  [yellow]SP-API Catalog returned no title (endpoint stall or denied). "
                      f"Full PDP scrape.[/yellow]")
        scraped_cat, scraped_price = await scrape_pdp(comp_asin)
        if not comp_data.get("title") and scraped_cat.get("title"):
            # Preserve product_type from SP-API if it somehow returned one even on partial failure
            if comp_data.get("product_type") and not scraped_cat.get("product_type"):
                scraped_cat["product_type"] = comp_data["product_type"]
                scraped_cat["item_type_keyword"] = comp_data.get("item_type_keyword", "")
            comp_data = scraped_cat
        if pricing.get("buy_box_price", 0) == 0 and scraped_price.get("buy_box_price", 0) > 0:
            pricing = scraped_price
            pricing["price_source"] = "PDP scrape (Buy Box)"
    elif not _price_ok:
        # Title/specs from Catalog are good -- KEEP THEM. Only the price is
        # missing (competitor likely OOS or Buy Box suppressed). Try scraping
        # ONLY for a price, and only merge that price back. Do NOT overwrite
        # Catalog's rich title/specs with a thin scrape.
        _offer_n = pricing.get("offer_count", 0)
        _msg = ("competitor appears OUT OF STOCK (no Buy Box, no active offers)"
                if _offer_n == 0 else
                f"no Buy Box winner ({_offer_n} offer(s) exist but none won)")
        console.print(f"  [yellow]Catalog data is good, but no price -- {_msg}. "
                      f"Trying PDP for a last-known / strikethrough price...[/yellow]")
        try:
            _, scraped_price = await scrape_pdp(comp_asin)
            if scraped_price.get("buy_box_price", 0) > 0:
                pricing["buy_box_price"] = scraped_price["buy_box_price"]
                pricing["price_source"] = "PDP scrape (was-price / strikethrough)"
                console.print(f"  [green]Got a price from PDP: "
                              f"{_cur_sym()}{pricing['buy_box_price']:.2f} "
                              f"({pricing['price_source']}).[/green]")
            else:
                console.print(f"  [yellow]No price found on PDP either. Keeping Catalog data; "
                              f"the row will need a manual price -- generation continues.[/yellow]")
        except Exception as _pe:
            console.print(f"  [yellow]PDP price probe failed: {str(_pe)[:80]}. "
                          f"Keeping Catalog data; generation continues.[/yellow]")
    else:
        # Best case: title + Buy Box price both fine. Nothing to fall back to.
        console.print(f"  [green]SP-API OK (price source: {_price_src}).[/green]")

    product_type = comp_data.get("product_type", "") or ""
    # When neither SP-API nor the PDP scrape produced a product type (e.g. the
    # SP-API app lacks the Catalog Items role, so catalogue + scrape both came
    # back empty), infer one from the title/specs so downstream never sees a
    # blank or the invalid literal "PRODUCT".
    if not product_type or product_type.upper() == "PRODUCT":
        try:
            _vv = load_static_valid_values() if "load_static_valid_values" in globals() else None
        except Exception:
            _vv = None
        product_type = infer_product_type(comp_data, item_name=item_name, valid_types=_vv)
        comp_data["product_type"] = product_type
        console.print(f"  [cyan]Product type inferred from content: {product_type}[/cyan]")

    # --- eBay-first merge -----------------------------------------------------
    # When an eBay link drove this row (ebay_first), eBay content is AUTHORITATIVE:
    # its title, specs, and images become the primary product data. Amazon's
    # catalogue data stays attached for Claude to REFERENCE and to supply fields
    # eBay lacks (product type, schema). When there is no eBay link we keep the
    # original Amazon-primary behaviour (eBay only patches junk titles / missing
    # images, exactly as before).
    # PROVENANCE: remember which SOURCE supplied each attribute value so the
    # dashboard can tag each box (eBay / Amazon / AI). eBay specs are tagged
    # 'ebay'; Amazon-catalogue attrs already in comp_data are 'amazon'; anything
    # the AI adds later is tagged 'ai' at write time.
    _prov_map = {}
    _amazon_attr_keys = set((comp_data.get("attributes") or {}).keys())
    if ebay_supp.get("title"):
        comp_data["_ebay_supplement"] = ebay_supp
        if ebay_first:
            # eBay is the source of truth -> seed content from it.
            comp_data["title"] = ebay_supp["title"]
            # Merge specs: eBay specifics win; keep Amazon attrs as extra reference.
            _merged = dict(comp_data.get("attributes") or {})
            _ebay_keys = set()
            for _k, _v in (ebay_supp.get("item_specifics") or {}).items():
                if _v:
                    _merged[_k] = str(_v)[:400]   # eBay value takes precedence
                    _ebay_keys.add(_k)
            comp_data["attributes"] = _merged
            # tag sources: eBay keys -> ebay; remaining pre-existing -> amazon
            for _k in _merged:
                _prov_map[_k] = "ebay" if _k in _ebay_keys else ("amazon" if _k in _amazon_attr_keys else "amazon")
            if ebay_supp.get("images"):
                comp_data["images"] = ebay_supp["images"][:5]   # eBay images primary
            console.print(f"  [cyan]eBay-first: content/specs/images sourced from eBay; "
                           f"Amazon used for price/fees/type[/cyan]")
        else:
            # Amazon-primary fallback: original junk-patch behaviour.
            amzn_title = comp_data.get("title", "") or ""
            junk_re = re.compile(r"keyboard shortcut|shift\s*\+|alt\s*\+|opt\s*\+", re.IGNORECASE)
            if not amzn_title or len(amzn_title) < 25 or junk_re.search(amzn_title):
                console.print(f"  [yellow]Amazon title looks bad -- using eBay title instead[/yellow]")
                comp_data["title"] = ebay_supp["title"]
            if not comp_data.get("images") and ebay_supp.get("images"):
                comp_data["images"] = ebay_supp["images"][:5]
                console.print(f"  [yellow]No Amazon images -- using {len(comp_data['images'])} from eBay[/yellow]")
    elif ebay_url and ebay_app_id and ebay_cert_id:
        console.print(f"  [dim]eBay data unavailable for this URL[/dim]")

    # If we did NOT go eBay-first, every existing attribute came from Amazon.
    if not _prov_map:
        for _k in (comp_data.get("attributes") or {}):
            _prov_map[_k] = "amazon"
    comp_data["_provenance"] = _prov_map

    # Ensure we always have *some* title (eBay-only rows where Amazon was empty).
    if not comp_data.get("title"):
        comp_data["title"] = item_name or (ebay_supp.get("title") or "")

    # eBay is the source of truth for SOURCE PRICE: if the input sheet had no
    # source cost, take it from the eBay listing price.
    if source_cost <= 0 and ebay_supp.get("price"):
        _ep = re.sub(r"[^\d.]", "", str(ebay_supp["price"]))
        try:
            if _ep and float(_ep) > 0:
                source_cost = float(_ep)
                console.print(f"  [cyan]Source price taken from eBay listing: {source_cost}[/cyan]")
        except ValueError:
            pass

    console.print(f"  '{(comp_data.get('title') or '')[:55]}' | Type: {product_type} | "
                   f"{len(comp_data.get('attributes', {}))} specs | "
                   f"{len(comp_data.get('images', []))} images")

    console.print("[bold]PRE-FLIGHT 2/3:[/bold] Market pricing + fees")
    t       = Timer()
    # ------------------------------------------------------------------
    # User's pricing rule: max(floor, competitor Buy Box)
    #   floor = cost + Amazon fees + £3 ship + £2 ads + £1 profit
    # If the input sheet already has a manual selling_price, that WINS
    # (respects the user's explicit override).
    # ------------------------------------------------------------------
    if selling_price <= 0:
        _competitor = float(pricing.get("buy_box_price", 0) or 0)
        # SHIPPING LABEL: use the per-row 'shipping' variable that the rest of
        # the app already tracks so it isn't double-counted here AND in
        # calculate_financials below. Falls back to the constant only if the
        # row has no shipping cost set (rare).
        _ship_cost = float(shipping) if float(shipping) > 0 else PRICING_RULE_SHIPPING_LABEL
        # Two-pass fee convergence: Amazon's referral fee is a % of the selling
        # price, so we need to seed with a rough fee then recompute at the new
        # price. Two passes is enough for standard 15% referrals to converge to
        # within a penny.
        _seed_price = max(_competitor, source_cost + 6.0) if source_cost > 0 else max(_competitor, 10.0)
        _fees_seed  = get_fees(comp_asin, _seed_price, creds)
        _r1 = compute_selling_price(source_cost, _fees_seed["total_amazon_fees"],
                                    _competitor, shipping_label=_ship_cost)
        # Pass 2: refetch fees at r1's price, recompute
        _fees_final = get_fees(comp_asin, _r1["selling_price"], creds)
        _r2 = compute_selling_price(source_cost, _fees_final["total_amazon_fees"],
                                    _competitor, shipping_label=_ship_cost)
        selling_price = _r2["selling_price"]
        fees = _fees_final
        console.print(f"  [cyan]Pricing rule -> {_cur_code()}{selling_price:.2f} "
                      f"({_r2['rule_source']}) | {_r2['breakdown']}[/cyan]")
        if _r2["rule_source"].startswith("floor") and _competitor > 0:
            _gap = selling_price - _competitor
            console.print(f"  [yellow]  Note: competitor Buy Box is {_cur_code()}{_competitor:.2f}, "
                          f"but our floor is {_cur_code()}{selling_price:.2f} "
                          f"({_cur_code()}{_gap:.2f} higher). Won't be Buy-Box competitive at this price.[/yellow]")
    else:
        fees = get_fees(comp_asin, selling_price, creds)
        console.print(f"  [dim]Using input-sheet price {_cur_code()}{selling_price:.2f} (manual override)[/dim]")
    financials = calculate_financials(source_cost, selling_price, shipping, fees)
    vc         = "green" if financials["viable"] == "YES" else "red"
    console.print(
        f"  Buy Box: {_cur_code()}{pricing['buy_box_price']:.2f} | "
        f"Our Price: {_cur_code()}{selling_price:.2f} | "
        f"Fees: {_cur_code()}{fees['total_amazon_fees']:.2f} ({fees['fee_source']}) | "
        f"[{vc}]Profit: {_cur_code()}{financials['profit']:.2f} | "
        f"Margin: {financials['margin_pct']} [{financials['viable']}][/{vc}] "
        f"({t.elapsed()}s)")

    console.print("[bold]PRE-FLIGHT 3/3:[/bold] Attribute schema")
    t      = Timer()
    schema = get_product_type_schema(product_type, creds)
    if schema and (schema.get("required") or schema.get("properties")):
        console.print(f"  live schema loaded ({t.elapsed()}s)")
    else:
        # SP-API live schema is optional. Attribute enforcement still comes from
        # the local XLSM valid-values already loaded at startup, so generation
        # is unaffected -- this is an enhancement, not a requirement.
        console.print(f"  using local attribute data (live schema optional) ({t.elapsed()}s)")

    # --- Brand selection per product ----------------------------------------
    if user_brand:
        chosen_brand = user_brand
    else:
        # Prefer the seller's own configured brand; fall back to the schema's
        # enforced list (or Unbranded) only when no brand is configured.
        chosen_brand = config.get("brand_name", "").strip() or pick_brand_for_product(schema)
        console.print(f"  Brand: [bold]{chosen_brand}[/bold]")

    # --- Model Number: generate for own-brand listings ----------------------
    # Amazon flags Model Number / Model Name / Part Number as recommended even
    # when not strictly required. For own-brand products we generate a stable
    # model number ({Brand}-{Cat}-{seq:03d}) so those fields populate.
    model_number = ""
    _is_own_brand = bool(chosen_brand) and chosen_brand.strip().lower() not in ("unbranded", "generic")
    if is_model_number_required(schema) or _is_own_brand:
        model_number = next_model_number(chosen_brand, product_type, model_counter)
        save_model_counter(model_counter)
        console.print(f"  Model Number generated: [bold]{model_number}[/bold]")
    else:
        console.print("  Model Number not required -- leaving blank")

    console.print("[bold]STEP 1/3:[/bold] Keywords")
    t         = Timer()
    core_term = extract_core_search_term(comp_data.get("item_type_keyword", "") or item_name)
    keywords  = get_autocomplete_keywords(core_term)
    console.print(f"  {len(keywords)} autocomplete keywords for '{core_term}'")
    # Review scraping removed: crawl4ai never reliably returned Amazon reviews and
    # cost ~45s per listing. Claude writes from category knowledge + the listing
    # data instead, which is what actually happened on the timeouts anyway.
    voc_data = {"reviews": [], "source": "none", "review_count": 0}

    console.print("[bold]STEP 2/3:[/bold] Claude listing generation")
    try:
        listing = generate_listing(
            client, comp_data, pricing, financials,
            keywords, voc_data, schema,
            chosen_brand, manufacturer, selling_price,
            handling_str, item_name, static_vv)

        title = listing.get("title", "")
        if len(title) > TITLE_MAX_CHARS:
            listing["title"] = cap_chars(title, TITLE_MAX_CHARS)
            console.print(f"  [yellow]Title trimmed to {TITLE_MAX_CHARS} chars[/yellow]")

        # Item Highlights cap (new field, app-wide)
        if listing.get("item_highlights"):
            listing["item_highlights"] = cap_chars(listing["item_highlights"], HIGHLIGHTS_MAX)

        st = listing.get("search_terms", "")
        cleaned = clean_search_terms(st)
        if cleaned != st:
            listing["search_terms"] = cleaned
            console.print("  [yellow]Search terms cleaned (no punctuation) + 249-byte capped[/yellow]")

        kw_in_copy = len([k for k in keywords[:5]
                          if k["keyword"].lower() in
                          (listing.get("title","") + listing.get("bullet_1","")).lower()])
        console.print(f"  Category: {listing.get('amazon_category','')} | "
                       f"KW in copy: {kw_in_copy}")
        status = "NEEDS_REVIEW"
    except Exception as e:
        console.print(f"  [red]Claude failed: {str(e)[:100]}[/red]")
        listing = {"title": item_name,
                   "amazon_category": "", "amazon_subcategory": "",
                   "target_demographic": "", "pain_points": "",
                   "purchase_trigger": "", "bullet_1": "", "bullet_2": "",
                   "bullet_3": "", "bullet_4": "", "bullet_5": "",
                   "description": "", "search_terms": "", "material": "",
                   "colour": "", "size": "", "number_of_items": "",
                   "target_gender": "Unisex", "age_range": "Adult",
                   "compliance_notes": "None"}
        status = "ERROR"

    console.print("[bold]STEP 3/3:[/bold] Writing to Google Sheet")

    # --- Compliance check: keyword-match listing against compliance_rules ----
    comp_result = check_compliance(item_name, listing, compliance_rules)
    if comp_result["matched_categories"]:
        notes_parts.append(comp_result["summary"])
        # Append a short version of requirements (max 3 lines, full list in JSON for reference)
        top_reqs = comp_result["requirements"][:3]
        if top_reqs:
            notes_parts.append("Key reqs: " + " // ".join(top_reqs))
        console.print(f"  [yellow]Compliance: {comp_result['summary']}[/yellow]")
        if comp_result["highest_risk"] == "HIGH" and status == "NEEDS_REVIEW":
            status = "COMPLIANCE_HOLD"
            console.print(f"  [red]Status downgraded to COMPLIANCE_HOLD -- HIGH risk category[/red]")

    # --- IP / trademark check ------------------------------------------------
    ip_result = check_ip_violations(listing, chosen_brand, ip_rules)
    ip_risk_level = ""
    if ip_result["has_violations"]:
        notes_parts.append(ip_result["summary"])
        console.print(f"  [red]{ip_result['summary']}[/red]")
        ip_risk_level = "HIGH"
        # IP_HOLD supersedes COMPLIANCE_HOLD and NEEDS_REVIEW. ERROR stays.
        if status in ("NEEDS_REVIEW", "COMPLIANCE_HOLD"):
            status = "IP_HOLD"
            console.print(f"  [red]Status set to IP_HOLD -- brand/trademark risk[/red]")

    notes_text = " | ".join(notes_parts)
    row_data = build_sheet_row(
        comp_asin, row, listing, comp_data,
        financials, pricing, voc_data, keywords,
        handling_str, handling_days, sku, status,
        brand=chosen_brand, model_number=model_number, notes=notes_text,
        compliance_risk=comp_result["highest_risk"], ip_risk=ip_risk_level)
    ok = sheet_write_row(ws_out, row_data, comp_asin)
    if ok:
        console.print(f"  [green]OK[/green] Written | Total: {t_total.elapsed()}s")
    return ok


# =============================================================================
# MAIN
# =============================================================================

# =============================================================================
# SP-API DIRECT LISTING  (templateless)  -- VALIDATION_PREVIEW first, then submit
#   modes:  api          -> validate every APPROVED row, create NOTHING, write
#                           issues back to the sheet (Status API_READY / API_ERROR)
#           api submit   -> actually create/replace the listing for each row that
#                           is APPROVED or API_READY  (Status -> LIVE / API_ERROR)
#   Attributes are built straight from the generated data and shaped by the LIVE
#   getDefinitions schema, so only attributes that apply to the product type are
#   sent, with the exact value/unit/enum/language structure Amazon expects.
# =============================================================================

LANG = "en_GB"   # legacy default; real tag is derived per-marketplace via _lang_for()


# _lang_for moved to listing/shaper.py in Phase 5 (uses only the immutable US_MARKETPLACE_ID;
# behaviour unchanged). Imported so amazon_listing_generator._lang_for still resolves.
from listing.shaper import _lang_for


def _raw_schema(product_type: str, creds: dict):
    """Return (properties, required_set, raw_schema) for a product type, or ({},set(),{}).
    Two steps: (1) getDefinitions returns a CDN link to the schema JSON;
    (2) download that JSON. Step 2 can be slow/large, so it gets a generous
    timeout + one retry. Locale + marketplaceIds follow the ACTIVE marketplace."""
    _locale = "en_US" if MARKETPLACE_ID == US_MARKETPLACE_ID else "en_GB"
    try:
        ptd  = ProductTypeDefinitions(credentials=creds, marketplace=MARKETPLACE, timeout=30)
        resp = ptd.get_definitions_product_type(
            productType=product_type, requirements="LISTING",
            requirementsEnforced="ENFORCED", locale=_locale,
            marketplaceIds=[MARKETPLACE_ID])
        link = resp.payload.get("schema", {}).get("link", {}).get("resource", "")
        if not link:
            console.print(f"  [yellow]schema for {product_type}: definitions returned no link[/yellow]")
            return {}, set(), {}
    except Exception as e:
        # This is the metadata call -- if THIS fails it's a real auth/role issue.
        console.print(f"  [red]schema definitions call failed for {product_type}: {str(e)[:120]}[/red]")
        return {}, set(), {}

    # Validate the CDN link before we try to download it. A blank or malformed
    # host produces the SAME "getaddrinfo failed" error as a DNS block, so check
    # it up front and report which case it actually is.
    try:
        from urllib.parse import urlparse as _up0
        _lh = _up0(link).hostname or ""
    except Exception:
        _lh = ""
    if not _lh:
        console.print(f"  [red]schema link for {product_type} has no valid host "
                      f"(link='{str(link)[:80]}'). Amazon returned an unusable schema URL; "
                      f"retry the Preview -- this is usually transient.[/red]")
        return {}, set(), {}

    # Step 2: download the schema JSON from the CDN link. Large file + slow CDN
    # -> long timeout and a retry so a transient stall doesn't kill the row.
    last_err = ""
    for attempt in range(3):
        try:
            _req = urllib.request.Request(link, headers={"Accept": "application/json"})
            with urllib.request.urlopen(_req, timeout=60) as r:
                raw = json.loads(r.read().decode("utf-8"))
            _props = raw.get("properties", {}) or {}
            _required = set(raw.get("required", []) or [])
            # ENFORCED mode often returns a slim `properties` that omits the enum
            # definitions for required fields (battery, light_source, ...). Fetch
            # the schema AGAIN without enforcement to recover the FULL property
            # defs (with their enum lists) and merge any missing ones in, so the
            # backfill can snap to real allowed values instead of guessing.
            try:
                ptd2 = ProductTypeDefinitions(credentials=creds, marketplace=MARKETPLACE, timeout=30)
                resp2 = ptd2.get_definitions_product_type(
                    productType=product_type, requirements="LISTING",
                    locale=_locale, marketplaceIds=[MARKETPLACE_ID])
                link2 = resp2.payload.get("schema", {}).get("link", {}).get("resource", "")
                if link2:
                    _r2 = urllib.request.Request(link2, headers={"Accept": "application/json"})
                    with urllib.request.urlopen(_r2, timeout=60) as r2:
                        raw2 = json.loads(r2.read().decode("utf-8"))
                    full_props = raw2.get("properties", {}) or {}
                    for _k, _v in full_props.items():
                        # prefer the fuller def when the enforced one is missing/blank
                        if _k not in _props or not _props.get(_k):
                            _props[_k] = _v
                        elif isinstance(_props.get(_k), dict) and isinstance(_v, dict):
                            # merge enum info if the enforced def lacks it
                            if "items" not in _props[_k] and "items" in _v:
                                _props[_k] = _v
            except Exception:
                pass  # if the second fetch fails, proceed with the enforced props
            # PERMANENT FIX: pull allowed values out of the schema's conditional
            # branches (allOf/anyOf/oneOf/if-then-else) and merge them into props,
            # so fields like battery_installation_device_type expose their REAL
            # enum instead of looking like free-text. Done once here -> every part
            # of the app (dropdowns, snapping, hints) benefits automatically.
            try:
                _props = _merge_conditional_enums(_props, raw)
            except Exception as _e:
                console.print(f"  [yellow]conditional-enum merge skipped: {str(_e)[:80]}[/yellow]")
            return _props, _required, raw
        except Exception as e:
            last_err = str(e)
            # Extract the host we were trying to reach so a DNS failure points at
            # the EXACT host (the schema CDN), not a vague "network problem".
            _host = ""
            try:
                from urllib.parse import urlparse as _up
                _host = _up(link).hostname or ""
            except Exception:
                _host = ""
            _is_dns = ("getaddrinfo" in last_err or "11002" in last_err or "11001" in last_err
                       or "Name or service not known" in last_err or "nodename nor servname" in last_err)
            if _is_dns:
                # DNS failures are not transient speed issues; retrying rarely helps
                # within the same run. Report precisely and stop early.
                console.print(f"  [red]schema host DNS lookup failed: {_host or 'schema CDN'} "
                              f"(Errno 11002 getaddrinfo). The API host resolved but this CDN host did not.[/red]")
                console.print(f"  [yellow]Most likely: a VPN/proxy or DNS filter is blocking '{_host}'. "
                              f"Try: disable VPN/proxy, run 'ipconfig /flushdns', or switch DNS to 1.1.1.1 / 8.8.8.8.[/yellow]")
                break
            if attempt < 2:
                console.print(f"  [yellow]schema download slow for {product_type} "
                              f"(attempt {attempt+1}/3) -- retrying...[/yellow]")
                time.sleep(2)
                continue
    console.print(f"  [red]schema download failed for {product_type}: {last_err[:90]}[/red]")
    return {}, set(), {}


def _merge_conditional_enums(props: dict, raw: dict) -> dict:
    """Amazon hides many fields' REAL allowed values inside conditional branches
    (allOf / anyOf / oneOf / if-then-else) of the schema, NOT in top-level
    `properties`. The loader used to read only `properties`, so such fields looked
    like free-text (e.g. battery_installation_device_type) even though Amazon
    validates them server-side. This walks the WHOLE schema, collects every enum
    found for each field across ALL branches, and injects the union into
    props[field] so the rest of the app (dropdowns, snapping, hints) sees the real
    list. Purely additive: existing enums are preserved; we only fill gaps/extend.
    """
    # 1) gather: field_name -> set of allowed values (from anywhere in the doc)
    found = {}   # field -> list (order-preserving)

    def _add(field, values):
        if not values:
            return
        bucket = found.setdefault(field, [])
        for v in values:
            sv = str(v)
            if sv not in bucket:
                bucket.append(sv)

    def _enum_under_value(node):
        """Given a field-definition node, return enum at items.properties.value.enum
        (and a few variants), searching simple anyOf wrappers too."""
        out = []
        if not isinstance(node, dict):
            return out
        it = node.get("items", {})
        ip = it.get("properties", {}) if isinstance(it, dict) else {}
        vp = ip.get("value", {}) if isinstance(ip, dict) else {}
        # direct
        if isinstance(vp, dict) and isinstance(vp.get("enum"), list):
            out += vp["enum"]
        # anyOf/oneOf wrappers around value
        for key in ("anyOf", "oneOf", "allOf"):
            for sub in (vp.get(key) or []) if isinstance(vp, dict) else []:
                if isinstance(sub, dict) and isinstance(sub.get("enum"), list):
                    out += sub["enum"]
        # some defs put enum straight on items or the node
        if isinstance(it, dict) and isinstance(it.get("enum"), list):
            out += it["enum"]
        if isinstance(node.get("enum"), list):
            out += node["enum"]
        return out

    def _walk(node):
        if isinstance(node, dict):
            # if this dict is a `properties` map, each key is a field name
            props_map = node.get("properties")
            if isinstance(props_map, dict):
                for fname, fdef in props_map.items():
                    vals = _enum_under_value(fdef)
                    if vals:
                        _add(fname, vals)
            for v in node.values():
                _walk(v)
        elif isinstance(node, list):
            for v in node:
                _walk(v)

    _walk(raw)

    # 2) inject: ensure props[field] carries the discovered enum at the standard
    #    location the rest of the app reads (items.properties.value.enum).
    for field, values in found.items():
        if not values:
            continue
        cur = props.get(field)
        if not isinstance(cur, dict):
            cur = {}
        it = cur.setdefault("items", {})
        if not isinstance(it, dict):
            it = {}; cur["items"] = it
        ip = it.setdefault("properties", {})
        if not isinstance(ip, dict):
            ip = {}; it["properties"] = ip
        vp = ip.setdefault("value", {})
        if not isinstance(vp, dict):
            vp = {}; ip["value"] = vp
        existing = vp.get("enum")
        if isinstance(existing, list) and existing:
            # extend without dupes (existing wins ordering)
            merged = list(existing)
            for v in values:
                if v not in merged:
                    merged.append(v)
            vp["enum"] = merged
        else:
            vp["enum"] = values
        props[field] = cur
    return props


def _try_fetch_seller_id(creds: dict) -> str:
    """Best-effort auto-fetch of the merchant/seller id. Falls back to '' so the
    caller can ask for config['seller_id']."""
    try:
        from sp_api.api import Sellers
        s = Sellers(credentials=creds, marketplace=MARKETPLACE, timeout=30)
        resp = s.get_account()
        p = resp.payload if hasattr(resp, "payload") else {}
        if isinstance(p, dict):
            for k in ("sellerId", "merchantId", "amazonMerchantId", "merchant_id"):
                if p.get(k):
                    return str(p[k])
    except Exception:
        pass
    return ""


# ---- value shaping helpers ---------------------------------------------------

# _NA/_is_blank/_truthy moved to listing/builder.py in Phase 5 (self-contained; behaviour
# unchanged). _NA is private to _is_blank (used nowhere else), so it moved too.
from listing.builder import _is_blank, _truthy


# _split_value_unit moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _split_value_unit


# _norm_tok moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _norm_tok


# _snap_enum moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _snap_enum


# _coerce_value moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _coerce_value


# _item_props moved to listing/builder.py in Phase 5 (self-contained; behaviour unchanged).
from listing.builder import _item_props


# _shape_simple moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _shape_simple


# _shape_dimensions moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _shape_dimensions


# _shape_axes moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _shape_axes


def _shape_list_price(field_schema: dict, price, mid: str):
    """List Price (RRP). Defaults to the selling price; shape adapts to schema
    (value vs value_with_tax, optional currency / marketplace_id). When the schema
    omits list_price entirely, fall back to a safe UK default {currency, value}."""
    ip = _item_props(field_schema)
    try:
        val = round(float(str(price)), 2)
    except Exception:
        return []
    _def_cur = "USD" if MARKETPLACE_ID == US_MARKETPLACE_ID else "GBP"
    if not ip:                       # schema has no list_price shape -> safe default
        return [{"currency": _def_cur, "value": val}]
    o = {}
    if "marketplace_id" in ip:
        o["marketplace_id"] = mid
    if "currency" in ip:
        cp = ip["currency"]
        o["currency"] = _snap_enum(cp.get("enum"), _def_cur) if cp.get("enum") else _def_cur
    if "value_with_tax" in ip:
        o["value_with_tax"] = val
    else:
        o["value"] = val
    return [o] if ("value" in o or "value_with_tax" in o) else []


# _shape_weight moved to listing/shaper.py in Phase 5 (behaviour unchanged).
from listing.shaper import _shape_weight


# _offer/_fulfillment moved to listing/builder.py in Phase 5 (self-contained; _fulfillment
# calls the now-module-local _is_blank there; behaviour unchanged).
from listing.builder import _offer, _fulfillment


def _issue_str(issues, sent_attrs: dict = None) -> str:
    """Format Amazon's listing issues. Amazon reports a field with a MALFORMED
    value using the same "X is required but missing" text it uses for a truly
    empty required field -- which is misleading. When we know we actually sent a
    value for that field (it's in sent_attrs), we relabel it so the user isn't
    sent hunting for an empty field that isn't empty."""
    sent_attrs = sent_attrs or {}
    parts = []
    for x in issues:
        sev = str(x.get("severity", "?"))[:1].upper()
        an  = x.get("attributeNames") or []
        a   = an[0] if an else ""
        msg = x.get("message", "")
        # misleading-error rewrite: we DID send this attribute, yet Amazon says
        # "required but missing" -> it's really a structure/format problem.
        if a and a in sent_attrs and "required but missing" in msg.lower():
            msg = (f"value was sent but Amazon rejected its STRUCTURE/format "
                   f"(reported as '{msg.strip()}') -- the field is not actually "
                   f"empty; its shape didn't match Amazon's schema.")
        parts.append(f"[{sev}] {a} {msg}".strip())
    # Keep generous room so all errors are stored (was 1500 -> cut off ~8+ errors,
    # making the sheet/dashboard show fewer than the terminal).
    return "; ".join(parts)[:6000]


# ---- attribute payload builder ----------------------------------------------

_ATTR_DEFAULTS_CACHE = {"data": None}


def _load_attr_defaults() -> dict:
    """Per-product-type remembered defaults (attribute_defaults.json), so judgment
    fields (size, is_assembly_required, ...) don't have to be re-entered for every
    new listing of the same type. Shape: {product_type: {attr: value}}."""
    if _ATTR_DEFAULTS_CACHE["data"] is None:
        try:
            _ATTR_DEFAULTS_CACHE["data"] = json.load(open(CONFIG_PATH.parent / "attribute_defaults.json", encoding="utf-8"))
        except Exception:
            _ATTR_DEFAULTS_CACHE["data"] = {}
    return _ATTR_DEFAULTS_CACHE["data"] or {}


# _build_ghs_from_schema moved to listing/hazmat.py in Phase 5 (self-contained; behaviour
# unchanged). Imported so amazon_listing_generator._build_ghs_from_schema still resolves.
from listing.hazmat import _build_ghs_from_schema



def _verify_live_status(li, seller_id, sku, mid, locale="en_GB", settle=True):
    """After a SUBMIT is 'accepted', Amazon processes the listing ASYNCHRONOUSLY --
    'accepted' is NOT 'published'. Query the REAL listing state so a row is marked
    LIVE only when Amazon actually shows it BUYABLE/DISCOVERABLE, and reflects a
    downstream rejection (e.g. a blocked main image) instead of a false LIVE.
    Returns (status_list, error_issues); (None, None) if the check itself failed.

    settle=True waits a few seconds first (right after a fresh submit, Amazon needs a
    moment). Pass settle=False when RE-verifying an already-submitted listing minutes
    later -- there's nothing to wait for, so skip the delay and check immediately."""
    import time as _t
    for _attempt in range(2):
        try:
            if settle:
                _t.sleep(4)   # give Amazon a moment to process the submission
            resp = li.get_listings_item(seller_id, sku, marketplaceIds=[mid],
                                        issueLocale=locale,
                                        includedData=["summaries", "issues"])
            p = resp.payload if hasattr(resp, "payload") else (resp or {})
            summaries = (p or {}).get("summaries", []) or []
            status = summaries[0].get("status", []) if summaries else []
            issues = (p or {}).get("issues", []) or []
            errs = [x for x in issues if str(x.get("severity", "")).upper() == "ERROR"]
            return status, errs
        except Exception:
            continue
    return None, None


def _skus_across_all_tabs(ws_out) -> set:
    """Union of the SKU column across EVERY worksheet in ws_out's spreadsheet.

    Miles listing copies are split across category tabs (e.g. Sheet2, 'Hydraulic
    Fluids', 'Synthetic Gallon-ISEL'). Generation must skip an item whose copy
    already exists on ANY tab -- not just the one it writes to -- or a re-run
    duplicates a listing that was generated onto a different tab.

    Robust to malformed tabs: a sheet with duplicate/blank headers (e.g. an old
    'Sheet1' whose SKU column sits under blank leading headers) can't be read as
    records, so we fall back to the raw grid -- pulling the 'SKU' column by header
    position, or, if there's no SKU header, scanning for item-number-shaped cells
    so nothing is missed."""
    out = set()
    try:
        worksheets = ws_out.spreadsheet.worksheets()
    except Exception as e:
        console.print(f"  [yellow]Cross-tab SKU scan unavailable: {str(e)[:80]}[/yellow]")
        return out
    _itempat = re.compile(r"^[A-Za-z]{2,4}\d{4,}$")
    for ws in worksheets:
        try:
            for r in _safe_records(ws):
                v = str(r.get("SKU", "") or "").strip()
                if v:
                    out.add(v)
            continue
        except Exception:
            pass
        # Fallback: malformed/duplicate headers -> read the raw grid.
        try:
            vals = ws.get_all_values()
        except Exception:
            continue
        if not vals:
            continue
        hdr = [str(c).strip().lower() for c in vals[0]]
        if "sku" in hdr:
            ci = hdr.index("sku")
            for row in vals[1:]:
                if ci < len(row) and str(row[ci]).strip():
                    out.add(str(row[ci]).strip())
        else:
            for row in vals:
                for c in row:
                    if _itempat.match(str(c).strip()):
                        out.add(str(c).strip())
    return out


# shape_by_schema moved to listing/shaper.py in Phase 5 (self-contained; behaviour
# unchanged). Imported here so amazon_listing_generator.shape_by_schema still resolves
# for every internal caller and any external importer.
from listing.shaper import shape_by_schema

def build_api_attributes(row: dict, pt: str, props: dict, required: set, config: dict) -> dict:
    """Assemble the SP-API 'attributes' object for one listing, gated to `props`
    (the live schema for this product type) so nothing inapplicable is sent."""
    mid = MARKETPLACE_ID
    A   = {}
    g   = lambda k: str(row.get(k, "") or "").strip()

    try:
        pa = json.loads(row.get("Attributes JSON", "") or "{}")
    except Exception:
        pa = {}
    if not isinstance(pa, dict):
        pa = {}
    # _provenance is dashboard metadata (which source filled each field) -- never
    # send it to Amazon.
    pa.pop("_provenance", None)
    pa.pop("provenance", None)

    # Re-nest flat dot-keys saved by the dashboard's nested sub-field editor.
    # e.g. "battery.weight.value":"180" + "battery.weight.unit":"grams"
    #   -> "battery": {"weight": {"value":"180","unit":"grams"}}
    # The downstream wrapping (array-of-one + marketplace_id) is applied later by
    # the normal attribute handling; here we only rebuild the object shape.
    def _renest(flat: dict) -> dict:
        """Re-nest flat dot-keys into an object tree.

        Handles COLLISION between keys of different depths gracefully:
        e.g. if `flat` contains BOTH `leg.length = "feet"` (an old shallow key
        from a prior schema-extractor version) AND `leg.length.decimal_value =
        "50.0"` + `leg.length.unit = "feet"` (deeper keys from the current
        extractor version), the deeper keys win because they're strictly more
        specific -- the shallow scalar gets promoted to a dict node with the
        scalar preserved under a synthetic `.value` sub-key (so no data is
        silently dropped).

        Before this defensiveness, `cur.setdefault(p, {})` returned the
        existing scalar; the next iteration crashed with 'str object does
        not support item assignment' as soon as the sheet accumulated keys
        at multiple depths -- which was inevitable once the extractor
        started walking deeper on each fix. See the assert-strings that
        Amazon returned from prior runs mixed with the new decimal_value/
        unit sub-keys."""
        nested, plain = {}, {}
        # Iterate shortest-key-first so shallow entries are placed as leaves
        # first, then get PROMOTED to dicts when a deeper sibling arrives.
        # (If we ran longest-first, the deeper writes would land in fresh
        # dicts and the shallow scalar arriving later would overwrite the
        # whole subtree.)
        for k in sorted([x for x in flat.keys() if isinstance(x, str)], key=lambda s: s.count(".")):
            v = flat[k]
            if "." in k and not k.startswith("_"):
                top, rest = k.split(".", 1)
                # If `nested[top]` was previously set to a scalar (from an
                # even-shallower key like just "leg" = "feet"), promote it.
                if top in nested and not isinstance(nested[top], dict):
                    _prev = nested[top]
                    nested[top] = {"value": _prev}
                cur = nested.setdefault(top, {})
                parts = rest.split(".")
                for p in parts[:-1]:
                    if p in cur and not isinstance(cur[p], dict):
                        _prev = cur[p]
                        cur[p] = {"value": _prev}
                    cur = cur.setdefault(p, {})
                # Final leaf: if a dict is already there (deeper keys arrived
                # earlier despite sort, or a prior iteration created one),
                # don't overwrite it -- store under `.value` instead.
                _leaf = parts[-1]
                if _leaf in cur and isinstance(cur[_leaf], dict) and not isinstance(v, dict):
                    cur[_leaf].setdefault("value", v)
                else:
                    cur[_leaf] = v
            else:
                # Plain key (no dot). If nested already has this parent as a
                # dict from a deeper key that came earlier, don't overwrite
                # the dict -- fold the plain value into it as `.value`.
                if k in nested and isinstance(nested[k], dict) and not isinstance(v, dict):
                    nested[k].setdefault("value", v)
                else:
                    plain[k] = v
        # nested objects win where a flat parent also exists
        for top, obj in nested.items():
            if isinstance(plain.get(top), dict):
                plain[top].update(obj)
            else:
                plain[top] = obj
        return plain
    if any(isinstance(k, str) and "." in k for k in pa.keys()):
        pa = _renest(pa)

    # fill any attribute the row didn't set from remembered per-type defaults
    _defs = _load_attr_defaults().get(pt, {})
    if isinstance(_defs, dict) and _defs:
        pa = {**{k: v for k, v in _defs.items() if not _is_blank(v)}, **pa}

    has = lambda f: f in props

    def put(f, value):
        if value:
            A[f] = value

    # --- title / bullets / description / keywords (localised text) -----------
    if has("item_name") and g("Title"):
        put("item_name", _shape_simple(props["item_name"], g("Title"), mid))

    if has("bullet_point"):
        bl = []
        for i in range(1, 6):
            b = g(f"Bullet {i}")
            if b:
                bl += _shape_simple(props["bullet_point"], b, mid)
        if bl:
            A["bullet_point"] = bl

    if has("product_description") and g("Description (HTML)"):
        desc = re.sub(r"<[^>]+>", " ", g("Description (HTML)"))
        desc = re.sub(r"\s+", " ", desc).strip()
        put("product_description", _shape_simple(props["product_description"], desc, mid))

    if has("generic_keyword") and g("Search Terms / KW"):
        put("generic_keyword", _shape_simple(props["generic_keyword"], g("Search Terms / KW"), mid))

    # --- brand / condition ----------------------------------------------------
    # Brand = the ACCOUNT'S OWN TRADEMARK, which is the authority. Trust the Brand
    # column only when it is one of THIS account's registered brands (supports
    # multi-brand accounts); otherwise the column holds a stale/leaked value -> use
    # the account's primary trademark. NEVER the global config["brand_name"] -- that
    # leak is exactly how one account's brand ended up on another's listings.
    brand = g("Brand").strip()
    _acct_brands = [str(x).strip() for x in (config.get("_account_brands") or []) if str(x).strip()]
    if _acct_brands:
        if brand not in _acct_brands:
            brand = _acct_brands[0]
    elif config.get("_account_brand") is not None:
        brand = ""                       # account resolved but NO trademark set -> blank
    else:
        brand = brand or config.get("brand_name", "")   # legacy / no-account fallback
    if has("brand") and brand:
        put("brand", _shape_simple(props["brand"], brand, mid))
    if has("condition_type"):
        put("condition_type", _shape_simple(props["condition_type"], "new_new", mid))

    # --- manufacturer / model / part (required by many types) -----------------
    # Prefer a real value scraped from the competitor; otherwise the generated model
    # number from the "Model Number" column.
    #
    # It used to fall back to the SKU. Our SKU is price_handlingdays_competitorASIN
    # (e.g. "12.74_2Days_B00IE769RO"), so listings went live on Amazon with that string
    # published as their Model Number and Part Number -- exposing our pricing, handling
    # time and the competitor's ASIN on the public product page. Never publish the SKU
    # as product data. If we have no real model number, omit the field: it is only
    # written when the schema exposes it, and Amazon reports it as missing if required,
    # which is a fixable error rather than permanently-wrong public data.
    model_default = (g("Model Number")
                     or str(pa.get("model_number") or pa.get("part_number") or "").strip())
    if has("manufacturer"):
        put("manufacturer", _shape_simple(props["manufacturer"], pa.get("manufacturer") or brand, mid))
    if has("model_number") and (pa.get("model_number") or model_default):
        put("model_number", _shape_simple(props["model_number"], pa.get("model_number") or model_default, mid))
    if has("part_number") and (pa.get("part_number") or model_default):
        put("part_number", _shape_simple(props["part_number"], pa.get("part_number") or model_default, mid))

    # --- offer + fulfillment --------------------------------------------------
    price = _clean_price(g("Our Price (GBP)"))   # strip any "GBP"/symbol so float() works (was dropping list_price)
    if not _is_blank(price):
        try:
            A["purchasable_offer"] = _offer(price, mid)
        except Exception:
            pass
    _qty = pa.pop("fulfillment_quantity", None)            # per-listing stock from the dashboard; blank -> config default
    try:
        _qty = int(str(_qty).strip()) if str(_qty).strip() not in ("", "None") else int(config.get("default_quantity", 10))
    except Exception:
        _qty = int(config.get("default_quantity", 10))
    A["fulfillment_availability"] = _fulfillment(_qty, g("Handling Days"))

    # --- product images: main + additional ----------------------------------
    # Amazon must be able to FETCH the image over the public internet. A local
    # path like "/media/<sku>/generated.png" is not fetchable and HARD-FAILS the
    # whole listing ("Unable to Retrieve Media Content ... Invalid URL"). Images
    # are NOT required to create a listing, so only send a value when it's a real
    # http(s) URL; otherwise skip it (the listing still goes through; images can
    # be added later in Seller Central / via a hosted URL).
    def _is_public_url(u):
        u = str(u or "").strip()
        return u.lower().startswith("http://") or u.lower().startswith("https://")

    _main_img = pa.pop("main_product_image_locator", "")
    if _main_img and has("main_product_image_locator") and _is_public_url(_main_img):
        A["main_product_image_locator"] = [{"media_location": str(_main_img).strip(),
                                            "marketplace_id": mid}]
    elif _main_img and not _is_public_url(_main_img):
        console.print(f"  [yellow]Skipping main image (not a public URL): "
                      f"{str(_main_img)[:60]}[/yellow]")
    for _n in range(1, 9):
        _ik = f"other_product_image_locator_{_n}"
        _iv = pa.pop(_ik, "")
        if _iv and has(_ik) and _is_public_url(_iv):
            A[_ik] = [{"media_location": str(_iv).strip(), "marketplace_id": mid}]

    # list_price is structurally required for an offer -- build it whenever we have
    # a price, even if the schema's `properties` omits it (some product types list
    # it only under `required`, which would otherwise drop it -> "list_price
    # required but missing"). Uses the schema when present, else a safe UK default.
    if not _is_blank(price):
        d = _shape_list_price(props.get("list_price", {}), price, mid)
        if d:
            A["list_price"] = d
    if has("website_shipping_weight"):
        d = _shape_weight(props["website_shipping_weight"],
                          pa.get("item_package_weight") or pa.get("item_weight"), mid)
        if d:
            A["website_shipping_weight"] = d

    # Leave merchant_shipping_group blank in config.json to let Amazon apply the
    # account's DEFAULT shipping template. Pass a name only if you specifically
    # want a NON-default template -- a named value can be rejected post-migration
    # even when it exists, so blank -> default is the reliable choice.
    msg = (config.get("merchant_shipping_group", "") or "").strip()
    if msg.lower() == "migrated template":   # auto-migration default Amazon rejects by name -> use account default
        msg = ""
    if has("merchant_shipping_group") and msg:
        put("merchant_shipping_group", _shape_simple(props["merchant_shipping_group"], msg, mid))

    # --- product identifier: real barcode, else claim GTIN exemption ----------
    barcode = g("UPC")
    if barcode and has("externally_assigned_product_identifier"):
        typ = "ean" if len(barcode) == 13 else "upc"
        A["externally_assigned_product_identifier"] = [
            {"value": barcode, "type": typ, "marketplace_id": mid}]
    elif has("supplier_declared_has_product_identifier_exemption"):
        A["supplier_declared_has_product_identifier_exemption"] = [
            {"value": True, "marketplace_id": mid}]

    # --- dimensions (composite if the type uses it) ---------------------------
    if has("item_dimensions"):
        d = _shape_dimensions(props["item_dimensions"],
                              pa.get("item_length"), pa.get("item_width"), pa.get("item_height"), mid)
        if d:
            A["item_dimensions"] = d
    if has("item_package_dimensions"):
        d = _shape_dimensions(props["item_package_dimensions"],
                              pa.get("item_package_length"), pa.get("item_package_width"),
                              pa.get("item_package_height"), mid)
        if d:
            A["item_package_dimensions"] = d

    # composite dimension variants some categories require instead of item_dimensions.
    # Schema-driven: any field whose item properties contain axis sub-fields
    # (length/width/height/depth) with their own value+unit is a composite dim
    # attribute and gets routed through shape_by_schema. Previously this loop
    # hardcoded ("item_depth_width_height", "item_length_width_height") -- but
    # Amazon uses more variants than that (item_length_width for flat/flexible
    # products like expandable hoses is a notable example). Missing variants
    # were silently dropped by _shape_simple's fallback return-[] at line 4343,
    # causing "X Unit is required but missing" errors even after Applied values
    # were correctly saved. Now every axis-shaped field is handled by name.
    _axis_names = ("length", "width", "height", "depth")
    def _is_composite_dim(fname):
        if not isinstance(props.get(fname), dict):
            return False
        _fip = _item_props(props[fname])
        # must have at least ONE axis key, and no top-level value/unit (those
        # single-axis attributes are handled by _shape_simple already).
        _has_axis = any(a in _fip for a in _axis_names)
        _has_flat = ("value" in _fip)
        return _has_axis and not _has_flat
    _axes_src = {"length": pa.get("item_length"), "width": pa.get("item_width"),
                 "height": pa.get("item_height"),
                 "depth":  pa.get("item_depth") or pa.get("item_length")}
    # Merge in any user-supplied nested values from pa (e.g. _renest folded
    # item_length_width.width.value into pa["item_length_width"]["width"]).
    # These take priority over the generic item_length / item_width columns.
    def _from_user_composite(fname):
        v = pa.get(fname)
        if not isinstance(v, dict):
            return {}
        out = {}
        for axis in _axis_names:
            av = v.get(axis)
            if isinstance(av, dict):
                # {value|decimal_value: "0.0", unit: "centimeters"} -> "0.0 centimeters"
                # Amazon nests some axes as `decimal_value` (leg/cable/HARDWARE_TUBING)
                # and others as `value`. _renest stores whichever the AI applied, so
                # accept both or the applied measurement is silently dropped and the
                # field is shaped from empty generic columns (rejected as invalid).
                num = av.get("decimal_value", av.get("value"))
                unit = av.get("unit")
                if num not in (None, "") and unit:
                    out[axis] = f"{num} {unit}"
                elif num not in (None, ""):
                    out[axis] = str(num)
            elif isinstance(av, list) and av and isinstance(av[0], dict):
                # already-shaped list: preserve as-is by picking the first entry
                num = av[0].get("decimal_value", av[0].get("value"))
                unit = av[0].get("unit")
                if num not in (None, "") and unit:
                    out[axis] = f"{num} {unit}"
        return out
    for _fname in list(props.keys()):
        if not _is_composite_dim(_fname):
            continue
        if _fname in A:
            continue
        _user_axes = _from_user_composite(_fname)
        _merged_axes = dict(_axes_src)
        _merged_axes.update({k: v for k, v in _user_axes.items() if v})
        # Schema-driven shaping. shape_by_schema reads the LIVE schema at every
        # level -- array-vs-object wrapping, `value` vs `decimal_value` leaf, and
        # enum units are all READ, never assumed. This is what leg/cable need:
        # their `length` axis is itself a `type: array` whose item leaf key is
        # `decimal_value` (one level deeper than _shape_axes looked, so _shape_axes
        # defaulted to a flat `value` object and Amazon rejected it as invalid).
        # We pass ONLY the axes the field's own schema declares AND that carry a
        # value, so absent axes never fold into empty {}/[{}] wrappers.
        _fip = _item_props(props[_fname])
        _raw_axes = {k: v for k, v in _merged_axes.items()
                     if k in _fip and not _is_blank(v)}
        if not _raw_axes:
            continue
        d = shape_by_schema(props[_fname], _raw_axes, mid, _lang_for(mid))
        if d:
            A[_fname] = d

    # --- write the flat product_attributes (pa) into A ----------------------
    # pa holds BOTH the generator's attributes AND any values the user applied
    # via "Suggest missing fields"/the editor (saved to Attributes JSON). Each
    # is written if the schema lists it OR it's a required field (Amazon's
    # ENFORCED schema omits some required fields' defs, so `has(f)` alone would
    # wrongly drop user-applied values like material/color/light_source).
    skip_axes = {"item_length", "item_width", "item_height",
                 "item_package_length", "item_package_width", "item_package_height"}
    # These need a SPECIAL structure (nested composite / integer / strict enum),
    # not a flat value. Let the specialized backfill below shape them so a plain
    # text value applied in the editor (e.g. battery="Lithium Ion") doesn't get
    # written in the wrong shape and then rejected as "missing".
    _special_shape = {"battery", "num_batteries", "light_source", "power_source_type",
                      "has_multiple_battery_powered_components", "supplier_declared_dg_hz_regulation",
                      "special_feature", "warranty_description", "safety_data_sheet_url", "ghs"}
    alias     = {"colour": "color"}
    for k, v in pa.items():
        if k in skip_axes or _is_blank(v):
            continue
        f = alias.get(k, k)
        if f in A or f in _special_shape:
            continue
        # Write the value. The user/generator put it in pa deliberately, so even
        # if Amazon's slim ENFORCED schema doesn't list this field (has(f) False)
        # and it's not in the static required set, we still send it -- dropping a
        # value the user applied is what caused "X required but missing" for
        # special_feature / warranty_description / safety_data_sheet_url.
        _fprop = props.get(f) if isinstance(props.get(f), dict) else {}
        shaped = _shape_simple(_fprop, v, mid) if _fprop else [{"value": str(v), "marketplace_id": mid}]
        if shaped:
            A[f] = shaped

    # --- SPECIAL NESTED FIELDS (always shaped to Amazon's exact structure) ----
    # FLASHLIGHT (and similar electronics) need these in a specific nested shape.
    # Amazon reports a malformed value as "required but missing", and these are
    # conditionally required for battery products even though they're NOT in the
    # static `required` list -- so shape them whenever the row/applied data
    # references them or it's clearly a battery-powered item. Structures verified
    # against getDefinitions.
    _hay_sf = (g("Item Name") + " " + g("Title") + " " + g("Product Description")).lower()
    # Only treat this as a battery product if the SCHEMA actually requires the
    # battery fields OR the user already supplied them. Keyword-only guessing
    # forced FLASHLIGHT battery structures onto unrelated product types (e.g. a
    # MASSAGER wants different sub-fields) and created errors. Be conservative:
    # follow what THIS product type's schema asks for.
    def _req_or_present(name):
        return name in required or name in pa or isinstance(props.get(name), dict)
    _kw_batt = any(w in _hay_sf for w in ["battery", "rechargeable", "lithium",
                                          "li-ion", "usb", "torch", "flashlight", "led"])
    # battery group only fires when the schema/user signals it, not on keywords alone
    _is_batt = _kw_batt and (_req_or_present("battery") or _req_or_present("num_batteries")
                             or _req_or_present("power_source_type"))
    _lang_sf = "en_US" if mid == US_MARKETPLACE_ID else "en_GB"

    def _has_prop(name):
        return isinstance(props.get(name), dict)

    # GLOBAL SAFE-DEFAULTS: neutralise the hazard/regulatory compliance fields to
    # their 'not applicable / none' option from the live schema, so a non-chemical
    # gadget never errors on a compliance dropdown and never trips a cascade like
    # the dg-regulation "ghs" trap. Records what it set so the dashboard can show
    # the user (choice 2b).
    try:
        _compliance_notes = apply_compliance_safe_defaults(A, props, required, mid, _is_batt)
    except Exception:
        _compliance_notes = []
    if _compliance_notes:
        console.print(f"  [cyan]Compliance auto-set ({len(_compliance_notes)} field(s)) "
                      f"-- shown so you can override:[/cyan]")
        for _cf, _cv, _cr in _compliance_notes:
            console.print(f"    [dim]\u2022 {_cf} = \"{_cv}\"  ({_cr})[/dim]")
        # stash for any downstream reporter (dashboard reads the run log)
        try:
            globals().setdefault("_LAST_COMPLIANCE_NOTES", {})
            _LAST_COMPLIANCE_NOTES[row.get("SKU", "") or row.get("Sku", "")] = _compliance_notes
        except Exception:
            pass

    # UK RESPONSIBLE PERSON: for Amazon.co.uk listings, fill the responsible-person
    # / manufacturer-contact compliance fields from the account's saved RP details.
    # Only for UK/GB runs (mid != US) and only when the schema actually declares
    # the field, so US listings are untouched and we never send a field Amazon
    # doesn't expect.
    if mid != US_MARKETPLACE_ID:
        _rp = (config.get("_uk_responsible_person") or {}) if isinstance(config, dict) else {}
        if isinstance(_rp, dict) and (_rp.get("name") or _rp.get("address")):
            _rp_name = str(_rp.get("name", "")).strip()
            _rp_addr = str(_rp.get("address", "")).strip()
            _rp_email = str(_rp.get("email", "")).strip()
            _rp_phone = str(_rp.get("phone", "")).strip()
            # Amazon UK uses a handful of possible field names across product types.
            # Fill whichever the live schema declares.
            for _rpf in ("manufacturer_contact_information", "responsible_person_address",
                         "eu_responsible_person", "uk_responsible_person"):
                if _rpf in A or not isinstance(props.get(_rpf), dict):
                    continue
                _block = ", ".join([p for p in (_rp_name, _rp_addr, _rp_email, _rp_phone) if p])
                A[_rpf] = [{"value": _block[:500], "marketplace_id": mid}]
                _compliance_notes.append((_rpf, _block[:60] + ("…" if len(_block) > 60 else ""),
                                          "auto: UK Responsible Person from account settings"))

    # light_source -> [{type: [{value, language_tag}]}]
    if ("light_source" in pa or _has_prop("light_source") or "led" in _hay_sf) and "light_source" not in A:
        _ls_val = str(pa.get("light_source", "")).strip() or "LED"
        A["light_source"] = [{"type": [{"value": _ls_val, "language_tag": _lang_sf}],
                              "marketplace_id": mid}]

    # num_batteries -> [{quantity:int, type:enum}]
    if ("num_batteries" in pa or _has_prop("num_batteries") or _is_batt) and "num_batteries" not in A:
        _bt_enum = ["12v", "9v", "a", "aa", "aaa", "aaaa", "c", "d", "nonstandard_battery"]
        try:
            _bt_enum = props["num_batteries"]["items"]["properties"]["type"].get("enum") or _bt_enum
        except Exception:
            pass
        _applied_bt = str(pa.get("num_batteries", "")).strip().lower()
        _bt_val = next((o for o in _bt_enum if o == _applied_bt), "nonstandard_battery")
        try:
            _qty = max(0, int(float(str(pa.get("num_batteries", "1")).strip() or "1")))
        except Exception:
            _qty = 1
        A["num_batteries"] = [{"quantity": _qty, "type": _bt_val, "marketplace_id": mid}]

    # battery -> [{cell_composition:[{value}], average_life:[{value,unit}]}]
    # Amazon's error names "Battery Cell Composition" -> cell_composition is the
    # part it wants. Include both; rechargeable torch -> lithium_ion.
    if ("battery" in pa or _has_prop("battery") or _is_batt) and "battery" not in A:
        # If _renest folded user-supplied sub-field values into pa["battery"] as
        # a nested dict (e.g. {"capacity":{"value":"2000","unit":"milliamp_hour"}}),
        # capture them so we merge OVER our defaults instead of throwing them
        # away. Without this, the hardcoded defaults below always win and the
        # user's Applied values silently vanish.
        _user_bat = pa.get("battery")
        _user_bat_dict = _user_bat if isinstance(_user_bat, dict) else {}
        _life = 6.0
        try:
            import re as _re_sf
            _m = _re_sf.search(r"(\d+(?:\.\d+)?)\s*(?:hours?|hrs?|h)\b", _hay_sf)
            if _m:
                _life = float(_m.group(1))
        except Exception:
            _life = 6.0
        # cell_composition enum (snap if schema provides it); lithium_ion default
        _cc_enum = []
        try:
            _ccp = props["battery"]["items"]["properties"]["cell_composition"]
            _cc_enum = (_ccp.get("items", {}).get("properties", {}).get("value", {}).get("enum")
                        or _ccp.get("items", {}).get("enum") or _ccp.get("enum") or [])
        except Exception:
            _cc_enum = []
        # Prefer a user-supplied cell_composition (from sub-field editor) over
        # a string applied to the parent 'battery' key.
        _user_cc = ""
        if _user_bat_dict.get("cell_composition"):
            _uc = _user_bat_dict["cell_composition"]
            if isinstance(_uc, list) and _uc:
                _user_cc = str(_uc[0].get("value") if isinstance(_uc[0], dict) else _uc[0])
            elif isinstance(_uc, dict):
                _user_cc = str(_uc.get("value", ""))
            else:
                _user_cc = str(_uc)
        _applied_cc = (_user_cc or str(pa.get("battery", "") if not isinstance(pa.get("battery"), dict) else "")).strip().lower().replace(" ", "_").replace("-", "_")
        _rech_sf = any(w in _hay_sf for w in ["rechargeable", "usb", "li-ion", "lithium", "type-c"])
        _cc_val = "lithium_ion" if _rech_sf else "alkaline"
        if _applied_cc in ("lithium_ion", "lithium", "li_ion", "lithium_polymer", "alkaline",
                           "nickel_metal_hydride", "lithium_metal"):
            _cc_val = "lithium_ion" if _applied_cc in ("lithium", "li_ion") else _applied_cc
        if _cc_enum and _cc_val not in _cc_enum:
            _cc_val = "lithium_ion" if "lithium_ion" in _cc_enum else _cc_enum[0]
        # SCHEMA-AWARE: only include the sub-fields THIS product type's battery
        # object actually declares. Different product types want different
        # sub-fields (e.g. MASSAGER requires charge_time; FLASHLIGHT doesn't).
        _bat_subs = {}
        try:
            _bat_subs = props["battery"]["items"]["properties"] or {}
        except Exception:
            _bat_subs = {}

        # Helper: extract a user-supplied value+unit sub-field from _user_bat_dict.
        # Returns (value, unit) or (None, None). Handles the shape _renest
        # produces: {"capacity": {"value": "2000", "unit": "milliamp_hour"}}.
        def _user_vu(subname):
            n = _user_bat_dict.get(subname)
            if isinstance(n, dict):
                return n.get("value"), n.get("unit")
            if isinstance(n, list) and n and isinstance(n[0], dict):
                return n[0].get("value"), n[0].get("unit")
            return None, None

        _bat_obj = {"marketplace_id": mid}
        if not _bat_subs or "cell_composition" in _bat_subs:
            _bat_obj["cell_composition"] = [{"value": _cc_val}]
        if not _bat_subs or "average_life" in _bat_subs:
            _uv, _uu = _user_vu("average_life")
            try:    _v = float(_uv) if _uv not in (None, "") else _life
            except Exception: _v = _life
            _bat_obj["average_life"] = [{"value": _v, "unit": (_uu or "hours")}]
        if not _bat_subs or "weight" in _bat_subs:
            _uv, _uu = _user_vu("weight")
            try:    _v = float(_uv) if _uv not in (None, "") else 50.0
            except Exception: _v = 50.0
            _bat_obj["weight"] = [{"value": _v, "unit": (_uu or "grams")}]
        # product-type-specific sub-fields, added ONLY when the schema declares them
        if "charge_time" in _bat_subs:
            _uv, _uu = _user_vu("charge_time")
            try:    _v = float(_uv) if _uv not in (None, "") else 3.0
            except Exception: _v = 3.0
            _bat_obj["charge_time"] = [{"value": _v, "unit": (_uu or "hours")}]
        if "capacity" in _bat_subs:
            _uv, _uu = _user_vu("capacity")
            try:    _v = float(_uv) if _uv not in (None, "") else 1000.0
            except Exception: _v = 1000.0
            _bat_obj["capacity"] = [{"value": _v, "unit": (_uu or "milliamp_hour")}]
        A["battery"] = [_bat_obj]

    # --- LITHIUM BATTERY GROUP (required once a lithium cell is declared) -----
    # Declaring a lithium-ion battery triggers Amazon's hazmat group. These are
    # best-effort structures for a built-in rechargeable lithium-ion torch.
    # SCHEMA-DRIVEN HAZMAT NET (independent of keyword detection).
    # Some product types (e.g. UNMANNED_AERIAL_VEHICLE / drones) REQUIRE
    # contains_battery_or_cell + number_of_lithium_ion_cells even when the
    # title/description contain none of the battery keywords, so _is_batt stays
    # False and the keyword-gated block below never runs. The dashboard, however,
    # still marks these two fields [CODE-OWNED] "filled on Preview" purely by
    # field NAME -- so the AI is told not to fill them and Preview never fills
    # them either. Result: both sit missing and the auto-fix loop stalls at
    # IDENTICAL forever. Ground truth is the schema, not keywords: if THIS
    # product type declares/requires these fields, fill their safe defaults
    # regardless of _is_batt. Only the two flagged fields -- we do NOT force the
    # full lithium composite block onto a non-keyword product.
    def _schema_wants(_f):
        return (_f in required) or isinstance(props.get(_f), dict)
    def _cbc_value(_prop):
        # contains_battery_or_cell is an ENUM (e.g. "Yes"/"No") for some product
        # types (UNMANNED_AERIAL_VEHICLE) and a BOOLEAN for others. Sending JSON
        # `true` to an enum field fails with "select an approved value from the
        # list". Inspect the schema: if it declares an enum, pick the allowed
        # value meaning "yes"; otherwise fall back to boolean True.
        _enum = []
        if isinstance(_prop, dict):
            _it  = _prop.get("items", {}) if isinstance(_prop.get("items"), dict) else {}
            _itp = _it.get("properties", {}) if isinstance(_it, dict) else {}
            _vpp = _itp.get("value", {}) if isinstance(_itp, dict) else {}
            _enum = [str(x) for x in (_vpp.get("enum") or _itp.get("enum")
                                      or _it.get("enum") or _prop.get("enum") or [])]
        if _enum:
            for _e in _enum:
                if str(_e).strip().lower() in ("yes", "true", "1"):
                    return _e
            return _enum[0]
        return True
    if _schema_wants("contains_battery_or_cell") and "contains_battery_or_cell" not in A:
        A["contains_battery_or_cell"] = [{"value": _cbc_value(props.get("contains_battery_or_cell", {})),
                                          "marketplace_id": mid}]
    if _schema_wants("number_of_lithium_ion_cells") and "number_of_lithium_ion_cells" not in A:
        A["number_of_lithium_ion_cells"] = [{"value": 1, "marketplace_id": mid}]

    _is_lithium = _is_batt and any(w in _hay_sf for w in ["lithium", "li-ion", "li_ion", "rechargeable", "usb"])
    if _is_lithium:
        # contains_battery_or_cell -> boolean (yes, it does)
        if "contains_battery_or_cell" not in A:
            A["contains_battery_or_cell"] = [{"value": True, "marketplace_id": mid}]
        # number_of_lithium_ion_cells -> integer
        if "number_of_lithium_ion_cells" not in A:
            A["number_of_lithium_ion_cells"] = [{"value": 1, "marketplace_id": mid}]
        # number_of_lithium_metal_cells -> 0 (it's ion, not metal)
        if "number_of_lithium_metal_cells" not in A:
            A["number_of_lithium_metal_cells"] = [{"value": 0, "marketplace_id": mid}]
        # lithium_battery -> composite: packaging, energy_content (value+unit),
        # weight. Built-in cell -> "batteries_contained_in_equipment".
        if "lithium_battery" not in A:
            A["lithium_battery"] = [{
                "packaging": [{"value": "batteries_contained_in_equipment"}],
                "energy_content": [{"value": 10.0, "unit": "watt_hours"}],
                "weight": [{"value": 50.0, "unit": "grams"}],
                "marketplace_id": mid,
            }]

    # power_source_type -> simple enum [{value}]
    if ("power_source_type" in pa or _has_prop("power_source_type") or _is_batt) and "power_source_type" not in A:
        _ps_enum = []
        try:
            _psp = props["power_source_type"]
            _ps_enum = (_psp.get("items", {}).get("properties", {}).get("value", {}).get("enum")
                        or _psp.get("items", {}).get("enum") or _psp.get("enum") or [])
        except Exception:
            _ps_enum = []
        _applied_ps = str(pa.get("power_source_type", "")).strip().lower().replace(" ", "_")
        _syn = {"usb": "battery_powered", "usb-c": "battery_powered", "usb_c": "battery_powered",
                "rechargeable": "battery_powered", "battery": "battery_powered",
                "corded": "corded_electric", "mains": "corded_electric"}
        _ps_val = _syn.get(_applied_ps, _applied_ps) or ("battery_powered" if _is_batt else "")
        if _ps_enum and _ps_val not in _ps_enum:
            _ps_val = "battery_powered" if "battery_powered" in _ps_enum else _ps_enum[0]
        if _ps_val:
            A["power_source_type"] = [{"value": _ps_val, "marketplace_id": mid}]

    # has_multiple_battery_powered_components -> boolean
    if ("has_multiple_battery_powered_components" in pa or _has_prop("has_multiple_battery_powered_components")) \
            and "has_multiple_battery_powered_components" not in A:
        A["has_multiple_battery_powered_components"] = [{"value": False, "marketplace_id": mid}]

    # ghs (Globally Harmonized System hazard labelling). Amazon models this as a
    # nested object, NOT a flat value -- a flat string like "not_applicable" is
    # rejected ("GHS Class is required but missing"). For most non-chemical retail
    # items GHS is optional and best OMITTED. BUT some product types (e.g.
    # FLASHLIGHT in some marketplaces) list `ghs` as REQUIRED -- there we must send
    # a real structure built from the schema's own allowed values, or Amazon
    # rejects the listing for the missing required field.
    # GHS (Globally Harmonized System chemical hazard labelling). Amazon models it
    # as a NESTED object and makes it REQUIRED only when
    # supplier_declared_dg_hz_regulation is set to "ghs". There is NO "not
    # applicable" GHS class -- the only values are real chemical hazards
    # (explosive, flammable, corrosive, toxic, ...), so a non-chemical product like
    # a flashlight can never legitimately satisfy a GHS requirement. Strategy:
    #   1. Drop any flat/garbage ghs value the AI may have written.
    #   2. Work out whether GHS is actually being demanded (static required OR
    #      dg_regulation == ghs).
    #   3. If demanded: build a valid structure from the schema. If the schema has
    #      no genuinely-applicable "no real hazard" class, prefer to flip
    #      dg_regulation AWAY from ghs to "not_applicable" so GHS is no longer
    #      required -- correct for a non-chemical item -- rather than mislabel the
    #      product with a real hazard class.
    def _dg_value():
        v = A.get("supplier_declared_dg_hz_regulation")
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return str(v[0].get("value", "")).lower()
        return str(v or "").lower()
    # 1) drop flat/garbage ghs
    if "ghs" in A and not (isinstance(A.get("ghs"), list)
                           and A["ghs"] and isinstance(A["ghs"][0], dict)):
        del A["ghs"]
    # 2) is GHS demanded?
    _ghs_demanded = ("ghs" in required) or (_dg_value() == "ghs")
    if _ghs_demanded and "ghs" not in A:
        _ghs_obj = _build_ghs_from_schema(props.get("ghs", {}), mid)
        # _build_ghs_from_schema only returns a value if the schema offered a
        # "no real hazard"-style option. For FLASHLIGHT it won't (all classes are
        # real hazards) -> _ghs_obj is None -> flip dg_regulation off ghs instead.
        if _ghs_obj is not None:
            A["ghs"] = _ghs_obj
        else:
            # no honest GHS class -> stop declaring GHS as the DG regulation
            _dg_enum = []
            try:
                _dgp = props.get("supplier_declared_dg_hz_regulation", {})
                _dgi = _dgp.get("items", {}) if isinstance(_dgp.get("items"), dict) else {}
                _dgip = _dgi.get("properties", {}) if isinstance(_dgi, dict) else {}
                _dgvp = _dgip.get("value", {}) if isinstance(_dgip, dict) else {}
                _dg_enum = (_dgvp.get("enum") or _dgip.get("enum") or _dgi.get("enum") or [])
            except Exception:
                _dg_enum = []
            _safe = "not_applicable"
            if _dg_enum:
                _low = {str(e).lower(): e for e in _dg_enum}
                _safe = _low.get("not_applicable") or next(
                    (e for e in _dg_enum if str(e).lower() != "ghs"), _dg_enum[0])
            A["supplier_declared_dg_hz_regulation"] = [{"value": _safe, "marketplace_id": mid}]

    # Fields whose items REQUIRE language_tag + value (per schema):
    # special_feature, warranty_description, safety_data_sheet_url. Build them
    # with language_tag so Amazon accepts them (missing language_tag reads as
    # "required but missing").
    def _put_lang(field, value, split=False):
        if split:
            _parts = [s.strip() for s in str(value).replace(";", ",").split(",") if s.strip()]
            if _parts:
                A[field] = [{"value": p, "language_tag": _lang_sf, "marketplace_id": mid}
                            for p in _parts[:5]]
        else:
            if str(value).strip():
                A[field] = [{"value": str(value).strip(), "language_tag": _lang_sf,
                             "marketplace_id": mid}]

    if "special_feature" in pa and str(pa.get("special_feature", "")).strip():
        _put_lang("special_feature", pa["special_feature"], split=True)
    if "warranty_description" in pa and str(pa.get("warranty_description", "")).strip():
        _put_lang("warranty_description", pa["warranty_description"])
    if "safety_data_sheet_url" in pa and str(pa.get("safety_data_sheet_url", "")).strip():
        _put_lang("safety_data_sheet_url", pa["safety_data_sheet_url"])

    # --- REQUIRED-FIELD BACKFILL --------------------------------------------
    # Amazon rejects a listing when a category-required attribute is missing
    # ("X is required but missing"). For any field the schema marks required but
    # we still haven't set, fill a safe, valid value derived from what we know,
    # or snap to the first allowed enum value. This makes VALIDATION_PREVIEW pass
    # on required-but-unmapped fields instead of erroring.
    _row_title = g("Item Name") or g("Title") or ""
    _row_model = g("Model Number")
    _row_brand = g("Brand Name") or g("Brand") or g("Manufacturer") or ""
    _row_country = g("Country of Origin") or g("Country/Region of Origin") or ""
    _backfilled = []
    for _rf in required:
        if _rf in A:
            continue
        _before_keys = set(A.keys())
        # Amazon often lists a field under `required` without including its full
        # property definition in `properties` (ENFORCED mode returns a slim set,
        # or the def lives in a referenced sub-schema). Don't skip those -- we
        # still fill them with a sensible value. `_prop` may be {} in that case.
        _prop = props.get(_rf) if isinstance(props.get(_rf), dict) else {}
        # Pull the allowed enum list (if any) so we can snap to a legal value.
        _items   = _prop.get("items", {}) if isinstance(_prop.get("items"), dict) else {}
        _ip      = _items.get("properties", {}) if isinstance(_items, dict) else {}
        _vp      = _ip.get("value", {}) if isinstance(_ip, dict) else {}
        _enum    = (_vp.get("enum") or _ip.get("enum") or _items.get("enum") or _prop.get("enum") or [])
        # Sensible content-derived defaults for the common required offenders.
        _default = None
        _hay = (_row_title + " " + g("Product Description")).lower()
        _is_rechargeable = any(w in _hay for w in ["rechargeable", "usb", "usb-c", "type-c", "li-ion", "lithium"])
        if _rf in ("model_name", "model"):
            _default = _row_model or _row_title[:60] or _row_brand or "Standard"
        elif _rf == "part_number":
            _default = _row_model or "NA"
        elif _rf in ("manufacturer",):
            _default = _row_brand or "Generic"
        elif _rf == "num_batteries":
            _default = "1"
        elif _rf == "battery_type":
            _default = "battery_type_lithium_ion" if _is_rechargeable else "battery_type_a"
        elif _rf == "power_source_type":
            _default = "battery_powered" if _is_rechargeable else "corded_electric"
        elif _rf == "warranty_description":
            _default = "No warranty"
        elif _rf in ("number_of_items", "unit_count"):
            _default = "1"
        elif _rf == "country_of_origin":
            _default = _row_country or "CN"
        elif _rf in ("included_components",):
            _default = _row_title[:60] or "Main unit"
        elif _rf in ("specific_uses_for_product", "recommended_uses_for_product"):
            _default = "General use"
        elif _rf == "lithium_battery_packaging":
            _default = "batteries_contained_in_equipment" if _is_rechargeable else None
        elif _rf == "material":
            _default = g("Material") or ("Aluminum Alloy" if ("flashlight" in _hay or "torch" in _hay) else "Plastic")
        elif _rf == "color":
            _default = g("Colour") or g("Color") or "Black"
        elif _rf == "item_type_keyword":
            # short keyword describing the item; derive from product type/title
            _default = (g("Product Type") or "").replace("_", " ").lower() or _row_title[:30] or "flashlight"
        elif _rf == "special_feature":
            _default = "Rechargeable" if _is_rechargeable else "Portable"
        elif _rf == "light_source":
            # enum field -> Amazon expects values like 'led'. Prefer the user's
            # applied value (normalised) and snap to the enum; default 'led'.
            _applied = str(pa.get("light_source", "")).strip().lower().replace(" ", "_")
            if not _enum:
                _enum = ["led", "incandescent", "fluorescent", "halogen",
                         "xenon", "neon", "laser", "lcd", "oled", "solar_powered"]
            _default = _applied if (_applied and _applied in _enum) else (
                "led" if (_applied in ("", "led") or "led" in _applied) else
                (_applied if _applied else "led"))
            if _default not in _enum:
                _default = "led" if "led" in _enum else _enum[0]
        elif _rf == "power_source_type":
            # enum field. Prefer applied value; map common synonyms.
            _applied = str(pa.get("power_source_type", "")).strip().lower().replace(" ", "_")
            if not _enum:
                _enum = ["battery_powered", "corded_electric", "ac_dc",
                         "solar_powered", "hand_powered", "usb"]
            # USB-charged rechargeable torch -> battery_powered (most accurate)
            _syn = {"usb": "battery_powered", "usb-c": "battery_powered",
                    "rechargeable": "battery_powered", "battery": "battery_powered",
                    "corded": "corded_electric", "mains": "corded_electric"}
            _cand = _syn.get(_applied, _applied)
            _default = _cand if (_cand and _cand in _enum) else (
                "battery_powered" if _is_rechargeable else "corded_electric")
            if _default not in _enum:
                _default = _enum[0]
        elif _rf == "ghs":
            # GHS hazard classification -> for a non-chemical retail item.
            if not _enum:
                _enum = ["not_applicable"]
            _default = "not_applicable"
        elif _rf in ("safety_data_sheet_url", "msds_url"):
            # not a chemical product -> no SDS; write empty so the field is present
            _default = ""
        elif _rf in ("included_in_warranty",):
            _default = "No warranty"
        elif _rf in ("style", "style_name"):
            _default = _row_title[:40] or "Standard"
        elif _rf in ("wattage",):
            _default = None
        elif _rf in ("is_assembly_required",):
            _default = None

        # --- nested composite fields: shape EXACTLY per the FLASHLIGHT schema ---
        # (verified against getDefinitions: battery>average_life{value,unit};
        #  num_batteries>{quantity:int, type:enum}; light_source>type>{value,language_tag})
        if _rf == "battery" and _rf not in A:
            # battery.average_life -> [{value: <hours>, unit: "hours"}]
            _life = 6.0
            try:
                import re as _re
                m = _re.search(r"(\d+(?:\.\d+)?)\s*(?:hours?|hrs?|h)\b", _hay)
                if m:
                    _life = float(m.group(1))
            except Exception:
                _life = 6.0
            A["battery"] = [{
                "average_life": [{"value": _life, "unit": "hours"}],
                "marketplace_id": mid,
            }]
            continue
        if _rf == "num_batteries" and _rf not in A:
            # num_batteries -> [{quantity: <int>, type: <enum>}]; built-in
            # lithium cell isn't a standard AA/AAA size -> "nonstandard_battery".
            _bt_enum = ["12v", "9v", "a", "aa", "aaa", "aaaa", "c", "d", "nonstandard_battery"]
            try:
                _bt = props["num_batteries"]["items"]["properties"]["type"]
                _bt_enum = _bt.get("enum") or _bt_enum
            except Exception:
                pass
            _applied_bt = str(pa.get("num_batteries", "")).strip().lower()
            _bt_val = "nonstandard_battery"
            for _opt in _bt_enum:
                if _opt == _applied_bt:
                    _bt_val = _opt
                    break
            _qty = 1
            try:
                _qty = int(float(str(pa.get("num_batteries", "1")).strip() or "1"))
            except Exception:
                _qty = 1
            if _qty < 0:
                _qty = 1
            A["num_batteries"] = [{
                "quantity": _qty,
                "type": _bt_val,
                "marketplace_id": mid,
            }]
            continue
        if _rf == "light_source" and _rf not in A:
            # light_source -> [{type: [{value: <str>, language_tag: <locale>}]}]
            _ls_val = str(pa.get("light_source", "")).strip() or "LED"
            _lang = "en_US" if mid == US_MARKETPLACE_ID else "en_GB"
            A["light_source"] = [{
                "type": [{"value": _ls_val, "language_tag": _lang}],
                "marketplace_id": mid,
            }]
            continue
        if _rf == "has_multiple_battery_powered_components" and _rf not in A:
            A[_rf] = [{"value": False, "marketplace_id": mid}]
            continue
        if _rf in ("supplier_declared_dg_hz_regulation",) and _rf not in A:
            # Transport DG regulation. CRITICAL: if this is set to "ghs", Amazon
            # then REQUIRES a GHS hazard class (explosive/flammable/etc.) -- and
            # there is NO "not applicable" GHS class, so a non-chemical product
            # like a flashlight can never satisfy it. So we must NEVER let this
            # field fall back to "ghs". Prefer the lithium-battery value when the
            # schema offers one (a rechargeable torch), else "not_applicable".
            _pref = []
            if _is_rechargeable:
                _pref = ["battery_lithium_ion", "lithium_ion", "battery", "transportation"]
            _pref += ["not_applicable", "not_app", "none"]
            _val = None
            if _enum:
                # pick the first preferred value that actually exists in the enum
                _enum_low = {str(e).lower(): e for e in _enum}
                for _p in _pref:
                    if _p in _enum_low:
                        _val = _enum_low[_p]; break
                if _val is None:
                    # last resort: ANY enum value that is NOT ghs (never trigger GHS)
                    _val = next((e for e in _enum if str(e).lower() != "ghs"), None)
            if _val is None:
                _val = ("battery_lithium_ion" if _is_rechargeable else "not_applicable")
            A[_rf] = [{"value": _val, "marketplace_id": mid}]
            continue
        # Build the field value: numeric vs simple vs enum. Build the structure
        # DIRECTLY (array-of-one + marketplace_id) so it works even when `_prop`
        # is empty (field required but no property def returned by Amazon).
        def _put_simple(val):
            A[_rf] = [{"value": val, "marketplace_id": mid}]

        # CONSERVATIVE GUARD: never inject a guessed text value into a field that
        # has a format/pattern/numeric constraint we can't satisfy. Filling these
        # with the product title creates "does not meet pattern" errors that are
        # worse than leaving the field for the user. Skip them entirely.
        _pattern = ""
        _ptype = ""
        try:
            _vp2 = (_prop.get("items", {}).get("properties", {}).get("value", {})
                    if isinstance(_prop, dict) else {})
            _pattern = _vp2.get("pattern", "") or _prop.get("pattern", "")
            _ptype   = _vp2.get("type", "") or _prop.get("type", "")
        except Exception:
            _pattern, _ptype = "", ""
        # fields that are IDs / numeric / pattern-constrained -> don't guess
        _NO_GUESS = ("browse_node" in _rf or _rf.endswith("_id") or "url" in _rf
                     or _rf in ("recommended_browse_nodes", "external_product_id",
                                "gtin", "ean", "upc", "isbn", "model_number"))
        try:
            if _rf in ("num_batteries", "number_of_items", "unit_count"):
                _put_simple(int(_default) if _default else 1)
            elif _enum:
                # snap to first allowed value (or a content default if it's in the enum)
                _pick = _default if (_default and _default in _enum) else _enum[0]
                _put_simple(str(_pick))
            elif _default is not None and str(_default) != "":
                # respect a numeric type / pattern: only write if the default fits
                if _pattern:
                    import re as _re_pat
                    if _re_pat.match(_pattern.replace("\\A", "^").replace("\\z", "$"), str(_default)):
                        _put_simple(str(_default))
                    # else: skip -- a guessed value won't match the pattern
                elif _ptype in ("integer", "number"):
                    try:
                        _put_simple(float(_default) if "." in str(_default) else int(_default))
                    except Exception:
                        pass  # not numeric -> skip rather than send bad data
                else:
                    _put_simple(str(_default))
            elif _default == "":
                # explicitly-empty default (e.g. SDS url for a non-chemical item)
                _put_simple("")
            elif _NO_GUESS or _pattern or _ptype in ("integer", "number"):
                # required, no safe default, and we MUST NOT guess (ID/numeric/
                # pattern field) -> leave it for the user instead of injecting junk
                pass
            else:
                # CATCH-ALL: plain free-text field, still required, no default ->
                # a neutral value is acceptable here.
                _put_simple(_row_title[:30] or "Standard")
        except Exception:
            # never let backfill crash the build; just skip the field
            pass
        if _rf in A and _rf not in _before_keys:
            _backfilled.append(_rf)

    # FINAL GHS SAFETY NET (runs after dg_regulation is fully resolved above).
    # If, after everything, the DG regulation is "ghs" but we have no valid ghs
    # object, the listing WILL be rejected ("GHS Class is required but missing").
    # A flashlight has no honest GHS hazard class, so flip the regulation to a
    # non-ghs value instead of mislabelling the product.
    def _dg_value_final():
        v = A.get("supplier_declared_dg_hz_regulation")
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return str(v[0].get("value", "")).lower()
        return str(v or "").lower()
    if _dg_value_final() == "ghs":
        _has_valid_ghs = (isinstance(A.get("ghs"), list) and A.get("ghs")
                          and isinstance(A["ghs"][0], dict) and A["ghs"][0].get("classification"))
        if not _has_valid_ghs:
            _ghs_obj2 = _build_ghs_from_schema(props.get("ghs", {}), mid)
            if _ghs_obj2 is not None:
                A["ghs"] = _ghs_obj2
            else:
                _dg_enum2 = []
                try:
                    _dgp = props.get("supplier_declared_dg_hz_regulation", {})
                    _dgi = _dgp.get("items", {}) if isinstance(_dgp.get("items"), dict) else {}
                    _dgip = _dgi.get("properties", {}) if isinstance(_dgi, dict) else {}
                    _dgvp = _dgip.get("value", {}) if isinstance(_dgip, dict) else {}
                    _dg_enum2 = (_dgvp.get("enum") or _dgip.get("enum") or _dgi.get("enum") or [])
                except Exception:
                    _dg_enum2 = []
                _safe2 = "not_applicable"
                if _dg_enum2:
                    _low2 = {str(e).lower(): e for e in _dg_enum2}
                    _safe2 = _low2.get("not_applicable") or next(
                        (e for e in _dg_enum2 if str(e).lower() != "ghs"), _dg_enum2[0])
                A["supplier_declared_dg_hz_regulation"] = [{"value": _safe2, "marketplace_id": mid}]
                A.pop("ghs", None)   # not needed once regulation isn't ghs

    # CONDITIONALLY-REQUIRED SAFETY NET ------------------------------------------
    # Amazon does NOT list these in the static `required` set, so the backfill
    # loop above never fills them -- yet VALIDATION_PREVIEW demands them anyway for
    # battery/electronic items. That is the exact "required but missing" cycle on
    # model_name / special_feature / warranty_description /
    # battery_installation_device_type, plus the hazmat structure error. Fill them
    # here, schema-driven, so the listing validates on the FIRST preview.
    def _enum_of_prop(_p):
        if not isinstance(_p, dict):
            return []
        _it  = _p.get("items", {}) if isinstance(_p.get("items"), dict) else {}
        _itp = _it.get("properties", {}) if isinstance(_it, dict) else {}
        _vpp = _itp.get("value", {}) if isinstance(_itp, dict) else {}
        return [str(x) for x in (_vpp.get("enum") or _itp.get("enum") or _it.get("enum") or _p.get("enum") or [])]

    _cond_title = g("Item Name") or g("Title") or ""
    _cond_hay   = (_cond_title + " " + g("Product Description")).lower()
    _cond_rech  = any(w in _cond_hay for w in ["rechargeable", "usb", "usb-c", "type-c", "li-ion", "lithium"])

    # --- ALWAYS-REBUILD-CLEAN for the conditionally-required fields ----------
    # ROOT CAUSE of the recurring "X is required but missing" while the box looks
    # filled: earlier layers (the _put_lang fills and the required-backfill loop)
    # may put a HALF-BUILT or EMPTY value into A under these keys. The old guards
    # here were `if "field" not in A:` -- so when a broken value already existed,
    # the safety net SKIPPED it ("already present") and the broken value shipped,
    # and Amazon reported it missing/invalid. Fix: don't trust an existing value
    # -- validate it, and rebuild into Amazon's exact array-of-one structure
    # whenever it's absent, empty, or malformed.
    _lang_c = "en_US" if mid == US_MARKETPLACE_ID else "en_GB"

    def _valid_text_attr(_v):
        # Valid = non-empty list whose every entry is a dict with a non-empty
        # `value`. Anything else (missing, "", [], flat string, dict missing
        # value) is treated as broken and rebuilt.
        if not isinstance(_v, list) or not _v:
            return False
        for _e in _v:
            if not isinstance(_e, dict):
                return False
            if not str(_e.get("value", "")).strip():
                return False
        return True

    # model_name (free text): mirror the generated model number, else short title.
    if not _valid_text_attr(A.get("model_name")):
        _mn = g("Model Number") or (_cond_title[:60].strip()) or "Standard"
        A["model_name"] = [{"value": _mn, "marketplace_id": mid}]

    # special_feature (SINGULAR is what Amazon wants; the AI often writes the
    # PLURAL `special_features`). Reconcile: prefer an already-valid singular,
    # else pull from singular/plural in the row JSON, else a sensible default.
    if not _valid_text_attr(A.get("special_feature")):
        _sf_src = ""
        for _k in ("special_feature", "special_features"):
            _v = pa.get(_k)
            if isinstance(_v, str) and _v.strip():
                _sf_src = _v.strip(); break
            if isinstance(_v, list) and _v:
                _sf_src = ", ".join(str(x) for x in _v if str(x).strip()); break
        if not _sf_src:
            _sf_src = "Rechargeable" if _cond_rech else "Portable"
        _sf_vals = [s.strip() for s in _sf_src.replace(";", ",").split(",") if s.strip()][:5] or [_sf_src]
        A["special_feature"] = [{"value": v, "language_tag": _lang_c, "marketplace_id": mid} for v in _sf_vals]
    # never ship the plural variant -- Amazon ignores it and it confuses audits
    A.pop("special_features", None)

    # warranty_description (free text). Use ONE consistent default everywhere
    # (the required-backfill loop used "No warranty"; reconcile to the real one).
    if not _valid_text_attr(A.get("warranty_description")):
        A["warranty_description"] = [{"value": "1 Year Manufacturer Warranty",
                                      "language_tag": _lang_c, "marketplace_id": mid}]

    # battery_installation_device_type: this field is NOT really free-text even
    # when the schema hides its enum -- Amazon validates it SERVER-SIDE against
    # battery.cell_composition. Sending the product name ("Flashlight"/"flashlight")
    # or "Installed in device" is REJECTED ("not a valid value"). The accepted
    # values are device-CATEGORY tokens (underscored). For a consumer torch the
    # correct token is "installed_in_equipment" (verified via getDefinitions:
    # allowed = installed_in_equipment | installed_in_vehicle | installed_in_vessel
    # | not_installed; a built-in battery = installed_in_equipment). Override via config
    # so it can be changed without editing code.
    # battery_installation_device_type: the allowed value depends on the battery
    # CHEMISTRY (Amazon's allOf[98] conditional). VERIFIED via getDefinitions:
    #   - lithium chemistry (lithium_ion/metal/polymer, etc.) -> the THEN branch
    #     allows ONLY: installed_in_vehicle | installed_in_vessel | not_installed
    #     (installed_in_equipment is NOT allowed for lithium!)
    #   - non-lithium -> the ELSE branch also allows installed_in_equipment.
    # A built-in lithium torch that isn't a vehicle/vessel -> "not_installed".
    # Detect the chemistry we actually send in A["battery"].
    _cc_sent = ""
    try:
        _cc_sent = str(A["battery"][0]["cell_composition"][0]["value"]).strip().lower()
    except Exception:
        _cc_sent = ""
    _is_lith = ("lithium" in _cc_sent) or (not _cc_sent and (_cond_rech or "lithium" in _cond_hay))
    _bidt_default = (config.get("battery_installation_device_type_default")
                     or ("not_installed" if _is_lith else "installed_in_equipment"))
    _bidt_enum = _enum_of_prop(props.get("battery_installation_device_type", {}))
    if _bidt_enum:
        # Pick the chemistry-correct default FIRST (for lithium that's
        # not_installed; installed_in_equipment is rejected for lithium). Only
        # fall back to other tokens if the default isn't an allowed option.
        _pick = None
        for _cand in _bidt_enum:
            if str(_cand).strip().lower() == _bidt_default.lower():
                _pick = _cand; break
        if not _pick:
            # ordered preference that is SAFE for lithium first
            _pref = (["not_installed", "installed_in_vehicle", "installed_in_vessel"]
                     if _is_lith else
                     ["installed_in_equipment", "not_installed"])
            for _want in _pref:
                for _cand in _bidt_enum:
                    if str(_cand).strip().lower() == _want:
                        _pick = _cand; break
                if _pick: break
        A["battery_installation_device_type"] = [{"value": _pick or _bidt_enum[0], "marketplace_id": mid}]
    else:
        # no enum exposed -> use the chemistry-correct token (NOT the product name)
        _cur = ""
        if isinstance(A.get("battery_installation_device_type"), list) and A["battery_installation_device_type"]:
            _cur = str(A["battery_installation_device_type"][0].get("value", "")).strip()
        # replace any known-bad value. For lithium, installed_in_equipment is BAD.
        _bad_tokens = ["flashlight", "torch", "installed in device", "installed_in_device",
                       "flash light", "consumer_electronics"]
        if _is_lith:
            _bad_tokens.append("installed_in_equipment")
        _bad = (not _cur) or _cur.lower() in _bad_tokens
        _val = _bidt_default if _bad else _cur
        A["battery_installation_device_type"] = [{"value": _val, "marketplace_id": mid}]

    # --- SCHEMA-INDEPENDENT COMPLIANCE HARDENING -----------------------------
    # The fixes above lean on the live schema (enum snapping). But the schema
    # call can intermittently fail ("Amazon's value lists haven't loaded"), and
    # when props is empty the enum branches do nothing, so raw bad values (e.g.
    # the word "cell", or "Flashlight") sail through to Amazon. These fields have
    # KNOWN-GOOD values for a battery item regardless of schema, so set them
    # deterministically here. This is what makes the listing pass even on a run
    # where the schema didn't load.
    _bs_hay = (_cond_title + " " + g("Product Description") + " "
               + g("Included Components") + " " + g("Bullet Point 1")).lower()
    _has_battery = (
        _truthy(g("Batteries Included") or g("Are Batteries Included") or g("batteries_included"))
        or _cond_rech or "battery" in _bs_hay or "lithium" in _bs_hay
    )

    if _has_battery:
        # contains_battery_or_cell: the CORRECT value depends on the field's schema
        # type, which differs by product type/marketplace:
        #   - if it's an ENUM (dropdown), Amazon wants the allowed STRING (e.g. "Yes")
        #   - if it's a BOOLEAN, Amazon wants JSON true/false
        # The old code always sent boolean True, which fails when the field is an
        # enum ("select an approved value from the list"). Detect and match.
        _cbc_prop = props.get("contains_battery_or_cell", {})
        _cbc_enum = _enum_of_prop(_cbc_prop)
        if _cbc_enum:
            # pick the allowed value meaning "yes"
            _yes = None
            for _e in _cbc_enum:
                if str(_e).strip().lower() in ("yes", "true", "1"):
                    _yes = _e; break
            A["contains_battery_or_cell"] = [{"value": _yes or _cbc_enum[0], "marketplace_id": mid}]
        else:
            # boolean type, or schema not loaded -> JSON boolean true is the
            # documented default shape for this attribute.
            A["contains_battery_or_cell"] = [{"value": True, "marketplace_id": mid}]

        # battery_installation_device_type: the allowed value depends on battery
        # CHEMISTRY. For lithium, installed_in_equipment is REJECTED; valid are
        # not_installed | installed_in_vehicle | installed_in_vessel. Detect what
        # chemistry we actually sent and choose accordingly.
        _cc_sent2 = ""
        try:
            _cc_sent2 = str(A["battery"][0]["cell_composition"][0]["value"]).strip().lower()
        except Exception:
            _cc_sent2 = ""
        _is_lith2 = ("lithium" in _cc_sent2) or (not _cc_sent2 and _has_battery)
        _bidt_default2 = (config.get("battery_installation_device_type_default")
                          or ("not_installed" if _is_lith2 else "installed_in_equipment"))
        _bidt_enum2 = _enum_of_prop(props.get("battery_installation_device_type", {}))
        _cur_bidt = ""
        if isinstance(A.get("battery_installation_device_type"), list) and A["battery_installation_device_type"]:
            _cur_bidt = str(A["battery_installation_device_type"][0].get("value", "")).strip()
        # for lithium, treat installed_in_equipment in the current value as BAD
        _cur_is_bad = _cur_bidt.lower() in (
            "flashlight", "torch", "installed in device", "installed_in_device",
            "flash light", "consumer_electronics") or (_is_lith2 and _cur_bidt.lower() == "installed_in_equipment")
        if _bidt_enum2:
            if _cur_bidt not in _bidt_enum2 or _cur_is_bad:
                _pick2 = None
                for _cand in _bidt_enum2:
                    if str(_cand).strip().lower() == _bidt_default2.lower():
                        _pick2 = _cand; break
                if not _pick2:
                    _pref2 = (["not_installed", "installed_in_vehicle", "installed_in_vessel"]
                              if _is_lith2 else ["installed_in_equipment", "not_installed"])
                    for _want in _pref2:
                        for _cand in _bidt_enum2:
                            if str(_cand).strip().lower() == _want:
                                _pick2 = _cand; break
                        if _pick2: break
                A["battery_installation_device_type"] = [{"value": _pick2 or _bidt_enum2[0], "marketplace_id": mid}]
        else:
            A["battery_installation_device_type"] = [
                {"value": (_bidt_default2 if (not _cur_bidt or _cur_is_bad) else _cur_bidt), "marketplace_id": mid}]

    # wattage: RECURRING PROBLEM. Amazon needs wattage as {value:<number>,
    # unit:<watts>} together. But the value/unit often arrive as separate nested
    # keys, and when the schema fails to load we can't confirm the unit token --
    # so a number ships with no unit and Amazon rejects it ("None ... Wattage").
    # Wattage is OPTIONAL for a flashlight/torch. The safe, permanent fix is:
    # only KEEP wattage if we have BOTH a real number AND can pair a unit with it;
    # otherwise DROP it entirely. A torch listing is valid without wattage.
    _watt_in_a = A.get("wattage")

    def _watt_number(_v):
        """Return the numeric part of a wattage value in any shape, or '' if none."""
        cand = ""
        if isinstance(_v, list) and _v:
            f = _v[0]
            cand = (f.get("value") if isinstance(f, dict) else f)
        elif isinstance(_v, dict):
            cand = _v.get("value")
        else:
            cand = _v
        if cand is None:
            return ""
        m = re.search(r"-?\d+(?:\.\d+)?", str(cand))
        return m.group(0) if m else ""

    # Determine if this product type even declares wattage (when schema loaded).
    _watt_declared = isinstance(props.get("wattage"), dict) and bool(props.get("wattage"))
    _wnum = _watt_number(_watt_in_a) if _watt_in_a is not None else ""

    if _wnum and _watt_declared:
        # we have a number AND the schema is present -> ship value+unit together
        try:
            _wval = float(_wnum) if ("." in _wnum) else int(_wnum)
        except Exception:
            _wval = _wnum
        A["wattage"] = [{"value": _wval, "unit": "watts", "marketplace_id": mid}]
    else:
        # no number, OR schema not loaded (can't confirm the unit) -> drop it.
        # Torches don't require wattage, so this never blocks the listing.
        if "wattage" in A:
            A.pop("wattage", None)
            try:
                console.print("  [dim]wattage dropped (optional for this product; "
                              "avoids the missing-unit rejection)[/dim]")
            except Exception:
                pass

    # FINAL wattage guard (bulletproof): never return an empty/None/partial wattage.
    def _has_real_number(_v):
        cand = None
        if isinstance(_v, list) and _v:
            first = _v[0]
            cand = first.get("value") if isinstance(first, dict) else first
        elif isinstance(_v, dict):
            cand = _v.get("value")
        else:
            cand = _v
        if cand is None:
            return False
        s = str(cand).strip().lower()
        if s in ("", "none", "null"):
            return False
        return bool(re.search(r"-?\d", s))

    if "wattage" in A:
        _wf = A.get("wattage")
        _ok = _has_real_number(_wf) and isinstance(_wf, list) and _wf and isinstance(_wf[0], dict) and str(_wf[0].get("unit", "")).strip()
        if not _ok:
            A.pop("wattage", None)
            try:
                console.print("  [dim]wattage dropped (no real value/unit) -- optional for this product[/dim]")
            except Exception:
                pass

    # never ship the plural special_features (belt-and-braces; also done above)
    A.pop("special_features", None)

    # hazmat: build from the LIVE schema structure, ALWAYS rebuilt clean.
    # Two real Amazon errors this fixes (seen on FLASHLIGHT/US):
    #   1) "Hazmat Aspect does not have the expected value(s)" -> the `aspect`
    #      sub-field must be a value from its enum. For flashlights the schema's
    #      ONLY allowed aspect is 'united_nations_regulatory_id'.
    #   2) "field 'value' ... does not have enough values (min 1)" -> the `value`
    #      sub-field is FREE TEXT (no enum), so the old loop skipped it and left it
    #      empty. We must fill it. For a lithium battery packed inside the device
    #      the correct UN id is UN3481 (lithium-ion batteries contained in
    #      equipment). We only set hazmat when the product actually has a battery.
    #
    # Rebuild policy: do NOT trust a pre-existing hazmat value (it may be a flat
    # {value:..} or half-built). Validate it; if any enum sub-field is wrong or the
    # required free-text `value` is blank, rebuild the whole object.
    _hz_prop = props.get("hazmat", {}) if isinstance(props.get("hazmat"), dict) else {}
    if _hz_prop:
        _hz_items = _hz_prop.get("items", {}) if isinstance(_hz_prop.get("items"), dict) else {}
        _hz_props = _hz_items.get("properties", {}) if isinstance(_hz_items, dict) else {}
        # One-time visibility: print hazmat's real sub-fields + their allowed values.
        try:
            _hz_report = {}
            for _sk, _sv in _hz_props.items():
                _hz_report[_sk] = [str(x) for x in (_sv.get("enum") or [])] if isinstance(_sv, dict) else "free-text"
            console.print(f"  [dim]hazmat schema sub-fields: {_hz_report}[/dim]")
        except Exception:
            pass

        # Does this product carry a (lithium) battery? hazmat is only meaningful then.
        _hz_has_batt = (
            _truthy(g("Batteries Included") or g("Are Batteries Included") or g("batteries_included"))
            or _cond_rech
            or "battery" in _cond_hay or "lithium" in _cond_hay
        )

        # Build the correct object from the live sub-fields.
        _hz_obj = {"marketplace_id": mid}
        for _sk, _sv in _hz_props.items():
            if _sk in ("marketplace_id", "language_tag"):
                continue
            _senum = [str(x) for x in (_sv.get("enum") or [])] if isinstance(_sv, dict) else []
            if _senum:
                # enum sub-field (e.g. `aspect`): prefer a not-applicable style
                # value if the schema offers one; otherwise take the only/first
                # allowed value (for flashlights that's united_nations_regulatory_id).
                _low = {e.lower(): e for e in _senum}
                _val = (_low.get("not_applicable") or _low.get("none")
                        or _low.get("no_warning_applicable")
                        or next((e for e in _senum if "not_applic" in e.lower() or "no_haz" in e.lower()), None)
                        or _senum[0])
                _hz_obj[_sk] = _val
            elif _sk == "value":
                # FREE-TEXT required value. For a battery-in-equipment the correct
                # UN id is UN3481. (Only reached when hazmat is being built, which
                # is itself gated on a battery being present.)
                _hz_obj[_sk] = "UN3481"

        # If aspect resolved to the UN regulatory id but value somehow didn't get
        # set (schema variation), guarantee the UN number is present.
        if str(_hz_obj.get("aspect", "")).lower() == "united_nations_regulatory_id" and not _hz_obj.get("value"):
            _hz_obj["value"] = "UN3481"

        # Decide whether the existing value is already valid (so we don't churn).
        _existing = A.get("hazmat")
        _needs = True
        if isinstance(_existing, list) and _existing and isinstance(_existing[0], dict):
            _needs = False
            _ex0 = _existing[0]
            for _sk, _sv in _hz_props.items():
                _senum = [str(x).lower() for x in (_sv.get("enum") or [])] if isinstance(_sv, dict) else []
                if _senum:
                    if str(_ex0.get(_sk, "")).lower() not in _senum:
                        _needs = True; break
                elif _sk == "value":
                    # required free-text value must be non-empty
                    if not str(_ex0.get(_sk, "")).strip():
                        _needs = True; break

        _has_real_subfield = any(k not in ("marketplace_id", "language_tag") for k in _hz_obj)
        if _needs:
            if _has_real_subfield and _hz_obj.get("value"):
                A["hazmat"] = [_hz_obj]
            else:
                # Couldn't build a valid hazmat object -> drop it rather than ship
                # an invalid shape. hazmat is conditionally-required and the lithium
                # info is also carried by the dangerous-goods regulation field.
                A.pop("hazmat", None)
    else:
        # SCHEMA DIDN'T LOAD for hazmat (props empty / value lists failed). We
        # still must not ship a half-built hazmat that a manual edit may have left
        # in the row (e.g. {aspect: united_nations_regulatory_id} with no value,
        # or a flat scalar "not_applicable"). Repair it deterministically.
        _hz_batt = (
            _truthy(g("Batteries Included") or g("Are Batteries Included") or g("batteries_included"))
            or _cond_rech or "battery" in _cond_hay or "lithium" in _cond_hay
        )
        _ex = A.get("hazmat")
        _ex0 = _ex[0] if (isinstance(_ex, list) and _ex and isinstance(_ex[0], dict)) else {}
        _aspect = str(_ex0.get("aspect", "")).strip()
        _value  = str(_ex0.get("value", "")).strip()
        if _hz_batt:
            # Known-correct hazmat for a lithium battery packed in equipment.
            A["hazmat"] = [{
                "aspect": _aspect or "united_nations_regulatory_id",
                "value":  _value or "UN3481",
                "marketplace_id": mid,
            }]
        else:
            # No battery and no schema to validate against -> a flat scalar or
            # partial object is risky; drop it (dg regulation carries any info).
            if not (_aspect and _value):
                A.pop("hazmat", None)

    # MINIMAL MODE: keep only what Amazon strictly requires + offer essentials,
    # so a listing can be created now and enriched later in Seller Central.
    if MINIMAL_MODE:
        _keep = set(required) | {
            # offer / identity essentials needed for any buyable listing
            "item_name", "brand", "product_description", "bullet_point",
            "list_price", "purchasable_offer", "fulfillment_availability",
            "condition_type", "merchant_suggested_asin", "externally_assigned_product_identifier",
            "supplier_declared_dg_hz_regulation", "country_of_origin",
            "main_product_image_locator",
        }
        # CONDITIONAL requirements: Amazon escalates these to hard errors once it
        # detects certain content (e.g. a lithium battery), even though they are
        # NOT in the static `required` list. Minimal mode must KEEP them, or the
        # submission fails for "required but missing" -- the bug you hit. We keep
        # any safety/battery/compliance field the builder already populated.
        _conditional_keep = {
            "battery", "num_batteries", "number_of_batteries", "power_source_type",
            "lithium_battery", "number_of_lithium_ion_cells", "number_of_lithium_metal_cells",
            "contains_battery_or_cell", "battery_installation_device_type",
            "safety_data_sheet_url", "special_feature", "warranty_description",
            "is_heat_sensitive", "light_source",
            "supplier_declared_material_regulation", "ghs", "hazmat",
            "batteries_required", "batteries_included",
        }
        for _k in list(A.keys()):
            if _k in _conditional_keep:
                _keep.add(_k)
            # keep any field the builder added that's a safety/battery sub-structure
            if _k.startswith("other_product_image_locator_"):
                _keep.add(_k)
        _before = len(A)
        for _k in list(A.keys()):
            if _k not in _keep:
                A.pop(_k, None)
        console.print(f"  [cyan]Minimal mode: kept {len(A)} of {_before} fields "
                      f"(required + offer essentials + safety/battery groups Amazon "
                      f"still enforces)[/cyan]")

    if _backfilled:
        console.print(f"  [dim]auto-filled {len(_backfilled)} required field(s): "
                      f"{', '.join(_backfilled[:12])}[/dim]")

    # FINAL CLEANUP (last word before submit):
    # 1) ghs -> a flashlight has no GHS hazard class; sending any ghs value is
    #    rejected. Drop it no matter what re-added it above.
    # 2) Fields Amazon warns "don't belong to this product type" -> strip them so
    #    the submission is clean (these were leftovers / wrong-name variants).
    _STRIP_ALWAYS = {"ghs", "ghs_classification_class",
                     "special_features",        # plural variant; the valid one is special_feature
                     "item_condition_type"}
    for _bad in _STRIP_ALWAYS:
        A.pop(_bad, None)
    # Also drop any attribute that is NOT part of THIS product type's schema --
    # Amazon warns "attribute does not belong to the product type" and ignores
    # them, but they clutter the response and can confuse validation. Only strip
    # when we actually have the schema props (don't strip blindly if props empty).
    #
    # CRITICAL: preserve any field the user or AI deliberately put in `pa` (the
    # row's Attributes JSON). Amazon's enforced schema drops some conditionally-
    # required fields from the static required-set at fetch time -- but Amazon's
    # live-submit validator STILL requires them (VALIDATION_PREVIEW is looser
    # than live submit and gives a false-clean; auto-fix's Preview accepts but
    # Submit rejects the same payload with "required but missing"). The
    # symmetric write-side comment above at line 4853-4857 explains this same
    # pattern. If pa has the field, the AI or user put it there because Amazon
    # asked for it at some prior Preview -- don't strip on the way out.
    if isinstance(props, dict) and props:
        _keep_unknown = {"merchant_suggested_asin", "external_product_id",
                         "externally_assigned_product_identifier", "list_price",
                         "purchasable_offer", "fulfillment_availability", "condition_type",
                         "main_product_image_locator"}
        _pa_keys = set(pa.keys()) if isinstance(pa, dict) else set()
        for _k in list(A.keys()):
            if _k in _keep_unknown:
                continue
            if _k.startswith("other_product_image_locator_"):
                continue
            if _k in _pa_keys:                # user/AI put this in deliberately
                continue
            if _k not in props and _k not in required:
                A.pop(_k, None)

    # ABSOLUTE-LAST wattage sweep: only keep wattage if it has BOTH a real number
    # AND a unit; otherwise drop (Amazon rejects partial wattage; it's optional).
    if "wattage" in A:
        _wf = A.get("wattage")
        _ok = (_has_real_number(_wf) and isinstance(_wf, list) and _wf
               and isinstance(_wf[0], dict) and str(_wf[0].get("unit", "")).strip())
        if not _ok:
            A.pop("wattage", None)

    # --- PRODUCT IDENTIFIER: single AUTHORITATIVE pass (Rule 1) ----------------
    # The UPC column (the dashboard's "Barcode / GTIN" box) is the ONE source of
    # truth for the product identifier. Recompute it here, at the very end, so
    # nothing folded in earlier -- an AI auto-fix suggestion left in Attributes
    # JSON, or a stale value -- can leave a half-filled or conflicting identifier:
    #   * real barcode present -> send it, and DROP any GTIN-exemption claim.
    #   * no real barcode      -> DROP any (possibly AI-guessed / value-less)
    #                             externally_assigned_product_identifier and claim
    #                             the GTIN exemption instead.
    # This guarantees Amazon never receives an AI-guessed or value-less barcode,
    # and that the owner's real purchased EAN in the UPC box is what gets sent.
    _barcode = g("UPC")
    if _barcode and has("externally_assigned_product_identifier"):
        _typ = "ean" if len(_barcode) == 13 else "upc"
        A["externally_assigned_product_identifier"] = [
            {"value": _barcode, "type": _typ, "marketplace_id": mid}]
        A.pop("supplier_declared_has_product_identifier_exemption", None)
    else:
        A.pop("externally_assigned_product_identifier", None)
        if has("supplier_declared_has_product_identifier_exemption"):
            A["supplier_declared_has_product_identifier_exemption"] = [
                {"value": True, "marketplace_id": mid}]

    return A


# ---- runner ------------------------------------------------------------------

def run_api(config: dict, gc, creds: dict, submit: bool = False,
            marketplace: str = "UK", output_tab: str = None,
            spreadsheet_id: str = None, only_skus: set = None,
            verify: bool = False):
    mkt = str(marketplace or "UK").upper()
    mkt_id = marketplace_id_for(mkt)
    issue_locale = "en_US" if mkt == "US" else "en_GB"
    try:
        mkt_enum = Marketplaces.US if mkt == "US" else Marketplaces.UK
    except Exception:
        mkt_enum = MARKETPLACE
    # creds for the marketplace (US block if publishing to US)
    creds = sp_creds(config, mkt)
    seller_id = seller_id_for(config, mkt)
    if not seller_id:
        seller_id = _try_fetch_seller_id(creds)
    if not seller_id:
        console.print("[red]No seller_id available.[/red] Add  \"seller_id\": \"<Merchant Token>\"  "
                      "to config.json.\n  Find it in Seller Central -> Settings -> Account Info -> "
                      "Merchant Token (looks like A2XXXXXXXXXXXX).")
        return

    sh      = gc.open_by_key(spreadsheet_id or config["google_spreadsheet_id"])
    ws      = sh.worksheet(output_tab or OUTPUT_TAB)
    records = _safe_records(ws)
    headers = ws.row_values(1)
    col     = lambda h: (headers.index(h) + 1) if h in headers else None
    status_col, notes_col = col("Status"), col("Notes")
    payload_col = col("API Payload JSON")   # exact body sent to Amazon (debug view)

    try:
        from sp_api.api import ListingsItemsV20210801
    except ImportError:
        console.print("[red]Your python-amazon-sp-api is too old for the Listings API.[/red] "
                      "Update it:  pip install --upgrade python-amazon-sp-api")
        return
    from gspread.utils import rowcol_to_a1
    # Build the Listings client with a generous timeout. The default in python-amazon-
    # sp-api is short (~15s); the UK/EU endpoint (sellingpartnerapi-eu) round-trip from
    # Pakistan, combined with heavy product types like GARDEN_TOOL_SET, can exceed it
    # and throw "The read operation timed out" -- which is what failed in UK but not US
    # (the US endpoint path was faster). A longer timeout + a retry (below) fixes it.
    try:
        li = ListingsItemsV20210801(credentials=creds, marketplace=mkt_enum, timeout=90)
    except TypeError:
        # older sp_api builds don't accept a timeout kwarg -> fall back gracefully
        li = ListingsItemsV20210801(credentials=creds, marketplace=mkt_enum)
    schema_cache = {}
    # preview validates fresh + previously-errored rows; submit publishes vetted rows
    # PREVIEW (validation only, read-only) should work on ANY row that has a SKU +
    # product type -- including NEEDS_REVIEW / COMPLIANCE_HOLD / IP_HOLD -- so you can
    # check what Amazon needs BEFORE approving. SUBMIT stays strict (APPROVED only).
    if submit:
        eligible = {"APPROVED", "API_READY"}
    else:
        eligible = {"APPROVED", "API_READY", "API_ERROR", "NEEDS_REVIEW",
                    "COMPLIANCE_HOLD", "IP_HOLD", "HOLD", ""}

    # batch sheet writes so we never hit the Google Sheets write-rate quota
    _updates = []
    def queue(r, c, val):
        if c:
            _updates.append({"range": rowcol_to_a1(r, c), "values": [[val]]})
    def flush():
        if _updates:
            ws.batch_update(_updates, value_input_option="RAW")
            _updates.clear()

    # ---- VERIFY mode: re-check already-submitted rows and flip them to LIVE ---------
    # A fresh submit is only 'accepted'; Amazon publishes ASYNCHRONOUSLY (minutes). The
    # submit-time check runs seconds later, so rows usually land on SUBMITTED/pending.
    # This mode re-queries Amazon (no wait) for every not-yet-LIVE row and marks it LIVE
    # once Amazon shows it BUYABLE/DISCOVERABLE -- so the app reflects reality without the
    # user guessing. Read-only on Amazon; only writes the sheet Status/Notes.
    if verify:
        console.print("\n[bold cyan]API mode: RE-VERIFY LIVE STATUS (read-only; updates status only)[/bold cyan]")
        console.print(f"  seller: [bold]{seller_id}[/bold]   marketplace: {mkt_id} ({mkt})\n")
        _checked = _wentlive = 0
        for i, row in enumerate(records, start=2):
            sku = str(row.get("SKU", "")).strip()
            if not sku:
                continue
            if only_skus and sku not in only_skus:
                continue
            _st = str(row.get("Status", "")).strip().upper()
            if _st == "LIVE":
                continue                      # already LIVE -- nothing to do
            # only re-verify rows that were actually submitted or were ready to be
            if _st not in ("SUBMITTED", "API_ERROR", "API_READY", "APPROVED", "PENDING", ""):
                continue
            _rstatus, _rerrs = _verify_live_status(li, seller_id, sku, mkt_id, issue_locale, settle=False)
            _checked += 1
            if _rstatus and any(str(s).upper() in ("BUYABLE", "DISCOVERABLE") for s in _rstatus):
                queue(i, status_col, "LIVE")
                queue(i, notes_col, f"RE-VERIFIED -- LIVE ({', '.join(_rstatus)})")
                _wentlive += 1
                console.print(f"  [green]row {i} {sku}: now LIVE ({', '.join(_rstatus)})[/green]")
            elif _rerrs:
                queue(i, status_col, "API_ERROR")
                queue(i, notes_col, f"RE-VERIFIED -- not live yet: {_issue_str(_rerrs, {})}")
                console.print(f"  [red]row {i} {sku}: not live -- {len(_rerrs)} issue(s)[/red]")
            elif _rstatus is None and _rerrs is None:
                console.print(f"  [dim]row {i} {sku}: could not check (Amazon didn't return the listing)[/dim]")
            else:
                console.print(f"  [yellow]row {i} {sku}: still pending Amazon processing (not yet buyable)[/yellow]")
            time.sleep(0.3)                   # gentle on the listings rate limit
        flush()
        console.print(f"\n[bold]Live re-verify complete[/bold] -- checked: {_checked}   flipped to LIVE: {_wentlive}")
        if only_skus and _checked == 0:
            console.print("  [yellow]None of the requested SKU(s) needed re-verifying[/yellow] "
                          f"(already LIVE, or not in a submitted state). Looked for: {', '.join(sorted(only_skus))}.")
        return

    label = "SUBMIT (creates / replaces live listings)" if submit \
            else "VALIDATION_PREVIEW (validates only -- creates nothing)"
    console.print(f"\n[bold cyan]API mode: {label}[/bold cyan]")
    console.print(f"  seller: [bold]{seller_id}[/bold]   marketplace: {mkt_id} ({mkt})\n")

    ok = err = skip = 0
    _requested_status = {}   # requested SKU -> its Status in the sheet, for an accurate
                             # "not processed" message (distinguish missing vs status-gated)
    for i, row in enumerate(records, start=2):        # row 1 = headers
        _st = str(row.get("Status", "")).strip().upper()
        sku = str(row.get("SKU", "")).strip()
        if only_skus and sku in only_skus:
            _requested_status[sku] = _st
        if _st not in eligible:
            continue
        if only_skus and sku not in only_skus:
            continue
        pt  = str(row.get("Product Type", "")).strip()
        if not sku or not pt:
            _miss = []
            if not sku: _miss.append("SKU")
            if not pt: _miss.append("Product Type")
            console.print(f"  row {i}: missing {' + '.join(_miss)} -- skip "
                          f"(SKU='{sku[:30]}', Product Type='{pt[:30]}')")
            skip += 1; continue

        # BRAND GUARD (submit only): a listing must go out under THIS account's own
        # trademark. Resolve it the same way build_api_attributes does, and BLOCK the
        # submit (with an actionable message) if the account has no trademark set.
        # Preview is left alone so the user can still validate other fields.
        if submit:
            _rb = str(row.get("Brand", "") or "").strip()
            _acct_brands = [str(x).strip() for x in (config.get("_account_brands") or []) if str(x).strip()]
            if _acct_brands:
                if _rb not in _acct_brands:
                    _rb = _acct_brands[0]
            elif config.get("_account_brand") is not None:
                _rb = ""
            if not _rb:
                err += 1
                queue(i, status_col, "API_ERROR")
                queue(i, notes_col,
                      "[E] brand -- no trademark set for this account. Add your brand/trademark "
                      "in account Settings, or in the product card on the Listings page, then submit.")
                console.print(f"  [red]row {i} {sku}: SUBMIT BLOCKED -- no trademark set for this "
                              f"account (add it in Settings or the product card)[/red]")
                continue

        if pt not in schema_cache:
            console.print(f"  fetching schema: [bold]{pt}[/bold]")
            schema_cache[pt] = _raw_schema(pt, creds)
        props, required, _ = schema_cache[pt]
        if not props:
            console.print(f"  row {i} {sku}: no schema for {pt} -- skip"); skip += 1; continue

        attrs = build_api_attributes(row, pt, props, required, config)
        body  = {"productType": pt, "requirements": "LISTING", "attributes": attrs}
        # Save the EXACT payload we are about to send, so the dashboard can show
        # the literal wire data (not just the field view). Pretty-printed for the
        # human reader; this is read-only debug info and does not affect the call.
        if payload_col:
            try:
                queue(i, payload_col, json.dumps(body, ensure_ascii=False, indent=2))
            except Exception:
                pass
        kwargs = {"marketplaceIds": [mkt_id], "issueLocale": issue_locale, "body": body}
        if not submit:
            kwargs["mode"] = "VALIDATION_PREVIEW"

        try:
            # Retry the validation/submit call on a TIMEOUT only (not on real
            # validation errors). UK/EU calls from Pakistan can stall transiently;
            # one retry with a short backoff turns a "read operation timed out"
            # into a successful preview most of the time.
            resp = None
            _last_to = None
            for _try in range(3):
                try:
                    resp = li.put_listings_item(seller_id, sku, **kwargs)
                    break
                except Exception as _te:
                    _msg = str(_te).lower()
                    if ("timed out" in _msg or "timeout" in _msg or "read operation" in _msg) and _try < 2:
                        _last_to = _te
                        console.print(f"  [yellow]row {i} {sku}: validation call slow "
                                      f"(attempt {_try+1}/3) -- retrying...[/yellow]")
                        time.sleep(3)
                        continue
                    raise
            if resp is None and _last_to is not None:
                raise _last_to
            payload = resp.payload if hasattr(resp, "payload") else {}
            payload = payload or {}
            issues  = payload.get("issues", []) or []
            errors  = [x for x in issues if str(x.get("severity", "")).upper() == "ERROR"]
            msgs    = _issue_str(issues, attrs)

            if submit:
                # 'Accepted' != 'published'. Amazon processes asynchronously, so the
                # submit-time issues above are NOT the final outcome (a submit that
                # reported errors can still be live; a clean submit can be rejected
                # downstream, e.g. main-image compliance). Set status from the REAL
                # listing state via getListingsItem.
                _rstatus, _rerrs = _verify_live_status(li, seller_id, sku, mkt_id, issue_locale)
                if _rstatus is None and _rerrs is None:
                    # couldn't verify -> fall back to submit-time verdict, flagged unverified
                    if errors:
                        err += 1
                        queue(i, status_col, "API_ERROR")
                        queue(i, notes_col, f"API SUBMIT - {len(errors)} error(s) [live status unverified]: {msgs}")
                        console.print(f"  [red]row {i} {sku}: {len(errors)} submit error(s) (unverified)[/red]")
                    else:
                        ok += 1
                        queue(i, status_col, "SUBMITTED")
                        queue(i, notes_col, "API SUBMITTED -- accepted; live status could not be verified (re-check shortly).")
                        console.print(f"  [yellow]row {i} {sku}: SUBMITTED (status unverified)[/yellow]")
                elif _rstatus and any(str(s).upper() in ("BUYABLE", "DISCOVERABLE") for s in _rstatus):
                    # Amazon shows it live. If the main image (or other) is flagged,
                    # keep it LIVE but note it -- the user fixes the image later.
                    ok += 1
                    queue(i, status_col, "LIVE")
                    _rmsg = _issue_str(_rerrs, attrs) if _rerrs else ""
                    queue(i, notes_col, f"API SUBMITTED -- LIVE ({', '.join(_rstatus)})"
                                        + (f" | needs attention (fix later): {_rmsg}" if _rmsg else ""))
                    console.print(f"  [green]row {i} {sku}: LIVE ({', '.join(_rstatus)})[/green]"
                                  + ("  [yellow](flagged -- fix later)[/yellow]" if _rerrs else ""))
                elif _rerrs:
                    # Accepted but Amazon rejected it in processing -> NOT live.
                    err += 1
                    queue(i, status_col, "API_ERROR")
                    queue(i, notes_col, f"API SUBMIT accepted but NOT live -- Amazon rejected in processing: {_issue_str(_rerrs, attrs)}")
                    console.print(f"  [red]row {i} {sku}: NOT live -- {len(_rerrs)} issue(s) after processing[/red]")
                    for _em in (_issue_str(_rerrs, attrs).split("; ")):
                        if _em.strip():
                            console.print(f"      [red]- {_em.strip()}[/red]")
                else:
                    # accepted, no errors yet, but not yet BUYABLE -> still processing
                    ok += 1
                    queue(i, status_col, "SUBMITTED")
                    queue(i, notes_col, "API SUBMITTED -- accepted, pending Amazon processing (not yet live). Re-check shortly.")
                    console.print(f"  [yellow]row {i} {sku}: SUBMITTED (pending Amazon processing)[/yellow]")
            else:
                # PREVIEW: the submit-time validation IS the answer.
                if errors:
                    err += 1
                    queue(i, status_col, "API_ERROR")
                    queue(i, notes_col, f"API PREVIEW - {len(errors)} error(s): {msgs}")
                    console.print(f"  [red]row {i} {sku}: {len(errors)} error(s)[/red]")
                    for _em in (msgs.split("; ") if isinstance(msgs, str) else []):
                        if _em.strip():
                            console.print(f"      [red]- {_em.strip()}[/red]")
                else:
                    ok += 1
                    queue(i, status_col, "API_READY")
                    queue(i, notes_col, "API PREVIEW clean" + (f" | warnings: {msgs}" if msgs else ""))
                    tag = " (with warnings)" if msgs else ""
                    console.print(f"  [green]row {i} {sku}: API_READY[/green]{tag}")
        except Exception as e:
            err += 1
            em = str(e)[:300]
            _eml = em.lower()
            if "timed out" in _eml or "timeout" in _eml or "read operation" in _eml:
                # A timeout is NOT a listing problem -- don't make the user hunt for
                # fields to fix. Mark it clearly as a transient connection slowness.
                queue(i, notes_col, f"API call TIMED OUT (connection too slow to Amazon EU) -- not a listing problem; Preview again. ({em})")
                console.print(f"  [yellow]row {i} {sku}: validation call TIMED OUT after retries[/yellow] {em[:140]}")
            else:
                queue(i, notes_col, f"API call failed: {em}")
                console.print(f"  [red]row {i} {sku}: API call failed[/red] {em[:170]}")

        if (ok + err) % 20 == 0:
            flush()                       # periodic save -> stays under the write quota
        time.sleep(0.5)        # gentle on the 5 req/s listings limit

    flush()                               # write any remaining results
    console.print(f"\n[bold]API {'submit' if submit else 'preview'} complete[/bold] -- "
                  f"ok: {ok}   errors: {err}   skipped: {skip}")
    if only_skus and (ok + err) == 0:
        _blocked = {s: st for s, st in _requested_status.items() if st not in eligible}
        _elig = ", ".join(sorted(e for e in eligible if e))
        if _blocked:
            # the SKU(s) ARE in this tab -- they were skipped by the status gate, not missing.
            _lines = ", ".join(f"{s} (status '{st or 'blank'}')" for s, st in sorted(_blocked.items()))
            if submit:
                console.print(f"  [yellow]None of the requested SKU(s) were submitted.[/yellow] "
                              f"They ARE in this tab, but Submit only publishes {_elig} rows. "
                              f"Skipped: {_lines}. Fix any flagged errors, then click Approve, then Submit.")
            else:
                console.print(f"  [yellow]None of the requested SKU(s) were processed.[/yellow] "
                              f"They ARE in this tab but their status isn't eligible ({_elig}). "
                              f"Skipped: {_lines}.")
        else:
            console.print(f"  [yellow]None of the requested SKU(s) were found in this tab.[/yellow] "
                          f"Looked for: {', '.join(sorted(only_skus))}. "
                          "Check the row actually has a 'SKU' and 'Product Type' value in this tab, "
                          "and that the column headers are exactly 'SKU' and 'Product Type'.")
    if not submit:
        console.print("  Review flags in the sheet / dashboard. When happy, publish with:  "
                      "[bold]python amazon_listing_generator.py api submit[/bold]")



def run_miles(config: dict, gc, creds: dict, ws_out=None):
    """MILES mode: read harvested supplier bundles (miles_bundles.json) and turn
    each into an Amazon DRAFT listing via the same compliance/IP/copy engine the
    brand path uses. The SDS text drives hazmat/GHS fields. Files are already in
    Drive; here we only generate the listing rows."""
    import sys as _sys, json as _json
    from pathlib import Path as _P
    base_dir = CONFIG_PATH.parent

    def _argval(flag):
        """Read a --flag value from argv (run_miles can't see main()'s helper)."""
        try:
            i = _sys.argv.index(flag)
            return _sys.argv[i + 1] if i + 1 < len(_sys.argv) else ""
        except ValueError:
            return ""

    # Prefer the PERMANENT store (all items ever harvested, keyed by item
    # number); fall back to the latest-run file for back-compat.
    store_path  = base_dir / "miles_bundles_store.json"
    bundle_path = base_dir / "miles_bundles.json"
    items_path  = base_dir / "miles_items.json"

    # Load the permanent store as a DICT keyed by item number (so we can test
    # membership and merge back-filled entries), not just its values.
    _store = {}
    if store_path.exists():
        try:
            _loaded = _json.load(open(store_path, encoding="utf-8"))
            if isinstance(_loaded, dict):
                _store = _loaded
        except Exception:
            _store = {}

    # The SKU list the user uploaded in Step 1 (persisted by the dashboard when
    # 'Generate drafts' is clicked). When present we generate for THIS list and,
    # crucially, BACK-FILL any listed SKU that was already harvested to Drive but
    # is missing from the local store -- reading its PDFs straight from Drive.
    # Drive is the source of truth; we do NOT re-scrape the Miles website. This is
    # what stops the "148 harvested in Drive but only 14 in the store -> 134
    # silently ignored" bug.
    _items = []
    if items_path.exists():
        try:
            _il = _json.load(open(items_path, encoding="utf-8"))
            if isinstance(_il, list):
                _items = [str(x).strip() for x in _il if str(x).strip()]
        except Exception:
            _items = []

    # If the user didn't upload a list, treat DRIVE as the source of truth: every item
    # folder under the master Drive folder is a harvested product. We take that whole set;
    # generation then SKIPS any SKU already in the output sheet (taken_skus below), so it
    # only fills the gap -- items harvested to Drive that don't have a listing copy yet.
    if not _items:
        try:
            import miles_import as _MI0
            _drv_scan, _derr_scan = _MI0.build_drive_rw(config, base_dir)
            if _drv_scan:
                _items = _MI0.list_all_item_folders(
                    _drv_scan, with_files_only=True,
                    log=lambda m: console.print(m, markup=False))
                console.print(f"[cyan]  No item list uploaded -- scanned Drive: "
                              f"{len(_items)} harvested folder(s) found. Generation will skip "
                              f"any already in the output sheet and build the rest.[/cyan]")
            else:
                console.print(f"[yellow]  Drive scan unavailable ({_derr_scan}); "
                              f"falling back to the local store only.[/yellow]")
        except Exception as _e0:
            console.print(f"[yellow]  Drive scan failed ({type(_e0).__name__}: "
                          f"{str(_e0)[:100]}); falling back to the local store only.[/yellow]")

    if _items:
        _missing = [s for s in _items if s not in _store]
        if _missing:
            console.print(f"[cyan]  {len(_missing)} listed SKU(s) not in the local store -- "
                          f"back-filling from Drive (reusing the harvest PDF logic)...[/cyan]")
            try:
                import miles_import as _MI
                _drv, _derr = _MI.build_drive_rw(config, base_dir)
            except Exception as _de:
                _drv, _derr = None, f"{type(_de).__name__}: {str(_de)[:120]}"
            if not _drv:
                console.print(f"[yellow]  Drive unavailable -- cannot back-fill "
                              f"({_derr}); those SKUs will be skipped.[/yellow]")
            else:
                _added = 0
                for _sku in _missing:
                    try:
                        _b = _MI.bundle_from_drive(_drv, _sku,
                                                   log=lambda m: console.print(m, markup=False))
                    except Exception as _be:
                        console.print(f"[yellow]  {_sku}: Drive back-fill error: "
                                      f"{type(_be).__name__}: {str(_be)[:100]}[/yellow]")
                        _b = None
                    # keep only entries that actually carry document text to ground on
                    if _b and (_b.get("sds_text") or _b.get("spec_text")):
                        _store[_sku] = _b
                        _added += 1
                if _added:
                    try:
                        _json.dump(_store, open(store_path, "w", encoding="utf-8"))
                        console.print(f"[green]  Back-filled {_added} SKU(s) from Drive into "
                                      f"miles_bundles_store.json (merged, existing kept).[/green]")
                    except Exception as _se:
                        console.print(f"[yellow]  Could not save store: {_se}[/yellow]")
        # Generate for the uploaded list, in order, for SKUs that now have data.
        bundles = [_store[s] for s in _items if s in _store]
    else:
        # No uploaded list -> keep the prior behaviour: generate everything in the
        # store, falling back to the latest-run file for back-compat.
        bundles = list(_store.values())
        if not bundles and bundle_path.exists():
            try:
                bundles = _json.load(open(bundle_path, encoding="utf-8"))
            except Exception:
                bundles = []

    if not bundles:
        console.print("[red]No harvested bundles found. Run a harvest first.[/red]")
        return

    # Only generate from bundles that actually have document text (skip the
    # "0 files" harvest failures -- nothing to ground a listing on).
    _with_text = [b for b in bundles if (b.get("sds_text") or b.get("spec_text"))]
    _skipped_empty = len(bundles) - len(_with_text)
    bundles = _with_text

    # --limit N: generate at most N listings this run (for controlled testing).
    try:
        _limit = int(_argval("--limit") or "0")
    except ValueError:
        _limit = 0
    if _limit > 0:
        bundles = bundles[:_limit]

    if not bundles:
        console.print("[yellow]No bundles with document text to generate.[/yellow]")
        return

    console.print(f"\n[bold magenta]{'='*55}[/bold magenta]")
    console.print(f"[bold magenta]  MILES LISTING MODE -- {len(bundles)} product(s)"
                  + (f" (limit {_limit})" if _limit else "")
                  + (f" | {_skipped_empty} skipped (no text)" if _skipped_empty else "")
                  + "[/bold magenta]")
    console.print(f"[bold magenta]{'='*55}[/bold magenta]\n")

    if ws_out is None:
        _gc, _wsin, ws_out = init_sheets(config)

    # A minimal pseudo-profile so process_brand_row treats Miles as a brand.
    profile = {
        "brand_name":      "Miles Lubricants",
        "marketplace":     _argval("--marketplace") or "US",
        "voice_mode":      "regenerate",
        "source_language": "en",
        "tone_language":   "en",
        "country_of_origin": "US",
        "handling_time":   "5",
        "lead_with_brand": True,
        "forbidden_brands": [],
        "price_markup":    config.get("miles_price_markup", 1.0),
        "replace_existing": False,   # skip SKUs already in the output sheet
        # Miles-specific Amazon format limits:
        "title_max_chars":  75,
        "description_spec": "up to 2000 characters including HTML tags",
        "keyword_boxes":    2,
        "miles_sheet_format": True,   # write the Miles column layout, not FIXED_HEADERS
        # Auto main image: when the account has the "image_template" feature on,
        # generate a templated main image per listing from the harvested product
        # photo + the brand's saved image recipe. Passed via --auto-image.
        "auto_image": ("--auto-image" in sys.argv),
        "_config": config,   # so the auto-image step can reach API keys
        "_account_id": (_argval("--account-id") or ""),  # scope generated media
        # Regulatory/technical terms that appear in lubricant SDS copy but are
        # NOT competitor brand names. Added to the IP scanner's safe list so
        # they don't trigger false-positive "suspected brand words" violations.
        "safe_words_extra": [
            # Regulatory frameworks
            "GHS", "HazCom", "OSHA", "NFPA", "HMIS", "SDS", "CFR", "REACH",
            "SARA", "CERCLA", "RCRA", "DOT", "IATA", "IMDG", "TLV", "PEL",
            "UN", "CAS", "WHMIS", "EC", "EU",
            # SDS section terms
            "Safety", "Data", "Sheet", "Section", "Health", "Reactivity",
            "Flammability", "Hazard", "Exposure", "Regulation", "Classification",
            "Fire", "Stability", "Disposal", "Transport", "Information",
            "Physical", "Chemical", "First", "Aid", "Emergency", "Response",
            # Lubricant chemistry
            "Diester", "Diester-Based", "Polyglycol", "Polyol", "Ester",
            "Synthetic", "Hydrocarbon", "Naphthenic", "Paraffinic",
            "Hydrotreated", "Mineral", "Petroleum", "Silicone", "PTFE",
            "Polymer", "Additive", "Inhibitor",
            # Viscosity / standards
            "ISO", "AGMA", "SAE", "ASTM", "API", "ACEA", "NLGI", "DIN",
            "Viscosity", "Kinematic", "Centistoke", "Centipoise", "Grade",
            "VG", "Index", "Pour", "Flash", "Point", "Density",
            # Lubricant product types / applications
            "Compressor", "Hydraulic", "Turbine", "Gearbox", "Slideway",
            "Circulating", "Coolant", "Metalworking", "Cutting", "Way",
            "Rotary", "Reciprocating", "Screw", "Vane", "Piston",
            "Air", "Gas", "Steam", "Industrial", "Commercial", "Mobile",
            "Heavy", "Duty", "Premium", "Professional",
            # Equipment descriptors
            "OEM", "OEMs", "System", "Equipment", "Component", "Service",
            "Interval", "Maintenance", "Operation", "Application",
            "Bearing", "Gear", "Pump", "Motor", "Seal", "Gasket", "Valve",
            # Action words capitalised in bullets/descriptions
            "Supplied", "Formulated", "Engineered", "Designed", "Developed",
            "Manufactured", "Produced", "Blended", "Tested", "Rated",
            "Classified", "Recognised", "Recognized", "Certified", "Approved",
            "Recommended", "Specified", "Matched", "Targeted",
            # Miles brand + products
            "Farmingdale", "Lubricants", "Miles", "Stratus", "Nimbus",
            "Voltage", "SXR", "UNIV", "COMP", "FG",
            # Common sentence-interior caps
            "US", "USA", "NY", "LLC", "Inc", "Corp", "Ltd",
            "Always", "Note", "SKU", "MPN", "Brand", "Country", "Origin",
            "Manufacturer", "Supplier",
        ],
        "allowed_phrases_override": [
            "produced by", "made by", "manufactured by", "supplied by",
            "backed by", "developed by",
        ],
    }

    try:
        import brand_listing
    except Exception as e:
        console.print(f"[red]brand_listing unavailable: {e}[/red]")
        return

    # --- Miles-specific compliance (from the original Miles workflow) ----------
    # Lubricant spec sheets constantly reference OEM brands/specs (John Deere
    # J-20C, Caterpillar TO-4, etc). Listing those on Amazon = IP takedown risk.
    # Load the forbidden brands/specs + safe rephrasings and feed them into the
    # generation context so Claude scrubs them out and uses compliant language.
    def _load_lines(fn):
        p = base_dir / "miles_compliance" / fn
        try:
            return [l.strip() for l in open(p, encoding="utf-8") if l.strip()]
        except Exception:
            return []
    _forbidden_brands = _load_lines("forbidden_brands.txt")
    _forbidden_specs  = _load_lines("forbidden_specs.txt")
    def _load_text(fn):
        p = base_dir / "miles_compliance" / fn
        try:
            return open(p, encoding="utf-8").read().strip()
        except Exception:
            return ""
    _safe_alts = _load_text("safe_alternatives.txt")
    _miles_addl = _load_text("additional_instructions.txt")

    _miles_compliance_block = ""

    # Load the full Amazon policies document for comprehensive compliance
    _amazon_policies = _load_text("amazon_policies.txt")

    _miles_compliance_block += (
        "YOUR ROLE: You are an expert Amazon listing optimization specialist with "
        "deep knowledge of Amazon's A10 search algorithm, conversion-focused "
        "copywriting, and industrial lubricant customer behavior. Your goal is to "
        "create HIGH-CONVERTING, fully compliant Amazon listings — not just spec "
        "sheets. Every word must earn its place by either ranking for a keyword or "
        "moving the buyer toward a purchase decision.\n\n"

        "CRITICAL KEYWORD PRIORITIZATION RULE (mandatory for all content):\n"
        "- TITLE: Highest-value search keywords FIRST. Buyers search 'synthetic "
        "compressor oil iso 46' not 'Miles Lubricants product'. Lead with what "
        "buyers type into Amazon search.\n"
        "- BULLETS: Put high-ranking keywords in the first 3-5 words of each bullet. "
        "Amazon indexes the first ~1000 bytes across all bullets.\n"
        "- DESCRIPTION: Descending order of keyword importance.\n"
        "- BACKEND: Remaining keywords by search volume priority. Fill to 249 bytes.\n\n"

        "LISTING COPY RULES:\n"
        "1. TITLE (max 75 chars incl spaces): Product type + key feature + viscosity "
        "grade + size. Highest-value keywords FIRST. Title case. No brand violations.\n"
        "2. BULLETS (5, max 500 chars each): Each bullet MUST start with a short "
        "BENEFIT DESCRIPTOR (1-3 words, caps) then a dash, then explain the customer "
        "benefit — WHY should the buyer care? Don't just list specs. Convert features "
        "into benefits. Example: 'EXTENDED EQUIPMENT LIFE — The ISO 46 viscosity grade "
        "provides optimal film thickness...'\n"
        "3. DESCRIPTION (max 2000 chars, HTML: <p><ul><li><b><br>): Narrative format "
        "answering: What is it? Who is it for? How does it help them? Key specs? "
        "What applications? Write for the maintenance manager who needs to solve a "
        "problem, not an engineer reading a datasheet.\n"
        "4. BACKEND KEYWORDS (max 249 bytes): Fill completely. Lowercase, single "
        "spaces, NO punctuation. Include: synonyms (oil/fluid/lubricant), viscosity "
        "variants (iso 46/iso vg 46/vg 46), applications, equipment types, standards. "
        "Do NOT repeat title words, brand name, or SKU.\n"
        "5. NO phone numbers, email addresses, URLs, or contact info anywhere.\n"
        "6. NO SKU/MPN inside description or bullet text.\n"
        "7. NO emojis, NO promotional language ('best ever', '#1'), NO unsubstantiated "
        "claims.\n"
        "8. Professional tone — never casual. Focus on performance, reliability, and "
        "quality. Emphasize technical specifications from the SDS/TDS data.\n"
        "9. Be specific about viscosity grades, temperature ranges, NFPA ratings, "
        "and applications when the data supports it.\n\n"

        "CUSTOMER INTENT — WRITE FOR THE BUYER:\n"
        "Your buyer is a maintenance manager, shop foreman, fleet operator, or "
        "equipment technician searching Amazon for a specific lubricant. They care "
        "about: Does it fit my equipment? What viscosity? Will it protect my "
        "investment? Is it safe to handle? Can I trust the brand? Address these "
        "questions directly in the copy.\n\n"
    )

    # Inject the full Amazon policies document (compressed to key rules)
    if _amazon_policies:
        _miles_compliance_block += (
            "AMAZON POLICY COMPLIANCE (key rules from Amazon's official guidelines):\n"
            + _amazon_policies[:2000] + "\n\n"
        )
    if _forbidden_brands:
        _miles_compliance_block += (
            "\n\nCRITICAL IP RULES FOR THIS LUBRICANT (ZERO TOLERANCE -- the spec "
            "sheets WILL mention these; you must NOT put any of them in the listing):\n"
            "FORBIDDEN BRANDS: " + ", ".join(_forbidden_brands[:120]) + "\n")
    if _forbidden_specs:
        _miles_compliance_block += ("FORBIDDEN OEM SPEC CODES: "
                                    + ", ".join(_forbidden_specs[:60]) + "\n")
    if _safe_alts:
        _miles_compliance_block += "\nUSE THESE SAFE ALTERNATIVES INSTEAD:\n" + _safe_alts[:1200] + "\n"
    if _miles_addl:
        _miles_compliance_block += "\n" + _miles_addl[:600]

    # -------------------------------------------------------------------------
    # PHASE 1: Brand Analytics category search-terms seeding (pre-launch).
    # Pull the marketplace top search terms (category-wide, not ASIN-specific),
    # filter to lubricant terms, and inject the REAL high-volume keywords so the
    # AI prioritises them. Works with Shee'lady's Brand Analytics access since
    # the compressor-oil search terms are the same regardless of which brand
    # pulls them. Gated by --use-brand-analytics; non-fatal if it fails.
    _ba_block = ""
    if "--use-brand-analytics" in sys.argv:
        try:
            import brand_analytics as _BA
            _ba_mkt = (_argval("--marketplace") or "US").upper()
            _ba_creds = creds
            console.print("  [cyan][BA] pulling category search terms (Phase 1)...[/cyan]")
            _terms = _BA.fetch_search_terms(_ba_creds, marketplace=_ba_mkt,
                                            log=lambda m: console.print(m))
            # Compliance + relevance gate: every surviving term must pass the SAME
            # forbidden-brand/spec check the listing copy passes, AND be product
            # relevant (contain a lubricant core word, no junk/accessory words).
            _filt = _BA.filter_terms(
                _terms, _BA.LUBRICANT_INCLUDE, top_n=40,
                forbidden_brands=_forbidden_brands,
                forbidden_specs=_forbidden_specs)
            if _filt:
                _ba_block = _BA.build_keyword_context(search_terms=_filt)
                console.print(f"  [green][BA] {len(_filt)} compliant, relevant "
                              f"search terms seeded (forbidden brands/specs "
                              f"filtered out)[/green]")
            else:
                console.print("  [yellow][BA] no compliant lubricant terms "
                              "after filtering[/yellow]")
        except Exception as _bae:
            console.print(f"  [yellow][BA] search-terms pull skipped: "
                          f"{type(_bae).__name__}: {str(_bae)[:160]}[/yellow]")
    if _ba_block:
        _miles_compliance_block = _ba_block + "\n\n" + _miles_compliance_block

    client = anthropic.Anthropic(api_key=config["anthropic_api_key"])
    taken_skus, _ = load_existing_skus_and_asins(ws_out)
    # Miles copies are spread across MANY tabs in the same spreadsheet -- skip an
    # item if its SKU exists on ANY tab, not just the one we write to, so a re-run
    # (or the one-click Harvest->Generate flow) never duplicates a listing that
    # was generated onto a different tab.
    try:
        _cross = _skus_across_all_tabs(ws_out)
        _extra = len(_cross - taken_skus)
        taken_skus |= _cross
        console.print(f"  [cyan]Cross-tab dedup: {len(_cross)} SKU(s) already exist across all tabs "
                      f"(+{_extra} beyond the target tab) -- these will be skipped.[/cyan]")
    except Exception as _e:
        console.print(f"  [yellow]Cross-tab SKU scan failed: {str(_e)[:80]}[/yellow]")
    compliance_rules = load_compliance_rules()
    ip_rules = load_ip_rules()
    static_vv = load_static_valid_values()

    profile["forbidden_brands"] = _forbidden_brands + _forbidden_specs

    ok = 0
    total = len(bundles)
    for idx, b in enumerate(bundles, 1):
        # Map the Miles bundle into the product shape process_brand_row expects.
        # The item number is the model/SKU. SDS + spec text become the
        # "competitor_specs" extra context that grounds the copy + compliance.
        product = {
            "title":        b.get("title", "") or b.get("item_number", ""),
            "description":  b.get("description", ""),
            "vendor":       "Miles Lubricants",
            "sku":          b.get("item_number", ""),
            "model_number": b.get("item_number", ""),
            "barcode":      "",
            # Industrial lubricants -> Amazon product type LUBRICANT (generic
            # PRODUCT isn't a real Amazon type and fails schema lookup). Use the
            # harvested type if it looks specific, else default to LUBRICANT.
            "product_type": (b.get("product_type") or "").strip().upper() or "LUBRICANT",
            "attributes":   b.get("attributes", {}),
            "images":       b.get("images", []),
            # size/pack label harvested from the PDP (e.g. "5 Gal. / PAIL").
            # Used by the title rule to append a normalised size suffix.
            "volume":       b.get("volume", "") or b.get("pack", ""),
        }
        # Combine the PDF-derived text as grounding context (SDS first -- it drives
        # the hazmat/compliance fields; then the technical spec text). The Miles
        # IP-compliance block is prepended so the forbidden OEM brands/specs are
        # scrubbed from the copy.
        specs_ctx = _miles_compliance_block + "\n\n" if _miles_compliance_block else ""
        if b.get("sds_text"):
            specs_ctx += "SAFETY DATA SHEET (SDS):\n" + b["sds_text"][:4000] + "\n\n"
        if b.get("spec_text"):
            specs_ctx += "TECHNICAL DATA SHEET (TDS):\n" + b["spec_text"][:3000] + "\n\n"
        if b.get("other_pdf_text"):
            specs_ctx += "ADDITIONAL:\n" + b["other_pdf_text"][:1500]

        try:
            done = brand_listing.process_brand_row(
                product, profile, host=_sys.modules[__name__], client=client,
                ws_out=ws_out, creds=creds, config=config, idx=idx, total=total,
                taken_skus=taken_skus, compliance_rules=compliance_rules,
                ip_rules=ip_rules, static_vv=static_vv, claim_docs=[],
                competitor_specs=specs_ctx)
            if done:
                ok += 1
                try:
                    console.print(f"  [green]WROTE draft for {product['sku']} -> "
                                  f"sheet '{ws_out.spreadsheet.title}' / tab '{ws_out.title}'[/green]")
                except Exception:
                    console.print(f"  [green]WROTE draft for {product['sku']}[/green]")
        except Exception as e:
            console.print(f"  [red]{b.get('item_number','?')}: {type(e).__name__}: {str(e)[:120]}[/red]")

    console.print(f"\n[green]Miles generation complete: {ok}/{total} draft(s) written.[/green]")
    if ok:
        try:
            console.print(f"[cyan]Check tab '{ws_out.title}' in sheet '{ws_out.spreadsheet.title}'. "
                          f"New rows are appended at the BOTTOM.[/cyan]")
        except Exception:
            pass


def run_miles_optimize(config: dict, gc, creds: dict, ws_out=None):
    """PHASE 2: Optimise EXISTING live Miles listings using real SQP data.

    For each row in the Miles sheet that has an ASIN, pull that ASIN's Search
    Query Performance (the actual search queries that drove impressions/clicks/
    purchases), then rewrite the title + bullets + backend keywords to front-load
    the converting queries. Writes the optimised copy back, marking the row.

    Requires: the ASIN is LIVE with search history (~1-4 weeks), and the calling
    account has the Brand Analytics role + Brand Registry.

    CLI:  miles-optimize --account-id <acct> --marketplace US
                         [--sheet <id>] [--tab <name>] [--asin <ASIN>]
    """
    def _argval(flag):
        if flag in sys.argv:
            i = sys.argv.index(flag)
            if i + 1 < len(sys.argv):
                return sys.argv[i + 1]
        return None

    try:
        import brand_analytics as _BA
    except Exception as e:
        console.print(f"[red]brand_analytics module unavailable: {e}[/red]")
        return

    # Load the same forbidden brand/spec lists used at generation time so the
    # SQP queries are gated through the IP layer before entering the prompt.
    def _load_lines(fname):
        p = Path(__file__).parent / "miles_compliance" / fname
        if p.exists():
            return [ln.strip() for ln in p.read_text(encoding="utf-8").splitlines()
                    if ln.strip() and not ln.strip().startswith("#")]
        return []
    _fb = _load_lines("forbidden_brands.txt")
    _fs = _load_lines("forbidden_specs.txt")

    mkt = (_argval("--marketplace") or "US").upper()
    only_asin = (_argval("--asin") or "").strip().upper()

    if ws_out is None:
        console.print("[red]No output sheet bound; cannot optimise.[/red]")
        return

    rows = ws_out.get_all_values()
    if not rows or len(rows) < 2:
        console.print("[yellow]Sheet has no listing rows to optimise.[/yellow]")
        return
    header = rows[0]

    def _col(name):
        for i, h in enumerate(header):
            if (h or "").strip().lower() == name.lower():
                return i
        return -1

    c_sku   = _col("SKU")
    c_title = _col("Title")
    c_asin  = _col("ASIN")
    c_b1    = _col("Bullet Point 1")
    c_kw    = _col("Backend Keywords")
    c_rep   = _col("Compliance Report")

    # ASIN may live in a column named ASIN, or be supplied via --asin for a
    # single-row optimise. If there's no ASIN column we can only do --asin mode.
    client = anthropic.Anthropic(api_key=config["anthropic_api_key"])
    optimised = 0

    for ridx in range(1, len(rows)):
        row = rows[ridx]
        sku = row[c_sku] if c_sku >= 0 and c_sku < len(row) else ""
        asin = ""
        if c_asin >= 0 and c_asin < len(row):
            asin = (row[c_asin] or "").strip().upper()
        if only_asin:
            if asin != only_asin:
                continue
        if not asin:
            continue

        console.print(f"\n[cyan]Optimising {sku} (ASIN {asin})...[/cyan]")
        try:
            sqp = _BA.fetch_sqp_for_asin(creds, asin, marketplace=mkt,
                                         log=lambda m: console.print(m))
        except Exception as e:
            console.print(f"  [yellow]SQP pull failed for {asin}: "
                          f"{type(e).__name__}: {str(e)[:160]}[/yellow]")
            continue
        if not sqp:
            console.print(f"  [yellow]No SQP data yet for {asin} "
                          f"(needs more live history).[/yellow]")
            continue

        kw_ctx = _BA.build_keyword_context(sqp=sqp, forbidden_brands=_fb,
                                           forbidden_specs=_fs)
        cur_title = row[c_title] if c_title >= 0 and c_title < len(row) else ""
        cur_bullets = []
        for bi in range(5):
            ci = c_b1 + bi
            cur_bullets.append(row[ci] if 0 <= ci < len(row) else "")

        prompt = (
            "You are optimising an EXISTING live Amazon listing using REAL search "
            "query performance data. Rewrite ONLY to front-load the proven "
            "converting queries while keeping all facts accurate and compliant.\n\n"
            f"{kw_ctx}\n\n"
            f"CURRENT TITLE: {cur_title}\n"
            f"CURRENT BULLETS:\n" + "\n".join(f"- {b}" for b in cur_bullets if b)
            + "\n\nRULES: Title max 75 chars, highest-converting query FIRST. "
            "Keep 5 bullets, each starting with a benefit descriptor. Backend "
            "keywords max 249 bytes, lowercase, no punctuation, no SKU/brand.\n\n"
            "Return JSON ONLY: {\"title\":\"\",\"bullet_1\":\"\",\"bullet_2\":\"\","
            "\"bullet_3\":\"\",\"bullet_4\":\"\",\"bullet_5\":\"\","
            "\"search_terms\":\"\"}")

        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6", max_tokens=2000,
                messages=[{"role": "user", "content": prompt}])
            raw = resp.content[0].text.strip()
            raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.MULTILINE).strip()
            new = json.loads(raw)
        except Exception as e:
            console.print(f"  [red]Optimise generation failed: {e}[/red]")
            continue

        # Write the optimised fields back to the row
        updates = []
        if c_title >= 0 and new.get("title"):
            updates.append((ridx + 1, c_title + 1, new["title"][:75]))
        for bi in range(5):
            key = f"bullet_{bi+1}"
            if new.get(key):
                updates.append((ridx + 1, c_b1 + bi + 1, new[key]))
        if c_kw >= 0 and new.get("search_terms"):
            updates.append((ridx + 1, c_kw + 1, new["search_terms"]))
        if c_rep >= 0:
            updates.append((ridx + 1, c_rep + 1, "Optimised (SQP)"))

        for (r1, c1, val) in updates:
            try:
                ws_out.update_cell(r1, c1, val)
            except Exception as e:
                console.print(f"  [yellow]cell write failed: {e}[/yellow]")
        optimised += 1
        console.print(f"  [green]Optimised {sku} from {len(sqp)} real queries[/green]")

    console.print(f"\n[bold green]miles-optimize complete: {optimised} "
                  f"listing(s) refined from real SQP data.[/bold green]")


def run_brand(config: dict, gc, creds: dict, ws_out=None,
              brand_name: str = "", csv_path: str = "",
              status_filter=("active", ""), test_limit: int = 0):
    """BRAND mode: Shopify export -> brand-voice Amazon listings into the same
    output sheet. ws_out is passed in from main() so we never re-auth Sheets.
    test_limit: if > 0, only generate the first N products (saves credits)."""
    import sys as _sys
    from pathlib import Path as _P
    base_dir = CONFIG_PATH.parent

    if ws_out is None:                      # safety net: open ONCE if not given
        _gc, _wsin, ws_out = init_sheets(config)

    console.print(f"\n[bold magenta]{'='*55}[/bold magenta]")
    console.print(f"[bold magenta]  BRAND LISTING MODE[/bold magenta]")
    console.print(f"[bold magenta]{'='*55}[/bold magenta]\n")

    brand_name = (brand_name or config.get("active_brand")
                  or config.get("brand_name", "")).strip()
    if not brand_name:
        console.print("[red]No brand specified.[/red] Pass a brand name or set "
                      "'active_brand' in config.json.")
        return

    profile = brand_profile.load_profile(config, base_dir, brand_name)
    console.print(f"  Brand profile: [bold]{profile.get('brand_name')}[/bold] "
                  f"| vendor_mode={profile.get('vendor_mode')} "
                  f"| voice={profile.get('voice_mode')} "
                  f"| marketplace={profile.get('marketplace')}")

    # --- per-brand sheet routing (runs AFTER profile is loaded) ----------------
    # If the brand profile names its own output sheet, write there instead of the
    # global one (e.g. a US brand -> its US sheet). Accepts either a bare sheet ID
    # or a full Google Sheets URL (we extract the ID).
    _brand_mkt = profile.get("marketplace", "UK")
    # Use the brand's marketplace credentials (US block for US brands)
    try:
        creds = sp_creds(config, _brand_mkt)
    except Exception:
        pass
    def _extract_sheet_id(s):
        s = (s or "").strip()
        if not s:
            return ""
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", s)
        return m.group(1) if m else s
    prof_sheet = _extract_sheet_id(profile.get("output_spreadsheet_id"))
    prof_tab   = (profile.get("output_tab") or "").strip()
    if prof_sheet:
        try:
            _sh = gc.open_by_key(prof_sheet)
            _tab = prof_tab or OUTPUT_TAB
            try:
                ws_out = _sh.worksheet(_tab)
            except Exception:
                ws_out = _sh.add_worksheet(title=_tab, rows=2000, cols=100)
                ws_out.append_row(FIXED_HEADERS, value_input_option="RAW")
                ws_out.freeze(rows=1)
            console.print(f"  [cyan]Output -> brand sheet {prof_sheet[:12]}... / "
                          f"tab '{_tab}'[/cyan]")
        except Exception as e:
            console.print(f"  [yellow]Could not open brand sheet ({str(e)[:80]}); "
                          f"using default sheet.[/yellow]")

    csv_path = (csv_path or profile.get("shopify_export_path")
                or config.get("shopify_export_path", "")).strip()
    if not csv_path:
        console.print("[red]No Shopify export path.[/red] Set it in the brand "
                      "profile or config['shopify_export_path'].")
        return
    if not _P(csv_path).is_absolute():
        csv_path = str(base_dir / csv_path)
    if not _P(csv_path).exists():
        console.print(f"[red]Shopify export not found:[/red] {csv_path}")
        return

    console.print("[bold]Step 1:[/bold] Parsing Shopify export")
    sf_status = (None if str(profile.get("import_all_statuses")).lower() == "true"
                 else status_filter)
    products = shopify_import.load_shopify_products(csv_path, include_statuses=sf_status)
    catlang  = shopify_import.detect_catalogue_language(products)
    if profile.get("source_language") != catlang["code"]:
        profile["source_language"] = catlang["code"]
        brand_profile.save_profile(config, base_dir, profile)
    console.print(f"  {len(products)} product(s) | source language: "
                  f"{catlang['name']} ({catlang['code']})")
    if not products:
        console.print("[yellow]No products after status filter.[/yellow]")
        return

    only_vendor = (profile.get("only_vendor") or "").strip()
    if only_vendor:
        products = [p for p in products if p.get("vendor", "") == only_vendor]
        console.print(f"  Reseller filter: vendor='{only_vendor}' -> {len(products)} product(s)")

    # --- TEST LIMIT: only generate the first N (protects API credits) ---------
    if not test_limit:
        try:
            test_limit = int(profile.get("test_limit") or config.get("test_limit") or 0)
        except (ValueError, TypeError):
            test_limit = 0
    console.print(f"  [dim]test_limit received: {test_limit}[/dim]")
    if test_limit and test_limit > 0:
        products = products[:test_limit]
        console.print(f"  [bold yellow]TEST MODE: limited to first {len(products)} "
                      f"product(s)[/bold yellow]")

    console.print("[bold]Step 2:[/bold] Loading claim-support documents")
    try:
        docpack = brand_profile.fetch_claim_documents(config, profile, base_dir)
    except Exception as e:
        console.print(f"  [yellow]claim-doc fetch skipped: {str(e)[:120]}[/yellow]")
        docpack = {"docs": [], "office_skipped": [], "warnings": []}
    for w in docpack.get("warnings", []):
        console.print(f"  [yellow]{w}[/yellow]")
    claim_docs = docpack.get("docs", [])
    console.print(f"  {len(claim_docs)} readable claim doc(s) attached")

    competitor_specs = ""
    pool = profile.get("competitor_asins") or []
    if pool:
        console.print(f"[bold]Step 3:[/bold] Competitor enrichment ({len(pool)} ASIN)")
        bits = []
        for asin in pool[:5]:
            try:
                cd = get_competitor_asin_data(asin, creds)
                if cd.get("title"):
                    specs = "; ".join(f"{k}: {v}" for k, v in
                                      list(cd.get("attributes", {}).items())[:20])
                    bits.append(f"[{asin}] {cd['title'][:80]} :: {specs[:400]}")
            except Exception:
                continue
        competitor_specs = "\n".join(bits)

    console.print(f"[bold]Step 4:[/bold] Generating {len(products)} brand listing(s)")
    client           = anthropic.Anthropic(api_key=config["anthropic_api_key"])
    taken_skus, _    = load_existing_skus_and_asins(ws_out)
    compliance_rules = load_compliance_rules()
    ip_rules         = load_ip_rules()
    static_vv        = load_static_valid_values()

    ok_count = 0
    for i, product in enumerate(products, 1):
        # HARD STOP: never exceed the test limit, whatever happened upstream.
        if test_limit and test_limit > 0 and i > test_limit:
            console.print(f"  [bold yellow]Test limit ({test_limit}) reached -- stopping.[/bold yellow]")
            break
        try:
            ok = brand_listing.process_brand_row(
                product, profile,
                host=_sys.modules[__name__], client=client, ws_out=ws_out,
                creds=creds, config=config, idx=i, total=len(products),
                taken_skus=taken_skus, compliance_rules=compliance_rules,
                ip_rules=ip_rules, static_vv=static_vv, claim_docs=claim_docs,
                competitor_specs=competitor_specs)
            if ok:
                ok_count += 1
        except Exception as e:
            console.print(f"  [red]error on '{product.get('title','')[:40]}': {str(e)[:140]}[/red]")
        if i < len(products):
            time.sleep(2)

    console.print(f"\n[bold green]Brand run complete:[/bold green] {ok_count}/{len(products)} written")
    console.print("  Review in the dashboard, set Status=APPROVED, then run export or api.")


async def main():
    # The first positional arg is the MODE word (generate/api/export/...). When
    # the dashboard runs generate it passes only --flags, so argv[1] may be a
    # flag -- in that case the mode is the default "generate".
    mode = "generate"
    if len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        mode = sys.argv[1].lower()

    config = load_config()
    fallback_brand = config.get("brand_name", "Pollinecfecto")
    creds  = sp_creds(config)

    # ACCOUNT/WORKSPACE SCOPING: if the dashboard passed --sheet / --tab, use them
    # for ALL modes (including generate) so listings go to the CORRECT account's
    # sheet instead of the default config spreadsheet. This is what stops e.g.
    # a jack_richard run from writing into the dropshipping sheet.
    global OUTPUT_TAB
    def _early_argval(flag):
        try:
            i = sys.argv.index(flag)
            return sys.argv[i + 1] if i + 1 < len(sys.argv) else None
        except ValueError:
            return None

    # --minimal: create listings with only required + offer-essential fields.
    global MINIMAL_MODE
    if "--minimal" in sys.argv:
        MINIMAL_MODE = True
        console.print("  [cyan]MINIMAL MODE: sending only required + offer-essential "
                      "fields (enrich the rest later in Seller Central).[/cyan]")
    _cli_sheet = _early_argval("--sheet")
    _cli_tab   = _early_argval("--tab")
    _cli_tab_gid = _early_argval("--tab-gid")
    _cli_in_sheet = _early_argval("--input-sheet")
    _cli_in_gid = _early_argval("--input-tab-gid")
    _cli_mkt = (_early_argval("--marketplace") or "").strip().upper()
    # ACCOUNT-SCOPED CREDENTIALS: load the active account's SP-API creds so submit
    # / preview publish to the CORRECT seller account in this workspace.
    _cli_account_id = _early_argval("--account-id")
    _acc_default_mkt = ""   # marketplace the account itself declares (authority)
    _acc_brand = ""         # brand this account sells under (its own, not global)
    if _cli_account_id:
        try:
            import accounts as _acc_mod
            _acc_obj = _acc_mod.get_account(config, _cli_account_id, "config.json")
            if _acc_obj:
                # the account's own brand (first of its 'brands' list) is the
                # authority for this run -- NOT the global config["brand_name"]
                # (which was a single-account leftover). Falls back to auto-pick.
                _abr = _acc_obj.get("brands") or []
                if isinstance(_abr, list) and _abr:
                    _acc_brand = str(_abr[0]).strip()
                elif _acc_obj.get("brand"):
                    _acc_brand = str(_acc_obj.get("brand")).strip()
                # Expose the account's OWN trademark(s) to the API/submit path so
                # build_api_attributes uses THIS account's brand -- never the global
                # default, and never another account's brand.
                config["_account_brand"]  = _acc_brand
                config["_account_brands"] = [str(x).strip() for x in (_abr or []) if str(x).strip()]
                # UK Responsible Person (for Amazon.co.uk compliance fields). Only
                # used when the run targets a UK/GB marketplace; harmless otherwise.
                _rp = _acc_obj.get("uk_responsible_person") or {}
                if isinstance(_rp, dict) and any(_rp.values()):
                    config["_uk_responsible_person"] = _rp
                # Per-account eBay override: if THIS account carries its own eBay
                # Browse-API creds (both App ID + Cert ID), they override the global
                # config values for this run so fetch_ebay_supplement scrapes under
                # the account's own eBay app. A half-filled pair is ignored (would
                # break OAuth) -> global creds stand. Mirrors the dashboard's
                # _ebay_creds() resolver so CLI and UI resolve identically.
                _eb_app  = str(_acc_obj.get("ebay_app_id", "") or "").strip()
                _eb_cert = str(_acc_obj.get("ebay_cert_id", "") or "").strip()
                if _eb_app and _eb_cert:
                    config["ebay_app_id"]  = _eb_app
                    config["ebay_cert_id"] = _eb_cert
                    console.print(f"  [yellow]eBay creds: account override[/yellow] "
                                  f"(app {_eb_app[:14]}…)")
                _ac = _acc_mod.account_creds(_acc_obj)
                if _ac.get("lwa_client_secret") and _ac.get("refresh_token"):
                    config["_account_creds"] = _ac
                    # CRITICAL: `creds` was computed earlier (line ~4627) BEFORE
                    # _account_creds existed, so it held the default (UK) creds.
                    # Recompute it now so the whole run uses THIS account's creds.
                    creds = sp_creds(config)
                    console.print(f"  [yellow]Account-scoped creds:[/yellow] {_acc_obj.get('label','?')} "
                                  f"(seller {_ac.get('seller_id','?')})")
                else:
                    # This account has no Amazon app of its own. It used to fall through
                    # to the global sp_api_* block -- which is jack_uk's credentials -- so
                    # a run (and a SUBMIT) from Miles published into Jack's catalogue.
                    # Instead: borrow a nominated account's app for CATALOGUE lookups only,
                    # and refuse to write anything.
                    _lender_creds, _lender = None, None
                    try:
                        _lender_creds, _lender = _acc_mod.resolve_catalog_creds(
                            config, _acc_obj, str(CONFIG_PATH))
                    except LookupError as _le:
                        console.print(f"  [red]{_le}[/red]")
                        sys.exit(2)
                    # Strip the LENDER's seller_id. Every seller-scoped call keys off it,
                    # so leaving it here would let this run read (and putListingsItem
                    # against) the lender's catalogue. Catalogue calls don't need it.
                    _lender_creds = dict(_lender_creds)
                    _lender_creds["seller_id"] = ""
                    config["_account_creds"] = _lender_creds
                    config["_read_only"] = True
                    creds = sp_creds(config)
                    console.print(
                        f"  [yellow]READ-ONLY workspace:[/yellow] '{_acc_obj.get('label','?')}' has no "
                        f"Amazon app of its own. Borrowing '{(_lender or {}).get('label','?')}'s app for "
                        f"catalogue lookups (product types, item type keyword, valid values, fees).")
                    console.print(
                        "  [yellow]It cannot read or change that account's listings, and cannot "
                        "publish.[/yellow]")
                    # `api` mode covers preview, verify AND submit. Even VALIDATION_PREVIEW
                    # calls putListingsItem against a seller_id -- which here would be the
                    # LENDER's. Refuse the whole mode; generate/miles need no seller.
                    # (NB: `submit` isn't parsed until later in main(), so test argv.)
                    if mode == "api":
                        console.print(f"  [bold red]REFUSED: '{_acc_obj.get('label','?')}' is a read-only "
                                      f"workspace — it has no Seller Central account, so it cannot "
                                      f"preview, verify or publish listings. Connect its own SP-API "
                                      f"credentials to enable this.[/bold red]")
                        sys.exit(3)
                # The account is the source of truth for its own marketplace.
                # default_marketplace first; else the FIRST entry that is US/UK/GB
                # (never blindly marketplaces[0] -- that can be MX/CA/BR and would
                # wrongly fall through to the UK default, denying a US token).
                _acc_default_mkt = (_acc_obj.get("default_marketplace") or "").strip().upper()
                if _acc_default_mkt not in ("US", "UK", "GB"):
                    for _m in (_acc_obj.get("marketplaces") or []):
                        _mu = str(_m).strip().upper()
                        if _mu in ("US", "UK", "GB"):
                            _acc_default_mkt = _mu
                            break
        except Exception as _e:
            console.print(f"  [yellow]Could not load account creds: {_e}[/yellow]")
    # If the dashboard didn't pass --marketplace (or passed a non-UK/US value like
    # MX), fall back to what the account declares. This is the fix for the silent
    # UK-default bug: a US account would otherwise hit UK endpoints with a US token
    # and get 'Access to requested resource is denied' on catalog/pricing/fees.
    if _cli_mkt not in ("US", "UK", "GB") and _acc_default_mkt in ("US", "UK", "GB"):
        _cli_mkt = _acc_default_mkt
        console.print(f"  [yellow]Marketplace not on command -- using account default: "
                      f"{_cli_mkt}[/yellow]")
    # MARKETPLACE OVERRIDE: a US account must price/fee/validate against the US
    # marketplace, not the UK default. Flip the module globals so the whole
    # generate path (catalog, pricing, fees, SP-API) uses the right marketplace.
    global MARKETPLACE, MARKETPLACE_ID
    if _cli_mkt == "US":
        MARKETPLACE = Marketplaces.US
        MARKETPLACE_ID = US_MARKETPLACE_ID
        if not _cli_tab:
            OUTPUT_TAB = "Listings v7.0 US"
    elif _cli_mkt in ("UK", "GB"):
        MARKETPLACE = Marketplaces.UK
        MARKETPLACE_ID = "A1F83G8C2ARO7P"
    if _cli_sheet:
        config["google_spreadsheet_id"] = _cli_sheet
    if _cli_tab:
        OUTPUT_TAB = _cli_tab
    if _cli_in_sheet:
        config["input_spreadsheet_id"] = _cli_in_sheet
    # tab gids resolve a worksheet by its numeric id (from the sheet URL); stash
    # them on config so init_sheets can resolve the exact tab.
    config["_output_tab_gid"] = _cli_tab_gid or ""
    config["_input_tab_gid"] = _cli_in_gid or ""

    console.print(f"\n[bold cyan]{'='*55}[/bold cyan]")
    console.print(f"[bold cyan]  Listing Generator[/bold cyan]")
    console.print(f"[bold cyan]  SP-API + Flat File Export[/bold cyan]")
    console.print(f"[bold cyan]  MODE: {mode.upper()}[/bold cyan]")
    console.print(f"[bold cyan]{'='*55}[/bold cyan]\n")

    console.print("Connecting to Google Sheets...")
    if mode in ("miles", "miles-optimize"):
        config["_miles_mode"] = True   # don't clobber the Miles tab's own headers
    if _cli_sheet or _cli_tab:
        console.print(f"  [yellow]Scoped to account sheet/tab:[/yellow] {_cli_sheet or '(config default)'} / '{OUTPUT_TAB}'")
    gc, ws_in, ws_out = init_sheets(config)
    console.print("[green]OK[/green]")

    if mode == "export":
        run_export_unified(config, gc)
        return

    if mode == "api":
        submit = any(a.lower() == "submit" for a in sys.argv[2:])
        verify = any(a.lower() == "verify" for a in sys.argv[2:])
        if verify:
            submit = False       # verify is read-only -- it never publishes
        def _argval(flag):
            try:
                i = sys.argv.index(flag)
                return sys.argv[i + 1] if i + 1 < len(sys.argv) else None
            except ValueError:
                return None
        _sheet = _argval("--sheet")
        _tab   = _argval("--tab")
        _mkt   = _argval("--marketplace") or "UK"
        _skus_raw = (_argval("--skus") or "").strip()
        _only = set(s.strip() for s in _skus_raw.split(",") if s.strip()) or None
        if _only:
            console.print(f"  [yellow]Scoped to {len(_only)} SKU(s) only:[/yellow] {', '.join(sorted(_only))}")
        run_api(config, gc, creds, submit=submit,
                marketplace=_mkt, output_tab=_tab, spreadsheet_id=_sheet, only_skus=_only,
                verify=verify)
        return

    if mode == "regen":
        # Regenerate listing copy for a specific set of SKUs only, on the active
        # sheet/tab/marketplace. Lower-risk: reuses run_api's scoping plumbing for
        # sheet/tab/marketplace and passes a SKU filter through.
        def _argval(flag):
            try:
                i = sys.argv.index(flag)
                return sys.argv[i + 1] if i + 1 < len(sys.argv) else None
            except ValueError:
                return None
        _skus  = (_argval("--skus") or "").strip()
        _sheet = _argval("--sheet")
        _tab   = _argval("--tab")
        _mkt   = _argval("--marketplace") or "UK"
        sku_list = [s.strip() for s in _skus.split(",") if s.strip()]
        if not sku_list:
            console.print("[regen] no --skus given; nothing to do.")
            return
        console.print(f"[regen] regenerating {len(sku_list)} SKU(s) on "
                      f"{_tab or 'default tab'} ({_mkt}): {', '.join(sku_list)}")
        try:
            run_regen(config, gc, creds, skus=sku_list, marketplace=_mkt,
                      output_tab=_tab, spreadsheet_id=_sheet)
        except NameError:
            console.print("[regen] This generator build does not yet include "
                          "run_regen(). Per-listing regeneration via the dashboard "
                          "editor still works; batch copy-regen needs run_regen wired "
                          "into the generator.")
        return

    if mode == "brand":
        bname = sys.argv[2] if len(sys.argv) > 2 else ""
        csvp  = sys.argv[3] if len(sys.argv) > 3 else ""
        try:
            tlim = int(sys.argv[4]) if len(sys.argv) > 4 else 0
        except (ValueError, IndexError):
            tlim = 0
        run_brand(config, gc, creds, ws_out=ws_out, brand_name=bname,
                  csv_path=csvp, test_limit=tlim)
        return

    if mode == "miles":
        run_miles(config, gc, creds, ws_out=ws_out)
        return

    if mode == "miles-optimize":
        run_miles_optimize(config, gc, creds, ws_out=ws_out)
        return

    # --- Brand: resolve WITHOUT blocking on input(). Priority:
    #   1) --brand CLI arg  2) the active account's own brand  3) auto-pick.
    # The old interactive prompt was a single-account (UK template) leftover and
    # would hang when the dashboard runs this headless as a subprocess.
    user_brand = (_early_argval("--brand") or "").strip()
    if not user_brand:
        user_brand = (_acc_brand or "").strip()
    if not user_brand and sys.stdin and sys.stdin.isatty():
        # only prompt if a human is actually at a terminal
        user_brand = prompt_for_brand()
    if user_brand:
        console.print(f"Brand for this run: [bold]{user_brand}[/bold]")
    else:
        console.print("Brand for this run: [bold]<auto-pick per product>[/bold]")

    console.print("Reading input sheet...", end=" ")
    products = read_input_sheet(ws_in)
    if not products:
        console.print("[red]No products found in input sheet.[/red]")
        return
    console.print(f"[green]OK[/green] {len(products)} product(s)")

    # --- Row selection: generate only the chosen row(s) ----------------------
    # --select carries a row-number list, an ASIN, an eBay item number, or a URL.
    # --select-type tells us how to read a BARE value (row/asin/ebay_item); a URL
    # auto-detects. Empty -> generate all (unchanged).
    _sel_raw  = _early_argval("--select") or ""
    _sel_type = (_early_argval("--select-type") or "auto").strip().lower()
    if _sel_raw.strip():
        _picked, _err = select_rows(products, _sel_raw, _sel_type)
        if _err:
            console.print(f"[red]Selection: {_err}[/red]")
            return
        products = _picked
        console.print(f"  [cyan]Selection -> {len(products)} row(s) chosen for this run[/cyan]")

    if mode == "retry":
        products = get_retry_rows(ws_out, products)
        if not products:
            return

    # Load duplicate-detection state and the per-category model counter.
    taken_skus, seen_asins = load_existing_skus_and_asins(ws_out)
    console.print(f"  Existing rows -- SKUs: {len(taken_skus)} | ASINs: {len(seen_asins)}")
    model_counter    = load_model_counter()
    compliance_rules = load_compliance_rules()
    if compliance_rules:
        cats = [k for k in compliance_rules if k != "general"]
        console.print(f"  Compliance rules loaded: {len(cats)} categories")
    else:
        console.print(f"  [yellow]No compliance rules active[/yellow]")
    ip_rules = load_ip_rules()
    if ip_rules:
        console.print(f"  IP rules loaded: {len(ip_rules.get('forbidden_phrases', []))} phrases + "
                       f"{len(ip_rules.get('safe_capitalised', set()))} allowlisted words "
                       f"(threshold: {ip_rules.get('max_unrecognised', 2)})")
    else:
        console.print(f"  [yellow]No IP rules active[/yellow]")
    static_vv = load_static_valid_values()
    if static_vv:
        total_vals = sum(sum(len(v) for v in attrs.values()) for attrs in static_vv.values())
        console.print(f"  Valid values loaded: {len(static_vv)} product types, "
                       f"{total_vals} enforced strings (from Amazon UK XLSM templates)")
    else:
        console.print(f"  [yellow]valid_values.json not found -- Claude will guess attribute formats[/yellow]")

    total   = len(products)
    client  = anthropic.Anthropic(api_key=config["anthropic_api_key"])
    success = 0
    skipped = 0

    for idx, row in enumerate(products, 1):
        try:
            ok = await process_row(row, client, ws_out, creds, config, idx, total,
                                   user_brand, taken_skus, seen_asins, model_counter,
                                   compliance_rules, ip_rules, static_vv,
                                   skip_existing=(mode != "retry"))
            if ok:
                success += 1
            elif ok is None:
                skipped += 1
        except Exception as e:
            console.print(f"  [red]Unexpected error: {str(e)[:120]}[/red]")
        if idx < total:
            console.print("[dim]Pausing 3s...[/dim]")
            await asyncio.sleep(3)

    console.print(f"\n[bold cyan]{'='*55}[/bold cyan]")
    console.print(f"  Total: {total} | [green]Success: {success}[/green] | "
                   f"[yellow]Skipped (already done): {skipped}[/yellow] | "
                   f"[red]Failed: {total-success-skipped}[/red]")
    console.print(f"[bold cyan]{'='*55}[/bold cyan]")

    sheet_url = f"https://docs.google.com/spreadsheets/d/{config['google_spreadsheet_id']}"
    console.print(f"\n[bold green]Results -> '{OUTPUT_TAB}' tab[/bold green]")
    console.print(f"  {sheet_url}\n")
    console.print("[bold]Next:[/bold]")
    console.print("  1. Open the sheet, review each row (esp. any with notes in the Notes column)")
    console.print("  2. Set Status = APPROVED for rows you want to list")
    console.print("  3. Run: py -3.11 amazon_listing_generator.py export\n")


# =============================================================================
# SCHEMA GATE -- never fill template fields that are grey/not-applicable
# =============================================================================

# Always written even if the product-type schema omits them (offer/control/
# identity/image fields the schema skips or that the template needs structurally).
_ALWAYS_WRITE_TOKENS = (
    "contribution_sku", "record_action", "product_type", "parent_sku",
    "child_parent_sku_relationship", "variation_theme", "purchasable_offer",
    "list_price", "fulfillment_availability", "merchant_shipping_group",
    "product_tax_code", "image_locator", "product_id", "condition_type",
)


def build_col_attr_map(template_path: str) -> dict:
    """1-based column index -> base attribute key (text before '[' or '#'),
    read from the template's field-ID row (row 5). Used by the schema gate."""
    import openpyxl
    wb = openpyxl.load_workbook(template_path, read_only=True, keep_vba=True)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb[wb.sheetnames[0]]
    out = {}
    for c in range(1, ws.max_column + 1):
        fid = ws.cell(row=5, column=c).value
        if fid:
            out[c] = re.split(r"[\[#]", str(fid))[0].strip().lower()
    wb.close()
    return out


def gate_built_row(built_row: list, col_attr_map: dict, applicable: set) -> int:
    """Clear cells whose attribute is NOT in the product type's schema, so grey/
    not-applicable template fields never get filled. Fail-open: if `applicable`
    is empty (schema fetch failed) nothing is cleared -- we never silently drop
    data, worst case is the old behaviour. Returns count of cells cleared."""
    if not applicable:
        return 0
    cleared = 0
    for c, attr in col_attr_map.items():
        i = c - 1
        if i >= len(built_row) or built_row[i] in (None, ""):
            continue
        if any(tok in attr for tok in _ALWAYS_WRITE_TOKENS):
            continue
        if attr not in applicable:
            built_row[i] = ""
            cleared += 1
    return cleared


def run_export_unified(config: dict, gc, status_filter: str = "APPROVED"):
    """
    UNIFIED-template export. Reuses detect_route() and build_flat_row() from this
    script; only the column map (read dynamically from the template) and the write
    target (a local .xlsm preserving row-1 signature + macros) differ from the
    legacy run_export(). Generation is unaffected.
    """
    console.print(f"\n[bold cyan]{'='*55}[/bold cyan]")
    console.print(f"[bold cyan]  UNIFIED FLAT FILE EXPORT -- Status: {status_filter}[/bold cyan]")
    console.print(f"[bold cyan]{'='*55}[/bold cyan]\n")

    brand        = config.get("brand_name", "")
    manufacturer = config.get("manufacturer", brand)

    template_path = config.get("unified_template_path")
    output_path   = config.get("unified_output_path") or str(CONFIG_PATH.parent / "filled_unified_template.xlsm")
    if not Path(output_path).is_absolute():
        output_path = str(CONFIG_PATH.parent / output_path)

    if not template_path:
        console.print("[red]unified_template_path missing in config.json[/red]")
        console.print("[yellow]Set it to the blank template downloaded from the "
                      "CORRECT Seller Central account.[/yellow]")
        return

    if not Path(template_path).exists():
        console.print(f"[red]Template not found: {template_path}[/red]")
        console.print("[yellow]Copy the new unified template into this folder, "
                      "then set unified_template_path to its exact filename.[/yellow]")
        return

    # Step 1: dynamic column map from the template's field-ID row.
    console.print("[bold]Step 1:[/bold] Mapping template columns by field ID")
    cols_map = unified_export.build_field_map(template_path)
    console.print(f"  Template width: {cols_map['TOTAL_COLS']} columns; "
                  f"{len([k for k in cols_map if k != 'TOTAL_COLS'])} fields located")

    # Step 2: valid values for dropdown snapping (reuse static XLSM values).
    console.print("\n[bold]Step 2:[/bold] Loading valid values")
    static_vv = load_static_valid_values()
    valid_all = merge_static_into_runtime({}, static_vv) if static_vv else {}

    # Step 3: read APPROVED rows from the same output tab as always.
    console.print("\n[bold]Step 3:[/bold] Reading approved listings")
    sh       = gc.open_by_key(config["google_spreadsheet_id"])
    ws       = sh.worksheet(OUTPUT_TAB)
    all_rows = _safe_records(ws)
    rows     = [r for r in all_rows
                if str(r.get("Status", "")).upper().startswith(status_filter.upper())]
    console.print(f"  {len(all_rows)} total -> [bold]{len(rows)}[/bold] '{status_filter}'")
    if not rows:
        console.print(f"[yellow]No '{status_filter}' rows. Set Status=APPROVED in sheet.[/yellow]")
        return

    # Step 4: route + build each row's values, gated by the product-type schema.
    console.print(f"\n[bold]Step 4:[/bold] Routing + building {len(rows)} product(s)")
    creds        = sp_creds(config)
    col_attr_map = build_col_attr_map(template_path)
    schema_cache = {}          # product_type -> set(applicable base attrs)
    mapped       = []          # rows whose exact type wasn't in the template
    built = []
    for row in rows:
        title  = str(row.get("Title",          ""))
        cat    = str(row.get("Amazon Category", ""))
        pt_raw = str(row.get("Product Type",   ""))
        _file_id, prod_type, node = detect_route(title, cat, pt_raw)

        pt_norm = _norm_pt(pt_raw)
        if pt_norm and pt_norm not in TEMPLATE_PRODUCT_TYPES:
            sku = str(row.get("SKU", "") or title[:40])
            mapped.append((sku, pt_raw, prod_type))
            console.print(f"  [yellow]MAP[/yellow] {sku}: '{pt_raw}' -> '{prod_type}' (nearest available)")

        if prod_type not in schema_cache:                    # fetch schema once per type
            sch = get_product_type_schema(prod_type, creds)
            schema_cache[prod_type] = set(sch.get("all", {}).keys())
        applicable = schema_cache[prod_type]

        vv      = valid_all.get(prod_type, {})
        rowvals = build_flat_row(row, brand, manufacturer, cols_map, vv, prod_type, node,
                                 config.get("merchant_shipping_group", ""))
        nclr    = gate_built_row(rowvals, col_attr_map, applicable)
        if nclr:
            console.print(f"  gated {prod_type}: cleared {nclr} non-applicable field(s)")
        built.append(rowvals)

    if mapped:
        console.print(f"\n  [yellow]{len(mapped)} product(s) had no exact template type -> mapped to nearest:[/yellow]")
        for sku, orig, used in mapped:
            console.print(f"     - {sku}: {orig} -> {used}")
        console.print("  [yellow]These will list under a generic/nearest category. Review if discoverability matters.[/yellow]")
    console.print(f"  Built [bold]{len(built)}[/bold] row(s)")

    # Step 5: write the local filled .xlsm (signature + macros preserved).
    console.print("\n[bold]Step 5:[/bold] Writing local filled template")
    out = unified_export.write_local_template(template_path, output_path, built)

    console.print(f"\n[bold green]Export complete![/bold green]")
    console.print(f"  Filled file: [bold]{out}[/bold]")
    console.print("  Next: open it in Excel -> File -> Save As -> "
                  "Text (Tab delimited) (*.txt) -> upload to Seller Central.\n")


if __name__ == "__main__":
    asyncio.run(main())
